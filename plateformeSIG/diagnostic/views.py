import os
import json
import zipfile
import tempfile
from functools import wraps

from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_GET
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.contrib.gis.gdal import DataSource
from django.contrib.gis.geos import GEOSGeometry

from .models import (
    Perimetre, Assolement, TourEau, OrganisationAgriculteur,
    Seuil, MurProtection, Seguias, TronconSeguia, BarrageRetenue,
    Khettara, ForagePuits, PriseLocale,
    EtatSeuil, EtatTronconSeguia, EtatMurProtection, EtatKhettara, EtatForagePuits,
    EtatBarrageRetenue, EtatPriseLocale,
)
from .forms import (
    PerimetreForm, SeuilForm, MurProtectionForm, SeguiasForm, TronconSeguiaForm,
    BarrageRetenueForm, KhettaraForm, ForagePuitsForm, PriseLocaleForm,
    ShpImportForm, ShpImportUnifiedForm,
    EtatSeuilForm, EtatTronconSeguiaForm, EtatMurProtectionForm, EtatKhettaraForm,
    EtatForagePuitsForm, EtatBarrageRetenueForm, EtatPriseLocaleForm,
)


# ── Decorators ───────────────────────────────────────────────────────────────

def editeur_ou_operateur_requis(view_func):
    @wraps(view_func)
    @login_required
    def wrapper(request, *args, **kwargs):
        if request.user.role not in ('editeur', 'operateur', 'administrateur'):
            messages.error(request, "Accès réservé aux éditeurs, opérateurs et administrateurs.")
            return redirect('accueil')
        return view_func(request, *args, **kwargs)
    return wrapper


def operateur_requis(view_func):
    @wraps(view_func)
    @login_required
    def wrapper(request, *args, **kwargs):
        if request.user.role not in ('operateur', 'administrateur'):
            messages.error(request, "Action réservée aux opérateurs et administrateurs.")
            return redirect('accueil')
        return view_func(request, *args, **kwargs)
    return wrapper


# ── SHP helpers ──────────────────────────────────────────────────────────────

def _set_saisie_infos(obj, user):
    if not getattr(obj, 'saisi_par_id', None):
        obj.saisi_par = user


def _toggle_validation(obj, user):
    if obj.statut != 'valide':
        obj.statut = 'valide'
        obj.valide_par = user
    else:
        obj.statut = 'non_valide'
        obj.valide_par = None
    obj.save()


def _extract_shp(zip_file):
    """Extract zip to temp dir, return (tmp_dir, shp_path)."""
    tmp = tempfile.mkdtemp()
    with zipfile.ZipFile(zip_file, 'r') as z:
        z.extractall(tmp)
    for root, _, files in os.walk(tmp):
        for f in files:
            if f.lower().endswith('.shp'):
                return tmp, os.path.join(root, f)
    return tmp, None


def _map_field(shp_name, field_map):
    return field_map.get(shp_name.lower())


def _safe_float(val):
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def _safe_int(val):
    try:
        return int(val)
    except (TypeError, ValueError):
        return None


# Field maps for each model (SHP attr name → model field name)
SEUIL_FIELD_MAP = {
    'nom': 'nom_du_seuil', 'nom_seuil': 'nom_du_seuil', 'name': 'nom_du_seuil',
    'localisa': 'localisation_du_seuil', 'localisation': 'localisation_du_seuil',
    'coord_x': 'coordonnes_x', 'x': 'coordonnes_x',
    'coord_y': 'coordonnes_y', 'y': 'coordonnes_y',
    'nature': 'nature_du_seuil', 'type': 'type_du_seuil',
    'materiaux': 'materiaux_de_construction', 'materiau': 'materiaux_de_construction',
    'debit': 'debit_mobilise', 'debit_mob': 'debit_mobilise',
    'longueur': 'longueur', 'largeur': 'largeur_de_base', 'larg_base': 'largeur_de_base',
    'hauteur': 'hauteur', 'l_tapis': 'largeur_tapis_amortissement',
    'larg_tap': 'largeur_tapis_amortissement',
    'l_prise_d': 'longueur_prise_droit', 'larg_pr_d': 'largeur_prise_droit',
    'nbr_pr_d': 'nbr_pertuis_prise_droit',
    'l_prise_g': 'longueur_prise_gauche', 'larg_pr_g': 'largeur_prise_gauche',
    'nbr_pr_g': 'nbr_pertuis_prise_gauche',
    'l_deg_d': 'longueur_degrevement_droit', 'larg_dg_d': 'largeur_degrevement_droit',
    'nbr_dg_d': 'nbr_pertuis_degrevement_droit',
    'l_deg_g': 'longueur_degrevement_gauche', 'larg_dg_g': 'largeur_degrevement_gauche',
    'nbr_dg_g': 'nbr_pertuis_degrevement_gauche',
    'etat': 'etat_construction_fonctionnement', 'etat_mat': 'etat_materiel_hydromecanique',
    'annee_reh': 'annee_derniere_rehabilitation', 'annee': 'annee_derniere_rehabilitation',
}

MUR_FIELD_MAP = {
    'rive': 'rive', 'position': 'position',
    'nature': 'nature_materiaux', 'materiaux': 'nature_materiaux',
    'longueur': 'longueur', 'hauteur': 'hauteur',
    'ep_sup': 'epaisseur_superieure', 'ep_inf': 'epaisseur_inferieure',
    'etat': 'etat_construction',
}

SEGUIA_FIELD_MAP = {
    'nom': 'nom_de_la_seguia', 'nom_seg': 'nom_de_la_seguia', 'name': 'nom_de_la_seguia',
    'type': 'type_deguia', 'type_seg': 'type_deguia',
}

BARRAGE_FIELD_MAP = {
    'nom': 'nom', 'name': 'nom',
    'coord_x': 'coordonnees_lambert_x', 'lambert_x': 'coordonnees_lambert_x', 'x': 'coordonnees_lambert_x',
    'coord_y': 'coordonnees_lambert_y', 'lambert_y': 'coordonnees_lambert_y', 'y': 'coordonnees_lambert_y',
    'debit': 'debit_derive', 'debit_der': 'debit_derive',
    'vol_irrig': 'volume_attribue_irrigation', 'vol_irr': 'volume_attribue_irrigation',
    'capacite': 'capacite_retenue', 'cap_ret': 'capacite_retenue',
    'longueur': 'longueur', 'largeur': 'largeur', 'hauteur': 'hauteur',
    'materiaux': 'materiaux_de_construction',
    'etat': 'etat_construction_fonctionnement',
}

KHETTARA_FIELD_MAP = {
    'nom': 'nom', 'name': 'nom',
    'coord_x': 'coordonnees_lambert_x', 'x': 'coordonnees_lambert_x',
    'coord_y': 'coordonnees_lambert_y', 'y': 'coordonnees_lambert_y',
    'debit': 'debit', 'longueur': 'longueur', 'largeur': 'largeur', 'hauteur': 'hauteur',
    'materiaux': 'materiaux_de_construction',
    'etat': 'etat_construction_fonctionnement',
}

FORAGE_FIELD_MAP = {
    'nom': 'nom', 'name': 'nom',
    'coord_x': 'coordonnees_lambert_x', 'x': 'coordonnees_lambert_x',
    'coord_y': 'coordonnees_lambert_y', 'y': 'coordonnees_lambert_y',
    'debit': 'debit', 'profond': 'profondeur', 'profondeur': 'profondeur',
    'diametre': 'diametre', 'diam': 'diametre',
    'equipem': 'equipements_associes', 'equip': 'equipements_associes',
    'energie': 'source_energie_pompage', 'src_energ': 'source_energie_pompage',
    'etat': 'etat_construction_fonctionnement',
}

PERIMETRE_FIELD_MAP = {
    'province': 'province', 'coordin': 'coordination', 'coordination': 'coordination',
    'commune': 'commune_territoriale', 'ksar': 'ksar_village',
    'nb_benef': 'nombre_beneficiaires', 'nb_menage': 'nombre_menages',
    'sup_tot': 'superficie_totale', 'sup_sau': 'superficie_agricole_utile',
    'sup_irr': 'superficie_irriguee', 'sup_bour': 'superficie_en_bour',
    'type_sol': 'type_de_sol', 'fertilite': 'niveau_de_fertilite',
    'temp_moy': 'temperature_moyenne_annuelle', 'precip': 'precipitations_moyennes_annuelles',
}

FLOAT_FIELDS = {
    'debit_mobilise', 'longueur', 'largeur_de_base', 'hauteur', 'largeur_tapis_amortissement',
    'longueur_prise_droit', 'largeur_prise_droit', 'longueur_prise_gauche', 'largeur_prise_gauche',
    'longueur_degrevement_droit', 'largeur_degrevement_droit',
    'longueur_degrevement_gauche', 'largeur_degrevement_gauche',
    'nbr_pertuis_prise_droit', 'nbr_pertuis_prise_gauche',
    'nbr_pertuis_degrevement_droit', 'nbr_pertuis_degrevement_gauche',
    'coordonnes_x', 'coordonnes_y',
    'epaisseur_superieure', 'epaisseur_inferieure',
    'largeur_meroire', 'hauteur_eau', 'fruit_de_berge', 'epaisseur_parois', 'debit',
    'coordonnees_lambert_x', 'coordonnees_lambert_y', 'debit_derive',
    'volume_attribue_irrigation', 'capacite_retenue', 'largeur',
    'profondeur', 'diametre',
    'superficie_totale', 'superficie_agricole_utile', 'superficie_irriguee', 'superficie_en_bour',
    'temperature_moyenne_annuelle', 'precipitations_moyennes_annuelles',
}

INT_FIELDS = {'annee_derniere_rehabilitation', 'nombre_beneficiaires', 'nombre_menages'}


def _build_kwargs_from_feature(feature, field_map):
    kwargs = {}
    for shp_field in feature.fields:
        model_field = _map_field(shp_field, field_map)
        if not model_field:
            continue
        val = feature.get(shp_field)
        if model_field in INT_FIELDS:
            v = _safe_int(val)
            if v is not None:
                kwargs[model_field] = v
        elif model_field in FLOAT_FIELDS:
            v = _safe_float(val)
            if v is not None:
                kwargs[model_field] = v
        else:
            if val is not None:
                kwargs[model_field] = str(val)
    return kwargs


def _import_shp_generic(request, model_class, field_map, perimetre, shp_path,
                         required_fields, redirect_url):
    ds = DataSource(shp_path)
    layer = ds[0]
    created = 0
    errors = []
    for feature in layer:
        try:
            geom = GEOSGeometry(feature.geom.wkt, srid=feature.geom.srid or 4326)
            kwargs = _build_kwargs_from_feature(feature, field_map)
            missing = [f for f in required_fields if f not in kwargs]
            if missing:
                errors.append(f"Feature {feature.fid}: champs manquants {missing}")
                continue
            kwargs['geometrie'] = geom
            kwargs['perimetre'] = perimetre
            kwargs['statut'] = 'non_valide'
            kwargs['saisi_par'] = request.user
            model_class.objects.create(**kwargs)
            created += 1
        except Exception as e:
            errors.append(f"Feature {feature.fid}: {e}")

    if created:
        messages.success(request, f"{created} enregistrement(s) importé(s) avec succès.")
    if errors:
        for err in errors[:5]:
            messages.warning(request, err)
    return redirect(redirect_url)


# ── Accueil ──────────────────────────────────────────────────────────────────

@editeur_ou_operateur_requis
def accueil_diagnostic(request):
    stats = {
        'perimetres': Perimetre.objects.count(),
        'seuils': Seuil.objects.count(),
        'murs': MurProtection.objects.count(),
        'barrages': BarrageRetenue.objects.count(),
        'khettaras': Khettara.objects.count(),
        'forages': ForagePuits.objects.count(),
        'prises': PriseLocale.objects.count(),
        'seguias': Seguias.objects.count(),
    }
    return render(request, 'diagnostic/accueil.html', {'stats': stats})


# ── Suivi et évaluation ───────────────────────────────────────────────────────

def _ouvrage_stats(queryset):
    total = queryset.count()
    valides = queryset.filter(statut='valide').count()
    try:
        avec_diagnostic = queryset.filter(diagnostic_etat__isnull=False).count()
    except Exception:
        # Au cas où le modèle n'aurait pas encore de relation diagnostic_etat
        avec_diagnostic = 0
    return {
        'total': total,
        'valides': valides,
        'brouillons': total - valides,
        'avec_diagnostic': avec_diagnostic,
    }


@editeur_ou_operateur_requis
def suivi_evaluation(request):
    seuils_qs = Seuil.objects.all()
    barrages_qs = BarrageRetenue.objects.all()
    khettaras_qs = Khettara.objects.all()
    forages_qs = ForagePuits.objects.all()
    prises_qs = PriseLocale.objects.all()
    troncons_qs = TronconSeguia.objects.all()
    murs_qs = MurProtection.objects.all()

    all_counts = [
        seuils_qs.count(), barrages_qs.count(), khettaras_qs.count(),
        forages_qs.count(), prises_qs.count(), troncons_qs.count(), murs_qs.count(),
    ]
    total_all = sum(all_counts)
    valides_all = (
        seuils_qs.filter(statut='valide').count() +
        barrages_qs.filter(statut='valide').count() +
        khettaras_qs.filter(statut='valide').count() +
        forages_qs.filter(statut='valide').count() +
        prises_qs.filter(statut='valide').count() +
        troncons_qs.filter(statut='valide').count() +
        murs_qs.filter(statut='valide').count()
    )

    return render(request, 'diagnostic/suivi_evaluation.html', {
        'global': {
            'total': total_all,
            'valides': valides_all,
            'brouillons': total_all - valides_all,
            'seuils': seuils_qs.count(),
            'barrages': barrages_qs.count(),
            'khettaras': khettaras_qs.count(),
            'forages': forages_qs.count(),
            'prises': prises_qs.count(),
            'seguias': troncons_qs.count(),
            'murs': murs_qs.count(),
        },
        'stats_seuils': _ouvrage_stats(seuils_qs),
        'stats_barrages': _ouvrage_stats(barrages_qs),
        'stats_khettaras': _ouvrage_stats(khettaras_qs),
        'stats_forages': _ouvrage_stats(forages_qs),
        'stats_prises': _ouvrage_stats(prises_qs),
        'stats_seguias': _ouvrage_stats(troncons_qs),
        'stats_murs': _ouvrage_stats(murs_qs),
    })


@editeur_ou_operateur_requis
def suivi_seuils(request):
    seuils = Seuil.objects.select_related('perimetre', 'diagnostic_etat').order_by('-updated_at')
    diagnostics = EtatSeuil.objects.select_related('seuil__perimetre', 'editeur').order_by('-updated_at')
    return render(request, 'diagnostic/suivi/seuils.html', {
        'seuils': seuils,
        'diagnostics': diagnostics,
        'stats': _ouvrage_stats(seuils),
    })


@editeur_ou_operateur_requis
def suivi_barrages(request):
    barrages = BarrageRetenue.objects.select_related('perimetre', 'diagnostic_etat').order_by('-updated_at')
    diagnostics = EtatBarrageRetenue.objects.select_related('barrage__perimetre', 'editeur_operateur').order_by('-updated_at')
    return render(request, 'diagnostic/suivi/barrages.html', {
        'barrages': barrages,
        'diagnostics': diagnostics,
        'stats': _ouvrage_stats(barrages),
    })


@editeur_ou_operateur_requis
def suivi_khettaras(request):
    khettaras = Khettara.objects.select_related('perimetre', 'diagnostic_etat').order_by('-updated_at')
    diagnostics = EtatKhettara.objects.select_related('khettara__perimetre', 'editeur_operateur').order_by('-updated_at')
    return render(request, 'diagnostic/suivi/khettaras.html', {
        'khettaras': khettaras,
        'diagnostics': diagnostics,
        'stats': _ouvrage_stats(khettaras),
    })


@editeur_ou_operateur_requis
def suivi_forages(request):
    forages = ForagePuits.objects.select_related('perimetre', 'diagnostic_etat').order_by('-updated_at')
    diagnostics = EtatForagePuits.objects.select_related('forage__perimetre', 'editeur_operateur').order_by('-updated_at')
    return render(request, 'diagnostic/suivi/forages.html', {
        'forages': forages,
        'diagnostics': diagnostics,
        'stats': _ouvrage_stats(forages),
    })


@editeur_ou_operateur_requis
def suivi_prises(request):
    prises = PriseLocale.objects.select_related('perimetre', 'diagnostic_etat').order_by('-updated_at')
    diagnostics = EtatPriseLocale.objects.select_related('prise__perimetre', 'editeur_operateur').order_by('-updated_at')
    return render(request, 'diagnostic/suivi/prises.html', {
        'prises': prises,
        'diagnostics': diagnostics,
        'stats': _ouvrage_stats(prises),
    })


@editeur_ou_operateur_requis
def suivi_seguias(request):
    seguias = Seguias.objects.select_related('perimetre').prefetch_related('troncons').order_by('-updated_at')
    troncons = TronconSeguia.objects.select_related('seguia__perimetre').order_by('-updated_at')
    diagnostics = EtatTronconSeguia.objects.select_related('troncon__seguia__perimetre', 'editeur_operateur').order_by('-updated_at')
    total = troncons.count()
    valides = troncons.filter(statut='valide').count()
    return render(request, 'diagnostic/suivi/seguias.html', {
        'seguias': seguias,
        'troncons': troncons,
        'diagnostics': diagnostics,
        'stats': {
            'total': total,
            'valides': valides,
            'brouillons': total - valides,
            'avec_diagnostic': diagnostics.count(),
        },
    })


@editeur_ou_operateur_requis
def suivi_murs(request):
    murs = MurProtection.objects.select_related(
        'perimetre', 'ouvrage_associe', 'diagnostic_etat'
    ).order_by('-updated_at')
    diagnostics = EtatMurProtection.objects.select_related('mur__perimetre', 'editeur_operateur').order_by('-updated_at')
    return render(request, 'diagnostic/suivi/murs.html', {
        'murs': murs,
        'diagnostics': diagnostics,
        'stats': _ouvrage_stats(murs),
    })


# ── Périmètre ────────────────────────────────────────────────────────────────

@editeur_ou_operateur_requis
def perimetre_list(request):
    perimetres = Perimetre.objects.all().order_by('-updated_at')
    return render(request, 'diagnostic/perimetre_list.html', {'perimetres': perimetres})


@login_required
@require_GET
def communes_par_province(request):
    from carte.models import Commune as CarteCommune
    province_nom = request.GET.get('province', '').strip()
    if not province_nom:
        return JsonResponse({'communes': []})
    communes = list(
        CarteCommune.objects.filter(province__nom_fr=province_nom)
        .values_list('nom_fr', flat=True)
        .order_by('nom_fr')
    )
    return JsonResponse({'communes': communes})


@editeur_ou_operateur_requis
def perimetre_create(request):
    form = PerimetreForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        p = form.save(commit=False)
        p.statut = 'non_valide'
        p.save()
        # form.save_m2m() déclenche la création des tables enfants
        # (Assolement, TourEau, OrganisationAgriculteur, OuvrageTeteAssocie)
        # via le hook installé par PerimetreForm.save(commit=False).
        form.save_m2m()
        messages.success(request, "Périmètre créé avec succès.")
        return redirect('diagnostic:perimetre_list')
    return render(request, 'diagnostic/perimetre_form.html',
                  {'form': form, 'titre': 'Ajouter un périmètre'})


@editeur_ou_operateur_requis
def perimetre_edit(request, pk):
    perimetre = get_object_or_404(Perimetre, pk=pk)
    if perimetre.statut == 'valide' and request.user.role == 'editeur':
        messages.error(request, "Ce périmètre est validé. Seul un opérateur peut le modifier.")
        return redirect('diagnostic:perimetre_list')
    form = PerimetreForm(request.POST or None, instance=perimetre)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, "Périmètre mis à jour.")
        return redirect('diagnostic:perimetre_list')
    return render(request, 'diagnostic/perimetre_form.html',
                  {'form': form, 'titre': f'Modifier : {perimetre}', 'perimetre': perimetre})


@operateur_requis
def perimetre_valider(request, pk):
    perimetre = get_object_or_404(Perimetre, pk=pk)
    perimetre.statut = 'valide' if perimetre.statut != 'valide' else 'non_valide'
    perimetre.save()
    label = 'validé' if perimetre.statut == 'valide' else 'remis à non valide'
    messages.success(request, f"Périmètre {label}.")
    return redirect('diagnostic:perimetre_list')


@operateur_requis
def perimetre_delete(request, pk):
    perimetre = get_object_or_404(Perimetre, pk=pk)
    if request.method == 'POST':
        perimetre.delete()
        messages.success(request, "Périmètre supprimé.")
        return redirect('diagnostic:perimetre_list')
    return render(request, 'diagnostic/confirm_delete.html',
                  {'objet': perimetre, 'cancel_url': '/diagnostic/perimetres/'})


@editeur_ou_operateur_requis
def perimetre_shp_import(request):
    form = ShpImportForm(request.POST or None, request.FILES or None)
    if request.method == 'POST' and form.is_valid():
        perimetre = form.cleaned_data['perimetre']
        tmp_dir, shp_path = _extract_shp(request.FILES['fichier_zip'])
        if not shp_path:
            messages.error(request, "Aucun fichier .shp trouvé dans le zip.")
            return redirect('diagnostic:perimetre_shp_import')
        ds = DataSource(shp_path)
        layer = ds[0]
        created = 0
        for feature in layer:
            try:
                geom = GEOSGeometry(feature.geom.wkt, srid=feature.geom.srid or 4326)
                kwargs = _build_kwargs_from_feature(feature, PERIMETRE_FIELD_MAP)
                kwargs['geometrie'] = geom
                kwargs['statut'] = 'non_valide'
                for req in ['nombre_beneficiaires', 'nombre_menages', 'superficie_totale',
                             'superficie_agricole_utile', 'superficie_irriguee',
                             'parcelles_moins_1ha', 'parcelles_1_a_3ha', 'parcelles_plus_3ha']:
                    kwargs.setdefault(req, 0)
                Perimetre.objects.create(**kwargs)
                created += 1
            except Exception as e:
                messages.warning(request, f"Feature {feature.fid}: {e}")
        if created:
            messages.success(request, f"{created} périmètre(s) importé(s).")
        return redirect('diagnostic:perimetre_list')
    return render(request, 'diagnostic/shp_import.html',
                  {'form': form, 'titre': 'Importer des périmètres (.shp)',
                   'cancel_url': 'diagnostic:perimetre_list'})


# ── Ouvrages de tête — detail ────────────────────────────────────────────────

@editeur_ou_operateur_requis
def ouvrages_tete_detail(request, pk):
    perimetre = get_object_or_404(Perimetre, pk=pk)

    seuils    = perimetre.seuils.all()
    murs      = perimetre.murs_protection.all()
    barrages  = perimetre.barrages_retenue.all()
    khettaras = perimetre.khettaras.all()
    forages   = perimetre.forages_puits.all()
    prises    = perimetre.prises_locales.all()
    seguias   = perimetre.seguias.prefetch_related('troncons').all()

    # Build GeoJSON for Leaflet map
    def _etat(obj):
        try:
            return obj.etat.get_etat_general_display()
        except Exception:
            return '—'

    features_ouvrages = []
    for qs, label, name_attr in [
        (seuils,    'Seuil',             'nom_du_seuil'),
        (murs,      'Mur de protection', 'nom_mur_protection'),
        (barrages,  'Barrage',           'nom'),
        (khettaras, 'Khettara',          'nom'),
        (forages,   'Forage/Puits',      'nom'),
        (prises,    'Prise locale',      'nom'),
    ]:
        for o in qs:
            if not o.geometrie:
                continue
            features_ouvrages.append({
                'type': 'Feature',
                'geometry': json.loads(o.geometrie.geojson),
                'properties': {
                    'name': getattr(o, name_attr, '') or '—',
                    'type_label': label,
                    'etat': _etat(o),
                },
            })

    features_troncons = []
    for seg in seguias:
        for tr in seg.troncons.all():
            if not tr.geometrie:
                continue
            features_troncons.append({
                'type': 'Feature',
                'geometry': json.loads(tr.geometrie.geojson),
                'properties': {
                    'name': f'{seg.nom_de_la_seguia} — {tr.troncon}',
                    'type_label': 'Tronçon séguia',
                    'etat': '—',
                },
            })

    context = {
        'perimetre': perimetre,
        'seuils':    seuils,
        'murs':      murs,
        'barrages':  barrages,
        'khettaras': khettaras,
        'forages':   forages,
        'prises':    prises,
        'seguias':   seguias,
        'ouvrages_geojson': json.dumps({'type': 'FeatureCollection', 'features': features_ouvrages}),
        'troncons_geojson': json.dumps({'type': 'FeatureCollection', 'features': features_troncons}),
    }
    return render(request, 'diagnostic/ouvrages_tete/detail.html', context)


# ── Seuil ────────────────────────────────────────────────────────────────────

@editeur_ou_operateur_requis
def seuil_create(request, perimetre_pk):
    perimetre = get_object_or_404(Perimetre, pk=perimetre_pk)
    form = SeuilForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        s = form.save(commit=False)
        s.perimetre = perimetre
        s.statut = 'non_valide'
        _set_saisie_infos(s, request.user)
        s.save()
        messages.success(request, "Seuil ajouté.")
        return redirect('diagnostic:ouvrages_tete_detail', pk=perimetre_pk)
    return render(request, 'diagnostic/ouvrages_tete/seuil_form.html',
                  {'form': form, 'perimetre': perimetre, 'titre': 'Ajouter un seuil'})


@editeur_ou_operateur_requis
def seuil_edit(request, pk):
    seuil = get_object_or_404(Seuil, pk=pk)
    if seuil.statut == 'valide' and request.user.role == 'editeur':
        messages.error(request, "Ce seuil est validé. Seul un opérateur peut le modifier.")
        return redirect('diagnostic:ouvrages_tete_detail', pk=seuil.perimetre_id)
    form = SeuilForm(request.POST or None, instance=seuil)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, "Seuil mis à jour.")
        return redirect('diagnostic:ouvrages_tete_detail', pk=seuil.perimetre_id)
    return render(request, 'diagnostic/ouvrages_tete/seuil_form.html',
                  {'form': form, 'perimetre': seuil.perimetre, 'titre': f'Modifier : {seuil}'})


@operateur_requis
def seuil_valider(request, pk):
    seuil = get_object_or_404(Seuil, pk=pk)
    _toggle_validation(seuil, request.user)
    messages.success(request, f"Seuil {'validé' if seuil.statut == 'valide' else 'remis à non valide'}.")
    return redirect('diagnostic:ouvrages_tete_detail', pk=seuil.perimetre_id)


@operateur_requis
def seuil_delete(request, pk):
    seuil = get_object_or_404(Seuil, pk=pk)
    perimetre_pk = seuil.perimetre_id
    if request.method == 'POST':
        seuil.delete()
        messages.success(request, "Seuil supprimé.")
        return redirect('diagnostic:ouvrages_tete_detail', pk=perimetre_pk)
    return render(request, 'diagnostic/confirm_delete.html',
                  {'objet': seuil,
                   'cancel_url': f"/diagnostic/ouvrages-tete/{perimetre_pk}/"})


@editeur_ou_operateur_requis
def seuil_diagnostic(request, pk):
    """Vue diagnostic d'un seuil : lecture seule des données d'ouvrage + saisie EtatSeuil."""
    seuil = get_object_or_404(Seuil, pk=pk)
    etat, _ = EtatSeuil.objects.get_or_create(seuil=seuil)
    form = EtatSeuilForm(request.POST or None, instance=etat)
    if request.method == 'POST' and form.is_valid():
        etat = form.save(commit=False)
        etat.editeur = request.user
        etat.save()
        seuil.statut = 'non_valide'
        seuil.save(update_fields=['statut'])
        messages.success(request, "Diagnostic du seuil enregistré.")
        return redirect('diagnostic:seuil_diagnostic', pk=pk)
    return render(request, 'diagnostic/ouvrages_tete/seuil_diagnostic.html', {
        'seuil': seuil,
        'etat': etat,
        'form': form,
    })


@editeur_ou_operateur_requis
def seuil_shp_import(request):
    form = ShpImportForm(request.POST or None, request.FILES or None)
    if request.method == 'POST' and form.is_valid():
        perimetre = form.cleaned_data['perimetre']
        tmp_dir, shp_path = _extract_shp(request.FILES['fichier_zip'])
        if not shp_path:
            messages.error(request, "Aucun fichier .shp trouvé dans le zip.")
        else:
            return _import_shp_generic(
                request, Seuil, SEUIL_FIELD_MAP, perimetre, shp_path,
                required_fields=['nom_du_seuil'],
                redirect_url=f"/diagnostic/ouvrages-tete/{perimetre.pk}/",
            )
    return render(request, 'diagnostic/shp_import.html',
                  {'form': form, 'titre': 'Importer des seuils (.shp)',
                   'cancel_url': 'diagnostic:perimetre_list'})


# ── Mur de protection ────────────────────────────────────────────────────────

@editeur_ou_operateur_requis
def mur_create(request, perimetre_pk):
    perimetre = get_object_or_404(Perimetre, pk=perimetre_pk)
    form = MurProtectionForm(request.POST or None)
    form.fields['ouvrage_associe'].queryset = perimetre.seuils.all()
    if request.method == 'POST' and form.is_valid():
        m = form.save(commit=False)
        m.perimetre = perimetre
        m.statut = 'non_valide'
        _set_saisie_infos(m, request.user)
        m.save()
        messages.success(request, "Mur de protection ajouté.")
        return redirect('diagnostic:ouvrages_tete_detail', pk=perimetre_pk)
    return render(request, 'diagnostic/ouvrages_tete/mur_form.html',
                  {'form': form, 'perimetre': perimetre, 'titre': 'Ajouter un mur de protection'})


@editeur_ou_operateur_requis
def mur_edit(request, pk):
    mur = get_object_or_404(MurProtection, pk=pk)
    if mur.statut == 'valide' and request.user.role == 'editeur':
        messages.error(request, "Ce mur est validé. Seul un opérateur peut le modifier.")
        return redirect('diagnostic:ouvrages_tete_detail', pk=mur.perimetre_id)
    form = MurProtectionForm(request.POST or None, instance=mur)
    form.fields['ouvrage_associe'].queryset = mur.perimetre.seuils.all()
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, "Mur de protection mis à jour.")
        return redirect('diagnostic:ouvrages_tete_detail', pk=mur.perimetre_id)
    return render(request, 'diagnostic/ouvrages_tete/mur_form.html',
                  {'form': form, 'perimetre': mur.perimetre, 'titre': f'Modifier : {mur}'})


@operateur_requis
def mur_valider(request, pk):
    mur = get_object_or_404(MurProtection, pk=pk)
    _toggle_validation(mur, request.user)
    messages.success(request, f"Mur {'validé' if mur.statut == 'valide' else 'remis à non valide'}.")
    return redirect('diagnostic:ouvrages_tete_detail', pk=mur.perimetre_id)


@operateur_requis
def mur_delete(request, pk):
    mur = get_object_or_404(MurProtection, pk=pk)
    perimetre_pk = mur.perimetre_id
    if request.method == 'POST':
        mur.delete()
        messages.success(request, "Mur supprimé.")
        return redirect('diagnostic:ouvrages_tete_detail', pk=perimetre_pk)
    return render(request, 'diagnostic/confirm_delete.html',
                  {'objet': mur, 'cancel_url': f"/diagnostic/ouvrages-tete/{perimetre_pk}/"})


@editeur_ou_operateur_requis
def mur_diagnostic(request, pk):
    """Vue diagnostic d'un mur de protection : lecture seule + saisie EtatMurProtection."""
    mur = get_object_or_404(MurProtection, pk=pk)
    etat, _ = EtatMurProtection.objects.get_or_create(mur=mur)
    form = EtatMurProtectionForm(request.POST or None, instance=etat)
    if request.method == 'POST' and form.is_valid():
        etat = form.save(commit=False)
        etat.editeur_operateur = request.user
        etat.save()
        mur.statut = 'non_valide'
        mur.save(update_fields=['statut'])
        messages.success(request, "Diagnostic du mur enregistré.")
        return redirect('diagnostic:mur_diagnostic', pk=pk)
    return render(request, 'diagnostic/ouvrages_tete/mur_diagnostic.html', {
        'mur': mur,
        'etat': etat,
        'form': form,
    })


@editeur_ou_operateur_requis
def mur_shp_import(request):
    form = ShpImportForm(request.POST or None, request.FILES or None)
    if request.method == 'POST' and form.is_valid():
        perimetre = form.cleaned_data['perimetre']
        tmp_dir, shp_path = _extract_shp(request.FILES['fichier_zip'])
        if not shp_path:
            messages.error(request, "Aucun fichier .shp trouvé dans le zip.")
        else:
            return _import_shp_generic(
                request, MurProtection, MUR_FIELD_MAP, perimetre, shp_path,
                required_fields=['rive', 'position'],
                redirect_url=f"/diagnostic/ouvrages-tete/{perimetre.pk}/",
            )
    return render(request, 'diagnostic/shp_import.html',
                  {'form': form, 'titre': 'Importer des murs de protection (.shp)',
                   'cancel_url': 'diagnostic:perimetre_list'})


# ── Barrage de retenue ───────────────────────────────────────────────────────

@editeur_ou_operateur_requis
def barrage_create(request, perimetre_pk):
    perimetre = get_object_or_404(Perimetre, pk=perimetre_pk)
    form = BarrageRetenueForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        b = form.save(commit=False)
        b.perimetre = perimetre
        b.statut = 'non_valide'
        _set_saisie_infos(b, request.user)
        b.save()
        messages.success(request, "Barrage de retenue ajouté.")
        return redirect('diagnostic:ouvrages_tete_detail', pk=perimetre_pk)
    return render(request, 'diagnostic/ouvrages_tete/barrage_form.html',
                  {'form': form, 'perimetre': perimetre, 'titre': 'Ajouter un barrage de retenue'})


@editeur_ou_operateur_requis
def barrage_edit(request, pk):
    barrage = get_object_or_404(BarrageRetenue, pk=pk)
    if barrage.statut == 'valide' and request.user.role == 'editeur':
        messages.error(request, "Ce barrage est validé. Seul un opérateur peut le modifier.")
        return redirect('diagnostic:ouvrages_tete_detail', pk=barrage.perimetre_id)
    form = BarrageRetenueForm(request.POST or None, instance=barrage)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, "Barrage mis à jour.")
        return redirect('diagnostic:ouvrages_tete_detail', pk=barrage.perimetre_id)
    return render(request, 'diagnostic/ouvrages_tete/barrage_form.html',
                  {'form': form, 'perimetre': barrage.perimetre, 'titre': f'Modifier : {barrage}'})


@operateur_requis
def barrage_valider(request, pk):
    barrage = get_object_or_404(BarrageRetenue, pk=pk)
    _toggle_validation(barrage, request.user)
    messages.success(request, f"Barrage {'validé' if barrage.statut == 'valide' else 'remis à non valide'}.")
    return redirect('diagnostic:ouvrages_tete_detail', pk=barrage.perimetre_id)


@operateur_requis
def barrage_delete(request, pk):
    barrage = get_object_or_404(BarrageRetenue, pk=pk)
    perimetre_pk = barrage.perimetre_id
    if request.method == 'POST':
        barrage.delete()
        messages.success(request, "Barrage supprimé.")
        return redirect('diagnostic:ouvrages_tete_detail', pk=perimetre_pk)
    return render(request, 'diagnostic/confirm_delete.html',
                  {'objet': barrage, 'cancel_url': f"/diagnostic/ouvrages-tete/{perimetre_pk}/"})


@editeur_ou_operateur_requis
def barrage_diagnostic(request, pk):
    """Vue diagnostic d'un barrage de retenue : lecture seule + saisie EtatBarrageRetenue."""
    barrage = get_object_or_404(BarrageRetenue, pk=pk)
    etat, _ = EtatBarrageRetenue.objects.get_or_create(barrage=barrage)
    form = EtatBarrageRetenueForm(request.POST or None, instance=etat)
    if request.method == 'POST' and form.is_valid():
        etat = form.save(commit=False)
        etat.editeur_operateur = request.user
        etat.save()
        barrage.statut = 'non_valide'
        barrage.save(update_fields=['statut'])
        messages.success(request, "Diagnostic du barrage enregistré.")
        return redirect('diagnostic:barrage_diagnostic', pk=pk)
    return render(request, 'diagnostic/ouvrages_tete/barrage_diagnostic.html', {
        'barrage': barrage,
        'etat': etat,
        'form': form,
    })


@editeur_ou_operateur_requis
def barrage_shp_import(request):
    form = ShpImportForm(request.POST or None, request.FILES or None)
    if request.method == 'POST' and form.is_valid():
        perimetre = form.cleaned_data['perimetre']
        tmp_dir, shp_path = _extract_shp(request.FILES['fichier_zip'])
        if not shp_path:
            messages.error(request, "Aucun fichier .shp trouvé dans le zip.")
        else:
            return _import_shp_generic(
                request, BarrageRetenue, BARRAGE_FIELD_MAP, perimetre, shp_path,
                required_fields=['nom'],
                redirect_url=f"/diagnostic/ouvrages-tete/{perimetre.pk}/",
            )
    return render(request, 'diagnostic/shp_import.html',
                  {'form': form, 'titre': 'Importer des barrages de retenue (.shp)',
                   'cancel_url': 'diagnostic:perimetre_list'})


# ── Khettara ──────────────────────────────────────────────────────────────────

@editeur_ou_operateur_requis
def khettara_create(request, perimetre_pk):
    perimetre = get_object_or_404(Perimetre, pk=perimetre_pk)
    form = KhettaraForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        k = form.save(commit=False)
        k.perimetre = perimetre
        k.statut = 'non_valide'
        _set_saisie_infos(k, request.user)
        k.save()
        messages.success(request, "Khettara ajoutée.")
        return redirect('diagnostic:ouvrages_tete_detail', pk=perimetre_pk)
    return render(request, 'diagnostic/ouvrages_tete/khettara_form.html',
                  {'form': form, 'perimetre': perimetre, 'titre': 'Ajouter une khettara'})


@editeur_ou_operateur_requis
def khettara_edit(request, pk):
    khettara = get_object_or_404(Khettara, pk=pk)
    if khettara.statut == 'valide' and request.user.role == 'editeur':
        messages.error(request, "Cette khettara est validée. Seul un opérateur peut la modifier.")
        return redirect('diagnostic:ouvrages_tete_detail', pk=khettara.perimetre_id)
    form = KhettaraForm(request.POST or None, instance=khettara)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, "Khettara mise à jour.")
        return redirect('diagnostic:ouvrages_tete_detail', pk=khettara.perimetre_id)
    return render(request, 'diagnostic/ouvrages_tete/khettara_form.html',
                  {'form': form, 'perimetre': khettara.perimetre, 'titre': f'Modifier : {khettara}'})


@operateur_requis
def khettara_valider(request, pk):
    khettara = get_object_or_404(Khettara, pk=pk)
    _toggle_validation(khettara, request.user)
    messages.success(request, f"Khettara {'validée' if khettara.statut == 'valide' else 'remise à non valide'}.")
    return redirect('diagnostic:ouvrages_tete_detail', pk=khettara.perimetre_id)


@operateur_requis
def khettara_delete(request, pk):
    khettara = get_object_or_404(Khettara, pk=pk)
    perimetre_pk = khettara.perimetre_id
    if request.method == 'POST':
        khettara.delete()
        messages.success(request, "Khettara supprimée.")
        return redirect('diagnostic:ouvrages_tete_detail', pk=perimetre_pk)
    return render(request, 'diagnostic/confirm_delete.html',
                  {'objet': khettara, 'cancel_url': f"/diagnostic/ouvrages-tete/{perimetre_pk}/"})


@editeur_ou_operateur_requis
def khettara_diagnostic(request, pk):
    """Vue diagnostic d'une khettara : lecture seule + saisie EtatKhettara."""
    khettara = get_object_or_404(Khettara, pk=pk)
    etat, _ = EtatKhettara.objects.get_or_create(khettara=khettara)
    form = EtatKhettaraForm(request.POST or None, instance=etat)
    if request.method == 'POST' and form.is_valid():
        etat = form.save(commit=False)
        etat.editeur_operateur = request.user
        etat.save()
        khettara.statut = 'non_valide'
        khettara.save(update_fields=['statut'])
        messages.success(request, "Diagnostic de la khettara enregistré.")
        return redirect('diagnostic:khettara_diagnostic', pk=pk)
    return render(request, 'diagnostic/ouvrages_tete/khettara_diagnostic.html', {
        'khettara': khettara,
        'etat': etat,
        'form': form,
    })


@editeur_ou_operateur_requis
def khettara_shp_import(request):
    form = ShpImportForm(request.POST or None, request.FILES or None)
    if request.method == 'POST' and form.is_valid():
        perimetre = form.cleaned_data['perimetre']
        tmp_dir, shp_path = _extract_shp(request.FILES['fichier_zip'])
        if not shp_path:
            messages.error(request, "Aucun fichier .shp trouvé dans le zip.")
        else:
            return _import_shp_generic(
                request, Khettara, KHETTARA_FIELD_MAP, perimetre, shp_path,
                required_fields=['nom'],
                redirect_url=f"/diagnostic/ouvrages-tete/{perimetre.pk}/",
            )
    return render(request, 'diagnostic/shp_import.html',
                  {'form': form, 'titre': 'Importer des khettaras (.shp)',
                   'cancel_url': 'diagnostic:perimetre_list'})


# ── Forage / Puits ───────────────────────────────────────────────────────────

@editeur_ou_operateur_requis
def forage_create(request, perimetre_pk):
    perimetre = get_object_or_404(Perimetre, pk=perimetre_pk)
    form = ForagePuitsForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        f = form.save(commit=False)
        f.perimetre = perimetre
        f.statut = 'non_valide'
        _set_saisie_infos(f, request.user)
        f.save()
        messages.success(request, "Forage/Puits ajouté.")
        return redirect('diagnostic:ouvrages_tete_detail', pk=perimetre_pk)
    return render(request, 'diagnostic/ouvrages_tete/forage_form.html',
                  {'form': form, 'perimetre': perimetre, 'titre': 'Ajouter un forage / puits'})


@editeur_ou_operateur_requis
def forage_edit(request, pk):
    forage = get_object_or_404(ForagePuits, pk=pk)
    if forage.statut == 'valide' and request.user.role == 'editeur':
        messages.error(request, "Ce forage est validé. Seul un opérateur peut le modifier.")
        return redirect('diagnostic:ouvrages_tete_detail', pk=forage.perimetre_id)
    form = ForagePuitsForm(request.POST or None, instance=forage)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, "Forage/Puits mis à jour.")
        return redirect('diagnostic:ouvrages_tete_detail', pk=forage.perimetre_id)
    return render(request, 'diagnostic/ouvrages_tete/forage_form.html',
                  {'form': form, 'perimetre': forage.perimetre, 'titre': f'Modifier : {forage}'})


@operateur_requis
def forage_valider(request, pk):
    forage = get_object_or_404(ForagePuits, pk=pk)
    _toggle_validation(forage, request.user)
    messages.success(request, f"Forage {'validé' if forage.statut == 'valide' else 'remis à non valide'}.")
    return redirect('diagnostic:ouvrages_tete_detail', pk=forage.perimetre_id)


@operateur_requis
def forage_delete(request, pk):
    forage = get_object_or_404(ForagePuits, pk=pk)
    perimetre_pk = forage.perimetre_id
    if request.method == 'POST':
        forage.delete()
        messages.success(request, "Forage/Puits supprimé.")
        return redirect('diagnostic:ouvrages_tete_detail', pk=perimetre_pk)
    return render(request, 'diagnostic/confirm_delete.html',
                  {'objet': forage, 'cancel_url': f"/diagnostic/ouvrages-tete/{perimetre_pk}/"})


@editeur_ou_operateur_requis
def forage_diagnostic(request, pk):
    """Vue diagnostic d'un forage/puits : lecture seule + saisie EtatForagePuits."""
    forage = get_object_or_404(ForagePuits, pk=pk)
    etat, _ = EtatForagePuits.objects.get_or_create(forage=forage)
    form = EtatForagePuitsForm(request.POST or None, instance=etat)
    if request.method == 'POST' and form.is_valid():
        etat = form.save(commit=False)
        etat.editeur_operateur = request.user
        etat.save()
        forage.statut = 'non_valide'
        forage.save(update_fields=['statut'])
        messages.success(request, "Diagnostic du forage/puits enregistré.")
        return redirect('diagnostic:forage_diagnostic', pk=pk)
    return render(request, 'diagnostic/ouvrages_tete/forage_diagnostic.html', {
        'forage': forage,
        'etat': etat,
        'form': form,
    })


@editeur_ou_operateur_requis
def forage_shp_import(request):
    form = ShpImportForm(request.POST or None, request.FILES or None)
    if request.method == 'POST' and form.is_valid():
        perimetre = form.cleaned_data['perimetre']
        tmp_dir, shp_path = _extract_shp(request.FILES['fichier_zip'])
        if not shp_path:
            messages.error(request, "Aucun fichier .shp trouvé dans le zip.")
        else:
            return _import_shp_generic(
                request, ForagePuits, FORAGE_FIELD_MAP, perimetre, shp_path,
                required_fields=['nom'],
                redirect_url=f"/diagnostic/ouvrages-tete/{perimetre.pk}/",
            )
    return render(request, 'diagnostic/shp_import.html',
                  {'form': form, 'titre': 'Importer des forages/puits (.shp)',
                   'cancel_url': 'diagnostic:perimetre_list'})


# ── Prise locale ─────────────────────────────────────────────────────────────

PRISE_FIELD_MAP = {
    'nom': 'nom', 'name': 'nom',
    'forme': 'forme_pertuis', 'forme_per': 'forme_pertuis',
    'larg_mir': 'largeur_au_miroir', 'larg_mir': 'largeur_au_miroir',
    'haut_per': 'hauteur_pertuis', 'hauteur': 'hauteur_pertuis',
    'fruit': 'fruit_pente', 'fruit_pen': 'fruit_pente',
    'diametre': 'diametre', 'diam': 'diametre',
    'etat': 'etat_fonctionnement',
}

FLOAT_FIELDS.update({'largeur_au_miroir', 'hauteur_pertuis', 'fruit_pente'})


@editeur_ou_operateur_requis
def prise_create(request, perimetre_pk):
    perimetre = get_object_or_404(Perimetre, pk=perimetre_pk)
    form = PriseLocaleForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        p = form.save(commit=False)
        p.perimetre = perimetre
        p.statut = 'non_valide'
        _set_saisie_infos(p, request.user)
        p.save()
        messages.success(request, "Prise locale ajoutée.")
        return redirect('diagnostic:ouvrages_tete_detail', pk=perimetre_pk)
    return render(request, 'diagnostic/ouvrages_tete/prise_form.html',
                  {'form': form, 'perimetre': perimetre, 'titre': 'Ajouter une prise locale'})


@editeur_ou_operateur_requis
def prise_edit(request, pk):
    prise = get_object_or_404(PriseLocale, pk=pk)
    if prise.statut == 'valide' and request.user.role == 'editeur':
        messages.error(request, "Cette prise est validée. Seul un opérateur peut la modifier.")
        return redirect('diagnostic:ouvrages_tete_detail', pk=prise.perimetre_id)
    form = PriseLocaleForm(request.POST or None, instance=prise)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, "Prise locale mise à jour.")
        return redirect('diagnostic:ouvrages_tete_detail', pk=prise.perimetre_id)
    return render(request, 'diagnostic/ouvrages_tete/prise_form.html',
                  {'form': form, 'perimetre': prise.perimetre, 'titre': f'Modifier : {prise}'})


@operateur_requis
def prise_valider(request, pk):
    prise = get_object_or_404(PriseLocale, pk=pk)
    _toggle_validation(prise, request.user)
    messages.success(request, f"Prise locale {'validée' if prise.statut == 'valide' else 'remise à non valide'}.")
    return redirect('diagnostic:ouvrages_tete_detail', pk=prise.perimetre_id)


@operateur_requis
def prise_delete(request, pk):
    prise = get_object_or_404(PriseLocale, pk=pk)
    perimetre_pk = prise.perimetre_id
    if request.method == 'POST':
        prise.delete()
        messages.success(request, "Prise locale supprimée.")
        return redirect('diagnostic:ouvrages_tete_detail', pk=perimetre_pk)
    return render(request, 'diagnostic/confirm_delete.html',
                  {'objet': prise, 'cancel_url': f"/diagnostic/ouvrages-tete/{perimetre_pk}/"})


@editeur_ou_operateur_requis
def prise_diagnostic(request, pk):
    """Vue diagnostic d'une prise locale : lecture seule + saisie EtatPriseLocale."""
    prise = get_object_or_404(PriseLocale, pk=pk)
    etat, _ = EtatPriseLocale.objects.get_or_create(prise=prise)
    form = EtatPriseLocaleForm(request.POST or None, instance=etat)
    if request.method == 'POST' and form.is_valid():
        etat = form.save(commit=False)
        etat.editeur_operateur = request.user
        etat.save()
        prise.statut = 'non_valide'
        prise.save(update_fields=['statut'])
        messages.success(request, "Diagnostic de la prise locale enregistré.")
        return redirect('diagnostic:prise_diagnostic', pk=pk)
    return render(request, 'diagnostic/ouvrages_tete/prise_diagnostic.html', {
        'prise': prise,
        'etat': etat,
        'form': form,
    })


@editeur_ou_operateur_requis
def prise_shp_import(request):
    form = ShpImportForm(request.POST or None, request.FILES or None)
    if request.method == 'POST' and form.is_valid():
        perimetre = form.cleaned_data['perimetre']
        tmp_dir, shp_path = _extract_shp(request.FILES['fichier_zip'])
        if not shp_path:
            messages.error(request, "Aucun fichier .shp trouvé dans le zip.")
        else:
            return _import_shp_generic(
                request, PriseLocale, PRISE_FIELD_MAP, perimetre, shp_path,
                required_fields=['nom'],
                redirect_url=f"/diagnostic/ouvrages-tete/{perimetre.pk}/",
            )
    return render(request, 'diagnostic/shp_import.html',
                  {'form': form, 'titre': 'Importer des prises locales (.shp)',
                   'cancel_url': 'diagnostic:perimetre_list'})


# ── Réseaux d'irrigation — Séguias ───────────────────────────────────────────

@editeur_ou_operateur_requis
def reseaux_irrigation_detail(request, pk):
    perimetre = get_object_or_404(Perimetre, pk=pk)
    seguias = perimetre.seguias.prefetch_related('troncons').all()

    features = []
    for seg in seguias:
        for tr in seg.troncons.all():
            if not tr.geometrie:
                continue
            features.append({
                'type': 'Feature',
                'geometry': json.loads(tr.geometrie.geojson),
                'properties': {
                    'name': f'{seg.nom_de_la_seguia} — {tr.troncon}',
                    'seguia': seg.nom_de_la_seguia,
                    'longueur': tr.longueur,
                    'forme': tr.get_forme_display(),
                },
            })

    return render(request, 'diagnostic/reseaux_irrigation/detail.html', {
        'perimetre': perimetre,
        'seguias': seguias,
        'troncons_geojson': json.dumps({'type': 'FeatureCollection', 'features': features}),
    })


@editeur_ou_operateur_requis
def seguia_create(request, perimetre_pk):
    perimetre = get_object_or_404(Perimetre, pk=perimetre_pk)
    form = SeguiasForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        s = form.save(commit=False)
        s.perimetre = perimetre
        _set_saisie_infos(s, request.user)
        s.save()
        messages.success(request, "Séguia ajoutée. Ajoutez maintenant ses tronçons.")
        return redirect('diagnostic:reseaux_irrigation_detail', pk=perimetre_pk)
    return render(request, 'diagnostic/reseaux_irrigation/seguia_form.html',
                  {'form': form, 'perimetre': perimetre, 'titre': 'Ajouter une séguia'})


@editeur_ou_operateur_requis
def seguia_edit(request, pk):
    seguia = get_object_or_404(Seguias, pk=pk)
    form = SeguiasForm(request.POST or None, instance=seguia)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, "Séguia mise à jour.")
        return redirect('diagnostic:reseaux_irrigation_detail', pk=seguia.perimetre_id)
    return render(request, 'diagnostic/reseaux_irrigation/seguia_form.html',
                  {'form': form, 'perimetre': seguia.perimetre, 'titre': f'Modifier : {seguia}'})


@operateur_requis
def seguia_valider(request, pk):
    """Conservé pour compatibilité URL — redirige vers la liste du périmètre."""
    seguia = get_object_or_404(Seguias, pk=pk)
    messages.info(request, "La validation se fait maintenant au niveau de chaque tronçon.")
    return redirect('diagnostic:reseaux_irrigation_detail', pk=seguia.perimetre_id)


@operateur_requis
def seguia_delete(request, pk):
    seguia = get_object_or_404(Seguias, pk=pk)
    perimetre_pk = seguia.perimetre_id
    if request.method == 'POST':
        seguia.delete()
        messages.success(request, "Séguia supprimée.")
        return redirect('diagnostic:reseaux_irrigation_detail', pk=perimetre_pk)
    return render(request, 'diagnostic/confirm_delete.html',
                  {'objet': seguia, 'cancel_url': f"/diagnostic/reseaux-irrigation/{perimetre_pk}/"})


# ── Tronçons de séguia ────────────────────────────────────────────────────────

@editeur_ou_operateur_requis
def troncon_create(request, seguia_pk):
    seguia = get_object_or_404(Seguias, pk=seguia_pk)
    form = TronconSeguiaForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        t = form.save(commit=False)
        t.seguia = seguia
        t.statut = 'non_valide'
        t.save()
        messages.success(request, f"Tronçon {t.troncon} ajouté à {seguia}.")
        return redirect('diagnostic:reseaux_irrigation_detail', pk=seguia.perimetre_id)
    return render(request, 'diagnostic/reseaux_irrigation/troncon_form.html', {
        'form': form, 'seguia': seguia,
        'perimetre': seguia.perimetre,
        'titre': f'Ajouter un tronçon — {seguia}',
    })


@editeur_ou_operateur_requis
def troncon_edit(request, pk):
    troncon = get_object_or_404(TronconSeguia, pk=pk)
    seguia = troncon.seguia
    if troncon.statut == 'valide' and request.user.role == 'editeur':
        messages.error(request, "Ce tronçon est validé. Seul un opérateur peut le modifier.")
        return redirect('diagnostic:reseaux_irrigation_detail', pk=seguia.perimetre_id)
    form = TronconSeguiaForm(request.POST or None, instance=troncon)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, "Tronçon mis à jour.")
        return redirect('diagnostic:reseaux_irrigation_detail', pk=seguia.perimetre_id)
    return render(request, 'diagnostic/reseaux_irrigation/troncon_form.html', {
        'form': form, 'seguia': seguia,
        'perimetre': seguia.perimetre,
        'titre': f'Modifier : {troncon}',
    })


@operateur_requis
def troncon_delete(request, pk):
    troncon = get_object_or_404(TronconSeguia, pk=pk)
    perimetre_pk = troncon.seguia.perimetre_id
    if request.method == 'POST':
        troncon.delete()
        messages.success(request, "Tronçon supprimé.")
        return redirect('diagnostic:reseaux_irrigation_detail', pk=perimetre_pk)
    return render(request, 'diagnostic/confirm_delete.html', {
        'objet': troncon,
        'cancel_url': f"/diagnostic/reseaux-irrigation/{perimetre_pk}/",
    })


@operateur_requis
def troncon_valider(request, pk):
    troncon = get_object_or_404(TronconSeguia, pk=pk)
    _toggle_validation(troncon, request.user)
    label = 'validé' if troncon.statut == 'valide' else 'remis à non valide'
    messages.success(request, f"Tronçon {troncon.troncon} {label}.")
    return redirect('diagnostic:reseaux_irrigation_detail', pk=troncon.seguia.perimetre_id)


@editeur_ou_operateur_requis
def troncon_diagnostic(request, pk):
    """Diagnostic structuré d'un tronçon de séguia."""
    troncon = get_object_or_404(TronconSeguia, pk=pk)
    etat, _ = EtatTronconSeguia.objects.get_or_create(troncon=troncon)
    form = EtatTronconSeguiaForm(request.POST or None, instance=etat)
    if request.method == 'POST' and form.is_valid():
        etat = form.save(commit=False)
        etat.editeur_operateur = request.user
        etat.save()
        troncon.statut = 'non_valide'
        troncon.save(update_fields=['statut'])
        messages.success(request, "Diagnostic du tronçon enregistré.")
        return redirect('diagnostic:troncon_diagnostic', pk=pk)
    return render(request, 'diagnostic/reseaux_irrigation/troncon_diagnostic.html', {
        'troncon': troncon,
        'seguia': troncon.seguia,
        'etat': etat,
        'form': form,
    })


@editeur_ou_operateur_requis
def seguia_shp_import(request):
    form = ShpImportForm(request.POST or None, request.FILES or None)
    if request.method == 'POST' and form.is_valid():
        perimetre = form.cleaned_data['perimetre']
        tmp_dir, shp_path = _extract_shp(request.FILES['fichier_zip'])
        if not shp_path:
            messages.error(request, "Aucun fichier .shp trouvé dans le zip.")
        else:
            return _import_shp_generic(
                request, Seguias, SEGUIA_FIELD_MAP, perimetre, shp_path,
                required_fields=['nom_de_la_seguia'],
                redirect_url=f"/diagnostic/reseaux-irrigation/{perimetre.pk}/",
            )
    return render(request, 'diagnostic/shp_import.html',
                  {'form': form, 'titre': 'Importer des séguias (.shp)',
                   'cancel_url': 'diagnostic:perimetre_list'})


# ── Import unifié SHP — Wizard (étape 1 : upload, étape 2 : mapping) ─────────

# Champs disponibles par type (valeurs uniques des FIELD_MAP)
_WIZARD_FIELDS = {
    'seuil':    sorted(set(SEUIL_FIELD_MAP.values())),
    'mur':      sorted(set(MUR_FIELD_MAP.values())),
    'barrage':  sorted(set(BARRAGE_FIELD_MAP.values())),
    'khettara': sorted(set(KHETTARA_FIELD_MAP.values())),
    'forage':   sorted(set(FORAGE_FIELD_MAP.values())),
    'prise':    sorted(set(PRISE_FIELD_MAP.values())),
    'seguia':   sorted(set(SEGUIA_FIELD_MAP.values())),
    'perimetre':sorted(set(PERIMETRE_FIELD_MAP.values())),
}

_WIZARD_FIELD_MAPS = {
    'seuil': SEUIL_FIELD_MAP, 'mur': MUR_FIELD_MAP, 'barrage': BARRAGE_FIELD_MAP,
    'khettara': KHETTARA_FIELD_MAP, 'forage': FORAGE_FIELD_MAP,
    'prise': PRISE_FIELD_MAP, 'seguia': SEGUIA_FIELD_MAP, 'perimetre': PERIMETRE_FIELD_MAP,
}

_WIZARD_TYPE_LABELS = {
    'seuil': 'Seuils hydrauliques', 'mur': 'Murs de protection',
    'barrage': 'Barrages de retenue', 'khettara': 'Khettaras',
    'forage': 'Forages / Puits', 'prise': 'Prises locales',
    'seguia': 'Séguias', 'perimetre': 'Périmètres agricoles',
}


@editeur_ou_operateur_requis
def shp_import_unified(request):
    """Étape 1 : choix du type, périmètre et upload du .zip."""
    form = ShpImportUnifiedForm(request.POST or None, request.FILES or None)
    if request.method == 'POST' and form.is_valid():
        type_donnee = form.cleaned_data['type_donnee']
        perimetre = form.cleaned_data.get('perimetre')
        tmp_dir, shp_path = _extract_shp(request.FILES['fichier_zip'])
        if not shp_path:
            messages.error(request, "Aucun fichier .shp trouvé dans le zip.")
            return redirect('diagnostic:shp_import_unified')

        # Lire les colonnes du SHP
        try:
            ds = DataSource(shp_path)
            layer = ds[0]
            shp_columns = list(layer.fields)
            feature_count = len(layer)
        except Exception as e:
            messages.error(request, f"Erreur de lecture du shapefile : {e}")
            return redirect('diagnostic:shp_import_unified')

        # Stocker en session pour l'étape 2
        request.session['shp_wizard'] = {
            'type_donnee': type_donnee,
            'perimetre_id': perimetre.pk if perimetre else None,
            'shp_path': shp_path,
            'feature_count': feature_count,
        }
        request.session['shp_columns'] = shp_columns
        return redirect('diagnostic:shp_import_mapping')

    return render(request, 'diagnostic/shp_import_unified.html', {
        'form': form,
        'titre': 'Importer un shapefile — Étape 1 / 2',
    })


@editeur_ou_operateur_requis
def shp_import_mapping(request):
    """Étape 2 : mapping des colonnes SHP → champs modèle, puis import."""
    wizard = request.session.get('shp_wizard')
    if not wizard:
        messages.error(request, "Session expirée. Recommencez l'import.")
        return redirect('diagnostic:shp_import_unified')

    type_donnee = wizard['type_donnee']
    perimetre_id = wizard.get('perimetre_id')
    shp_path = wizard['shp_path']
    feature_count = wizard['feature_count']
    shp_columns = request.session.get('shp_columns', [])

    field_map = _WIZARD_FIELD_MAPS.get(type_donnee, {})
    available_fields = _WIZARD_FIELDS.get(type_donnee, [])
    # Suggestion automatique : col.lower() dans field_map → model field
    auto_mapping = {col: field_map.get(col.lower(), '') for col in shp_columns}

    if request.method == 'POST':
        # Construire le mapping personnalisé (clé = col en minuscules pour _build_kwargs_from_feature)
        custom_map = {}
        for col in shp_columns:
            field = request.POST.get(f'map_{col}', '').strip()
            if field:
                custom_map[col.lower()] = field

        # Récupérer le périmètre
        perimetre = Perimetre.objects.get(pk=perimetre_id) if perimetre_id else None

        # Nettoyer la session
        request.session.pop('shp_wizard', None)
        request.session.pop('shp_columns', None)

        if type_donnee == 'perimetre':
            ds = DataSource(shp_path)
            layer = ds[0]
            created = 0
            for feature in layer:
                try:
                    kwargs = _build_kwargs_from_feature(feature, custom_map)
                    kwargs['statut'] = 'non_valide'
                    for req in ['nombre_beneficiaires', 'nombre_menages', 'superficie_totale',
                                'superficie_agricole_utile', 'superficie_irriguee',
                                'parcelles_moins_1ha', 'parcelles_1_a_3ha', 'parcelles_plus_3ha']:
                        kwargs.setdefault(req, 0)
                    Perimetre.objects.create(**kwargs)
                    created += 1
                except Exception as e:
                    messages.warning(request, f"Feature {feature.fid}: {e}")
            if created:
                messages.success(request, f"{created} périmètre(s) importé(s).")
            return redirect('diagnostic:perimetre_list')

        dispatch = {
            'seuil':    (Seuil,          ['nom_du_seuil'], lambda p: f"/diagnostic/ouvrages-tete/{p.pk}/"),
            'mur':      (MurProtection,  [],               lambda p: f"/diagnostic/ouvrages-tete/{p.pk}/"),
            'barrage':  (BarrageRetenue, ['nom'],          lambda p: f"/diagnostic/ouvrages-tete/{p.pk}/"),
            'khettara': (Khettara,       ['nom'],          lambda p: f"/diagnostic/ouvrages-tete/{p.pk}/"),
            'forage':   (ForagePuits,    ['nom'],          lambda p: f"/diagnostic/ouvrages-tete/{p.pk}/"),
            'prise':    (PriseLocale,    ['nom'],          lambda p: f"/diagnostic/ouvrages-tete/{p.pk}/"),
            'seguia':   (Seguias,        ['nom_de_la_seguia', 'type_deguia'],
                         lambda p: f"/diagnostic/reseaux-irrigation/{p.pk}/"),
        }
        if type_donnee not in dispatch:
            messages.error(request, "Type de données inconnu.")
            return redirect('diagnostic:shp_import_unified')

        model_class, required_fields, get_url = dispatch[type_donnee]
        redirect_url = get_url(perimetre)
        return _import_shp_generic(request, model_class, custom_map, perimetre, shp_path,
                                   required_fields=required_fields, redirect_url=redirect_url)

    perimetre_label = ''
    if perimetre_id:
        try:
            p = Perimetre.objects.get(pk=perimetre_id)
            perimetre_label = str(p)
        except Perimetre.DoesNotExist:
            pass

    # Pre-process columns for template (avoids custom templatetag)
    columns_data = [
        {'col': col, 'auto_field': auto_mapping.get(col, '')}
        for col in shp_columns
    ]

    return render(request, 'diagnostic/shp_import_mapping.html', {
        'titre': 'Importer un shapefile — Étape 2 / 2',
        'type_donnee': type_donnee,
        'type_label': _WIZARD_TYPE_LABELS.get(type_donnee, type_donnee),
        'perimetre_label': perimetre_label,
        'columns_data': columns_data,
        'available_fields': available_fields,
        'feature_count': feature_count,
        'auto_mapping_json': json.dumps(auto_mapping),
    })


# ── Export Excel — Fiche enquête vide (template de saisie terrain) ────────────

@login_required
def export_enquete_global(request):
    """Génère un classeur Excel VIDE (aucune donnée DB) servant de fiche terrain.
    12 onglets : Périmètres + 3 tables enfants + 7 types d'ouvrages.
    Exclu : champs calculés, champs système, géométries."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    from datetime import date

    wb = Workbook()
    HDR_FILL  = PatternFill("solid", fgColor="1A1A2E")
    HDR_FONT  = Font(bold=True, color="FFFFFF")
    HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def ws_new(title):
        ws = wb.create_sheet(title)
        ws.sheet_view.showGridLines = True
        return ws

    def write_headers(ws, headers):
        ws.row_dimensions[1].height = 30
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = HDR_ALIGN

    def autowidth(ws):
        for col_cells in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col_cells), default=8)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 4, 42)

    # ── 01 — Périmètres ──────────────────────────────────────────────────────
    ws01 = wb.active
    ws01.title = "Périmètres"
    write_headers(ws01, [
        "Province", "Coordination", "Commune territoriale", "Ksar / Village",
        "Température moy. annuelle (°C)", "Précipitations moy. annuelles (mm)",
        "Vent", "Humidité",
        "Nombre de bénéficiaires", "Nombre de ménages",
        "Superficie totale (ha)", "Superficie agricole utile (ha)",
        "Superficie irriguée (ha)", "Superficie en bour (ha)",
        "Type de sol", "Niveau de fertilité",
        "% parcelles < 1 ha", "% parcelles 1-3 ha", "% parcelles > 3 ha",
        "% Melk", "% Collectif", "% Location", "% Guich", "% Habous",
        "Moyenne bovins (têtes/agric.)", "Moyenne ovins", "Moyenne caprins",
    ])
    autowidth(ws01)

    # ── 02 — Assolement (table enfant de Périmètre) ───────────────────────────
    ws02 = ws_new("Assolement")
    write_headers(ws02, [
        "Ksar / Village (périmètre)", "Culture",
        "% de la surface cultivée", "Surface (ha)",
        "Rendement", "Unité rendement (qx/ha ou kg/arbre)",
    ])
    autowidth(ws02)

    # ── 03 — Tours d'eau (table enfant de Périmètre) ─────────────────────────
    ws03 = ws_new("Tours d'eau")
    write_headers(ws03, [
        "Ksar / Village (périmètre)",
        "Ayant droit / Famille",
        "Cycle du tour d'eau (jours)",
        "Durée (heures)",
    ])
    autowidth(ws03)

    # ── 04 — Organisations (table enfant de Périmètre) ────────────────────────
    ws04 = ws_new("Organisations")
    write_headers(ws04, [
        "Ksar / Village (périmètre)",
        "Nom de l'organisation d'agriculteurs",
    ])
    autowidth(ws04)

    # ── 05 — Seuils ──────────────────────────────────────────────────────────
    ws05 = ws_new("Seuils")
    write_headers(ws05, [
        "Périmètre (Ksar / Village)",
        "Nom du seuil", "Localisation", "Coordonnée X (Nord Maroc, m)", "Coordonnée Y (Nord Maroc, m)",
        "Nature", "Type", "Matériaux de construction",
        "Débit mobilisé (l/s)", "Longueur (m)", "Largeur de base (m)",
        "Hauteur (m)", "Largeur tapis d'amortissement (m)",
        "Prise D — Longueur (m)", "Prise D — Largeur (m)", "Prise D — Nb pertuis",
        "Prise G — Longueur (m)", "Prise G — Largeur (m)", "Prise G — Nb pertuis",
        "Dégr. D — Longueur (m)", "Dégr. D — Largeur (m)", "Dégr. D — Nb pertuis",
        "Dégr. G — Longueur (m)", "Dégr. G — Largeur (m)", "Dégr. G — Nb pertuis",
        "Année dernière réhabilitation",
        "Date diagnostic", "Défauts observés", "Saisi par",
        # EtatSeuil
        "Diag. — État construction / fonctionnement",
        "Diag. — État matériel hydromécanique",
        "Diag. — État structurel digue (0-5)",
        "Diag. — Affouillement à l'aval (0-5)",
        "Diag. — Envasement de la retenue (0-5)",
        "Diag. — Murs guideaux (0-5)",
        "Diag. — Radier aval (0-5)",
        "Diag. — État des vannes (0-5)",
        "Diag. — Dessableur (0-5)",
        "Diag. — Dégradation du béton (0-5)",
        "Diag. — Infiltration / fuite (0-5)",
        "Diag. — Limiteur de débit (0-5)",
    ])
    autowidth(ws05)

    # ── 06 — Barrages ────────────────────────────────────────────────────────
    ws06 = ws_new("Barrages")
    write_headers(ws06, [
        "Périmètre (Ksar / Village)",
        "Nom", "Coordonnée X (m)", "Coordonnée Y (m)",
        "Débit dérivé (m³/s)", "Volume attribué irrigation (m³)", "Capacité retenue (m³)",
        "Longueur (m)", "Largeur (m)", "Hauteur (m)", "Matériaux de construction",
        "Bassin versant associé",
        "Date diagnostic", "Défauts observés", "Saisi par",
        "Diag. — État général",
        "Diag. — Affouillement au pied de digue aval (0-5)",
        "Diag. — Taux d'envasement de la retenue (0-5)",
        "Diag. — Régulation des débits aval (0-5)",
        "Diag. — Fonctionnement ouvrages prise d'eau (0-5)",
    ])
    autowidth(ws06)

    # ── 07 — Khettaras ───────────────────────────────────────────────────────
    ws07 = ws_new("Khettaras")
    write_headers(ws07, [
        "Périmètre (Ksar / Village)",
        "Nom", "Coordonnée X (m)", "Coordonnée Y (m)",
        "Débit (m³/s)", "Longueur (m)", "Largeur (m)", "Hauteur (m)",
        "Matériaux de construction",
        "Date diagnostic", "Défauts observés", "Saisi par",
        "Diag. — État général",
        "Diag. — Envasement / ensablement du fond (0-5)",
        "Diag. — Dégradation du béton (0-5)",
        "Diag. — Accessibilité pour l'entretien (0-5)",
        "Diag. — Stabilité de la galerie principale (0-5)",
    ])
    autowidth(ws07)

    # ── 08 — Forages / Puits ─────────────────────────────────────────────────
    ws08 = ws_new("Forages")
    write_headers(ws08, [
        "Périmètre (Ksar / Village)",
        "Nom", "Coordonnée X (m)", "Coordonnée Y (m)",
        "Débit (m³/h)", "Profondeur (m)", "Diamètre (m)",
        "Équipements associés", "Source d'énergie pompage",
        "Date diagnostic", "Défauts observés", "Saisi par",
        "Diag. — État général",
        "Diag. — Qualité physico-chimique de l'eau (0-5)",
        "Diag. — Dégradation structurelle du forage (0-5)",
        "Diag. — Colmatage du forage (0-5)",
        "Diag. — État des équipements (0-5)",
    ])
    autowidth(ws08)

    # ── 09 — Murs de protection ──────────────────────────────────────────────
    ws09 = ws_new("Murs")
    write_headers(ws09, [
        "Périmètre (Ksar / Village)",
        "Nom", "Rive (droite / gauche)", "Position (amont / aval)",
        "Nature des matériaux", "Longueur (m)", "Hauteur (m)",
        "Épaisseur supérieure (m)", "Épaisseur inférieure (m)",
        "Seuil associé (nom)",
        "Date diagnostic", "Défauts observés", "Saisi par",
        "Diag. — État général",
        "Diag. — Fissures du revêtement (0-5)",
        "Diag. — Dégradation du béton (0-5)",
        "Diag. — Risque de contournement (0-5)",
    ])
    autowidth(ws09)

    # ── 10 — Prises locales ──────────────────────────────────────────────────
    ws10 = ws_new("Prises locales")
    write_headers(ws10, [
        "Périmètre (Ksar / Village)",
        "Nom", "Coordonnée X (m)", "Coordonnée Y (m)",
        "Matériaux de construction", "Forme du pertuis",
        "Largeur au miroir (m)", "Hauteur du pertuis (m)",
        "Fruit / pente", "Diamètre (m)",
        "Débit dérivé (m³/s)", "Bassin versant associé",
        "Date diagnostic", "Défauts observés", "Saisi par",
        "Diag. — État général",
        "Diag. — Envasement / sédimentation à l'entrée (0-5)",
        "Diag. — Dégradation du revêtement (0-5)",
        "Diag. — Accumulation débris / végétation (0-5)",
        "Diag. — État dispositifs régulation (vannes, masques) (0-5)",
        "Diag. — Protection contre crues / débordements (0-5)",
    ])
    autowidth(ws10)

    # ── 11 — Séguias ─────────────────────────────────────────────────────────
    ws11 = ws_new("Séguias")
    write_headers(ws11, [
        "Périmètre (Ksar / Village)",
        "Nom de la séguia", "Type (principale / secondaire / tertiaire)",
        "Date diagnostic", "Défauts observés", "Saisi par",
    ])
    autowidth(ws11)

    # ── 12 — Tronçons séguias ────────────────────────────────────────────────
    ws12 = ws_new("Tronçons séguias")
    write_headers(ws12, [
        "Nom de la séguia (référence)",
        "Tronçon (TR1-TR20)", "Forme", "Nature",
        "Longueur (m)", "Largeur miroir (m)", "Hauteur (m)",
        "Hauteur d'eau (m)", "Fruit de berge", "Épaisseur parois (m)",
        "Diamètre (m)", "Débit (m³/s)", "Type d'écoulement",
        "Date diagnostic", "Défauts observés",
        "Diag. — État général",
        "Diag. — Fissures du revêtement (0-5)",
        "Diag. — Infiltration / fuite (0-5)",
        "Diag. — Obstructions / débris (0-5)",
        "Diag. — Érosion des berges (0-5)",
        "Diag. — Sédimentation au fond (0-5)",
        "Diag. — Ouvrages de régulation (0-5)",
        "Diag. — Spalling du béton (0-5)",
    ])
    autowidth(ws12)

    # ── Réponse HTTP ─────────────────────────────────────────────────────────
    filename = f"fiche_enquete_diagnostic_{date.today().strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ── Export Excel — Données d'un périmètre ─────────────────────────────────────

@login_required
def perimetre_export_excel(request, pk):
    """Classeur Excel avec toutes les données d'un seul périmètre (réel, avec données DB)."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    from datetime import date as date_cls

    p = get_object_or_404(Perimetre, pk=pk)
    wb = Workbook()

    HDR_FILL   = PatternFill("solid", fgColor="1A1A2E")
    HDR_FONT   = Font(bold=True, color="FFFFFF")
    HDR_ALIGN  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    SEC_FILL   = PatternFill("solid", fgColor="2C3E6B")
    SEC_FONT   = Font(bold=True, color="FFFFFF", size=10)
    LABEL_FONT = Font(bold=True, color="1A1A2E")

    def write_headers(ws, headers, row=1):
        ws.row_dimensions[row].height = 28
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = HDR_ALIGN

    def write_section(ws, row, title):
        """En-tête de section (fond bleu foncé secondaire)."""
        cell = ws.cell(row=row, column=1, value=title)
        cell.fill = SEC_FILL
        cell.font = SEC_FONT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.row_dimensions[row].height = 22

    def write_kv(ws, row, label, value):
        lbl = ws.cell(row=row, column=1, value=label)
        lbl.font = LABEL_FONT
        ws.cell(row=row, column=2, value=value if value is not None else '')

    def autowidth(ws):
        for col_cells in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col_cells), default=8)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 4, 42)

    def v(val):
        return val if val is not None else ''

    def dt(val):
        return val.strftime('%Y-%m-%d') if val else ''

    nom_perim = str(p)

    # ── Onglet 1 : Périmètre (format clé-valeur) ──────────────────────────────
    ws1 = wb.active
    ws1.title = "Périmètre"
    ws1.column_dimensions['A'].width = 36
    ws1.column_dimensions['B'].width = 32

    # Identité
    write_section(ws1, 1, "IDENTITÉ DU PÉRIMÈTRE")
    kv = [
        ("Province",                     v(p.province)),
        ("Coordination",                 v(p.coordination)),
        ("Commune territoriale",          v(p.commune_territoriale)),
        ("Ksar / Village",               v(p.ksar_village)),
    ]
    for i, (lbl, val) in enumerate(kv, 2):
        write_kv(ws1, i, lbl, val)

    # Climatologie
    r = 7
    write_section(ws1, r, "DONNÉES CLIMATIQUES")
    clim = [
        ("Température moyenne annuelle (°C)",   v(p.temperature_moyenne_annuelle)),
        ("Précipitations moyennes annuelles (mm)", v(p.precipitations_moyennes_annuelles)),
        ("Vent",                                v(p.vent)),
        ("Humidité",                            v(p.humidite)),
    ]
    for i, (lbl, val) in enumerate(clim, r + 1):
        write_kv(ws1, i, lbl, val)

    # Démographie & superficies
    r = 13
    write_section(ws1, r, "DÉMOGRAPHIE ET SUPERFICIES")
    demo = [
        ("Nombre de bénéficiaires",          p.nombre_beneficiaires),
        ("Nombre de ménages",                p.nombre_menages),
        ("Superficie totale (ha)",           p.superficie_totale),
        ("Superficie agricole utile (ha)",   p.superficie_agricole_utile),
        ("Superficie irriguée (ha)",         p.superficie_irriguee),
        ("Superficie en bour (ha)",          p.superficie_en_bour),
    ]
    for i, (lbl, val) in enumerate(demo, r + 1):
        write_kv(ws1, i, lbl, val)

    # Pédologie & parcellaire
    r = 21
    write_section(ws1, r, "PÉDOLOGIE ET PARCELLAIRE")
    pedo = [
        ("Type de sol",                      v(p.type_de_sol)),
        ("Niveau de fertilité",              v(p.niveau_de_fertilite)),
        ("% parcelles < 1 ha",               v(p.parcelles_moins_1ha)),
        ("% parcelles 1-3 ha",               v(p.parcelles_1_a_3ha)),
        ("% parcelles > 3 ha",               v(p.parcelles_plus_3ha)),
    ]
    for i, (lbl, val) in enumerate(pedo, r + 1):
        write_kv(ws1, i, lbl, val)

    # Statut juridique
    r = 28
    write_section(ws1, r, "STATUT JURIDIQUE DES TERRES (%)")
    jur = [
        ("% Melk",       v(p.statut_juridique_melk)),
        ("% Collectif",  v(p.statut_juridique_collectif)),
        ("% Location",   v(p.statut_juridique_location)),
        ("% Guich",      v(p.statut_juridique_guiche)),
        ("% Habous",     v(p.statut_juridique_habousse)),
    ]
    for i, (lbl, val) in enumerate(jur, r + 1):
        write_kv(ws1, i, lbl, val)

    # Cheptel
    r = 35
    write_section(ws1, r, "CHEPTEL (têtes/agriculteur)")
    chep = [
        ("Bovins",   v(p.moyenne_bovins)),
        ("Ovins",    v(p.moyenne_ovins)),
        ("Caprins",  v(p.moyenne_caprins)),
    ]
    for i, (lbl, val) in enumerate(chep, r + 1):
        write_kv(ws1, i, lbl, val)

    # ── Onglet 2 : Assolement ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Assolement")
    write_headers(ws2, ["Culture", "% surface cultivée", "Surface (ha)", "Rendement", "Unité rendement"])
    for a in p.assolement.all().order_by('ordre'):
        ws2.append([v(a.culture), v(a.pourcentage), v(a.surface_ha), v(a.rendement), v(a.unite_rendement)])
    autowidth(ws2)

    # ── Onglet 3 : Tours d'eau ────────────────────────────────────────────────
    ws3 = wb.create_sheet("Tours d'eau")
    write_headers(ws3, ["Ayant droit / Famille", "Cycle (jours)", "Durée (heures)"])
    for t in p.tours_eau.all().order_by('ordre'):
        ws3.append([v(t.ayant_droit), v(t.cycle_jours), v(t.duree_heures)])
    autowidth(ws3)

    # ── Onglet 4 : Organisations ──────────────────────────────────────────────
    ws4 = wb.create_sheet("Organisations")
    write_headers(ws4, ["Nom de l'organisation d'agriculteurs"])
    for org in p.organisations.all().order_by('ordre'):
        ws4.append([org.nom])
    autowidth(ws4)

    # Helper : pré-chargement des états pour ce périmètre
    seuils   = list(p.seuils.all().order_by('id'))
    barrages = list(p.barrages_retenue.all().order_by('id'))
    khettaras= list(p.khettaras.all().order_by('id'))
    forages  = list(p.forages_puits.all().order_by('id'))
    murs     = list(p.murs_protection.all().order_by('id'))
    prises   = list(p.prises_locales.all().order_by('id'))
    seguias  = list(p.seguias.prefetch_related('troncons').order_by('id'))

    etat_s  = {e.seuil_id:    e for e in EtatSeuil.objects.filter(seuil__in=seuils)}
    etat_b  = {e.barrage_id:  e for e in EtatBarrageRetenue.objects.filter(barrage__in=barrages)}
    etat_k  = {e.khettara_id: e for e in EtatKhettara.objects.filter(khettara__in=khettaras)}
    etat_f  = {e.forage_id:   e for e in EtatForagePuits.objects.filter(forage__in=forages)}
    etat_m  = {e.mur_id:      e for e in EtatMurProtection.objects.filter(mur__in=murs)}
    etat_p  = {e.prise_id:    e for e in EtatPriseLocale.objects.filter(prise__in=prises)}
    all_tr  = [tr for seg in seguias for tr in seg.troncons.all()]
    etat_tr = {e.troncon_id:  e for e in EtatTronconSeguia.objects.filter(troncon__in=all_tr)}

    # ── Onglet 5 : Seuils ────────────────────────────────────────────────────
    ws5 = wb.create_sheet("Seuils")
    write_headers(ws5, [
        "Nom", "Localisation", "Coord X (m)", "Coord Y (m)",
        "Nature", "Type", "Matériaux",
        "Débit mobilisé (l/s)", "Longueur (m)", "Largeur base (m)", "Hauteur (m)", "Largeur tapis (m)",
        "Prise D — L", "Prise D — l", "Prise D — nb",
        "Prise G — L", "Prise G — l", "Prise G — nb",
        "Dégr. D — L", "Dégr. D — l", "Dégr. D — nb",
        "Dégr. G — L", "Dégr. G — l", "Dégr. G — nb",
        "Année réhabilitation", "Date diagnostic", "Défauts", "Saisi par",
        "Diag. — État construction", "Diag. — État matériel hydromécanique",
        "Diag. — Structurel digue (0-5)", "Diag. — Affouillement aval (0-5)",
        "Diag. — Envasement retenue (0-5)", "Diag. — Murs guideaux (0-5)",
        "Diag. — Radier aval (0-5)", "Diag. — Vannes (0-5)",
        "Diag. — Dessableur (0-5)", "Diag. — Dégradation béton (0-5)",
        "Diag. — Infiltration/fuite (0-5)", "Diag. — Limiteur débit (0-5)",
    ])
    for s in seuils:
        e = etat_s.get(s.id)
        ws5.append([
            s.nom_du_seuil, v(s.localisation_du_seuil), v(s.coordonnes_x), v(s.coordonnes_y),
            s.nature_du_seuil, s.type_du_seuil, s.materiaux_de_construction,
            s.debit_mobilise, s.longueur, s.largeur_de_base, s.hauteur, s.largeur_tapis_amortissement,
            v(s.longueur_prise_droit), v(s.largeur_prise_droit), v(s.nbr_pertuis_prise_droit),
            v(s.longueur_prise_gauche), v(s.largeur_prise_gauche), v(s.nbr_pertuis_prise_gauche),
            v(s.longueur_degrevement_droit), v(s.largeur_degrevement_droit), v(s.nbr_pertuis_degrevement_droit),
            v(s.longueur_degrevement_gauche), v(s.largeur_degrevement_gauche), v(s.nbr_pertuis_degrevement_gauche),
            v(s.annee_derniere_rehabilitation), dt(s.date_diagnostic), v(s.defaut_ouvrage),
            str(s.saisi_par) if s.saisi_par else '',
            v(getattr(e, 'etat_construction_fonctionnement', '')),
            v(getattr(e, 'etat_materiel_hydromecanique', '')),
            v(getattr(e, 'etat_structurel_digue', '')),
            v(getattr(e, 'affouillement_aval', '')),
            v(getattr(e, 'envasement_retenue', '')),
            v(getattr(e, 'murs_guideaux', '')),
            v(getattr(e, 'radier_aval', '')),
            v(getattr(e, 'etat_vannes', '')),
            v(getattr(e, 'dessableur', '')),
            v(getattr(e, 'degradation_beton', '')),
            v(getattr(e, 'infiltration_fuite', '')),
            v(getattr(e, 'limiteur_debit', '')),
        ])
    autowidth(ws5)

    # ── Onglet 6 : Barrages ──────────────────────────────────────────────────
    ws6 = wb.create_sheet("Barrages")
    write_headers(ws6, [
        "Nom", "Coord X (m)", "Coord Y (m)",
        "Débit dérivé (m³/s)", "Volume attribué (m³)", "Capacité retenue (m³)",
        "Longueur (m)", "Largeur (m)", "Hauteur (m)", "Matériaux",
        "Date diagnostic", "Défauts", "Saisi par",
        "Diag. — État général",
        "Diag. — Affouillement pied digue (0-5)", "Diag. — Envasement retenue (0-5)",
        "Diag. — Régulation débits aval (0-5)", "Diag. — Fonctionnement prises (0-5)",
    ])
    for b in barrages:
        e = etat_b.get(b.id)
        ws6.append([
            b.nom, v(b.coordonnees_lambert_x), v(b.coordonnees_lambert_y),
            b.debit_derive, b.volume_attribue_irrigation, b.capacite_retenue,
            b.longueur, b.largeur, b.hauteur, b.materiaux_de_construction,
            dt(b.date_diagnostic), v(b.defaut_ouvrage),
            str(b.saisi_par) if b.saisi_par else '',
            v(getattr(e, 'etat_general', '')),
            v(getattr(e, 'affouillement_pied_digue_aval', '')),
            v(getattr(e, 'taux_envasement_retenue', '')),
            v(getattr(e, 'regulation_debits_aval', '')),
            v(getattr(e, 'fonctionnement_ouvrages_prise_eau', '')),
        ])
    autowidth(ws6)

    # ── Onglet 7 : Khettaras ─────────────────────────────────────────────────
    ws7 = wb.create_sheet("Khettaras")
    write_headers(ws7, [
        "Nom", "Coord X (m)", "Coord Y (m)",
        "Débit (m³/s)", "Longueur (m)", "Largeur (m)", "Hauteur (m)", "Matériaux",
        "Date diagnostic", "Défauts", "Saisi par",
        "Diag. — État général",
        "Diag. — Envasement fond (0-5)", "Diag. — Dégradation béton (0-5)",
        "Diag. — Accessibilité entretien (0-5)", "Diag. — Stabilité galerie (0-5)",
    ])
    for k in khettaras:
        e = etat_k.get(k.id)
        ws7.append([
            k.nom, v(k.coordonnees_lambert_x), v(k.coordonnees_lambert_y),
            k.debit, k.longueur, k.largeur, k.hauteur, k.materiaux_de_construction,
            dt(k.date_diagnostic), v(k.defaut_ouvrage),
            str(k.saisi_par) if k.saisi_par else '',
            v(getattr(e, 'etat_general', '')),
            v(getattr(e, 'envasement_ensablement_fond', '')),
            v(getattr(e, 'degradation_beton', '')),
            v(getattr(e, 'accessibilite_entretien', '')),
            v(getattr(e, 'stabilite_galerie_principale', '')),
        ])
    autowidth(ws7)

    # ── Onglet 8 : Forages ────────────────────────────────────────────────────
    ws8 = wb.create_sheet("Forages")
    write_headers(ws8, [
        "Nom", "Coord X (m)", "Coord Y (m)",
        "Débit (m³/h)", "Profondeur (m)", "Diamètre (m)",
        "Équipements", "Source énergie",
        "Date diagnostic", "Défauts", "Saisi par",
        "Diag. — État général",
        "Diag. — Qualité eau (0-5)", "Diag. — Dégradation structurelle (0-5)",
        "Diag. — Colmatage (0-5)", "Diag. — État équipements (0-5)",
    ])
    for f in forages:
        e = etat_f.get(f.id)
        ws8.append([
            f.nom, v(f.coordonnees_lambert_x), v(f.coordonnees_lambert_y),
            f.debit, f.profondeur, f.diametre,
            v(f.equipements_associes), v(f.source_energie_pompage),
            dt(f.date_diagnostic), v(f.defaut_ouvrage),
            str(f.saisi_par) if f.saisi_par else '',
            v(getattr(e, 'etat_general', '')),
            v(getattr(e, 'qualite_physico_chimique_eau', '')),
            v(getattr(e, 'degradation_structurelle_forage', '')),
            v(getattr(e, 'colmatage_forage', '')),
            v(getattr(e, 'etat_equipements', '')),
        ])
    autowidth(ws8)

    # ── Onglet 9 : Murs ──────────────────────────────────────────────────────
    ws9 = wb.create_sheet("Murs")
    write_headers(ws9, [
        "Nom", "Rive", "Position", "Matériaux",
        "Longueur (m)", "Hauteur (m)", "Épaisseur sup. (m)", "Épaisseur inf. (m)",
        "Seuil associé",
        "Date diagnostic", "Défauts", "Saisi par",
        "Diag. — État général",
        "Diag. — Fissures revêtement (0-5)", "Diag. — Dégradation béton (0-5)",
        "Diag. — Risque contournement (0-5)",
    ])
    for m in murs:
        e = etat_m.get(m.id)
        ws9.append([
            v(m.nom_mur_protection), v(m.rive), v(m.position), m.nature_materiaux,
            m.longueur, m.hauteur, m.epaisseur_superieure, m.epaisseur_inferieure,
            str(m.ouvrage_associe) if m.ouvrage_associe else '',
            dt(m.date_diagnostic), v(m.defaut_ouvrage),
            str(m.saisi_par) if m.saisi_par else '',
            v(getattr(e, 'etat_general', '')),
            v(getattr(e, 'fissures_revetement', '')),
            v(getattr(e, 'degradation_beton', '')),
            v(getattr(e, 'risque_contournement', '')),
        ])
    autowidth(ws9)

    # ── Onglet 10 : Prises locales ────────────────────────────────────────────
    ws10 = wb.create_sheet("Prises")
    write_headers(ws10, [
        "Nom", "Coord X (m)", "Coord Y (m)",
        "Matériaux", "Forme pertuis",
        "Largeur miroir (m)", "Hauteur pertuis (m)", "Fruit pente", "Diamètre (m)",
        "Débit dérivé (m³/s)",
        "Date diagnostic", "Défauts", "Saisi par",
        "Diag. — État général",
        "Diag. — Envasement entrée (0-5)", "Diag. — Dégradation revêtement (0-5)",
        "Diag. — Débris / végétation (0-5)", "Diag. — Dispositifs régulation (0-5)",
        "Diag. — Protection crues (0-5)",
    ])
    for pr in prises:
        e = etat_p.get(pr.id)
        ws10.append([
            pr.nom, v(pr.coordonnee_x), v(pr.coordonnee_y),
            v(pr.materiaux_construction), v(pr.forme_pertuis),
            v(pr.largeur_au_miroir), v(pr.hauteur_pertuis), v(pr.fruit_pente), v(pr.diametre),
            v(pr.debit_derive),
            dt(pr.date_diagnostic), v(pr.defaut_ouvrage),
            str(pr.saisi_par) if pr.saisi_par else '',
            v(getattr(e, 'etat_general', '')),
            v(getattr(e, 'envasement_sedimentation_entree', '')),
            v(getattr(e, 'degradation_revetement', '')),
            v(getattr(e, 'accumulation_debris_vegetation', '')),
            v(getattr(e, 'etat_dispositifs_regulation', '')),
            v(getattr(e, 'protection_crues_debordements', '')),
        ])
    autowidth(ws10)

    # ── Onglet 11 : Séguias + Tronçons ───────────────────────────────────────
    ws11 = wb.create_sheet("Séguias")
    write_headers(ws11, [
        "Nom séguia", "Type",
        "Tronçon", "Forme", "Nature",
        "Longueur (m)", "Largeur miroir (m)", "Hauteur (m)",
        "Hauteur eau (m)", "Fruit berge", "Épaisseur parois (m)",
        "Diamètre (m)", "Débit (m³/s)", "Type écoulement",
        "Date diagnostic séguia", "Défauts séguia",
        "Diag. — État général",
        "Diag. — Fissures revêtement (0-5)", "Diag. — Infiltration/fuite (0-5)",
        "Diag. — Obstructions/débris (0-5)", "Diag. — Érosion berges (0-5)",
        "Diag. — Sédimentation fond (0-5)", "Diag. — Ouvrages régulation (0-5)",
        "Diag. — Spalling béton (0-5)",
    ])
    for seg in seguias:
        for tr in seg.troncons.all():
            e = etat_tr.get(tr.id)
            ws11.append([
                seg.nom_de_la_seguia, v(seg.type_deguia),
                v(tr.troncon), v(tr.forme), v(tr.nature),
                tr.longueur, v(tr.largeur_meroire), v(tr.hauteur),
                tr.hauteur_eau, v(tr.fruit_de_berge), tr.epaisseur_parois,
                v(tr.diametre), tr.debit, v(tr.type_decoulement),
                dt(seg.date_diagnostic), v(seg.defaut_ouvrage),
                v(getattr(e, 'etat_general', '')),
                v(getattr(e, 'fissures_revetement', '')),
                v(getattr(e, 'infiltration_fuite', '')),
                v(getattr(e, 'obstructions_debris', '')),
                v(getattr(e, 'erosion_berges', '')),
                v(getattr(e, 'sedimentation_fond', '')),
                v(getattr(e, 'ouvrages_regulation', '')),
                v(getattr(e, 'spalling_beton', '')),
            ])
    autowidth(ws11)

    # ── Réponse HTTP ─────────────────────────────────────────────────────────
    safe_nom = nom_perim.replace('/', '-').replace(' ', '_')[:30]
    filename = f"perimetre_{safe_nom}_{date_cls.today().strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
