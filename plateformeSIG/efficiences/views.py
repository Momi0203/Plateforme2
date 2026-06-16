from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Prefetch
from django.http import JsonResponse, Http404
from django.shortcuts import render, get_object_or_404
from django.template.loader import render_to_string
from django.views.decorators.http import require_POST

from diagnostic.models import (
    Perimetre, Seuil, PriseLocale, Khettara, ForagePuits, BarrageRetenue,
    Seguias, TronconSeguia, SguiaAssocie_OuvrageTete,
)

from .models import Efficience
from .services.orchestrateur import (
    OUVRAGE_TETE_FK_MAP,
    calculer_efficience_complete,
    seguias_liees_a_ouvrage,
)


OUVRAGE_TETE_MODEL_MAP = {
    'seuil':           Seuil,
    'prise_locale':    PriseLocale,
    'khettara':        Khettara,
    'forage_puits':    ForagePuits,
    'barrage_retenue': BarrageRetenue,
}

OUVRAGE_TETE_LABEL = {
    'seuil':           'Seuil',
    'prise_locale':    'Prise locale',
    'khettara':        'Khettara',
    'forage_puits':    'Forage / Puits',
    'barrage_retenue': 'Barrage de retenue',
}

OUVRAGE_TETE_ICON = {
    'seuil':           'fa-water',
    'prise_locale':    'fa-faucet',
    'khettara':        'fa-stream',
    'forage_puits':    'fa-bullseye',
    'barrage_retenue': 'fa-mountain',
}

OUVRAGE_TETE_COLOR = {
    'seuil':           '#0c7a8a',
    'prise_locale':    '#1e8449',
    'khettara':        '#c0392b',
    'forage_puits':    '#7d3c98',
    'barrage_retenue': '#f0a500',
}


def _get_ouvrages_par_type(perimetre):
    return [
        ('seuil',           'Seuils',                perimetre.seuils.all()),
        ('prise_locale',    'Prises locales',        perimetre.prises_locales.all()),
        ('khettara',        'Khettaras',             perimetre.khettaras.all()),
        ('forage_puits',    'Forages / Puits',       perimetre.forages_puits.all()),
        ('barrage_retenue', 'Barrages de retenue',   perimetre.barrages_retenue.all()),
    ]


def _get_ouvrage(ouvrage_type, ouvrage_id):
    model_cls = OUVRAGE_TETE_MODEL_MAP.get(ouvrage_type)
    if model_cls is None:
        raise Http404("Type d'ouvrage de tête inconnu")
    return get_object_or_404(model_cls, pk=ouvrage_id)


@login_required(login_url='connexion')
def liste_perimetres(request):
    perimetres = Perimetre.objects.prefetch_related(
        Prefetch('efficiences', queryset=Efficience.objects.order_by('-date_calcul').select_related('operateur'))
    ).order_by('ksar_village')
    nb_perimetres = perimetres.count()
    nb_seguias = TronconSeguia.objects.count()
    nb_efficiences = Efficience.objects.count()
    nb_ouvrages = sum([
        Seuil.objects.count(),
        PriseLocale.objects.count(),
        Khettara.objects.count(),
        ForagePuits.objects.count(),
        BarrageRetenue.objects.count(),
    ])
    dernieres = (
        Efficience.objects
        .select_related('perimetre', 'operateur')
        .order_by('-date_calcul')[:5]
    )
    return render(request, 'efficiences/liste_perimetres.html', {
        'perimetres': perimetres,
        'nb_perimetres': nb_perimetres,
        'nb_seguias': nb_seguias,
        'nb_efficiences': nb_efficiences,
        'nb_ouvrages': nb_ouvrages,
        'dernieres': dernieres,
    })


@login_required(login_url='connexion')
def formulaire_efficience(request, perimetre_id):
    perimetre = get_object_or_404(Perimetre, pk=perimetre_id)
    groupes_ouvrages = _get_ouvrages_par_type(perimetre)
    nb_troncons = TronconSeguia.objects.filter(seguia__perimetre=perimetre).count()
    dernier_calcul = Efficience.objects.filter(perimetre=perimetre).order_by('-date_calcul').first()
    return render(request, 'efficiences/formulaire_efficience.html', {
        'perimetre': perimetre,
        'groupes_ouvrages': groupes_ouvrages,
        'nb_troncons': nb_troncons,
        'OUVRAGE_TETE_LABEL': OUVRAGE_TETE_LABEL,
        'OUVRAGE_TETE_ICON': OUVRAGE_TETE_ICON,
        'dernier_calcul': dernier_calcul,
    })


@login_required(login_url='connexion')
def api_perimetre_carte(request, perimetre_id):
    """GeoJSON du périmètre + ouvrages + séguias pour rendu Leaflet."""
    perimetre = get_object_or_404(Perimetre, pk=perimetre_id)
    payload = {
        'perimetre': {
            'id': perimetre.id,
            'nom': str(perimetre),
            'geometrie': perimetre.geometrie.geojson if getattr(perimetre, 'geometrie', None) else None,
        },
        'ouvrages': [],
        'seguias': [],
        'styles': {
            'colors': OUVRAGE_TETE_COLOR,
            'icons':  OUVRAGE_TETE_ICON,
            'labels': OUVRAGE_TETE_LABEL,
        },
    }
    for type_code, type_label, qs in _get_ouvrages_par_type(perimetre):
        for o in qs:
            payload['ouvrages'].append({
                'type': type_code,
                'type_label': type_label,
                'id': o.id,
                'nom': str(o),
                'geometrie': o.geometrie.geojson if getattr(o, 'geometrie', None) else None,
            })
    for tr in TronconSeguia.objects.filter(seguia__perimetre=perimetre).select_related('seguia'):
        payload['seguias'].append({
            'id': tr.id,
            'nom': f"{tr.seguia.nom_de_la_seguia} — {tr.troncon}",
            'type_decoulement': tr.type_decoulement,
            'type_deguia': tr.seguia.type_deguia,
            'geometrie': tr.geometrie.geojson if tr.geometrie else None,
        })
    return JsonResponse(payload)


@login_required(login_url='connexion')
def api_seguias_disponibles(request, perimetre_id, ouvrage_type, ouvrage_id):
    """Liste des séguias du périmètre avec flag is_linked pour l'ouvrage sélectionné."""
    perimetre = get_object_or_404(Perimetre, pk=perimetre_id)
    ouvrage = _get_ouvrage(ouvrage_type, ouvrage_id)
    if getattr(ouvrage, 'perimetre_id', None) != perimetre.id:
        return JsonResponse(
            {'success': False, 'error': "L'ouvrage n'appartient pas au périmètre."},
            status=400,
        )

    fk_lookup = f'{OUVRAGE_TETE_FK_MAP[ouvrage_type]}_id'
    linked_ids = set(
        SguiaAssocie_OuvrageTete.objects
        .filter(**{fk_lookup: ouvrage_id})
        .values_list('FK_nom_sguia_id', flat=True)
    )

    seguias = (
        perimetre.seguias
        .prefetch_related('troncons')
        .all()
        .order_by('nom_de_la_seguia')
    )
    data = []
    for s in seguias:
        troncons = list(s.troncons.order_by('troncon').all())
        premier = troncons[0] if troncons else None
        data.append({
            'id': s.id,
            'nom': s.nom_de_la_seguia,
            'type': s.get_type_deguia_display() if s.type_deguia else '',
            'nb_troncons': len(troncons),
            'nature': premier.get_nature_display() if premier else '',
            'longueur': sum(tr.longueur or 0 for tr in troncons),
            'debit': premier.debit if premier else None,
            'type_decoulement': premier.type_decoulement if premier else '',
            'is_linked': s.id in linked_ids,
        })
    return JsonResponse({
        'success': True,
        'seguias': data,
        'nb_total': len(data),
        'nb_lies': len(linked_ids),
    })


@require_POST
@login_required(login_url='connexion')
def enregistrer_liaisons(request, perimetre_id):
    """Synchronise les liaisons séguia ↔ ouvrage pour l'ouvrage choisi.

    Reçoit : ouvrage_type, ouvrage_id, seguia_ids[].
    Crée les liaisons manquantes, supprime celles décochées.
    """
    perimetre = get_object_or_404(Perimetre, pk=perimetre_id)
    ouvrage_type = request.POST.get('ouvrage_type')
    ouvrage_id_raw = request.POST.get('ouvrage_id')
    seguia_ids_raw = request.POST.getlist('seguia_ids')

    if ouvrage_type not in OUVRAGE_TETE_FK_MAP:
        return JsonResponse({'success': False, 'error': "Type d'ouvrage invalide."}, status=400)
    try:
        ouvrage_id = int(ouvrage_id_raw)
    except (TypeError, ValueError):
        return JsonResponse({'success': False, 'error': "ID d'ouvrage invalide."}, status=400)

    ouvrage = _get_ouvrage(ouvrage_type, ouvrage_id)
    if getattr(ouvrage, 'perimetre_id', None) != perimetre.id:
        return JsonResponse(
            {'success': False, 'error': "L'ouvrage n'appartient pas au périmètre."},
            status=400,
        )

    seguia_ids = set()
    for sid in seguia_ids_raw:
        try:
            seguia_ids.add(int(sid))
        except (TypeError, ValueError):
            continue

    valides = set(
        Seguias.objects
        .filter(perimetre=perimetre, id__in=seguia_ids)
        .values_list('id', flat=True)
    )
    invalides = seguia_ids - valides
    if invalides:
        return JsonResponse(
            {'success': False, 'error': f"Séguia(s) hors périmètre : {sorted(invalides)}"},
            status=400,
        )

    fk_lookup = f'{OUVRAGE_TETE_FK_MAP[ouvrage_type]}_id'
    existing_qs = SguiaAssocie_OuvrageTete.objects.filter(**{fk_lookup: ouvrage_id})
    existing_ids = set(existing_qs.values_list('FK_nom_sguia_id', flat=True))

    to_create = valides - existing_ids
    to_delete = existing_ids - valides

    with transaction.atomic():
        SguiaAssocie_OuvrageTete.objects.bulk_create([
            SguiaAssocie_OuvrageTete(
                FK_nom_sguia_id=sid,
                **{fk_lookup: ouvrage_id},
            )
            for sid in to_create
        ])
        existing_qs.filter(FK_nom_sguia_id__in=to_delete).delete()

    return JsonResponse({
        'success': True,
        'created': len(to_create),
        'deleted': len(to_delete),
        'total_lies': len(valides),
    })


@login_required(login_url='connexion')
def troncons_par_ouvrage(request, perimetre_id, ouvrage_type, ouvrage_id):
    """Tableau HTML des séguias liées (vue read-only conservée)."""
    perimetre = get_object_or_404(Perimetre, pk=perimetre_id)
    _get_ouvrage(ouvrage_type, ouvrage_id)
    troncons = seguias_liees_a_ouvrage(perimetre, ouvrage_type, ouvrage_id)
    html = render_to_string(
        'efficiences/partials/tableau_troncons.html',
        {'troncons': troncons, 'perimetre': perimetre},
        request=request,
    )
    return JsonResponse({
        'success': True,
        'nb_troncons': troncons.count(),
        'html': html,
    })


@require_POST
@login_required(login_url='connexion')
def lancer_calcul(request, perimetre_id):
    """Lance le calcul complet + persistance + renvoie HTML résultats."""
    perimetre = get_object_or_404(Perimetre, pk=perimetre_id)
    ouvrage_type = request.POST.get('ouvrage_type')
    ouvrage_id = request.POST.get('ouvrage_id')

    if ouvrage_type not in OUVRAGE_TETE_FK_MAP:
        return JsonResponse({'success': False, 'error': "Type d'ouvrage invalide."}, status=400)
    try:
        ouvrage_id_int = int(ouvrage_id)
    except (TypeError, ValueError):
        return JsonResponse({'success': False, 'error': "ID d'ouvrage invalide."}, status=400)

    ouvrage = _get_ouvrage(ouvrage_type, ouvrage_id_int)

    try:
        result = calculer_efficience_complete(
            perimetre=perimetre,
            ouvrage_tete_type=ouvrage_type,
            ouvrage_tete_id=ouvrage_id_int,
            operateur=request.user if request.user.is_authenticated else None,
        )
    except Exception as exc:  # noqa: BLE001
        return JsonResponse({'success': False, 'error': f"Erreur de calcul : {exc}"}, status=500)

    html = render_to_string(
        'efficiences/partials/resultats.html',
        {
            'efficience': result['efficience'],
            'details_par_troncon': result['details_par_troncon'],
            'tableau_par_type': result['tableau_par_type'],
            'tableau_par_ouvrage': result['tableau_par_ouvrage'],
            'efficiences_par_categorie': result['efficiences_par_categorie'],
            'compteurs': result['compteurs'],
            'efficience_globale_pourcent': result['efficience_globale_pourcent'],
            'ouvrage_label': OUVRAGE_TETE_LABEL[ouvrage_type],
            'ouvrage': ouvrage,
            'perimetre': perimetre,
        },
        request=request,
    )
    return JsonResponse({
        'success': True,
        'efficience_id': result['efficience'].id,
        'efficience_globale_pourcent': result['efficience_globale_pourcent'],
        'efficiences_par_categorie': result['efficiences_par_categorie'],
        'compteurs': result['compteurs'],
        'nb_troncons': len(result['details_par_troncon']),
        'html': html,
    })


@login_required(login_url='connexion')
@require_POST
def valider_efficience(request, pk):
    eff = get_object_or_404(Efficience, pk=pk)
    eff.statut = 'valide'
    eff.save(update_fields=['statut'])

    # Répercuter l'efficience globale validée sur le périmètre et l'ouvrage de tête
    eff_decimal = eff.efficience_globale / 100.0  # % → 0-1

    perimetre = eff.perimetre
    perimetre.efficiance_reseau = eff_decimal
    perimetre.save(update_fields=['efficiance_reseau'])

    model_cls = OUVRAGE_TETE_MODEL_MAP.get(eff.ouvrage_tete_type)
    if model_cls:
        try:
            ouvrage = model_cls.objects.get(pk=eff.ouvrage_tete_id)
            if hasattr(ouvrage, 'efficience_reseaux'):
                ouvrage.efficience_reseaux = eff_decimal
                ouvrage.save(update_fields=['efficience_reseaux'])
        except model_cls.DoesNotExist:
            pass

    return JsonResponse({'ok': True, 'statut': 'valide'})


@login_required(login_url='connexion')
def exporter_excel_efficience(request, pk):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse

    eff = get_object_or_404(Efficience, pk=pk)
    perimetre = eff.perimetre
    ouvrage_label = OUVRAGE_TETE_LABEL.get(eff.ouvrage_tete_type, eff.ouvrage_tete_type)
    ouvrage = _get_ouvrage(eff.ouvrage_tete_type, eff.ouvrage_tete_id)

    result = calculer_efficience_complete(perimetre, eff.ouvrage_tete_type, eff.ouvrage_tete_id)

    wb = Workbook()
    C_DARK, C_GOLD, C_GREY = "1A1A2E", "F0A500", "F8F4EE"
    thin = Side(style='thin', color="D0C8BC")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
    LEFT   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    F_HDR  = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    F_LBL  = Font(name="Calibri", bold=True, size=10, color=C_DARK)
    F_VAL  = Font(name="Calibri",             size=10, color=C_DARK)
    F_GOLD = Font(name="Calibri", bold=True, size=11, color=C_GOLD)
    F_TTL  = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    FILL_D = PatternFill("solid", fgColor=C_DARK)
    FILL_G = PatternFill("solid", fgColor=C_GREY)
    FILL_OK  = PatternFill("solid", fgColor="D4EDDA")
    FILL_MID = PatternFill("solid", fgColor="FFF8EE")
    FILL_BAD = PatternFill("solid", fgColor="FDEDEC")

    def _ttl(ws, text, ncols, row=1):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        c = ws.cell(row=row, column=1, value=text)
        c.font = F_TTL; c.fill = FILL_D; c.alignment = CENTER
        ws.row_dimensions[row].height = 24

    def _sec(ws, text, ncols, row):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        c = ws.cell(row=row, column=1, value=text)
        c.font = F_HDR; c.fill = FILL_D; c.alignment = CENTER
        ws.row_dimensions[row].height = 18

    def _hdr(ws, row, col, text):
        c = ws.cell(row=row, column=col, value=text)
        c.font = F_HDR; c.fill = FILL_D; c.alignment = CENTER; c.border = BORDER

    def _lbl(ws, row, col, text, fill=None):
        c = ws.cell(row=row, column=col, value=text)
        c.font = F_LBL; c.fill = fill or FILL_G; c.alignment = LEFT; c.border = BORDER

    def _val(ws, row, col, v, fill=None, gold=False):
        c = ws.cell(row=row, column=col, value=v)
        c.font = F_GOLD if gold else F_VAL
        c.fill = fill or PatternFill(); c.alignment = CENTER; c.border = BORDER

    def eff_fill(pct):
        if pct is None: return FILL_G
        return FILL_OK if pct >= 80 else (FILL_MID if pct >= 50 else FILL_BAD)

    # ── Feuille 1 : Synthèse ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Synthèse"
    ws1.column_dimensions["A"].width = 36
    ws1.column_dimensions["B"].width = 28
    _ttl(ws1, f"RAPPORT EFFICIENCE DU RÉSEAU — HydroPlan SIG", 2)
    _sec(ws1, "Informations générales", 2, 3)
    meta = [
        ("Périmètre irrigué",   str(perimetre)),
        ("Ouvrage de tête",     f"{ouvrage_label} — {ouvrage}"),
        ("Date de calcul",      eff.date_calcul.strftime("%d/%m/%Y  %H:%M") if eff.date_calcul else "—"),
        ("Statut",              "Validé" if eff.statut == 'valide' else "Calculé"),
        ("Opérateur",           str(eff.operateur) if eff.operateur else "—"),
    ]
    for ri, (k, v) in enumerate(meta, start=4):
        _lbl(ws1, ri, 1, k); _val(ws1, ri, 2, str(v))
    _sec(ws1, "KPI — Efficiences par catégorie", 2, 10)
    epc = result.get('efficiences_par_categorie', {})
    kpi = [
        ("Efficience principale (%)",  epc.get('principale')),
        ("Efficience secondaire (%)",  epc.get('secondaire')),
        ("Efficience tertiaire (%)",   epc.get('tertiaire')),
        ("Efficience globale (%)",     result.get('efficience_globale_pourcent')),
    ]
    for ri, (k, v) in enumerate(kpi, start=11):
        _lbl(ws1, ri, 1, k)
        v_disp = round(v, 2) if v is not None else "—"
        _val(ws1, ri, 2, v_disp, gold=True)

    # ── Feuille 2 : Tronçons ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("Tronçons")
    cols2 = [("Séguia", 20), ("Code tronçon", 14), ("Type", 16), ("Nature", 12),
             ("Débit amont (m³/s)", 18), ("Pi (m³/s)", 14), ("Pv (m³/s)", 14),
             ("Efficience (%)", 14), ("Dalot", 10)]
    for ci, (_, w) in enumerate(cols2, start=1):
        ws2.column_dimensions[chr(64+ci)].width = w
    _ttl(ws2, f"Détail par tronçon — Périmètre : {perimetre}", len(cols2))
    for ci, (hd, _) in enumerate(cols2, start=1):
        _hdr(ws2, 3, ci, hd)
    for ri, d in enumerate(result.get('details_par_troncon', []), start=4):
        pct = d.get('efficience_pourcent')
        fill = eff_fill(pct)
        _val(ws2, ri, 1, d.get('seguia_nom', '—'), fill=fill)
        _val(ws2, ri, 2, d.get('troncon_code', '—'), fill=fill)
        _val(ws2, ri, 3, d.get('type_seguia_label', '—'), fill=fill)
        _val(ws2, ri, 4, d.get('nature_label', '—'), fill=fill)
        _val(ws2, ri, 5, round(d.get('debit_amont', 0) or 0, 6), fill=fill)
        _val(ws2, ri, 6, round(d.get('perte_infiltration_m3s', 0) or 0, 8), fill=fill)
        _val(ws2, ri, 7, round(d.get('perte_vaporisation_m3s', 0) or 0, 8), fill=fill)
        _val(ws2, ri, 8, round(pct, 2) if pct is not None else "—", fill=fill, gold=True)
        _val(ws2, ri, 9, "Oui" if d.get('is_dalot') else "Non", fill=fill)

    # ── Feuille 3 : Par type ──────────────────────────────────────────────────
    ws3 = wb.create_sheet("Par type")
    cols3 = [("Type de séguia", 22), ("Nb tronçons", 13), ("Débit total (m³/s)", 18),
             ("Pi total (m³/s)", 16), ("Pv total (m³/s)", 16),
             ("Efficience moy. (%)", 18), ("Taux perte (%)", 14)]
    for ci, (_, w) in enumerate(cols3, start=1):
        ws3.column_dimensions[chr(64+ci)].width = w
    _ttl(ws3, f"Efficience par type de séguia — Périmètre : {perimetre}", len(cols3))
    for ci, (hd, _) in enumerate(cols3, start=1):
        _hdr(ws3, 3, ci, hd)
    for ri, t in enumerate(result.get('tableau_par_type', []), start=4):
        pct = t.get('efficience_moyenne_ponderee')
        fill = eff_fill(pct)
        _val(ws3, ri, 1, t.get('type_label', '—'), fill=fill)
        _val(ws3, ri, 2, t.get('nb_troncons', 0), fill=fill)
        _val(ws3, ri, 3, round(t.get('debit_total_m3s', 0) or 0, 6), fill=fill)
        _val(ws3, ri, 4, round(t.get('perte_infiltration_totale_m3s', 0) or 0, 8), fill=fill)
        _val(ws3, ri, 5, round(t.get('perte_vaporisation_totale_m3s', 0) or 0, 8), fill=fill)
        _val(ws3, ri, 6, round(pct, 2) if pct is not None else "—", fill=fill, gold=True)
        tp = t.get('taux_perte_global_pourcent')
        _val(ws3, ri, 7, round(tp, 2) if tp is not None else "—", fill=fill)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    fname = f"efficience_{perimetre.pk}_{eff.pk}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    wb.save(response)
    return response


@login_required(login_url='connexion')
def exporter_pdf_efficience(request, pk):
    import io as _io
    from datetime import date as _date
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    from django.http import HttpResponse

    eff = get_object_or_404(Efficience, pk=pk)
    perimetre = eff.perimetre
    ouvrage_label = OUVRAGE_TETE_LABEL.get(eff.ouvrage_tete_type, eff.ouvrage_tete_type)
    ouvrage = _get_ouvrage(eff.ouvrage_tete_type, eff.ouvrage_tete_id)

    result = calculer_efficience_complete(perimetre, eff.ouvrage_tete_type, eff.ouvrage_tete_id)

    buf = _io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
        leftMargin=1.8*cm, rightMargin=1.8*cm, topMargin=1.8*cm, bottomMargin=1.8*cm,
        title=f"Efficience — {perimetre}")
    PW = 16.9 * cm

    DARK = colors.HexColor("#1A1A2E")
    GOLD = colors.HexColor("#F0A500")
    GREY = colors.HexColor("#F8F4EE")
    C_OK  = colors.HexColor("#D4EDDA")
    C_MID = colors.HexColor("#FFF8EE")
    C_BAD = colors.HexColor("#FDEDEC")
    RED   = colors.HexColor("#C0392B")

    S_WHITE = ParagraphStyle("w", fontSize=14, fontName="Helvetica-Bold", textColor=colors.white, alignment=1)
    S_HEAD  = ParagraphStyle("h", fontSize=11, fontName="Helvetica-Bold", textColor=colors.white, alignment=1)
    S_LABEL = ParagraphStyle("lb",fontSize=8.5, fontName="Helvetica-Bold", textColor=DARK)
    S_VALUE = ParagraphStyle("v", fontSize=8.5, fontName="Helvetica",      textColor=colors.HexColor("#333333"))
    S_GOLD  = ParagraphStyle("g", fontSize=10,  fontName="Helvetica-Bold", textColor=GOLD, alignment=1)
    S_SMALL = ParagraphStyle("sm",fontSize=7.5, fontName="Helvetica",      textColor=colors.grey)
    S_CENT  = ParagraphStyle("c", fontSize=8,   fontName="Helvetica",      textColor=DARK, alignment=1)
    PAD = [("LEFTPADDING",(0,0),(-1,-1),5), ("RIGHTPADDING",(0,0),(-1,-1),5),
           ("TOPPADDING",(0,0),(-1,-1),4),  ("BOTTOMPADDING",(0,0),(-1,-1),4),
           ("VALIGN",(0,0),(-1,-1),"MIDDLE")]

    def sec(text):
        t = Table([[Paragraph(text, S_HEAD)]], colWidths=[PW])
        t.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),DARK), *PAD,
                                ("TOPPADDING",(0,0),(-1,-1),6), ("BOTTOMPADDING",(0,0),(-1,-1),6)]))
        return t

    def kv(rows, w1=6*cm):
        data = [[Paragraph(k, S_LABEL), Paragraph(str(v), S_VALUE)] for k, v in rows]
        t = Table(data, colWidths=[w1, PW - w1])
        t.setStyle(TableStyle([*PAD, ("BACKGROUND",(0,0),(0,-1),GREY),
                                ("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#E0D0C0"))]))
        return t

    def gh(txt, bg=DARK):
        return Paragraph(txt, ParagraphStyle("gh",fontSize=8,fontName="Helvetica-Bold",
               textColor=colors.white,alignment=1))

    def eff_bg(pct):
        if pct is None: return GREY
        return C_OK if pct >= 80 else (C_MID if pct >= 50 else C_BAD)

    fmt6 = lambda v: f"{v:.6f}" if v is not None else "—"
    fmt2 = lambda v: f"{v:.2f}" if v is not None else "—"
    story = []

    # Bandeau
    t = Table([[Paragraph("RAPPORT EFFICIENCE DU RÉSEAU", S_WHITE),
                Paragraph(f"Efficience #{pk}", S_SMALL)]],
              colWidths=[PW * 0.78, PW * 0.22])
    t.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),DARK), *PAD,
                            ("TOPPADDING",(0,0),(-1,-1),10), ("BOTTOMPADDING",(0,0),(-1,-1),10),
                            ("ALIGN",(-1,0),(-1,0),"RIGHT"), ("VALIGN",(0,0),(-1,-1),"MIDDLE")]))
    story += [t, Spacer(1, .35*cm)]

    # 1. Informations générales
    story += [sec("1. Informations générales"), Spacer(1, .1*cm)]
    story.append(kv([
        ("Périmètre irrigué",  str(perimetre)),
        ("Ouvrage de tête",    f"{ouvrage_label} — {ouvrage}"),
        ("Date de calcul",     eff.date_calcul.strftime("%d/%m/%Y  %H:%M") if eff.date_calcul else "—"),
        ("Statut",             "Validé" if eff.statut == 'valide' else "Calculé"),
        ("Opérateur",          str(eff.operateur) if eff.operateur else "—"),
    ]))

    # 2. KPI
    story += [Spacer(1, .3*cm), sec("2. Efficiences par catégorie"), Spacer(1, .1*cm)]
    epc = result.get('efficiences_par_categorie', {})
    eg  = result.get('efficience_globale_pourcent')
    kpi_data = [[gh("Catégorie"), gh("Efficience (%)"), gh("Nb tronçons")]]
    cptrs = result.get('compteurs', {})
    for cat, label in [('principale','Principale'), ('secondaire','Secondaire'), ('tertiaire','Tertiaire')]:
        v = epc.get(cat)
        kpi_data.append([
            Paragraph(label, S_LABEL),
            Paragraph(fmt2(v), ParagraphStyle("kv",fontSize=9,fontName="Helvetica-Bold",
                      textColor=RED if (v is not None and v < 50) else GOLD, alignment=1)),
            Paragraph(str(cptrs.get(cat, 0)), S_CENT),
        ])
    kpi_data.append([
        Paragraph("Efficience globale (P×S×T)", ParagraphStyle("gb",fontSize=9,fontName="Helvetica-Bold",textColor=DARK)),
        Paragraph(fmt2(eg), ParagraphStyle("gv",fontSize=11,fontName="Helvetica-Bold",textColor=GOLD,alignment=1)),
        Paragraph("—", S_CENT),
    ])
    t_kpi = Table(kpi_data, colWidths=[7*cm, 5.5*cm, 4.4*cm])
    t_kpi.setStyle(TableStyle([*PAD, ("BACKGROUND",(0,0),(-1,0),DARK),
                                ("BACKGROUND",(0,1),(0,-1),GREY),
                                ("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#E0D0C0")),
                                ("BACKGROUND",(0,4),(-1,4),colors.HexColor("#FFF8EE"))]))
    story.append(t_kpi)

    # 3. Tableau par tronçon
    story += [Spacer(1, .3*cm), sec("3. Détail par tronçon"), Spacer(1, .1*cm)]
    cw_trc = [3.8*cm, 2.4*cm, 2.4*cm, 2.4*cm, 2.4*cm, 2.4*cm, 1.5*cm]
    trc_hdr = [gh(h) for h in ["Séguia / Tronçon", "Type", "Débit am. (m³/s)",
                                "Pi (m³/s)", "Pv (m³/s)", "Efficience (%)", "Dalot"]]
    trc_rows = [trc_hdr]
    for d in result.get('details_par_troncon', []):
        pct = d.get('efficience_pourcent')
        bg  = eff_bg(pct)
        style_pct = ParagraphStyle("p",fontSize=8,fontName="Helvetica-Bold",
                    textColor=RED if (pct is not None and pct < 50) else GOLD, alignment=1)
        trc_rows.append([
            Paragraph(f"{d.get('seguia_nom','—')}\n{d.get('troncon_code','')}", S_LABEL),
            Paragraph(d.get('type_seguia_label','—'), S_CENT),
            Paragraph(fmt6(d.get('debit_amont')), S_CENT),
            Paragraph(f"{d.get('perte_infiltration_m3s',0):.8f}", S_CENT),
            Paragraph(f"{d.get('perte_vaporisation_m3s',0):.8f}", S_CENT),
            Paragraph(fmt2(pct), style_pct),
            Paragraph("O" if d.get('is_dalot') else "N", S_CENT),
        ])
    t_trc = Table(trc_rows, colWidths=cw_trc)
    t_trc.setStyle(TableStyle([*PAD, ("BACKGROUND",(0,0),(-1,0),DARK),
                                ("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#E0D0C0")),
                                ("FONTSIZE",(0,0),(-1,-1),7.5),
                                ("ALIGN",(0,0),(-1,-1),"CENTER")]))
    for ri, d in enumerate(result.get('details_par_troncon', []), start=1):
        pct = d.get('efficience_pourcent')
        t_trc.setStyle(TableStyle([("BACKGROUND",(0,ri),(-1,ri), eff_bg(pct))]))
    story.append(t_trc)

    # Pied de page
    story += [Spacer(1, .4*cm),
              HRFlowable(width="100%", thickness=0.5, color=GOLD),
              Spacer(1, .08*cm),
              Paragraph(f"HydroPlan SIG  ·  Généré le {_date.today().strftime('%d/%m/%Y')}  ·  Efficience #{pk}", S_SMALL)]

    doc.build(story)
    buf.seek(0)
    fname = f"efficience_{perimetre.pk}_{pk}_{_date.today().strftime('%Y%m%d')}.pdf"
    resp = HttpResponse(buf.read(), content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="{fname}"'
    return resp


@login_required(login_url='connexion')
def historique(request, perimetre_id=None):
    qs = Efficience.objects.select_related('perimetre', 'operateur').order_by('-date_calcul')
    perimetre = None
    if perimetre_id is not None:
        perimetre = get_object_or_404(Perimetre, pk=perimetre_id)
        qs = qs.filter(perimetre=perimetre)

    enrichis = []
    for eff in qs[:200]:
        ouvrage = None
        model_cls = OUVRAGE_TETE_MODEL_MAP.get(eff.ouvrage_tete_type)
        if model_cls is not None:
            ouvrage = model_cls.objects.filter(pk=eff.ouvrage_tete_id).first()
        enrichis.append({
            'efficience': eff,
            'ouvrage': ouvrage,
            'ouvrage_label': OUVRAGE_TETE_LABEL.get(eff.ouvrage_tete_type, eff.ouvrage_tete_type),
        })

    return render(request, 'efficiences/historique.html', {
        'perimetre': perimetre,
        'historique': enrichis,
    })
