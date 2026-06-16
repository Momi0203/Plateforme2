import io
import os
import json
import shutil
import zipfile
import tempfile
from datetime import date

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
from django.contrib.gis.gdal import DataSource
from django.contrib.gis.geos import GEOSGeometry

from .models import (
    BassinVersant, StationPluviometrique, StationHydrometrique,
    CoefficientMontana, ResultatAnalyseHydrologique,
)
from .forms import (
    BassinVersantForm, StationPluviometriqueForm, StationHydrometriqueForm,
    CoefficientMontanaForm, AnalyseParametresForm,
)
from .calculs import (
    bv_to_hydro,
    run_analyse,
    recalculer_depuis_tc,
    recalculer_depuis_formules_q,
    calculer_apports_crue_sans_prelevement,
)

# Réseaux hydrographiques « ouvrage de tête » (app carte) — un modèle par bassin.
from carte.models import (
    BassinVersant as CarteBassinVersant,
    ReseauOuvrageTeteZiz, ReseauOuvrageTeteMoulouya, ReseauOuvrageTeteGuir,
    ReseauOuvrageTeteRheris, ReseauOuvrageTeteMaider,
)

PERIODES = [10, 20, 50, 100]
SRID_NORD_MAROC = 26191

# Mapping : nom normalisé du carte.BassinVersant -> modèle réseau « ouvrage de tête »
_RESEAU_OT_PAR_BASSIN = {
    'ziz':      ReseauOuvrageTeteZiz,
    'moulouya': ReseauOuvrageTeteMoulouya,
    'guir':     ReseauOuvrageTeteGuir,
    'rheris':   ReseauOuvrageTeteRheris,
    'maider':   ReseauOuvrageTeteMaider,
}

SHP_FIELD_MAP = {
    # Surface / Perimetre
    's':           'surface',
    'surface':     'surface',
    'superficie':  'surface',
    'surface_m2':  'surface_m2',  # m² → conversion km² dans le traitement
    'shape_area':  'surface_shp',
    'p':           'perimetre',
    'perimetre':   'perimetre',
    'perimeter':   'perimetre',
    'perimetre_':  'perimetre',   # ArcGIS field name
    # Altitudes
    'zmax':        'z_max',
    'z_max':       'z_max',
    'altmax':      'z_max',
    'zmin':        'z_min',
    'z_min':       'z_min',
    'altmin':      'z_min',
    # Thalweg
    'thalweg':     'thalweg',
    'thalweg_m':   'thalweg',      # ArcGIS field name (thalweg en mètres)
    'long_thal':   'thalweg',
    'lthalweg':    'thalweg',
    # Coordinates
    'coordonnes':  'x_exutoire',
    'coordonne':   'x_exutoire',
    'coord_x':     'x_exutoire',
    'x':           'x_exutoire',
    'x_exutoire':  'x_exutoire',   # Exact field name from ArcGIS
    'coordonn_1':  'y_exutoire',
    'coord_y':     'y_exutoire',
    'y':           'y_exutoire',
    'y_exutoire':  'y_exutoire',   # Exact field name from ArcGIS
    # Station coords
    'x_station':   'x',
    'y_station':   'y',
    # Ouvrage
    'ouvrage_te':  'ouvrage_en_tete',
    'ouvrage':     'ouvrage_en_tete',
    'nom_ouvrage': 'ouvrage_en_tete',
    # Nom
    'nom':         'nom',
    'name':        'nom',
    'nom_bv':      'nom',
    'nom_st':      'nom',
    # Pluvio / Hydro
    'hmoy':        'hauteur_moyenne',
    'h_moy':       'hauteur_moyenne',
    'hmoyen':      'hauteur_moyenne',
    'grad':        'grad_exp_pluie',
    'gradex':      'grad_exp_pluie',
    'pj10':        'pjmax_t10',
    'pjmax10':     'pjmax_t10',
    'pj20':        'pjmax_t20',
    'pj50':        'pjmax_t50',
    'pj100':       'pjmax_t100',
}

BV_SOURCE_FIELDS = {
    'ouvrage_te', 's', 'p', 'zmax', 'zmin',
    'thalweg', 'thalweg_m', 'coordonnes', 'coordonn_1',
    'x_exutoire', 'y_exutoire', 'surface_m2', 'perimetre_',
}

BV_FORM_ATTR_FIELDS = {
    'nom', 'ouvrage_en_tete', 'surface', 'perimetre', 'z_max',
    'z_min', 'thalweg', 'x_exutoire', 'y_exutoire',
}

BV_BATCH_REQUIRED_ATTRS = {
    'surface', 'perimetre', 'z_max',
    'z_min', 'thalweg', 'x_exutoire', 'y_exutoire',
}


def _json_safe(obj):
    """Recursively convert int dict keys → str for JSON serialisation."""
    if isinstance(obj, dict):
        return {str(k): _json_safe(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [_json_safe(x) for x in obj]
    return obj


def _d(dct, key):
    """Get value from dict by int or str key (handles JSON round-trip)."""
    if dct is None:
        return None
    return dct.get(key, dct.get(str(key)))


def _bassin_versant_export_rows(bv):
    """Return detailed bassin versant rows for PDF/Excel exports."""
    if not bv:
        return []

    bv_obj = bv_to_hydro(bv)
    L_re, l_re = bv_obj.rectangle_equivalent()
    geom = bv.geometrie

    return [
        ("Nom du bassin versant", bv.nom, "—"),
        ("Ouvrage en tête", bv.ouvrage_en_tete or "—", "—"),
        ("X exutoire", round(bv.x_exutoire, 6), "coord."),
        ("Y exutoire", round(bv.y_exutoire, 6), "coord."),
        ("Surface du bassin versant", round(bv.surface, 4), "km²"),
        ("Périmètre", round(bv.perimetre, 4), "km"),
        ("Longueur du thalweg principal", round(bv.thalweg, 4), "km"),
        ("Altitude minimale Zmin", round(bv.z_min, 3), "m"),
        ("Altitude maximale Zmax", round(bv.z_max, 3), "m"),
        ("Dénivelé (Zmax - Zmin)", round(bv_obj.denivele(), 3), "m"),
        ("Pente du thalweg", round(bv_obj.pente() * 100, 4), "%"),
        ("Indice de Gravelius Kc", round(bv_obj.indice_gravelius(), 4), "—"),
        ("Rapport de forme", round(bv_obj.rapport_de_forme(), 4), "—"),
        ("Rectangle équivalent - grande dimension", round(L_re, 3), "km"),
        ("Rectangle équivalent - petite dimension", round(l_re, 3), "km"),
        ("Pente du rectangle équivalent", round(bv_obj.pente_rectangle_equivalent() * 100, 4), "%"),
        ("Géométrie disponible", "Oui" if geom else "Non", "—"),
        ("Type de géométrie", geom.geom_type if geom else "—", "—"),
        ("SRID de géométrie", geom.srid if geom else "—", "—"),
    ]


def _extract_shp_attrs(layer, feat):
    attrs = {}
    source_fields = {field_name.lower() for field_name in layer.fields}
    for field_name in layer.fields:
        key = field_name.lower()
        if key not in SHP_FIELD_MAP:
            continue
        raw_val = feat.get(field_name)
        if raw_val not in (None, '', 0.0) or raw_val == 0:
            try:
                mapped_key = SHP_FIELD_MAP[key]
                # Convert surface_m2 to km²
                if mapped_key == 'surface_m2':
                    attrs['surface'] = round(float(raw_val) / 1_000_000, 3)
                if isinstance(raw_val, (int, float)):
                    attrs[mapped_key] = round(float(raw_val), 6)
                else:
                    attrs[mapped_key] = str(raw_val)
            except (TypeError, ValueError):
                pass
    return attrs, source_fields


def _build_geom_result(geom, geom_type, attrs):
    result = {'geojson': json.loads(geom.geojson), 'wkt': geom.wkt}

    try:
        geom_nord = geom.transform(SRID_NORD_MAROC, clone=True)
        if geom_nord.geom_type == 'Point':
            cx, cy = geom_nord.x, geom_nord.y
        else:
            c = geom_nord.centroid
            cx, cy = c.x, c.y
        result['x'] = round(cx, 3)
        result['y'] = round(cy, 3)
    except Exception:
        # Keep processing even if projection to EPSG:26191 fails.
        pass

    if geom_type in ('polygon', 'linestring'):
        cx = geom.centroid.x
        cy = geom.centroid.y
        zone = int((cx + 180) / 6) + 1
        utm = 32600 + zone if cy >= 0 else 32700 + zone
        gm = geom.transform(utm, clone=True)
        if geom_type == 'polygon':
            result['surface_km2'] = round(gm.area / 1e6, 4)
            result['perimetre_km'] = round(gm.length / 1000, 4)
        else:
            result['longueur_km'] = round(gm.length / 1000, 4)

    if 'surface' in attrs and attrs['surface'] > 0:
        result['surface_km2'] = round(attrs['surface'], 4)
    if 'perimetre' in attrs and attrs['perimetre'] > 0:
        result['perimetre_km'] = round(attrs['perimetre'], 4)

    return result


# =============================================================================
# Accueil
# =============================================================================

@login_required
def accueil(request):
    context = {
        'nb_bv':      BassinVersant.objects.count(),
        'nb_pluvio':  StationPluviometrique.objects.count(),
        'nb_hydro':   StationHydrometrique.objects.count(),
        'nb_analyses': ResultatAnalyseHydrologique.objects.count(),
        'dernieres_analyses': (
            ResultatAnalyseHydrologique.objects
            .select_related('operateur', 'bassin_versant')
            .order_by('-date_analyse')[:5]
        ),
    }
    return render(request, 'analyse_hydrologique/accueil.html', context)


# =============================================================================
# Bassin Versant
# =============================================================================

@login_required
def liste_bv(request):
    q = request.GET.get('q', '').strip()
    bassins = BassinVersant.objects.all()
    if q:
        bassins = bassins.filter(nom__icontains=q)
    if request.GET.get('json'):
        data = list(bassins.values('pk', 'nom', 'surface'))
        return JsonResponse(data, safe=False)
    return render(request, 'analyse_hydrologique/bv/liste.html', {'bassins': bassins, 'q': q})


def _normaliser_nom_bassin(nom):
    """Minuscules + suppression des accents, pour apparier un nom de bassin."""
    s = (nom or '').strip().lower()
    for a, b in (('é', 'e'), ('è', 'e'), ('ê', 'e'), ('ë', 'e'), ('ï', 'i'),
                 ('î', 'i'), ('à', 'a'), ('â', 'a'), ('ô', 'o'), ('û', 'u'),
                 ('ü', 'u'), ('ç', 'c')):
        s = s.replace(a, b)
    return s


# Filtrage du réseau dense selon la SURFACE D'INTERSECTION BV∩bassin :
#   - petite intersection (< SEUIL) : on montre TOUT le réseau, grid faible inclus
#     (grid_code=0), car pour ces petits bassins de captage c'est le réseau local réel ;
#   - grande intersection (>= SEUIL) : on ne garde que les drains forts (grid_code >= 1)
#     pour rester léger côté navigateur (sinon jusqu'à ~65k tronçons / 15 Mo).
# Seuil calibré sur les données : gros BV intersectent 743–5802 km², les autres <= 421 km².
RESEAU_SEUIL_INTERSECTION_M2 = 500e6  # 500 km²


def _reseau_ouvrage_tete_pour_bv(bv, min_grid_code=None,
                                 seuil_intersection_m2=RESEAU_SEUIL_INTERSECTION_M2):
    """
    Logique d'appariement BV (analyse_hydrologique) -> réseau « ouvrage de tête » (carte) :

      1. Calcule la surface d'intersection (EPSG:26191, m²) entre `bv.geometrie` et
         chacun des carte.BassinVersant, et retient celui dont l'intersection est
         la plus grande.
      2. Mappe ce bassin vers son modèle ReseauOuvrageTete<X>.
      3. Renvoie le queryset des tronçons de ce réseau intersectant `bv.geometrie`.

    Filtrage grid faible / grid fort :
      - si `min_grid_code` est fourni (>= 0), il est appliqué tel quel (override) ;
      - sinon choix piloté par la SURFACE D'INTERSECTION : tout le réseau (grid faible
        inclus) si l'intersection BV∩bassin est < `seuil_intersection_m2`, sinon
        uniquement les drains forts grid_code >= 1 — avec repli sur tout le réseau
        si aucun drain fort n'existe dans le BV.

    Retourne (carte_bv, modele, queryset).
    (None, None, None) si pas de géométrie ou aucun bassin recoupé.
    (carte_bv, None, None) si le bassin retenu n'a pas de modèle associé.
    """
    if not bv.geometrie:
        return None, None, None

    meilleur_bv = None
    meilleure_aire = 0.0
    for cb in CarteBassinVersant.objects.exclude(geometrie__isnull=True):
        if not cb.geometrie.intersects(bv.geometrie):
            continue
        inter = cb.geometrie.intersection(bv.geometrie)
        if inter.empty:
            continue
        aire = inter.transform(SRID_NORD_MAROC, clone=True).area  # m²
        if aire > meilleure_aire:
            meilleure_aire = aire
            meilleur_bv = cb

    if meilleur_bv is None:
        return None, None, None

    modele = _RESEAU_OT_PAR_BASSIN.get(_normaliser_nom_bassin(meilleur_bv.nom))
    if modele is None:
        return meilleur_bv, None, None

    base = (
        modele.objects
        .filter(geometrie__bboverlaps=bv.geometrie)
        .filter(geometrie__intersects=bv.geometrie)
    )

    if min_grid_code is not None:
        return meilleur_bv, modele, base.filter(grid_code__gte=min_grid_code)

    # Choix selon la surface d'intersection déjà calculée : grande intersection
    # -> grid fort (>= 1) ; petite -> grid faible (tout le réseau).
    if meilleure_aire >= seuil_intersection_m2:
        forts = base.filter(grid_code__gte=1)
        # Repli sur grid faible si aucun drain fort (gros BV de plaine, tout en grid_code=0).
        if forts.exists():
            return meilleur_bv, modele, forts
    return meilleur_bv, modele, base


def _reseau_grid_max(bv):
    """Max grid_code du réseau « ouvrage de tête » du bassin apparié au BV (ou None).
    Sert à dimensionner le sélecteur « détail réseau » côté carte."""
    _, modele, _ = _reseau_ouvrage_tete_pour_bv(bv)
    if modele is None:
        return None
    from django.db.models import Max
    return modele.objects.aggregate(m=Max('grid_code'))['m']


@login_required
def detail_bv(request, pk):
    bv = get_object_or_404(BassinVersant, pk=pk)
    bv_obj = bv_to_hydro(bv)
    L_re, l_re = bv_obj.rectangle_equivalent()
    morpho = {
        'denivele':         round(bv_obj.denivele(), 2),
        'pente_pct':        round(bv_obj.pente() * 100, 4),
        'indice_gravelius': round(bv_obj.indice_gravelius(), 4),
        'rapport_forme':    round(bv_obj.rapport_de_forme(), 4),
        'L_rectangle':      round(L_re, 3),
        'l_rectangle':      round(l_re, 3),
        'pente_rect_pct':   round(bv_obj.pente_rectangle_equivalent() * 100, 4),
    }

    reseau_count = 0
    reseau_longueur_km = 0.0
    reseau_grid_max = None
    carte_bv, modele, reseau_qs = _reseau_ouvrage_tete_pour_bv(bv)
    if reseau_qs is not None:
        reseau_count = reseau_qs.count()
        if reseau_count:
            from django.contrib.gis.db.models.functions import Length, Transform
            from django.db.models import Sum
            agg = reseau_qs.annotate(
                l=Length(Transform('geometrie', SRID_NORD_MAROC))
            ).aggregate(total=Sum('l'))
            if agg['total']:
                reseau_longueur_km = round(agg['total'].km, 3)
    if modele is not None:
        from django.db.models import Max
        # Profondeur de hiérarchie du bassin -> borne haute du sélecteur grid_code.
        reseau_grid_max = modele.objects.aggregate(m=Max('grid_code'))['m']

    return render(request, 'analyse_hydrologique/bv/detail.html', {
        'bv':                 bv,
        'morpho':             morpho,
        'reseau_count':       reseau_count,
        'reseau_longueur_km': reseau_longueur_km,
        'reseau_bassin':      carte_bv.nom if carte_bv else None,
        'reseau_grid_max':    reseau_grid_max,
        'resultats':          bv.resultats_analyse.order_by('-date_analyse')[:5],
    })


@login_required
def bv_reseau_geojson(request, pk):
    """Réseau « ouvrage de tête » (app carte) du bassin le plus recouvrant,
    intersecté avec le BV, en GeoJSON."""
    bv = get_object_or_404(BassinVersant, pk=pk)
    # Override manuel facultatif ; sinon filtrage adaptatif (cf. helper).
    raw = request.GET.get('min_grid_code')
    min_gc = None
    if raw is not None:
        try:
            min_gc = int(raw)
        except ValueError:
            min_gc = None
    carte_bv, modele, qs = _reseau_ouvrage_tete_pour_bv(bv, min_grid_code=min_gc)
    if qs is None:
        return JsonResponse({'type': 'FeatureCollection', 'features': []})

    qs = qs.only('id', 'grid_code', 'geometrie')

    features = [
        {
            'type': 'Feature',
            'id': r.id,
            'properties': {'grid_code': r.grid_code},
            'geometry': json.loads(r.geometrie.geojson),
        }
        for r in qs
    ]
    return JsonResponse({'type': 'FeatureCollection', 'features': features})


# ESRI-specific WKIDs that GDAL returns but PROJ/GEOS can't transform directly.
# Maps ESRI code → standard EPSG equivalent.
_ESRI_TO_EPSG = {
    102191: 26191,  # Nord Maroc  (Clarke 1880 / Lambert Conformal Conic)
    102192: 26192,  # Sud Maroc
    102193: 26193,  # Sahara Maroc
    102100: 3857,   # Web Mercator (ESRI alias)
}

# Systèmes de coordonnées marocains usuels — proposés à l'utilisateur en cas de doute.
MOROCCO_CRS_CHOICES = [
    (4326,  "WGS 84 — géographique (degrés)"),
    (26191, "Merchich / Nord Maroc — Lambert (m)"),
    (26192, "Merchich / Sud Maroc — Lambert (m)"),
    (26194, "Merchich / Nord Maroc dégrées (m)"),
    (26195, "Merchich / Sud Maroc dégrées (m)"),
    (29701, "Nord Maroc historique (Voirol)"),
    (32629, "UTM zone 29 N — WGS84 (m)"),
    (32630, "UTM zone 30 N — WGS84 (m)"),
    (3857,  "Web Mercator (m)"),
]

# Enveloppe géographique du Maroc (WGS84). Sert de sanity check après reprojection.
MOROCCO_BBOX_4326 = (-17.5, 20.5, -0.5, 36.0)  # (lon_min, lat_min, lon_max, lat_max)


def _geom_in_morocco(geom):
    """True si l'enveloppe de la géométrie est dans la bbox du Maroc."""
    xmin, ymin, xmax, ymax = geom.extent
    return (
        MOROCCO_BBOX_4326[0] <= xmin <= MOROCCO_BBOX_4326[2] and
        MOROCCO_BBOX_4326[0] <= xmax <= MOROCCO_BBOX_4326[2] and
        MOROCCO_BBOX_4326[1] <= ymin <= MOROCCO_BBOX_4326[3] and
        MOROCCO_BBOX_4326[1] <= ymax <= MOROCCO_BBOX_4326[3]
    )


def _auto_correct_crs(raw_wkt, original_srid):
    """
    Méthode 3 (ArcGIS Pro « Define Projection ») appliquée automatiquement :
    essaie successivement chaque SRC candidat marocain, reprojette en 4326,
    et retourne le premier qui tombe dans la bbox du Maroc.

    Returns: (srid_correct, geom_4326) ou (None, None) si aucun candidat ne convient.
    """
    tried = []
    # On commence par le SRC d'origine puis on essaie les autres candidats marocains
    candidates = [original_srid] + [s for s, _ in MOROCCO_CRS_CHOICES if s != original_srid]
    for cand in candidates:
        try:
            g = GEOSGeometry(raw_wkt, srid=cand)
            if cand != 4326:
                g.transform(4326)
            tried.append(cand)
            if _geom_in_morocco(g):
                return cand, g, tried
        except Exception:
            tried.append(f"{cand}:err")
            continue
    return None, None, tried


def _zip_inventory(tmpdir):
    """Liste les extensions de fichiers présentes dans le ZIP décompressé."""
    found = set()
    for root, _, files in os.walk(tmpdir):
        for fn in files:
            ext = os.path.splitext(fn.lower())[1]
            if ext in ('.shp', '.dbf', '.shx', '.prj', '.cpg'):
                found.add(ext)
    return found


def _detect_srid(shp_path, layer):
    """
    Determine the SRID when GDAL returns 0.
    Strategy:
      1. Read the .prj file and look for known Moroccan projection keywords.
      2. Fall back to coordinate-range detection.
    Moroccan Lambert projections:
      EPSG:26191 — Nord Maroc  (Clarke 1880, central lat ~33.3°, Y ≈ 200k–600k)
      EPSG:26192 — Sud Maroc   (Clarke 1880, central lat ~29.7°, Y ≈ 0–300k)
    """
    # 1. .prj keyword scan
    prj_path = shp_path[:-4] + '.prj'
    if os.path.exists(prj_path):
        try:
            prj = open(prj_path, encoding='utf-8', errors='ignore').read().upper()
            if 'NORD_MAROC' in prj or 'NORD MAROC' in prj or '26191' in prj or '102191' in prj:
                return 26191
            if 'SUD_MAROC' in prj or 'SUD MAROC' in prj or '26192' in prj or '102192' in prj:
                return 26192
            # Generic Lambert Maroc — pick Nord by default
            if 'LAMBERT' in prj and 'MAROC' in prj:
                return 26191
            # UTM in Morocco: zones 28–30 N  (EPSG 326xx)
            if 'UTM' in prj:
                for zone in range(28, 31):
                    if f'ZONE {zone}' in prj or f'ZONE_{zone}' in prj:
                        return 32600 + zone
        except OSError:
            pass

    # 2. Coordinate-range detection
    env = layer.extent
    cx  = (env.min_x + env.max_x) / 2
    cy  = (env.min_y + env.max_y) / 2

    if -180 <= cx <= 180 and -90 <= cy <= 90:
        return 4326                              # WGS84 geographic
    if 100_000 <= cx <= 900_000 and 100_000 <= cy <= 700_000:
        return 26191 if cy >= 250_000 else 26192  # Moroccan Lambert
    if 100_000 <= cx <= 900_000 and -100_000 <= cy <= 100_000:
        return 26192                             # Sud Maroc
    # Last resort — assume WGS84
    return 4326


def _set_geom(instance, request, required=True):
    """Sets instance.geometrie from WKT in POST data. Returns error string or None."""
    wkt = request.POST.get('geometrie', '').strip()
    if wkt:
        try:
            instance.geometrie = GEOSGeometry(wkt, srid=4326)
        except Exception as e:
            return f"Géométrie invalide : {e}"
    elif required and not getattr(instance, 'pk', None):
        return "La géométrie est requise. Veuillez importer un fichier .shp."
    return None


@login_required
def creer_bv(request):
    form = BassinVersantForm(request.POST or None)
    if form.is_valid():
        instance = form.save(commit=False)
        geom_err = _set_geom(instance, request, required=False)
        if geom_err:
            messages.error(request, geom_err)
        else:
            instance.save()
            messages.success(request, "Bassin versant créé avec succès.")
            return redirect('liste_bv')
    return render(request, 'analyse_hydrologique/bv/form.html', {
        'form': form, 'titre': 'Nouveau bassin versant',
    })


@login_required
def modifier_bv(request, pk):
    bv = get_object_or_404(BassinVersant, pk=pk)
    form = BassinVersantForm(request.POST or None, instance=bv)
    if form.is_valid():
        instance = form.save(commit=False)
        geom_err = _set_geom(instance, request, required=False)
        if geom_err:
            messages.error(request, geom_err)
        else:
            instance.save()
            messages.success(request, "Bassin versant modifié.")
            return redirect('detail_bv', pk=pk)
    return render(request, 'analyse_hydrologique/bv/form.html', {
        'form': form, 'titre': f'Modifier — {bv.nom}',
    })


@login_required
def supprimer_bv(request, pk):
    bv = get_object_or_404(BassinVersant, pk=pk)
    if request.method == 'POST':
        bv.delete()
        messages.success(request, "Bassin versant supprimé.")
        return redirect('liste_bv')
    return render(request, 'analyse_hydrologique/confirmer_suppression.html', {
        'objet': bv, 'retour': 'liste_bv',
    })


@login_required
@require_POST
def supprimer_bv_multiple(request):
    """Supprimer plusieurs bassins versants via AJAX."""
    import json
    try:
        data = json.loads(request.body)
        ids = data.get('ids', [])
        if not ids:
            return JsonResponse({'error': 'Aucun bassin sélectionné.'}, status=400)

        count = BassinVersant.objects.filter(pk__in=ids).delete()[0]
        messages.success(request, f"{count} bassin(s) versant(s) supprimé(s).")
        return JsonResponse({'ok': True, 'deleted_count': count})
    except Exception as exc:
        return JsonResponse({'error': str(exc)}, status=400)
    


# =============================================================================
# Station Pluviométrique
# =============================================================================

@login_required
def liste_stations_pluvio(request):
    from django.db.models import Exists, OuterRef
    stations = StationPluviometrique.objects.annotate(
        has_montana=Exists(CoefficientMontana.objects.filter(station=OuterRef('pk')))
    )
    return render(request, 'analyse_hydrologique/pluvio/liste.html', {'stations': stations})


@login_required
def detail_station_pluvio(request, pk):
    station = get_object_or_404(StationPluviometrique, pk=pk)
    try:
        montana = station.coefficients_montana
    except CoefficientMontana.DoesNotExist:
        montana = None
    return render(request, 'analyse_hydrologique/pluvio/detail.html', {
        'station': station, 'montana': montana,
    })


@login_required
def creer_station_pluvio(request):
    form = StationPluviometriqueForm(request.POST or None)
    if form.is_valid():
        instance = form.save(commit=False)
        geom_err = _set_geom(instance, request, required=False)
        if geom_err:
            messages.error(request, geom_err)
        else:
            instance.save()
            messages.success(request, "Station pluviométrique créée.")
            return redirect('liste_stations_pluvio')
    return render(request, 'analyse_hydrologique/pluvio/form.html', {
        'form': form, 'titre': 'Nouvelle station pluviométrique',
    })


@login_required
def modifier_station_pluvio(request, pk):
    station = get_object_or_404(StationPluviometrique, pk=pk)
    form = StationPluviometriqueForm(request.POST or None, instance=station)
    if form.is_valid():
        instance = form.save(commit=False)
        geom_err = _set_geom(instance, request, required=False)
        if geom_err:
            messages.error(request, geom_err)
        else:
            instance.save()
            messages.success(request, "Station modifiée.")
            return redirect('detail_station_pluvio', pk=pk)
    return render(request, 'analyse_hydrologique/pluvio/form.html', {
        'form': form, 'titre': f'Modifier — {station.nom}',
    })


@login_required
def supprimer_station_pluvio(request, pk):
    station = get_object_or_404(StationPluviometrique, pk=pk)
    if request.method == 'POST':
        station.delete()
        messages.success(request, "Station supprimée.")
        return redirect('liste_stations_pluvio')
    return render(request, 'analyse_hydrologique/confirmer_suppression.html', {
        'objet': station, 'retour': 'liste_stations_pluvio',
    })


# =============================================================================
# Station Hydrométrique
# =============================================================================

@login_required
def liste_stations_hydro(request):
    return render(request, 'analyse_hydrologique/hydro/liste.html', {
        'stations': StationHydrometrique.objects.all(),
    })


_MOIS_HYDRO = ['Sep', 'Oct', 'Nov', 'Dec', 'Jan', 'Fev', 'Mar', 'Avr', 'Mai', 'Juin', 'Jul', 'Aou']


@login_required
def detail_station_hydro(request, pk):
    station = get_object_or_404(StationHydrometrique, pk=pk)
    lignes_tableau = [
        ("Debits annee normale (m3/s)", station.debits_mensuels_annee_normale or []),
        ("Frequence annee normale (jours)", station.frequences_mensuelles_annee_normale or []),
        ("Debits annee humide (m3/s)", station.debits_mensuels_annee_humide or []),
        ("Frequence annee humide (jours)", station.frequences_mensuelles_annee_humide or []),
    ]
    return render(request, 'analyse_hydrologique/hydro/detail.html', {
        'station': station,
        'mois_hydro': _MOIS_HYDRO,
        'lignes_tableau_debits': lignes_tableau,
    })


@login_required
def creer_station_hydro(request):
    form = StationHydrometriqueForm(request.POST or None)
    if form.is_valid():
        instance = form.save(commit=False)
        geom_err = _set_geom(instance, request, required=False)
        if geom_err:
            messages.error(request, geom_err)
        else:
            instance.save()
            messages.success(request, "Station hydrométrique créée.")
            return redirect('liste_stations_hydro')
    return render(request, 'analyse_hydrologique/hydro/form.html', {
        'form': form, 'titre': 'Nouvelle station hydrométrique',
    })


@login_required
def modifier_station_hydro(request, pk):
    station = get_object_or_404(StationHydrometrique, pk=pk)
    form = StationHydrometriqueForm(request.POST or None, instance=station)
    if form.is_valid():
        instance = form.save(commit=False)
        geom_err = _set_geom(instance, request, required=False)
        if geom_err:
            messages.error(request, geom_err)
        else:
            instance.save()
            messages.success(request, "Station modifiée.")
            return redirect('detail_station_hydro', pk=pk)
    return render(request, 'analyse_hydrologique/hydro/form.html', {
        'form': form, 'titre': f'Modifier — {station.nom}',
    })


@login_required
def supprimer_station_hydro(request, pk):
    station = get_object_or_404(StationHydrometrique, pk=pk)
    if request.method == 'POST':
        station.delete()
        messages.success(request, "Station supprimée.")
        return redirect('liste_stations_hydro')
    return render(request, 'analyse_hydrologique/confirmer_suppression.html', {
        'objet': station, 'retour': 'liste_stations_hydro',
    })



# =============================================================================
# Coefficients Montana
# =============================================================================

@login_required
def liste_coefficients(request):
    return render(request, 'analyse_hydrologique/montana/liste.html', {
        'coefficients': CoefficientMontana.objects.select_related('station').all(),
    })


@login_required
def creer_coefficient(request):
    form = CoefficientMontanaForm(request.POST or None)
    if form.is_valid():
        form.save()
        messages.success(request, "Coefficients Montana enregistrés.")
        return redirect('liste_coefficients')
    return render(request, 'analyse_hydrologique/montana/form.html', {
        'form': form, 'titre': 'Nouveaux coefficients Montana',
    })


@login_required
def modifier_coefficient(request, pk):
    coef = get_object_or_404(CoefficientMontana, pk=pk)
    form = CoefficientMontanaForm(request.POST or None, instance=coef)
    if form.is_valid():
        form.save()
        messages.success(request, "Coefficients modifiés.")
        return redirect('liste_coefficients')
    return render(request, 'analyse_hydrologique/montana/form.html', {
        'form': form, 'titre': str(coef),
    })


@login_required
def supprimer_coefficient(request, pk):
    coef = get_object_or_404(CoefficientMontana, pk=pk)
    if request.method == 'POST':
        coef.delete()
        messages.success(request, "Coefficients supprimés.")
        return redirect('liste_coefficients')
    return render(request, 'analyse_hydrologique/confirmer_suppression.html', {
        'objet': coef, 'retour': 'liste_coefficients',
    })


# =============================================================================
# Analyse Hydrologique
# =============================================================================

@login_required
def liste_analyses(request):
    return render(request, 'analyse_hydrologique/analyse/liste.html', {
        'analyses': (
            ResultatAnalyseHydrologique.objects
            .select_related('operateur', 'bassin_versant')
            .order_by('-date_analyse')
        ),
    })


@login_required
def lancer_analyse(request, bv_id):
    bv_model = get_object_or_404(BassinVersant, pk=bv_id)
    form = AnalyseParametresForm(request.POST or None)
    reseau_grid_max = _reseau_grid_max(bv_model)

    if request.method == 'POST' and form.is_valid():
        d        = form.cleaned_data
        sp_model = d['station_pluvio']
        sh_model = d.get('station_hydro')

        # Vérifier les coefficients Montana de la station choisie
        try:
            montana = sp_model.coefficients_montana
        except CoefficientMontana.DoesNotExist:
            messages.error(
                request,
                f"Aucun coefficient Montana pour '{sp_model.nom}'. "
                "Veuillez les saisir d'abord.",
            )
            return render(request, 'analyse_hydrologique/analyse/lancer.html', {
                'bv': bv_model, 'form': form, 'reseau_grid_max': reseau_grid_max,
            })

        champs_montana = [montana.a10, montana.a20, montana.a50, montana.a100,
                          montana.b10, montana.b20, montana.b50, montana.b100]
        if None in champs_montana:
            messages.error(
                request,
                f"Coefficients Montana incomplets pour '{sp_model.nom}' "
                "(tous les a/b pour T=10, 20, 50, 100 sont requis).",
            )
            return render(request, 'analyse_hydrologique/analyse/lancer.html', {
                'bv': bv_model, 'form': form, 'reseau_grid_max': reseau_grid_max,
            })

        # Lancer le calcul
        try:
            resultats = run_analyse(bv_model, sp_model, sh_model, d)
        except Exception as exc:
            messages.error(request, f"Erreur de calcul : {exc}")
            return render(request, 'analyse_hydrologique/analyse/lancer.html', {
                'bv': bv_model, 'form': form, 'reseau_grid_max': reseau_grid_max,
            })

        # Persister en base
        q = resultats['q_finaux']
        obj = ResultatAnalyseHydrologique.objects.create(
            operateur=request.user,
            bassin_versant=bv_model,
            methode='gradex',
            qcrue_t10=q.get(10),
            qcrue_t20=q.get(20),
            qcrue_t50=q.get(50),
            qcrue_t100=q.get(100),
            temps_concentration=resultats['Tc_min'],
            coefficient_ruissellement=d['C_rationnel'],
            observations=d.get('observations', ''),
            conclusions=d.get('conclusions', ''),
            details_calcul=_json_safe(resultats),
        )

        # Stocker les détails intermédiaires en session pour la page résultat
        request.session[f'analyse_{obj.pk}'] = resultats

        return redirect('resultat_analyse', pk=obj.pk)

    return render(request, 'analyse_hydrologique/analyse/lancer.html', {
        'bv': bv_model, 'form': form, 'reseau_grid_max': reseau_grid_max,
    })


@login_required
def resultat_analyse(request, pk):
    resultat = get_object_or_404(ResultatAnalyseHydrologique, pk=pk)
    details  = request.session.pop(f'analyse_{pk}', None) or resultat.details_calcul or {}

    station_hydro = None
    station_hydro_id = details.get('station_hydro_id')
    station_hydro_nom = details.get('station_hydro_nom') or details.get('station_hydro')
    if station_hydro_id:
        station_hydro = StationHydrometrique.objects.filter(pk=station_hydro_id).first()
    if station_hydro is None and station_hydro_nom and station_hydro_nom != '—':
        station_hydro = StationHydrometrique.objects.filter(nom=station_hydro_nom).first()

    tc_h = details.get('Tc_h')
    if tc_h is None and resultat.temps_concentration:
        tc_h = float(resultat.temps_concentration) / 60.0

    crue_graph = calculer_apports_crue_sans_prelevement(
        station_hydro, tc_h,
        getattr(resultat.bassin_versant, 'surface', None),
    ) if station_hydro and tc_h else None
    return render(request, 'analyse_hydrologique/analyse/resultat.html', {
        'resultat': resultat,
        'details':  details,
        'periodes': PERIODES,
        'crue_graph': crue_graph,
        'reseau_grid_max': _reseau_grid_max(resultat.bassin_versant),
    })


@login_required
@require_POST
def sauvegarder_annotations(request, pk):
    resultat = get_object_or_404(ResultatAnalyseHydrologique, pk=pk)
    try:
        data = json.loads(request.body)
        resultat.observations = data.get('observations', '')
        resultat.conclusions  = data.get('conclusions', '')
        resultat.save(update_fields=['observations', 'conclusions'])
        return JsonResponse({'ok': True})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=400)


@login_required
@require_POST
def valider_analyse(request, pk):
    resultat = get_object_or_404(ResultatAnalyseHydrologique, pk=pk)
    resultat.statut = 'valide'
    resultat.save(update_fields=['statut'])
    return JsonResponse({'ok': True, 'statut': 'valide'})


@login_required
@require_POST
def recalculer_analyse(request, pk):
    """
    AJAX : recalcule les résultats selon une personnalisation Tc ou Q.

    Body JSON :
      { "mode": "tc", "formules_tc": ["Kirpich", ...] }
      { "mode": "formules_q", "formules_q_incluses": ["Rationnelle", "Gradex", ...] }
    """
    resultat = get_object_or_404(ResultatAnalyseHydrologique, pk=pk)
    dc = resultat.details_calcul
    if not dc:
        return JsonResponse(
            {'error': 'Aucun détail de calcul disponible — veuillez relancer l\'analyse.'},
            status=400,
        )
    try:
        body = json.loads(request.body)
        mode = body.get('mode')

        if mode == 'tc':
            formules_tc = body.get('formules_tc', [])
            result = recalculer_depuis_tc(dc, formules_tc)
        elif mode == 'formules_q':
            formules_q_incluses = body.get('formules_q_incluses', [])
            result = recalculer_depuis_formules_q(dc, formules_q_incluses)
        else:
            return JsonResponse({'error': 'Mode invalide.'}, status=400)

        if result is None:
            return JsonResponse(
                {'error': 'Recalcul impossible — paramètres manquants dans cette analyse.'},
                status=400,
            )
        return JsonResponse(_json_safe(result))
    except Exception as exc:
        return JsonResponse({'error': str(exc)}, status=400)


@login_required
def exporter_excel(request, pk):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    r  = get_object_or_404(ResultatAnalyseHydrologique, pk=pk)
    bv = r.bassin_versant
    dc = r.details_calcul  # peut être None pour anciennes analyses

    # Station hydro pour la feuille Apports mensuels (même logique que resultat_analyse)
    _sh_id = dc.get('station_hydro_id') if dc else None
    _sh_nom = (dc.get('station_hydro_nom') or dc.get('station_hydro')) if dc else None
    _sh = StationHydrometrique.objects.filter(pk=_sh_id).first() if _sh_id else None
    if _sh is None and _sh_nom and _sh_nom != '—':
        _sh = StationHydrometrique.objects.filter(nom=_sh_nom).first()
    _tc_h = dc.get('Tc_h') if dc else None
    if _tc_h is None and r.temps_concentration:
        _tc_h = float(r.temps_concentration) / 60.0
    crue_data = calculer_apports_crue_sans_prelevement(
        _sh, _tc_h, getattr(bv, 'surface', None),
    ) if (_sh and _tc_h) else None

    wb = Workbook()

    # ── styles partagés ───────────────────────────────────────────────────────
    C_DARK, C_GOLD, C_LIGHT, C_GREEN, C_GREY = "1A1A2E", "F0A500", "FFFBF4", "D4EDDA", "F8F4EE"
    thin   = Side(style="thin", color="D0C8BC")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    F_TITLE = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    F_HEAD  = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    F_LABEL = Font(name="Calibri", bold=True, size=10, color=C_DARK)
    F_VALUE = Font(name="Calibri",             size=10, color=C_DARK)
    F_GOLD  = Font(name="Calibri", bold=True, size=11, color=C_GOLD)
    F_AVG   = Font(name="Calibri", bold=True, size=10, color="856404")

    FILL_DARK  = PatternFill("solid", fgColor=C_DARK)
    FILL_LIGHT = PatternFill("solid", fgColor=C_LIGHT)
    FILL_GREEN = PatternFill("solid", fgColor=C_GREEN)
    FILL_GREY  = PatternFill("solid", fgColor=C_GREY)
    FILL_AVG   = PatternFill("solid", fgColor="FFF3CD")
    FILL_FINAL = PatternFill("solid", fgColor="D4EDDA")

    def title_row(ws, text, ncols, row=1, height=26):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        c = ws.cell(row=row, column=1, value=text)
        c.font = F_TITLE; c.fill = FILL_DARK; c.alignment = CENTER
        ws.row_dimensions[row].height = height

    def section_row(ws, text, ncols, row, height=20):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        c = ws.cell(row=row, column=1, value=text)
        c.font = F_HEAD; c.fill = FILL_DARK; c.alignment = CENTER
        ws.row_dimensions[row].height = height

    def lbl(ws, row, col, text, fill=None):
        c = ws.cell(row=row, column=col, value=text)
        c.font = F_LABEL; c.fill = fill or FILL_GREY; c.alignment = LEFT; c.border = BORDER

    def val(ws, row, col, v, fill=None, gold=False, avg=False):
        c = ws.cell(row=row, column=col, value=v)
        c.font = F_AVG if avg else (F_GOLD if gold else F_VALUE)
        c.fill = fill or PatternFill(); c.alignment = CENTER; c.border = BORDER

    def head(ws, row, col, text, ncols=1):
        if ncols > 1:
            ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + ncols - 1)
        c = ws.cell(row=row, column=col, value=text)
        c.font = F_HEAD; c.fill = FILL_DARK; c.alignment = CENTER; c.border = BORDER

    fmt = lambda v: round(v, 3) if v is not None else "—"

    # ══════════════════════════════════════════════════════════════════════════
    # Feuille 1 — Résumé général
    # ══════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Résumé"
    ws1.column_dimensions["A"].width = 38
    ws1.column_dimensions["B"].width = 26

    title_row(ws1, "RAPPORT D'ANALYSE HYDROLOGIQUE — HydroPlan SIG", 2)

    r_ = 3
    section_row(ws1, "Informations générales", 2, r_)
    infos = [
        ("Bassin versant",              str(bv) if bv else "—"),
        ("Date d'analyse",              r.date_analyse.strftime("%d/%m/%Y  %H:%M") if r.date_analyse else "—"),
        ("Opérateur",                   str(r.operateur) if r.operateur else "—"),
        ("Station pluviométrique",      dc.get("station_pluvio", "—") if dc else "—"),
        ("Station hydrométrique",       dc.get("station_hydro",  "—") if dc else "—"),
        ("Temps de concentration (Tc)", f"{r.temps_concentration:.2f} min" if r.temps_concentration else "—"),
        ("Coeff. de ruissellement (C)", f"{r.coefficient_ruissellement:.3f}" if r.coefficient_ruissellement else "—"),
        ("Référence analyse",           f"Analyse #{pk}"),
    ]
    for lbl_txt, val_txt in infos:
        r_ += 1
        lbl(ws1, r_, 1, lbl_txt)
        val(ws1, r_, 2, val_txt)

    if bv:
        r_ += 2
        section_row(ws1, "Paramètres détaillés du bassin versant", 2, r_)
        for param, valeur, unite in _bassin_versant_export_rows(bv):
            r_ += 1
            lbl(ws1, r_, 1, param)
            val_txt = f"{valeur} {unite}" if unite and unite != "—" else str(valeur)
            val(ws1, r_, 2, val_txt)

    r_ += 2
    section_row(ws1, "Débits de crue retenus (m³/s)", 2, r_)
    for lbl_txt, qv in [("Q crue T = 10 ans", r.qcrue_t10), ("Q crue T = 20 ans", r.qcrue_t20),
                         ("Q crue T = 50 ans", r.qcrue_t50), ("Q crue T = 100 ans", r.qcrue_t100)]:
        r_ += 1
        lbl(ws1, r_, 1, lbl_txt, fill=FILL_LIGHT)
        val(ws1, r_, 2, fmt(qv), fill=FILL_GREEN, gold=True)

    if r.observations or r.conclusions:
        r_ += 2
        section_row(ws1, "Annotations", 2, r_)
        for annot_lbl, annot_txt in [("Observations", r.observations),
                                      ("Conclusions et recommandations", r.conclusions)]:
            if annot_txt:
                r_ += 1
                lbl(ws1, r_, 1, annot_lbl)
                c = ws1.cell(row=r_, column=2, value=annot_txt)
                c.font = F_VALUE; c.alignment = LEFT; c.border = BORDER
                ws1.row_dimensions[r_].height = max(16, len(annot_txt) // 3)

    # ══════════════════════════════════════════════════════════════════════════
    # Feuille 2 — Temps de concentration
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Temps de concentration")
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 18

    title_row(ws2, "Temps de concentration Tc — Par formule", 3)
    head(ws2, 3, 1, "Formule"); head(ws2, 3, 2, "Tc (min)"); head(ws2, 3, 3, "Tc (heures)")

    if dc and dc.get("tc"):
        row_tc = 4
        for formule, tc_val in dc["tc"].items():
            if formule == "Moyenne":
                continue
            lbl(ws2, row_tc, 1, formule)
            val(ws2, row_tc, 2, fmt(tc_val))
            val(ws2, row_tc, 3, fmt(tc_val / 60) if tc_val else "—")
            row_tc += 1
        # Ligne moyenne
        lbl(ws2, row_tc, 1, "Moyenne retenue (Tc)", fill=FILL_AVG)
        val(ws2, row_tc, 2, fmt(dc["tc"].get("Moyenne") or r.temps_concentration), fill=FILL_AVG, avg=True)
        val(ws2, row_tc, 3, fmt((dc["tc"].get("Moyenne") or r.temps_concentration or 0) / 60), fill=FILL_AVG, avg=True)
    else:
        ws2.cell(row=4, column=1, value="Détails non disponibles — relancer l'analyse pour les obtenir.").font = F_VALUE

    # ══════════════════════════════════════════════════════════════════════════
    # Feuille 3 — Intensités et pluies de projet
    # ══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Intensités et Pluies")
    ws3.column_dimensions["A"].width = 24
    for col, w in zip("BCDE", [18, 18, 18, 18]):
        ws3.column_dimensions[col].width = w

    title_row(ws3, "Intensités Montana et Pluies de projet Pj24h", 5)
    for col_i, T in enumerate(PERIODES, start=2):
        head(ws3, 3, col_i, f"T = {T} ans")
    head(ws3, 3, 1, "Paramètre")

    if dc:
        intensites = dc.get("intensites", {})
        pj24h      = dc.get("pj24h", {})
        for row_i, (row_lbl, dct) in enumerate(
            [("Intensité Montana I (mm/h)", intensites), ("Pluie journalière Pj24h (mm)", pj24h)],
            start=4
        ):
            lbl(ws3, row_i, 1, row_lbl)
            for col_i, T in enumerate(PERIODES, start=2):
                v = _d(dct, T)
                val(ws3, row_i, col_i, fmt(v))
    else:
        ws3.cell(row=4, column=1, value="Détails non disponibles.").font = F_VALUE

    # ══════════════════════════════════════════════════════════════════════════
    # Feuille 4 — Débits par formule (matrice complète)
    # ══════════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Débits par formule")
    ws4.column_dimensions["A"].width = 26
    for col, w in zip("BCDE", [16, 16, 16, 16]):
        ws4.column_dimensions[col].width = w

    title_row(ws4, "Débits de crue Q (m³/s) — Par formule et par période de retour", 5)
    head(ws4, 3, 1, "Formule")
    for col_i, T in enumerate(PERIODES, start=2):
        head(ws4, 3, col_i, f"T = {T} ans")

    fills_risk = {10: PatternFill("solid", fgColor="EBF5FB"),
                  20: PatternFill("solid", fgColor="FEF9E7"),
                  50: PatternFill("solid", fgColor="FDF2E9"),
                  100: PatternFill("solid", fgColor="FDEDEC")}

    if dc and dc.get("debits"):
        row_q = 4
        formules_incluses = set(dc.get("formules_q_incluses", []))
        for formule, qd in dc["debits"].items():
            incl = formule in formules_incluses
            lbl(ws4, row_q, 1, formule + (" ✓" if incl else ""), fill=FILL_GREY if incl else PatternFill())
            for col_i, T in enumerate(PERIODES, start=2):
                val(ws4, row_q, col_i, fmt(_d(qd, T)), fill=fills_risk[T] if incl else PatternFill())
            row_q += 1

        # Gradex
        if dc.get("debits_gradex"):
            lbl(ws4, row_q, 1, "Gradex", fill=PatternFill("solid", fgColor="D6EAF8"))
            for col_i, T in enumerate(PERIODES, start=2):
                val(ws4, row_q, col_i, fmt(_d(dc["debits_gradex"], T)),
                    fill=PatternFill("solid", fgColor="D6EAF8"))
            row_q += 1

        # Ligne finale retenue
        lbl(ws4, row_q, 1, "Débit retenu — moyenne (m³/s)", fill=FILL_FINAL)
        for col_i, T in enumerate(PERIODES, start=2):
            qv = _d(dc.get("q_finaux"), T)
            val(ws4, row_q, col_i, fmt(qv), fill=FILL_FINAL, gold=True)

        # Note formules incluses
        row_q += 2
        note = ws4.cell(row=row_q, column=1,
            value=f"✓ Formules incluses dans la moyenne : {', '.join(dc.get('formules_q_incluses', []))}")
        note.font = Font(name="Calibri", italic=True, size=9, color="888888")
    else:
        ws4.cell(row=4, column=1, value="Détails non disponibles.").font = F_VALUE

    # ══════════════════════════════════════════════════════════════════════════
    # Feuille 5 — Morphologie du bassin versant
    # ══════════════════════════════════════════════════════════════════════════
    ws5 = wb.create_sheet("Bassin versant")
    ws5.column_dimensions["A"].width = 38
    ws5.column_dimensions["B"].width = 20
    ws5.column_dimensions["C"].width = 10

    title_row(ws5, f"Paramètres détaillés du bassin versant — {bv if bv else 'Bassin versant'}", 3)
    head(ws5, 3, 1, "Paramètre"); head(ws5, 3, 2, "Valeur"); head(ws5, 3, 3, "Unité")

    if bv:
        for i, (param, valeur, unite) in enumerate(_bassin_versant_export_rows(bv), start=4):
            fl = FILL_LIGHT if i % 2 == 0 else PatternFill()
            lbl(ws5, i, 1, param, fill=FILL_GREY)
            val(ws5, i, 2, valeur, fill=fl)
            val(ws5, i, 3, unite,  fill=fl)
            ws5.row_dimensions[i].height = 16
    else:
        ws5.cell(row=4, column=1, value="Aucun bassin versant associé.").font = F_VALUE

    # ══════════════════════════════════════════════════════════════════════════
    # Feuille 6 — Paramètres de calcul
    # ══════════════════════════════════════════════════════════════════════════
    ws6 = wb.create_sheet("Paramètres de calcul")
    ws6.column_dimensions["A"].width = 42
    ws6.column_dimensions["B"].width = 22

    title_row(ws6, "Paramètres de calcul utilisés", 2)

    def fmt_list(lst):
        if not lst:
            return "—"
        return ", ".join(str(x) for x in lst)

    p_rows = []
    if dc:
        p_rows = [
            ("── Formules sélectionnées ──", ""),
            ("Formules Tc incluses",         fmt_list(dc.get("formules_tc_incluses"))),
            ("Formules Q incluses",          fmt_list(dc.get("formules_q_incluses"))),
            ("── Paramètres des formules ──", ""),
            ("C — coeff. ruissellement (Rationnelle)",
             dc.get("C_rationnel", r.coefficient_ruissellement or "—")),
            ("K (Mac-Math)",                 dc.get("K_macmath", "—")),
            ("A (Fuller II)",                dc.get("A_fuller", "—")),
            ("N (Fuller II)",                dc.get("N_fuller", "—")),
            ("k (Mallet-Gauthier)",          dc.get("k_mallet", "—")),
            ("a (Mallet-Gauthier)",          dc.get("a_mallet", "—")),
            ("K1 (Hazen-Lazervic)",          dc.get("K1_hl", "—")),
            ("K2 (Hazen-Lazervic)",          dc.get("K2_hl", "—")),
            ("a exposant (Hazen-Lazervic)",  dc.get("a_hl", "—")),
            ("── Coefficients Montana ──", ""),
            ("a  [T=10, 20, 50, 100]",       fmt_list(dc.get("montana_a"))),
            ("b  [T=10, 20, 50, 100]",       fmt_list(dc.get("montana_b"))),
            ("Gradex (g)",                   dc.get("grad_exp_pluie", "—")),
        ]
    else:
        p_rows = [("Paramètres non disponibles — relancer l'analyse.", "")]

    section_row(ws6, "Paramètres", 2, 3)
    r_p = 4
    for lbl_txt, val_txt in p_rows:
        if str(lbl_txt).startswith("──"):
            section_row(ws6, lbl_txt.replace("──", "").strip(), 2, r_p, height=16)
        else:
            lbl(ws6, r_p, 1, lbl_txt)
            val(ws6, r_p, 2, val_txt)
        r_p += 1

    # ══════════════════════════════════════════════════════════════════════════
    # Feuille 7 — Apports mensuels de crue (3 scénarios Sep→Aoû)
    # ══════════════════════════════════════════════════════════════════════════
    ws7 = wb.create_sheet("Apports mensuels")
    ws7.column_dimensions["A"].width = 24
    for _col in ["B","C","D","E","F","G","H","I","J","K","L","M"]:
        ws7.column_dimensions[_col].width = 10
    ws7.column_dimensions["N"].width = 14

    title_row(ws7, "Apports mensuels de crue par bassin versant (m³)", 14)
    head(ws7, 3, 1, "Scénario")
    for _ci, _m in enumerate(["Sep","Oct","Nov","Déc","Jan","Fév","Mar","Avr","Mai","Jun","Jul","Aoû"], start=2):
        head(ws7, 3, _ci, _m)
    head(ws7, 3, 14, "Total (m³)")

    if crue_data:
        _SCEN = [
            ("normale", "Année normale", PatternFill("solid", fgColor="FFF8EE")),
            ("humide",  "Année humide",  PatternFill("solid", fgColor="EBF5FB")),
            ("seche",   "Année sèche",   PatternFill("solid", fgColor="FDEDEC")),
        ]
        for _ri, (_sk, _slbl, _sfill) in enumerate(_SCEN, start=4):
            _sd = crue_data.get(_sk, {})
            _vols = _sd.get('volumes_m3', [0.0] * 12)
            _tot  = _sd.get('total_m3', sum(v for v in _vols if v))
            lbl(ws7, _ri, 1, _slbl, fill=_sfill)
            for _ci, _v in enumerate(_vols[:12], start=2):
                val(ws7, _ri, _ci, round(_v, 0) if _v else 0, fill=_sfill)
            val(ws7, _ri, 14, round(_tot, 0), fill=_sfill, gold=True)
            ws7.row_dimensions[_ri].height = 18
        # Ligne débits m³/s (info complémentaire)
        head(ws7, 8, 1, "Débit Qp moyen (m³/s) — Année normale")
        _nd = crue_data.get('normale', {})
        for _ci, _v in enumerate(_nd.get('debits_m3s', [])[:12], start=2):
            val(ws7, 8, _ci, round(_v, 3) if _v else 0)
        # Metadata
        ws7.cell(row=10, column=1,
            value=f"Tc utilisé : {crue_data.get('tc_h', '—')} h  |  Station hydro : {_sh_nom or '—'}").font = \
            Font(name="Calibri", italic=True, size=9, color="888888")
    else:
        ws7.cell(row=4, column=1,
            value="Aucune station hydrométrique associée — relancer l'analyse avec une station hydro.").font = \
            Font(name="Calibri", italic=True, size=9, color="888888")

    # ── Réponse HTTP ──────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    slug = f"analyse_{pk}_{date.today().strftime('%Y%m%d')}.xlsx"
    resp = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{slug}"'
    return resp


@login_required
def exporter_pdf(request, pk):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable,
    )

    r  = get_object_or_404(ResultatAnalyseHydrologique, pk=pk)
    bv = r.bassin_versant
    dc = r.details_calcul

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=1.8*cm, rightMargin=1.8*cm,
        topMargin=1.8*cm,  bottomMargin=1.8*cm,
        title=f"Analyse hydrologique — {bv}",
    )

    PW = 16.9 * cm  # usable page width

    # ── couleurs ──────────────────────────────────────────────────────────────
    DARK  = colors.HexColor("#1A1A2E")
    GOLD  = colors.HexColor("#F0A500")
    GREEN = colors.HexColor("#D4EDDA")
    GREY  = colors.HexColor("#F8F4EE")
    LIGHT = colors.HexColor("#FFFBF4")
    AVG   = colors.HexColor("#FFF3CD")
    FINAL = colors.HexColor("#D4EDDA")
    RED   = colors.HexColor("#FADBD8")
    ORAN  = colors.HexColor("#FDEBD0")

    # ── styles texte ──────────────────────────────────────────────────────────
    S_WHITE  = ParagraphStyle("w",  fontSize=14, fontName="Helvetica-Bold",
                               textColor=colors.white, alignment=1)
    S_HEAD   = ParagraphStyle("h",  fontSize=11, fontName="Helvetica-Bold",
                               textColor=colors.white, alignment=1)
    S_LABEL  = ParagraphStyle("lb", fontSize=8.5, fontName="Helvetica-Bold",  textColor=DARK)
    S_VALUE  = ParagraphStyle("v",  fontSize=8.5, fontName="Helvetica",        textColor=colors.HexColor("#333333"))
    S_GOLD   = ParagraphStyle("g",  fontSize=11,  fontName="Helvetica-Bold",   textColor=GOLD, alignment=1)
    S_AVG    = ParagraphStyle("av", fontSize=9,   fontName="Helvetica-Bold",   textColor=colors.HexColor("#856404"), alignment=1)
    S_SMALL  = ParagraphStyle("sm", fontSize=7.5, fontName="Helvetica",        textColor=colors.grey)

    PAD = [("LEFTPADDING",(0,0),(-1,-1),5), ("RIGHTPADDING",(0,0),(-1,-1),5),
           ("TOPPADDING", (0,0),(-1,-1),4), ("BOTTOMPADDING",(0,0),(-1,-1),4),
           ("VALIGN",     (0,0),(-1,-1),"MIDDLE")]

    def sec(text):
        t = Table([[Paragraph(text, S_HEAD)]], colWidths=[PW])
        t.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),DARK),
                                ("TOPPADDING",(0,0),(-1,-1),6),
                                ("BOTTOMPADDING",(0,0),(-1,-1),6)]))
        return t

    def kv(rows, w1=6*cm):
        data = [[Paragraph(k, S_LABEL), Paragraph(str(v), S_VALUE)] for k,v in rows]
        t = Table(data, colWidths=[w1, PW - w1])
        t.setStyle(TableStyle([
            *PAD,
            ("BACKGROUND",(0,0),(0,-1),GREY),
            ("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#E0D0C0")),
        ]))
        return t

    def grid_hdr(cols):
        return [Paragraph(c, ParagraphStyle("gh", fontSize=8, fontName="Helvetica-Bold",
                textColor=colors.white, alignment=1)) for c in cols]

    fmt = lambda v: f"{v:.3f}" if v is not None else "—"
    story = []

    # ══════════════════════════════════════════════════════════════════════════
    # BANDEAU TITRE
    # ══════════════════════════════════════════════════════════════════════════
    t = Table([[Paragraph("RAPPORT D'ANALYSE HYDROLOGIQUE", S_WHITE),
                Paragraph(f"Analyse #{pk}", S_SMALL)]],
              colWidths=[PW * 0.75, PW * 0.25])
    t.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),DARK),
                            *PAD,
                            ("TOPPADDING",(0,0),(-1,-1),10),
                            ("BOTTOMPADDING",(0,0),(-1,-1),10),
                            ("ALIGN",(-1,0),(-1,0),"RIGHT"),
                            ("VALIGN",(0,0),(-1,-1),"MIDDLE")]))
    story += [t, Spacer(1, .35*cm)]

    # ══════════════════════════════════════════════════════════════════════════
    # 1. INFORMATIONS GÉNÉRALES
    # ══════════════════════════════════════════════════════════════════════════
    story += [sec("1. Informations générales"), Spacer(1, .1*cm)]
    story.append(kv([
        ("Bassin versant",              str(bv) if bv else "—"),
        ("Date d'analyse",              r.date_analyse.strftime("%d/%m/%Y  %H:%M") if r.date_analyse else "—"),
        ("Opérateur",                   str(r.operateur) if r.operateur else "—"),
        ("Station pluviométrique",      dc.get("station_pluvio","—") if dc else "—"),
        ("Station hydrométrique",       dc.get("station_hydro","—")  if dc else "—"),
        ("Temps de concentration Tc",   f"{r.temps_concentration:.2f} min" if r.temps_concentration else "—"),
        ("Coeff. de ruissellement C",   f"{r.coefficient_ruissellement:.3f}" if r.coefficient_ruissellement else "—"),
    ]))

    # ══════════════════════════════════════════════════════════════════════════
    # 2. PARAMÈTRES DE CALCUL
    # ══════════════════════════════════════════════════════════════════════════
    story += [Spacer(1, .3*cm), sec("2. Paramètres de calcul utilisés"), Spacer(1, .1*cm)]
    if dc:
        def _fmtlist(lst):
            return ", ".join(str(x) for x in lst) if lst else "—"

        param_rows = [
            ("Formules Tc incluses",                    _fmtlist(dc.get("formules_tc_incluses"))),
            ("Formules Q incluses",                     _fmtlist(dc.get("formules_q_incluses"))),
            ("C — coeff. ruissellement (Rationnelle)",  str(dc.get("C_rationnel", r.coefficient_ruissellement or "—"))),
            ("K (Mac-Math)",                            str(dc.get("K_macmath", "—"))),
            ("A / N (Fuller II)",                       f"{dc.get('A_fuller','—')} / {dc.get('N_fuller','—')}"),
            ("k / a (Mallet-Gauthier)",                 f"{dc.get('k_mallet','—')} / {dc.get('a_mallet','—')}"),
            ("K1 / K2 / a (Hazen-Lazervic)",            f"{dc.get('K1_hl','—')} / {dc.get('K2_hl','—')} / {dc.get('a_hl','—')}"),
            ("Montana a [T=10,20,50,100]",              _fmtlist(dc.get("montana_a"))),
            ("Montana b [T=10,20,50,100]",              _fmtlist(dc.get("montana_b"))),
            ("Gradex g",                                str(dc.get("grad_exp_pluie", "—"))),
        ]
        story.append(kv(param_rows, w1=6.5*cm))
    else:
        story.append(Paragraph("Paramètres non disponibles — relancer l'analyse pour les obtenir.", S_SMALL))

    # ══════════════════════════════════════════════════════════════════════════
    # 3. PARAMETRES DETAILLES DU BASSIN VERSANT
    # ══════════════════════════════════════════════════════════════════════════
    if bv:
        story += [Spacer(1,.3*cm), sec("3. Paramètres détaillés du bassin versant"), Spacer(1,.1*cm)]
        rows = [grid_hdr(["Paramètre", "Valeur", "Unité"])]
        for param, valeur, unite in _bassin_versant_export_rows(bv):
            rows.append([
                Paragraph(str(param), S_LABEL),
                Paragraph(str(valeur), ParagraphStyle("bv_val", fontSize=8.5, fontName="Helvetica", alignment=1, textColor=DARK)),
                Paragraph(str(unite), ParagraphStyle("bv_unit", fontSize=8.5, fontName="Helvetica", alignment=1, textColor=DARK)),
            ])
        t = Table(rows, colWidths=[8.5*cm, 5.2*cm, 3.2*cm])
        t.setStyle(TableStyle([*PAD,
            ("BACKGROUND",(0,0),(-1,0),DARK),
            ("BACKGROUND",(0,1),(0,-1),GREY),
            ("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#E0D0C0"))]))
        story.append(t)

    # ══════════════════════════════════════════════════════════════════════════
    # 3. TEMPS DE CONCENTRATION
    # ══════════════════════════════════════════════════════════════════════════
    story += [Spacer(1,.3*cm), sec("4. Temps de concentration Tc — par formule"), Spacer(1,.1*cm)]
    if dc and dc.get("tc"):
        tc_data = [[Paragraph(c, ParagraphStyle("gh",fontSize=8,fontName="Helvetica-Bold",
                    textColor=colors.white,alignment=1)) for c in ["Formule","Tc (min)","Tc (h)"]]]
        for formule, tv in dc["tc"].items():
            if formule == "Moyenne":
                continue
            tc_data.append([Paragraph(formule, S_VALUE),
                             Paragraph(fmt(tv), ParagraphStyle("cv",fontSize=8.5,fontName="Helvetica",alignment=1,textColor=DARK)),
                             Paragraph(fmt(tv/60) if tv else "—", ParagraphStyle("cv",fontSize=8.5,fontName="Helvetica",alignment=1,textColor=DARK))])
        moy = dc["tc"].get("Moyenne") or r.temps_concentration
        tc_data.append([Paragraph("Moyenne retenue (Tc)", S_LABEL),
                        Paragraph(fmt(moy), S_AVG),
                        Paragraph(fmt(moy/60) if moy else "—", S_AVG)])
        t = Table(tc_data, colWidths=[8*cm, 4.4*cm, 4.5*cm])
        t.setStyle(TableStyle([*PAD,
            ("BACKGROUND",(0,0),(-1,0),DARK),
            ("BACKGROUND",(0,-1),(-1,-1),AVG),
            ("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#E0D0C0")),
            ("ALIGN",(1,0),(-1,-1),"CENTER")]))
        story.append(t)
    else:
        story.append(Paragraph("Détails non disponibles — relancer l'analyse.", S_SMALL))

    # ══════════════════════════════════════════════════════════════════════════
    # 4. INTENSITÉS MONTANA & PLUIES DE PROJET
    # ══════════════════════════════════════════════════════════════════════════
    story += [Spacer(1,.3*cm), sec("5. Intensités Montana et Pluies de projet Pj24h"), Spacer(1,.1*cm)]
    if dc and (dc.get("intensites") or dc.get("pj24h")):
        cw = [5*cm] + [2.9*cm]*4
        hdr = [grid_hdr(["Paramètre", "T=10 ans", "T=20 ans", "T=50 ans", "T=100 ans"])]
        i_row = [Paragraph("Intensité Montana I (mm/h)", S_LABEL)]
        p_row = [Paragraph("Pluie journalière Pj24h (mm)", S_LABEL)]
        for T in PERIODES:
            i_row.append(Paragraph(fmt(_d(dc.get("intensites"), T)), ParagraphStyle("c",fontSize=8.5,fontName="Helvetica",alignment=1)))
            p_row.append(Paragraph(fmt(_d(dc.get("pj24h"), T)),      ParagraphStyle("c",fontSize=8.5,fontName="Helvetica",alignment=1)))
        t = Table(hdr + [i_row, p_row], colWidths=cw)
        t.setStyle(TableStyle([*PAD,
            ("BACKGROUND",(0,0),(-1,0),DARK),
            ("BACKGROUND",(0,1),(0,-1),GREY),
            ("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#E0D0C0")),
            ("ALIGN",(1,0),(-1,-1),"CENTER")]))
        story.append(t)
    else:
        story.append(Paragraph("Détails non disponibles.", S_SMALL))

    # ══════════════════════════════════════════════════════════════════════════
    # 5. DÉBITS PAR FORMULE (matrice)
    # ══════════════════════════════════════════════════════════════════════════
    story += [Spacer(1,.3*cm), sec("6. Débits de crue Q (m³/s) — par formule et par période"), Spacer(1,.1*cm)]
    if dc and dc.get("debits"):
        formules_incl = set(dc.get("formules_q_incluses", []))
        cw = [5.2*cm] + [2.9*cm]*4
        hdr = [grid_hdr(["Formule", "T=10 ans", "T=20 ans", "T=50 ans", "T=100 ans"])]
        body = []
        for formule, qd in dc["debits"].items():
            incl = formule in formules_incl
            style_lbl = ParagraphStyle("fl", fontSize=8.5, fontName="Helvetica-Bold" if incl else "Helvetica", textColor=DARK)
            row = [Paragraph(formule + (" ✓" if incl else ""), style_lbl)]
            for T in PERIODES:
                row.append(Paragraph(fmt(_d(qd, T)), ParagraphStyle("c",fontSize=8.5,fontName="Helvetica",alignment=1,textColor=DARK)))
            body.append(row)
        # Gradex
        if dc.get("debits_gradex"):
            row_g = [Paragraph("Gradex", ParagraphStyle("gr",fontSize=8.5,fontName="Helvetica-Bold",textColor=colors.HexColor("#1A6FA8")))]
            for T in PERIODES:
                row_g.append(Paragraph(fmt(_d(dc["debits_gradex"], T)), ParagraphStyle("c",fontSize=8.5,fontName="Helvetica",alignment=1)))
            body.append(row_g)
        # Ligne finale
        row_f = [Paragraph("Débit retenu — moyenne ✓✓", ParagraphStyle("fr",fontSize=9,fontName="Helvetica-Bold",textColor=colors.HexColor("#1E8449")))]
        for T in PERIODES:
            row_f.append(Paragraph(fmt(_d(dc.get("q_finaux"), T)),
                ParagraphStyle("fv",fontSize=9,fontName="Helvetica-Bold",textColor=GOLD,alignment=1)))
        body.append(row_f)

        t = Table(hdr + body, colWidths=cw)
        # styles de base
        ts = [*PAD,
              ("BACKGROUND",(0,0),(-1,0),DARK),
              ("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#E0D0C0")),
              ("ALIGN",(1,0),(-1,-1),"CENTER")]
        # colorier formules incluses
        for row_i, formule in enumerate(list(dc["debits"].keys()), start=1):
            if formule in formules_incl:
                ts.append(("BACKGROUND",(0,row_i),(-1,row_i),LIGHT))
        # Gradex
        gr_row = len(dc["debits"]) + 1
        if dc.get("debits_gradex"):
            ts.append(("BACKGROUND",(0,gr_row),(-1,gr_row),colors.HexColor("#EBF5FB")))
        # Finale
        ts.append(("BACKGROUND",(0,-1),(-1,-1),FINAL))
        t.setStyle(TableStyle(ts))
        story.append(t)
        story.append(Spacer(1,.1*cm))
        story.append(Paragraph(f"✓ Formules incluses dans la moyenne : {', '.join(formules_incl)}", S_SMALL))
    else:
        story.append(Paragraph("Détails non disponibles.", S_SMALL))

    # ══════════════════════════════════════════════════════════════════════════
    # 6. DÉBITS FINAUX RETENUS
    # ══════════════════════════════════════════════════════════════════════════
    story += [Spacer(1,.3*cm), sec("7. Débits de crue finals retenus (m³/s)"), Spacer(1,.1*cm)]
    niveaux = {10:"Courant", 20:"Peu fréquent", 50:"Rare", 100:"Très rare"}
    fill_risk = {10:GREEN, 20:AVG, 50:ORAN, 100:RED}
    q_hdr = [grid_hdr(["Période de retour","Q retenu (m³/s)","Niveau de risque"])]
    q_body = []
    for T, qv in [(10,r.qcrue_t10),(20,r.qcrue_t20),(50,r.qcrue_t50),(100,r.qcrue_t100)]:
        q_body.append([
            Paragraph(f"T = {T} ans", ParagraphStyle("p",fontSize=9,fontName="Helvetica-Bold",alignment=1,textColor=DARK)),
            Paragraph(fmt(qv), S_GOLD),
            Paragraph(niveaux[T], ParagraphStyle("p",fontSize=9,fontName="Helvetica",alignment=1,textColor=DARK)),
        ])
    t = Table(q_hdr + q_body, colWidths=[5.5*cm, 5.7*cm, 5.7*cm])
    ts2 = [*PAD, ("BACKGROUND",(0,0),(-1,0),DARK),
           ("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#C3E6CB")),
           ("ALIGN",(0,0),(-1,-1),"CENTER")]
    for i,(T,_) in enumerate([(10,0),(20,0),(50,0),(100,0)], start=1):
        ts2.append(("BACKGROUND",(0,i),(-1,i),fill_risk[T]))
    t.setStyle(TableStyle(ts2))
    story.append(t)

    # ══════════════════════════════════════════════════════════════════════════
    # 7. ANNOTATIONS
    # ══════════════════════════════════════════════════════════════════════════
    if r.observations or r.conclusions:
        story += [Spacer(1,.3*cm), sec("8. Annotations"), Spacer(1,.1*cm)]
        annot = []
        if r.observations:  annot.append(("Observations", r.observations))
        if r.conclusions:   annot.append(("Conclusions et recommandations", r.conclusions))
        story.append(kv(annot, w1=5.5*cm))

    # ── Pied de page ──────────────────────────────────────────────────────────
    story += [Spacer(1,.4*cm),
              HRFlowable(width="100%", thickness=0.5, color=GOLD),
              Spacer(1,.08*cm),
              Paragraph(f"HydroPlan SIG  ·  Généré le {date.today().strftime('%d/%m/%Y')}  ·  Analyse #{pk}", S_SMALL)]

    doc.build(story)
    buf.seek(0)
    slug = f"analyse_{pk}_{date.today().strftime('%Y%m%d')}.pdf"
    resp = HttpResponse(buf.read(), content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="{slug}"'
    return resp


@login_required
def supprimer_analyse(request, pk):
    analyse = get_object_or_404(ResultatAnalyseHydrologique, pk=pk)
    if request.method == 'POST':
        analyse.delete()
        messages.success(request, "Analyse supprimée.")
        return redirect('liste_analyses')
    return render(request, 'analyse_hydrologique/confirmer_suppression.html', {
        'objet': analyse, 'retour': 'liste_analyses',
    })


# =============================================================================
# Upload SHP (AJAX)
# =============================================================================

@login_required
@require_POST
def importer_bv_multiple(request):
    shp_zip = request.FILES.get('shp_file')
    if not shp_zip:
        return JsonResponse({'error': 'Aucun fichier fourni.'}, status=400)

    tmpdir = tempfile.mkdtemp()
    try:
        zip_path = os.path.join(tmpdir, 'upload.zip')
        with open(zip_path, 'wb') as f:
            for chunk in shp_zip.chunks():
                f.write(chunk)

        try:
            with zipfile.ZipFile(zip_path, 'r') as z:
                z.extractall(tmpdir)
        except zipfile.BadZipFile:
            return JsonResponse({'error': 'Fichier ZIP invalide.'}, status=400)

        # Exigence ZIP : doit contenir au minimum .shp + .dbf + .shx
        inventory = _zip_inventory(tmpdir)
        missing_required = [ext for ext in ('.shp', '.dbf', '.shx') if ext not in inventory]
        if missing_required:
            return JsonResponse({
                'error': f"ZIP incomplet — fichiers manquants : {', '.join(missing_required)}. "
                         f"Un shapefile valide contient au minimum .shp + .dbf + .shx (et idéalement .prj)."
            }, status=400)

        shp_path = None
        for root, _, files in os.walk(tmpdir):
            for fn in files:
                if fn.lower().endswith('.shp'):
                    shp_path = os.path.join(root, fn)
                    break
            if shp_path:
                break

        for root, _, files in os.walk(tmpdir):
            for fn in files:
                if '.shp.' in fn and not fn.lower().endswith(('.shp', '.dbf', '.shx', '.prj', '.cpg')):
                    try:
                        os.remove(os.path.join(root, fn))
                    except OSError:
                        pass

        try:
            ds = DataSource(shp_path)
            layer = ds[0]
            if len(layer) == 0:
                return JsonResponse({'error': 'Le shapefile ne contient aucune géométrie.'}, status=400)

            srid = layer[0].geom.srid or 0
            srid = _ESRI_TO_EPSG.get(srid, srid)
            if srid == 0:
                srid = _detect_srid(shp_path, layer)

            # Override forcé par l'utilisateur (priorité maximale)
            forced_srid = request.POST.get('source_srid', '').strip()
            if forced_srid:
                try:
                    srid = int(forced_srid)
                except ValueError:
                    pass

            # ── Méthode 3 : auto-correction du SRC sur la 1ère forme ──
            # Si le SRC déclaré ne donne pas une géométrie au Maroc et que
            # l'utilisateur n'a pas forcé, on cherche un candidat valide.
            auto_corrected_from = None
            if not forced_srid:
                test_geom = GEOSGeometry(layer[0].geom.wkt, srid=srid)
                if test_geom.srid != 4326:
                    test_geom.transform(4326)
                if not _geom_in_morocco(test_geom):
                    fixed_srid, _, _ = _auto_correct_crs(layer[0].geom.wkt, srid)
                    if fixed_srid is not None:
                        auto_corrected_from = srid
                        srid = fixed_srid

            created = 0
            skipped = []

            for idx in range(len(layer)):
                feat = layer[idx]
                attrs, source_fields = _extract_shp_attrs(layer, feat)
                matched_bv_fields = BV_SOURCE_FIELDS.intersection(source_fields)

                if not matched_bv_fields:
                    skipped.append(f"forme {idx + 1}: schéma BV non reconnu")
                    continue

                missing = [field for field in BV_BATCH_REQUIRED_ATTRS if field not in attrs]
                if missing:
                    skipped.append(f"forme {idx + 1}: champs manquants ({', '.join(missing)})")
                    continue

                try:
                    geom = GEOSGeometry(feat.geom.wkt, srid=srid)
                    if geom.srid != 4326:
                        geom.transform(4326)

                    # Convertir 3D → 2D (supprimer coordonnée Z) si nécessaire
                    if geom.hasz:
                        # Utiliser make_valid pour obtenir une géométrie 2D propre
                        from django.contrib.gis.geos import Polygon, MultiPolygon, Point, LineString, LinearRing
                        if geom.geom_type == 'Polygon':
                            # Recréer le polygon en 2D
                            exterior = [(x, y) for x, y, *z in geom.exterior_ring.coords]
                            interiors = [[(x, y) for x, y, *z in ring.coords] for ring in geom.interior_rings]
                            geom = Polygon(LinearRing(exterior), *interiors, srid=4326)
                        elif geom.geom_type == 'MultiPolygon':
                            polygons_2d = []
                            for poly in geom:
                                exterior = [(x, y) for x, y, *z in poly.exterior_ring.coords]
                                interiors = [[(x, y) for x, y, *z in ring.coords] for ring in poly.interior_rings]
                                polygons_2d.append(Polygon(LinearRing(exterior), *interiors, srid=4326))
                            geom = MultiPolygon(*polygons_2d, srid=4326)

                    ouvrage = str(attrs.get('ouvrage_en_tete') or '').strip()
                    nom_base = ouvrage or str(attrs.get('nom') or f"BV {idx + 1}")

                    # Gérer les doublons de nom
                    nom = nom_base
                    compteur = 1
                    while BassinVersant.objects.filter(nom=nom).exists():
                        nom = f"{nom_base} (copie {compteur})"
                        compteur += 1

                    BassinVersant.objects.create(
                        nom=nom,
                        x_exutoire=attrs['x_exutoire'],
                        y_exutoire=attrs['y_exutoire'],
                        surface=attrs['surface'],
                        perimetre=attrs['perimetre'],
                        z_min=attrs['z_min'],
                        z_max=attrs['z_max'],
                        thalweg=attrs['thalweg'],
                        ouvrage_en_tete=ouvrage,
                        geometrie=geom,
                    )
                    created += 1
                except Exception as exc:
                    skipped.append(f"forme {idx + 1}: {exc}")

            del layer, ds

            if created == 0:
                return JsonResponse({
                    'error': 'Aucun bassin versant importé.',
                    'details': skipped[:10],
                }, status=400)

            messages.success(request, f"{created} bassin(s) versant(s) importé(s) avec succès.")
            if skipped:
                messages.warning(request, f"{len(skipped)} forme(s) ignorée(s) pendant l'import.")

            return JsonResponse({
                'ok': True,
                'created_count': created,
                'skipped_count': len(skipped),
                'details': skipped[:10],
                'source_srid': srid,
                'auto_corrected_from': auto_corrected_from,
            })
        except Exception as exc:
            return JsonResponse({'error': f'Erreur lecture SHP : {exc}'}, status=400)
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)


@login_required
@require_POST
def upload_shp(request):
    shp_zip  = request.FILES.get('shp_file')
    geom_type = request.POST.get('geom_type', 'polygon')  # polygon | point | linestring
    import_mode = request.POST.get('import_mode', '').strip()

    if not shp_zip:
        return JsonResponse({'error': 'Aucun fichier fourni.'}, status=400)

    tmpdir = tempfile.mkdtemp()
    try:
        zip_path = os.path.join(tmpdir, 'upload.zip')
        with open(zip_path, 'wb') as f:
            for chunk in shp_zip.chunks():
                f.write(chunk)

        try:
            with zipfile.ZipFile(zip_path, 'r') as z:
                z.extractall(tmpdir)
        except zipfile.BadZipFile:
            return JsonResponse({'error': 'Fichier ZIP invalide.'}, status=400)

        # Exigence : le ZIP doit contenir au minimum .shp, .dbf, .shx
        inventory = _zip_inventory(tmpdir)
        missing_required = [ext for ext in ('.shp', '.dbf', '.shx') if ext not in inventory]
        if missing_required:
            return JsonResponse({
                'error': f"ZIP incomplet — fichiers manquants : {', '.join(missing_required)}. "
                         f"Un shapefile valide contient au minimum .shp + .dbf + .shx (et idéalement .prj)."
            }, status=400)

        # Search recursively — file may be in a subdirectory inside the ZIP
        shp_path = None
        for root, _, files in os.walk(tmpdir):
            for fn in files:
                if fn.lower().endswith('.shp'):
                    shp_path = os.path.join(root, fn)
                    break
            if shp_path:
                break

        # Remove ArcGIS lock files that can interfere with GDAL
        for root, _, files in os.walk(tmpdir):
            for fn in files:
                if '.shp.' in fn and not fn.lower().endswith(('.shp', '.dbf', '.shx', '.prj', '.cpg')):
                    try:
                        os.remove(os.path.join(root, fn))
                    except OSError:
                        pass

        try:
            ds    = DataSource(shp_path)
            layer = ds[0]
            feat  = layer[0]

            # ── Geometry ────────────────────────────────────────────────────
            raw_wkt = feat.geom.wkt
            srid    = feat.geom.srid or 0

            # Remap ESRI-only codes to their EPSG equivalents
            srid = _ESRI_TO_EPSG.get(srid, srid)

            # If GDAL couldn't resolve SRID, try .prj file then coordinate range
            srid_source = 'shapefile'
            has_prj = '.prj' in inventory
            if srid == 0:
                srid = _detect_srid(shp_path, layer)
                srid_source = 'prj_keyword' if has_prj else 'coordinate_range'

            # Override forcé par l'utilisateur (priorité maximale)
            forced_srid = request.POST.get('source_srid', '').strip()
            if forced_srid:
                try:
                    srid = int(forced_srid)
                    srid_source = 'user_override'
                except ValueError:
                    pass

            geom = GEOSGeometry(raw_wkt, srid=srid)
            if geom.srid != 4326:
                geom.transform(4326)

            in_morocco = _geom_in_morocco(geom)
            auto_corrected_from = None
            tried_candidates = []

            # ── Méthode 3 : Define Projection automatique ──
            # Si le SRC déclaré donne une géométrie hors du Maroc et que l'utilisateur
            # n'a pas explicitement forcé un SRC, on essaie tous les candidats.
            if not in_morocco and srid_source != 'user_override':
                fixed_srid, fixed_geom, tried_candidates = _auto_correct_crs(raw_wkt, srid)
                if fixed_srid is not None:
                    auto_corrected_from = srid
                    srid = fixed_srid
                    geom = fixed_geom
                    srid_source = 'auto_corrected'
                    in_morocco = True

            attrs, source_fields = _extract_shp_attrs(layer, feat)
            result = _build_geom_result(geom, geom_type, attrs)
            result['source_srid']         = srid
            result['source_srid_origin']  = srid_source  # shapefile|prj_keyword|coordinate_range|user_override|auto_corrected
            result['has_prj']             = has_prj
            result['in_morocco']          = in_morocco
            if auto_corrected_from is not None:
                result['auto_corrected_from'] = auto_corrected_from
                result['tried_candidates']    = tried_candidates
            if not in_morocco:
                result['warning'] = (
                    "Aucun système de coordonnées candidat n'a permis d'aligner la géométrie sur le Maroc. "
                    "Vérifiez l'origine du fichier ou choisissez manuellement le SRC source."
                )

            if import_mode == 'bv_strict':
                matched_bv_fields = BV_SOURCE_FIELDS.intersection(source_fields)
                result['auto_populate_bv'] = bool(matched_bv_fields)
                result['matched_bv_fields'] = sorted(matched_bv_fields)
                result['attrs'] = {
                    key: value for key, value in attrs.items()
                    if key in BV_FORM_ATTR_FIELDS
                } if result['auto_populate_bv'] else {}
            else:
                result['attrs'] = attrs

            # Release GDAL file handles before temp dir cleanup (Windows)
            del feat, layer, ds

            return JsonResponse(result)

        except Exception as exc:
            return JsonResponse({'error': f'Erreur lecture SHP : {exc}'}, status=400)

    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)
