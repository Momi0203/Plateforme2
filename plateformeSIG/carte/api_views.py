"""Endpoints /carte/api/* — à implémenter (voir cc-projet/06_api_architecture.md §7)."""

import csv
import json

from django.apps import apps
from django.contrib.gis.geos import Polygon
from django.core import serializers as dj_serializers
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_GET, require_POST

from .layers import LAYER_REGISTRY
from . import tools as geotools
from compte.decorators import api_login_required, role_required

_EXP = ('operateur', 'editeur')   # shorthand — exports seulement (Phase 2 §5)


# ── Helpers partagés ──────────────────────────────────────────────────────────

def _build_q(lookup_field, operateur, valeur):
    """
    Construit un objet Q pour le filtre ORM à partir d'un opérateur textuel.
    Lève ValueError pour opérateur inconnu ou ENTRE avec valeur invalide.
    """
    if operateur == '=':
        return Q(**{f'{lookup_field}__exact': valeur})
    if operateur == '!=':
        return ~Q(**{f'{lookup_field}__exact': valeur})
    if operateur == '>':
        return Q(**{f'{lookup_field}__gt': valeur})
    if operateur == '>=':
        return Q(**{f'{lookup_field}__gte': valeur})
    if operateur == '<':
        return Q(**{f'{lookup_field}__lt': valeur})
    if operateur == '<=':
        return Q(**{f'{lookup_field}__lte': valeur})
    if operateur == 'CONTIENT':
        return Q(**{f'{lookup_field}__icontains': valeur})
    if operateur == 'COMMENCE_PAR':
        return Q(**{f'{lookup_field}__istartswith': valeur})
    if operateur == 'EST_NULL':
        return Q(**{f'{lookup_field}__isnull': True})
    if operateur == 'ENTRE':
        if not isinstance(valeur, list) or len(valeur) != 2:
            raise ValueError('ENTRE : valeur doit être [min, max]')
        return Q(**{f'{lookup_field}__range': (valeur[0], valeur[1])})
    raise ValueError(f'Opérateur inconnu : {operateur}')


def _resolve_champ(champ, meta):
    """
    Résout un nom de champ front-end vers le lookup ORM.
    Gère le champ virtuel 'etat_general' (RM-07) pour les couches avec join_etat.
    Retourne (lookup_field, erreur_ou_None).
    """
    if champ == 'etat_general':
        etat_lookup = meta.get('etat_lookup')
        if not etat_lookup:
            return None, 'Ce champ n\'est pas disponible pour cette couche'
        return etat_lookup, None
    if champ not in meta.get('fields', []):
        return None, f'Champ non exposé : {champ}'
    return champ, None


# ── §7.1 Couches GeoJSON ──────────────────────────────────────────────────────

@api_login_required
@require_GET
def liste_couches(request):
    data = [
        {
            "nom":       cle,
            "label":     meta["label"],
            "geom_type": meta["geom_type"],
            "groupe":    meta["groupe"],
            "fields":    meta.get("fields", []),
            # Champ affiché comme label au centre des polygones.
            # Par défaut le 1er champ déclaré (convention : c'est le « nom »).
            "label_field": meta.get("label_field") or (meta.get("fields") or [None])[0],
            # True si la couche expose le critère "état général" (RM-07)
            "has_etat":  bool(meta.get("etat_lookup")),
        }
        for cle, meta in LAYER_REGISTRY.items()
        # Couches `hidden` : absentes du panneau gauche, mais toujours servies
        # par geojson_couche/extent_couche aux outils (ex. bv_ouvrage_tete).
        if not meta.get("hidden")
    ]
    return JsonResponse(data, safe=False)


@api_login_required
@require_GET
def couches_activables(request):
    """
    GET /carte/api/couches/activables/
    → métadonnées des couches `groupe_activable` (box « Couches » / Hydrologie 2).
      Ces couches sont `hidden` (absentes de la liste principale) mais injectées
      à la demande dans un groupe du panneau gauche.
    """
    data = [
        {
            "nom":         cle,
            "label":       meta["label"],
            "geom_type":   meta["geom_type"],
            "groupe":      meta["groupe_activable"],
            "fields":      meta.get("fields", []),
            "label_field": meta.get("label_field") or (meta.get("fields") or [None])[0],
            # True → le panneau gauche remplace le bouton multicritère par
            # « Réseau du BV » (intersection forcée réseau ↔ BV sélectionné).
            "reseau_tete": bool(meta.get("reseau_tete")),
        }
        for cle, meta in LAYER_REGISTRY.items()
        if meta.get("groupe_activable")
    ]
    return JsonResponse(data, safe=False)


# Couche réseau « ouvrage de tête » → modèle. Mode FORCÉ (Q5-bis) : on respecte
# la couche cliquée, sans appariement automatique des 5 bassins.
_RESEAU_TETE_MODELES = {
    'reseau_tete_ziz':      'carte.ReseauOuvrageTeteZiz',
    'reseau_tete_moulouya': 'carte.ReseauOuvrageTeteMoulouya',
    'reseau_tete_guir':     'carte.ReseauOuvrageTeteGuir',
    'reseau_tete_rheris':   'carte.ReseauOuvrageTeteRheris',
    'reseau_tete_maider':   'carte.ReseauOuvrageTeteMaider',
}
# Même seuil que analyse_hydrologique : grand BV → drains forts (grid_code >= 1).
RESEAU_SEUIL_INTERSECTION_M2 = 500e6  # 500 km²


@api_login_required
@require_GET
def reseau_ouvrage_tete(request):
    """
    GET /carte/api/reseau-ouvrage-tete/?reseau=<couche>&bv=<pk>&min_grid_code=
    → réseau de la COUCHE choisie (mode forcé), clippé au BV sélectionné, avec
      filtrage adaptatif du grid_code (réseau complet pour les petits BV, drains
      forts grid_code>=1 pour les grands — repli sur tout si vide). Reprend
      l'étape 3 de analyse_hydrologique._reseau_ouvrage_tete_pour_bv.
      → GeoJSON {grid_code, geometry} + { reseau, bv, count, grid_max }.
    """
    from django.shortcuts import get_object_or_404
    from django.db.models import Max
    from analyse_hydrologique.models import BassinVersant

    reseau = request.GET.get('reseau', '').strip()
    model_path = _RESEAU_TETE_MODELES.get(reseau)
    if not model_path:
        return JsonResponse({'erreur': f'Couche réseau inconnue : {reseau}'}, status=404)

    bv_pk = request.GET.get('bv')
    if not bv_pk:
        return JsonResponse({'erreur': 'Paramètre bv (pk du bassin versant) requis'}, status=400)
    bv = get_object_or_404(BassinVersant, pk=bv_pk)
    if not bv.geometrie:
        return JsonResponse({'type': 'FeatureCollection', 'features': [],
                             'message': 'Bassin versant sans géométrie'})

    app_label, model_name = model_path.split('.')
    Model = apps.get_model(app_label, model_name)

    base = (Model.objects
            .filter(geometrie__bboverlaps=bv.geometrie)
            .filter(geometrie__intersects=bv.geometrie))

    raw = request.GET.get('min_grid_code')
    if raw not in (None, ''):
        try:
            qs = base.filter(grid_code__gte=int(raw))
        except ValueError:
            return JsonResponse({'erreur': 'min_grid_code invalide'}, status=400)
    else:
        try:
            aire = bv.geometrie.transform(26191, clone=True).area
        except Exception:
            aire = 0
        if aire >= RESEAU_SEUIL_INTERSECTION_M2:
            forts = base.filter(grid_code__gte=1)
            qs = forts if forts.exists() else base
        else:
            qs = base

    qs = qs.only('id', 'grid_code', 'geometrie')
    features = [
        {'type': 'Feature', 'id': r.id,
         'properties': {'grid_code': r.grid_code},
         'geometry': json.loads(r.geometrie.geojson)}
        for r in qs
    ]
    grid_max = Model.objects.aggregate(m=Max('grid_code'))['m']
    return JsonResponse({
        'type':     'FeatureCollection',
        'reseau':   reseau,
        'bv':       bv.nom,
        'count':    len(features),
        'grid_max': grid_max,
        'features': features,
    })


@api_login_required
@require_GET
def geojson_couche(request, nom):
    if nom not in LAYER_REGISTRY:
        return JsonResponse({"erreur": f"Couche inconnue : {nom}"}, status=404)

    meta = LAYER_REGISTRY[nom]
    app_label, model_name = meta["model"].split(".")
    Model = apps.get_model(app_label, model_name)
    geom_field = meta["geom_field"]
    join_etat  = meta.get("join_etat")

    # Paramètres GET
    try:
        limit  = min(int(request.GET.get("limit",  500)), 2000)
        offset = int(request.GET.get("offset", 0))
    except ValueError:
        return JsonResponse({"erreur": "limit/offset doivent être des entiers"}, status=400)

    # fields peut être surchargé via ?fields=f1,f2 ou ?fields=* (tous les champs déclarés)
    fields_param = request.GET.get("fields")
    if not fields_param or fields_param.strip() == '*':
        fields = meta.get("fields", [])
    else:
        fields = [f.strip() for f in fields_param.split(",") if f.strip()]

    qs = Model.objects.exclude(**{f"{geom_field}__isnull": True})

    # Filtre ?pks=1,2,3 — drill-down §5.1.6 (renvoi un sous-ensemble de PKs)
    pks_param = request.GET.get("pks")
    if pks_param:
        try:
            pk_list = [int(p.strip()) for p in pks_param.split(",") if p.strip()]
        except ValueError:
            return JsonResponse({"erreur": "pks invalide — entiers séparés par virgules attendu"}, status=400)
        qs = qs.filter(pk__in=pk_list)

    # Filtre bbox
    bbox_param = request.GET.get("bbox")
    if bbox_param:
        try:
            coords = [float(x) for x in bbox_param.split(",")]
            if len(coords) != 4:
                raise ValueError
        except ValueError:
            return JsonResponse({"erreur": "bbox invalide — attendu : xmin,ymin,xmax,ymax"}, status=400)
        bbox_poly = Polygon.from_bbox(coords)
        bbox_poly.srid = 4326
        qs = qs.filter(**{f"{geom_field}__bboverlaps": bbox_poly})

    # Join EtatX si la couche le déclare (ex. seuils → EtatSeuil)
    if join_etat:
        qs = qs.select_related(join_etat)

    objects = list(qs[offset: offset + limit])

    # Sérialisation GeoJSON via le sérialiseur GeoDjango
    geojson_str = dj_serializers.serialize(
        "geojson",
        objects,
        geometry_field=geom_field,
        fields=fields,
    )
    data = json.loads(geojson_str)

    # Enrichissement etat_general depuis EtatX (générique via join_etat)
    if join_etat and objects:
        etat_map = {}
        for obj in objects:
            etat = getattr(obj, join_etat, None)
            # EtatSeuil : champ fonctionnel = etat_construction_fonctionnement
            # Autres Etat* : champ standard = etat_general
            etat_map[obj.pk] = (
                getattr(etat, "etat_general", None)
                or getattr(etat, "etat_construction_fonctionnement", None)
            )
        for feature in data.get("features", []):
            pk = feature.get("id")
            try:
                pk = int(pk)
            except (TypeError, ValueError):
                pass
            feature["properties"]["etat_general"] = etat_map.get(pk)

    return JsonResponse(data)


@api_login_required
@require_GET
def geojson_entite(request, nom, pk):
    """Retourne le GeoJSON d'une seule entité par PK — drill-down §5.1.6."""
    if nom not in LAYER_REGISTRY:
        return JsonResponse({"erreur": f"Couche inconnue : {nom}"}, status=404)

    meta = LAYER_REGISTRY[nom]
    app_label, model_name = meta["model"].split(".")
    Model     = apps.get_model(app_label, model_name)
    geom_field = meta["geom_field"]
    join_etat  = meta.get("join_etat")
    fields     = meta.get("fields", [])

    qs = Model.objects.filter(pk=pk)
    if join_etat:
        qs = qs.select_related(join_etat)

    objects = list(qs)
    if not objects:
        return JsonResponse({"erreur": f"Entité introuvable : {nom}/{pk}"}, status=404)

    geojson_str = dj_serializers.serialize(
        "geojson", objects, geometry_field=geom_field, fields=fields,
    )
    data = json.loads(geojson_str)

    # Enrichissement etat_general (même logique que geojson_couche)
    if join_etat and objects:
        obj  = objects[0]
        etat = getattr(obj, join_etat, None)
        val  = (
            getattr(etat, "etat_general", None)
            or getattr(etat, "etat_construction_fonctionnement", None)
        )
        for feature in data.get("features", []):
            feature["properties"]["etat_general"] = val

    return JsonResponse(data)


@api_login_required
@require_GET
def extent_couche(request, nom):
    """
    GET /carte/api/couche/<nom>/extent/
    → { bbox: [xmin, ymin, xmax, ymax] }  (EPSG:4326)

    Retourne l'emprise géographique de toutes les entités de la couche.
    Utilisé par le bouton « Centrer » (BUG-L4-B).
    """
    if nom not in LAYER_REGISTRY:
        return JsonResponse({"erreur": f"Couche inconnue : {nom}"}, status=404)

    meta = LAYER_REGISTRY[nom]
    app_label, model_name = meta["model"].split(".")
    Model = apps.get_model(app_label, model_name)
    geom_field = meta["geom_field"]

    from django.contrib.gis.db.models import Extent as GISExtent

    qs = Model.objects.exclude(**{f"{geom_field}__isnull": True})

    # Filtrage optionnel par liste de PKs (utilisé par le masque)
    pks_param = request.GET.get("pks")
    if pks_param:
        try:
            pk_list = [int(p.strip()) for p in pks_param.split(",") if p.strip()]
            qs = qs.filter(pk__in=pk_list)
        except ValueError:
            pass

    result = qs.aggregate(extent=GISExtent(geom_field))
    bbox = result.get("extent")
    if not bbox:
        return JsonResponse({"erreur": "Aucune géométrie disponible pour cette couche"}, status=404)

    return JsonResponse({"bbox": list(bbox)})


# ── §7.2 Champs et valeurs ────────────────────────────────────────────────────

@api_login_required
@require_GET
def champs_couche(request, nom):
    """Liste des champs exposés pour une couche (depuis LAYER_REGISTRY)."""
    if nom not in LAYER_REGISTRY:
        return JsonResponse({'erreur': f'Couche inconnue : {nom}'}, status=404)
    return JsonResponse({'champs': LAYER_REGISTRY[nom].get('fields', [])})


@api_login_required
@require_GET
def valeurs_champ(request, nom, champ):
    """
    Valeurs distinctes (+ libellés) d'un champ pour une couche.

    Règle d'évolutivité §11.3 : si le champ a des CHOICES Django,
    retourne les couples (valeur, libellé) ; sinon, liste distincte en base.
    Seuls les champs déclarés dans LAYER_REGISTRY (+ 'etat_general' virtuel) sont autorisés.

    Paramètre optionnel ?pks=1,2,3 : restreint les valeurs distinctes au
    sous-ensemble issu d'une requête (filtre carte). Utilisé par la symbologie
    catégorisée pour ne proposer que les catégories réellement présentes dans
    le résultat de la requête, et non sur toute la couche.
    """
    if nom not in LAYER_REGISTRY:
        return JsonResponse({'erreur': f'Couche inconnue : {nom}'}, status=404)

    meta = LAYER_REGISTRY[nom]

    # ── Sous-ensemble optionnel issu d'une requête (?pks=...) ─────────────
    pk_list = None
    pks_param = request.GET.get('pks')
    if pks_param:
        try:
            pk_list = [int(p.strip()) for p in pks_param.split(',') if p.strip()]
        except ValueError:
            return JsonResponse(
                {'erreur': 'pks invalide — entiers séparés par virgules attendu'},
                status=400,
            )

    app_label, model_name = meta['model'].split('.')
    Model = apps.get_model(app_label, model_name)

    # ── Champ virtuel etat_general (RM-07) ────────────────────────────────
    if champ == 'etat_general':
        etat_lookup = meta.get('etat_lookup')
        if not etat_lookup:
            return JsonResponse(
                {'erreur': f'Champ etat_general non disponible pour la couche « {nom} »'},
                status=403,
            )
        from diagnostic.models import ETAT_CONSTRUCTION_DIAG_CHOICES
        if pk_list is not None:
            # Seules les valeurs d'état présentes dans le sous-ensemble de requête
            label_map = {v: str(l) for v, l in ETAT_CONSTRUCTION_DIAG_CHOICES}
            present = (
                Model.objects.filter(pk__in=pk_list)
                .exclude(**{f'{etat_lookup}__isnull': True})
                .values_list(etat_lookup, flat=True)
                .distinct()
            )
            valeurs = [
                {'valeur': v, 'label': label_map.get(v, str(v))}
                for v in present if v not in ('', None)
            ]
        else:
            valeurs = [
                {'valeur': v, 'label': str(l)}
                for v, l in ETAT_CONSTRUCTION_DIAG_CHOICES
            ]
        return JsonResponse({'valeurs': valeurs})

    # ── Champ standard : sécurité SEC-05 ─────────────────────────────────
    if champ not in meta.get('fields', []):
        return JsonResponse(
            {'erreur': f'Champ non exposé pour la couche « {nom} » : {champ}'},
            status=403,
        )

    try:
        field = Model._meta.get_field(champ)
    except Exception:
        return JsonResponse({'erreur': f'Champ introuvable sur le modèle : {champ}'}, status=404)

    # flatchoices gère les optgroups imbriqués
    choices = getattr(field, 'flatchoices', None) or getattr(field, 'choices', None)

    if pk_list is not None:
        # Valeurs réellement présentes dans le sous-ensemble de requête.
        # Un champ à choices conserve ses libellés (mappés depuis flatchoices).
        qs = (
            Model.objects
            .filter(pk__in=pk_list)
            .exclude(**{f'{champ}__isnull': True})
            .exclude(**{f'{champ}__exact': ''})
            .values_list(champ, flat=True)
            .distinct()
            .order_by(champ)
        )
        if choices:
            label_map = {v: str(l) for v, l in choices}
            valeurs = [{'valeur': v, 'label': label_map.get(v, str(v))} for v in qs]
        else:
            valeurs = [{'valeur': v, 'label': str(v)} for v in qs]
        return JsonResponse({'valeurs': valeurs})

    # ── Toute la couche (comportement historique, aucun filtre) ──────────
    if choices:
        valeurs = [{'valeur': v, 'label': str(l)} for v, l in choices if v not in ('', None)]
    else:
        qs = (
            Model.objects
            .exclude(**{f'{champ}__isnull': True})
            .exclude(**{f'{champ}__exact': ''})
            .values_list(champ, flat=True)
            .distinct()
            .order_by(champ)
        )
        valeurs = [{'valeur': v, 'label': str(v)} for v in qs]

    return JsonResponse({'valeurs': valeurs})


@api_login_required
@require_GET
def stats_champ(request, nom, champ):
    """
    Type et statistiques d'un champ pour la symbologie graduée (SY-02 quantitatif).

    GET /carte/api/couche/<nom>/champs/<champ>/stats/?classes=5[&pks=1,2,3]
    → qualitatif  : { type: 'qualitatif' }
    → quantitatif : { type: 'quantitatif', min, max, count,
                      breaks: { quantiles: [...], egaux: [...] } }

    Les breaks sont les BORNES SUPÉRIEURES des n classes (la dernière = max).
    Un champ à choices Django est toujours qualitatif, même s'il est numérique.

    Paramètre optionnel ?pks=1,2,3 : restreint min/max/breaks au sous-ensemble
    issu d'une requête (filtre carte), comme valeurs_champ.
    """
    from django.db import models as dj_models

    if nom not in LAYER_REGISTRY:
        return JsonResponse({'erreur': f'Couche inconnue : {nom}'}, status=404)

    meta = LAYER_REGISTRY[nom]

    # etat_general virtuel (RM-07) → toujours qualitatif
    if champ == 'etat_general':
        return JsonResponse({'type': 'qualitatif'})

    # Sécurité SEC-05 : seuls les champs déclarés sont autorisés
    if champ not in meta.get('fields', []):
        return JsonResponse(
            {'erreur': f'Champ non exposé pour la couche « {nom} » : {champ}'},
            status=403,
        )

    app_label, model_name = meta['model'].split('.')
    Model = apps.get_model(app_label, model_name)

    try:
        field = Model._meta.get_field(champ)
    except Exception:
        return JsonResponse({'erreur': f'Champ introuvable sur le modèle : {champ}'}, status=404)

    est_numerique = isinstance(
        field, (dj_models.IntegerField, dj_models.FloatField, dj_models.DecimalField)
    )
    a_choices = bool(getattr(field, 'flatchoices', None) or getattr(field, 'choices', None))

    if not est_numerique or a_choices or field.is_relation:
        return JsonResponse({'type': 'qualitatif'})

    try:
        n_classes = max(1, min(int(request.GET.get('classes', 5)), 10))
    except ValueError:
        return JsonResponse({'erreur': 'classes doit être un entier'}, status=400)

    # Sous-ensemble optionnel issu d'une requête (?pks=...)
    base_qs = Model.objects.exclude(**{f'{champ}__isnull': True})
    pks_param = request.GET.get('pks')
    if pks_param:
        try:
            pk_list = [int(p.strip()) for p in pks_param.split(',') if p.strip()]
        except ValueError:
            return JsonResponse(
                {'erreur': 'pks invalide — entiers séparés par virgules attendu'},
                status=400,
            )
        base_qs = base_qs.filter(pk__in=pk_list)

    valeurs = sorted(
        float(v) for v in base_qs.values_list(champ, flat=True)
    )
    if not valeurs:
        return JsonResponse({
            'type': 'quantitatif', 'min': None, 'max': None, 'count': 0,
            'breaks': {'quantiles': [], 'egaux': []},
        })

    vmin, vmax, count = valeurs[0], valeurs[-1], len(valeurs)

    if vmin == vmax:
        breaks_q = breaks_e = [vmax]
    else:
        # Quantiles (effectifs égaux) — interpolation linéaire des percentiles
        def _quantile(p):
            idx = p * (count - 1)
            lo  = int(idx)
            hi  = min(lo + 1, count - 1)
            return valeurs[lo] + (valeurs[hi] - valeurs[lo]) * (idx - lo)

        breaks_q = [_quantile((i + 1) / n_classes) for i in range(n_classes)]
        # Intervalles égaux entre min et max
        breaks_e = [vmin + (vmax - vmin) * (i + 1) / n_classes for i in range(n_classes)]
        # Les dernières bornes sont exactement le max
        breaks_q[-1] = breaks_e[-1] = vmax

        # Dédoublonnage croissant — les quantiles d'un champ entier à faible
        # cardinalité produisent des bornes identiques, inapplicables en step.
        def _dedup(seq):
            out = []
            for v in seq:
                if not out or v > out[-1]:
                    out.append(v)
            return out

        breaks_q = _dedup(breaks_q)
        breaks_e = _dedup(breaks_e)

    return JsonResponse({
        'type':   'quantitatif',
        'min':    vmin,
        'max':    vmax,
        'count':  count,
        'breaks': {'quantiles': breaks_q, 'egaux': breaks_e},
    })


@api_login_required
@require_GET
def criteres_scoring(request, nom):
    """
    GET /carte/api/couche/<nom>/criteres/
    → { criteres: [{champ, label}, ...] }

    Retourne les champs EtatX déclarés dans LAYER_REGISTRY.scoring_champs,
    enrichis du verbose_name du modèle correspondant.
    """
    if nom not in LAYER_REGISTRY:
        return JsonResponse({'erreur': f'Couche inconnue : {nom}'}, status=404)

    meta = LAYER_REGISTRY[nom]
    scoring_champs = meta.get('scoring_champs')
    if not scoring_champs:
        return JsonResponse({'erreur': f'Aucun critère de scoring pour la couche : {nom}'}, status=404)

    join_etat = meta.get('join_etat')
    if not join_etat:
        return JsonResponse({'erreur': 'join_etat manquant dans LAYER_REGISTRY'}, status=500)

    app_label, model_name = meta['model'].split('.')
    MainModel = apps.get_model(app_label, model_name)
    EtatModel = MainModel._meta.get_field(join_etat).related_model

    criteres = []
    for champ in scoring_champs:
        try:
            label = str(EtatModel._meta.get_field(champ).verbose_name)
        except Exception:
            label = champ
        criteres.append({'champ': champ, 'label': label})

    return JsonResponse({'criteres': criteres})


# ── §7.3 Requêtes ─────────────────────────────────────────────────────────────

@api_login_required
@require_POST
def requete_simple(request):
    """
    POST { couche, champ, operateur, valeur }
    →    { pks: [...], count: N }

    Opérateurs : = != > >= < <= CONTIENT COMMENCE_PAR EST_NULL ENTRE
    Champ virtuel 'etat_general' supporté pour les couches avec join_etat (RM-07).
    """
    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'erreur': 'Corps JSON invalide'}, status=400)

    couche    = body.get('couche', '').strip()
    champ     = body.get('champ',  '').strip()
    operateur = body.get('operateur', '').strip()
    valeur    = body.get('valeur')

    if couche not in LAYER_REGISTRY:
        return JsonResponse({'erreur': f'Couche inconnue : {couche}'}, status=404)

    meta = LAYER_REGISTRY[couche]
    lookup_field, err = _resolve_champ(champ, meta)
    if err:
        return JsonResponse({'erreur': err}, status=403)

    app_label, model_name = meta['model'].split('.')
    Model = apps.get_model(app_label, model_name)

    try:
        q  = _build_q(lookup_field, operateur, valeur)
        qs = Model.objects.filter(q)
    except ValueError as exc:
        return JsonResponse({'erreur': str(exc)}, status=400)
    except Exception as exc:
        return JsonResponse({'erreur': str(exc)}, status=400)

    pks = list(qs.values_list('pk', flat=True))
    return JsonResponse({'pks': pks, 'count': len(pks)})


@api_login_required
@require_POST
def requete_multicritere(request):
    """
    POST { couche, conditions: [{champ, operateur, valeur}, ...], logique: 'ET'|'OU' }
    →    { pks: [...], count: N }

    Les conditions sont combinées avec Q() & (ET) ou Q() | (OU).
    Champ virtuel 'etat_general' supporté pour les couches avec join_etat (RM-07).
    """
    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'erreur': 'Corps JSON invalide'}, status=400)

    couche     = body.get('couche', '').strip()
    conditions = body.get('conditions', [])
    logique    = body.get('logique', 'ET').strip().upper()

    # ── Validations de base ───────────────────────────────────────────────
    if couche not in LAYER_REGISTRY:
        return JsonResponse({'erreur': f'Couche inconnue : {couche}'}, status=404)

    if not isinstance(conditions, list) or not conditions:
        return JsonResponse({'erreur': 'conditions doit être une liste non vide'}, status=400)

    if logique not in ('ET', 'OU'):
        return JsonResponse({'erreur': "logique doit être 'ET' ou 'OU'"}, status=400)

    meta = LAYER_REGISTRY[couche]
    app_label, model_name = meta['model'].split('.')
    Model = apps.get_model(app_label, model_name)

    # ── Construction des Q objects ────────────────────────────────────────
    q_combined = None
    for i, cond in enumerate(conditions):
        if not isinstance(cond, dict):
            return JsonResponse({'erreur': f'Condition {i + 1} invalide'}, status=400)

        champ     = str(cond.get('champ',     '')).strip()
        operateur = str(cond.get('operateur', '')).strip()
        valeur    = cond.get('valeur')

        lookup_field, err = _resolve_champ(champ, meta)
        if err:
            return JsonResponse({'erreur': f'Condition {i + 1} : {err}'}, status=403)

        try:
            q = _build_q(lookup_field, operateur, valeur)
        except ValueError as exc:
            return JsonResponse({'erreur': f'Condition {i + 1} : {exc}'}, status=400)

        if q_combined is None:
            q_combined = q
        elif logique == 'ET':
            q_combined &= q
        else:
            q_combined |= q

    try:
        qs  = Model.objects.filter(q_combined)
        pks = list(qs.values_list('pk', flat=True))
    except Exception as exc:
        return JsonResponse({'erreur': str(exc)}, status=400)

    return JsonResponse({'pks': pks, 'count': len(pks)})


@api_login_required
@require_POST
def requete_spatiale(request):
    raise NotImplementedError


# ── §7.4 Outils géospatiaux ───────────────────────────────────────────────────

@api_login_required
@require_POST
def outil_buffer(request):
    """
    POST { couche, pks?, distance_m }
    →    GeoJSON FeatureCollection du tampon (union de la sélection, projeté en 26191,
         bufférisé, reprojeté en 4326).
    """
    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'erreur': 'Corps JSON invalide'}, status=400)

    couche     = body.get('couche', '').strip()
    pks        = body.get('pks')
    distance_m = body.get('distance_m')

    if couche not in LAYER_REGISTRY:
        return JsonResponse({'erreur': f'Couche inconnue : {couche}'}, status=404)

    try:
        distance_m = float(distance_m)
    except (TypeError, ValueError):
        return JsonResponse({'erreur': 'distance_m doit être un nombre'}, status=400)

    if distance_m <= 0:
        return JsonResponse({'erreur': 'distance_m doit être strictement positif'}, status=400)

    try:
        result = geotools.buffer(LAYER_REGISTRY, couche, distance_m, pks or None)
    except Exception as exc:
        return JsonResponse({'erreur': str(exc)}, status=500)

    return JsonResponse(result)


@api_login_required
@require_POST
def outil_intersection(request):
    """
    POST { couche_a, couche_b, pks_a? }
    →    GeoJSON FeatureCollection des entités de couche_b qui intersectent
         l'union géométrique de la sélection de couche_a.
    """
    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'erreur': 'Corps JSON invalide'}, status=400)

    couche_a = body.get('couche_a', '').strip()
    couche_b = body.get('couche_b', '').strip()
    pks_a    = body.get('pks_a')

    for key, val in (('couche_a', couche_a), ('couche_b', couche_b)):
        if val not in LAYER_REGISTRY:
            return JsonResponse({'erreur': f'Couche inconnue : {val} ({key})'}, status=404)

    try:
        result = geotools.intersection(LAYER_REGISTRY, couche_a, couche_b, pks_a or None)
    except Exception as exc:
        return JsonResponse({'erreur': str(exc)}, status=500)

    return JsonResponse(result)


@api_login_required
@require_POST
def outil_union(request):
    raise NotImplementedError


@api_login_required
@require_POST
def outil_dissolve(request):
    raise NotImplementedError


@api_login_required
@require_POST
def outil_near(request):
    raise NotImplementedError


@api_login_required
@require_POST
def outil_stats(request):
    raise NotImplementedError


@api_login_required
@require_POST
def outil_efficience(request):
    """
    POST { pks: [TronconSeguia.pk, ...] }
    →    { resultats: [{pk, troncon, efficience_pourcent, perte_infiltration_m3s,
                        perte_vaporisation_m3s, debit_amont, debit_aval}],
           erreurs: [{pk, erreur}], nb_calcules: N }

    Persiste efficience_calculee / perte_* / date_dernier_calcul sur chaque tronçon.
    """
    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'erreur': 'Corps JSON invalide'}, status=400)

    pks = body.get('pks', [])
    if not isinstance(pks, list) or not pks:
        return JsonResponse({'erreur': 'pks doit être une liste non vide'}, status=400)

    from diagnostic.models import TronconSeguia
    from efficiences.services.efficience_troncon import calculer_efficience_troncon

    troncons = list(
        TronconSeguia.objects
        .filter(pk__in=pks)
        .select_related('seguia__perimetre')
    )
    if not troncons:
        return JsonResponse({'erreur': 'Aucun tronçon trouvé pour les PKs fournis'}, status=404)

    resultats, erreurs = [], []
    for tr in troncons:
        try:
            detail = calculer_efficience_troncon(
                tr,
                perimetre=tr.seguia.perimetre,
                q_amont=None,
                persister=True,
            )
            resultats.append({
                'pk':                     tr.pk,
                'troncon':                tr.troncon,
                'debit_amont':            round(detail['debit_amont'], 6),
                'perte_infiltration_m3s': round(detail['perte_infiltration_m3s'], 6),
                'perte_vaporisation_m3s': round(detail['perte_vaporisation_m3s'], 6),
                'perte_totale_m3s':       round(detail['perte_totale_m3s'], 6),
                'debit_aval':             round(detail['debit_aval'], 6),
                'efficience_pourcent':    round(detail['efficience_pourcent'], 2),
                'coefficient_c':          detail['coefficient_c'],
            })
        except Exception as exc:
            erreurs.append({'pk': tr.pk, 'erreur': str(exc)})

    return JsonResponse({'resultats': resultats, 'erreurs': erreurs, 'nb_calcules': len(resultats)})


@api_login_required
@require_POST
def outil_manning(request):
    """
    POST { pks: [TronconSeguia.pk, ...], n_manning?: float, pente?: float }
    →    { resultats: [{pk, troncon, debit_calcule, n_utilise, forme, pente}] }

    Calcul en lecture seule — aucune écriture en base.
    pente (m/m, défaut 0.001) — non stockée sur le modèle, fournie par l'appelant.
    n_manning — si absent, déduit de la nature : béton 0.013, béton armé 0.014,
                terre 0.025, autre 0.020.
    """
    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'erreur': 'Corps JSON invalide'}, status=400)

    pks = body.get('pks', [])
    if not isinstance(pks, list) or not pks:
        return JsonResponse({'erreur': 'pks doit être une liste non vide'}, status=400)

    n_user = body.get('n_manning')
    if n_user is not None:
        try:
            n_user = float(n_user)
            if n_user <= 0:
                raise ValueError
        except (TypeError, ValueError):
            return JsonResponse({'erreur': 'n_manning doit être un flottant strictement positif'}, status=400)

    try:
        pente = float(body.get('pente', 0.001))
        if pente <= 0:
            raise ValueError
    except (TypeError, ValueError):
        return JsonResponse({'erreur': 'pente doit être un flottant strictement positif'}, status=400)

    from diagnostic.models import TronconSeguia

    troncons = list(TronconSeguia.objects.filter(pk__in=pks).select_related('seguia'))
    if not troncons:
        return JsonResponse({'erreur': 'Aucun tronçon trouvé pour les PKs fournis'}, status=404)

    resultats = [geotools.manning_troncon(tr, n_override=n_user, pente=pente) for tr in troncons]
    return JsonResponse({'resultats': resultats})


@api_login_required
@require_POST
def outil_scoring(request):
    """
    POST {
      couche       : str,
      pks          : [int, ...] | null,          ← null → toute la couche
      coefficients : {champ: float 0-5, ...},
      n_classes    : int (2-5, défaut 3),
      methode      : 'jenks' | 'quantile' (défaut 'jenks')
    }
    → {
        resultats : [{pk, score, classe}, ...],
        breaks    : [float, ...],
        n_classes : int,
        methode   : str
      }
    """
    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'erreur': 'Corps JSON invalide'}, status=400)

    couche       = body.get('couche', '').strip()
    pks          = body.get('pks') or None
    coefficients = body.get('coefficients')
    n_classes    = int(body.get('n_classes', 3))
    methode      = body.get('methode', 'jenks')

    if couche not in LAYER_REGISTRY:
        return JsonResponse({'erreur': f'Couche inconnue : {couche}'}, status=404)

    if not isinstance(coefficients, dict) or not coefficients:
        return JsonResponse({'erreur': 'coefficients doit être un objet {champ: valeur} non vide'}, status=400)

    try:
        coefficients = {k: float(v) for k, v in coefficients.items()}
    except (TypeError, ValueError):
        return JsonResponse({'erreur': 'Toutes les valeurs de coefficients doivent être des nombres'}, status=400)

    if not all(0 <= v <= 5 for v in coefficients.values()):
        return JsonResponse({'erreur': 'Les coefficients doivent être compris entre 0 et 5'}, status=400)

    if n_classes not in range(2, 6):
        return JsonResponse({'erreur': 'n_classes doit être compris entre 2 et 5'}, status=400)

    if methode not in ('jenks', 'quantile'):
        return JsonResponse({'erreur': "methode doit être 'jenks' ou 'quantile'"}, status=400)

    try:
        resultats, breaks = geotools.scoring(
            LAYER_REGISTRY, couche, coefficients, pks, n_classes, methode,
        )
    except Exception as exc:
        return JsonResponse({'erreur': str(exc)}, status=500)

    return JsonResponse({
        'resultats': resultats,
        'breaks':    breaks,
        'n_classes': n_classes,
        'methode':   methode,
    })


# ── §7.5 Exports ──────────────────────────────────────────────────────────────

@api_login_required
@role_required(*_EXP)
@require_POST
def export_csv(request):
    """
    POST { couche, pks?, champs? }
    → StreamingHttpResponse text/csv (UTF-8 BOM pour compatibilité Excel).

    champs : liste des colonnes à exporter (TA-13).
             Si absent/vide → toutes les colonnes déclarées dans LAYER_REGISTRY.
    pks    : liste de PKs pour export sélection (TA-12).
             Si absent/vide → export de toute la couche.
    """
    import io as _io
    from django.http import StreamingHttpResponse

    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'erreur': 'Corps JSON invalide'}, status=400)

    couche = body.get('couche', '').strip()
    pks    = body.get('pks')  or None
    champs = body.get('champs') or None

    if couche not in LAYER_REGISTRY:
        return JsonResponse({'erreur': f'Couche inconnue : {couche}'}, status=404)

    meta        = LAYER_REGISTRY[couche]
    app_label, model_name = meta['model'].split('.')
    Model       = apps.get_model(app_label, model_name)
    all_fields  = meta.get('fields', [])

    # Validation colonnes (TA-13) — seuls les champs déclarés sont autorisés
    if champs:
        fields = [f for f in champs if f in all_fields]
    else:
        fields = list(all_fields)

    qs = Model.objects.all()
    if pks:
        qs = qs.filter(pk__in=pks)
    qs = qs.order_by('pk')

    def _rows():
        buf = _io.StringIO()
        w   = csv.writer(buf)
        yield '﻿'                         # UTF-8 BOM
        w.writerow(['pk'] + fields)
        yield buf.getvalue(); buf.seek(0); buf.truncate()
        for obj in qs.values('pk', *fields).iterator(chunk_size=500):
            w.writerow(
                [obj.get('pk', '')] +
                ['' if obj.get(f) is None else obj.get(f) for f in fields]
            )
            yield buf.getvalue(); buf.seek(0); buf.truncate()

    suffix = '_selection' if pks else ''
    resp   = StreamingHttpResponse(_rows(), content_type='text/csv; charset=utf-8')
    resp['Content-Disposition'] = f'attachment; filename="{couche}{suffix}.csv"'
    return resp


@api_login_required
@role_required(*_EXP)
@require_POST
def export_excel(request):
    """
    POST { couche, pks?, champs? }
    → StreamingHttpResponse application/xlsx (openpyxl).

    Mise en forme :
      - En-tête : gras + fond sombre + texte blanc
      - Ligne alternative beige clair
      - Auto-filter sur l'en-tête
      - Première ligne figée (freeze_panes)
      - Largeurs de colonnes adaptées au contenu
    """
    import io as _io
    from django.http import StreamingHttpResponse

    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'erreur': 'Corps JSON invalide'}, status=400)

    couche = body.get('couche', '').strip()
    pks    = body.get('pks')   or None
    champs = body.get('champs') or None

    if couche not in LAYER_REGISTRY:
        return JsonResponse({'erreur': f'Couche inconnue : {couche}'}, status=404)

    meta       = LAYER_REGISTRY[couche]
    app_label, model_name = meta['model'].split('.')
    Model      = apps.get_model(app_label, model_name)
    all_fields = meta.get('fields', [])

    if champs:
        fields = [f for f in champs if f in all_fields]
    else:
        fields = list(all_fields)

    qs = Model.objects.all()
    if pks:
        qs = qs.filter(pk__in=pks)
    qs = qs.order_by('pk')

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return JsonResponse({'erreur': 'openpyxl non disponible'}, status=500)

    DARK  = '1A1A2E'
    GOLD  = 'F0A500'
    ALT   = 'FDF9F4'   # beige alternance
    BORDER_COLOR = 'E0D0C0'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = couche[:31]

    header_cols = ['pk'] + fields

    thin = Border(
        left  = Side(style='thin', color=BORDER_COLOR),
        right = Side(style='thin', color=BORDER_COLOR),
        top   = Side(style='thin', color=BORDER_COLOR),
        bottom= Side(style='thin', color=BORDER_COLOR),
    )

    # ── En-tête (ligne 1) ───────────────────────────────────────────────────
    hdr_font = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
    hdr_fill = PatternFill('solid', fgColor=DARK)
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=False)

    for col_idx, col_name in enumerate(header_cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font   = hdr_font
        cell.fill   = hdr_fill
        cell.alignment = hdr_align
        cell.border = thin

    # ── Données ────────────────────────────────────────────────────────────
    data_font  = Font(name='Calibri', size=9.5)
    alt_fill   = PatternFill('solid', fgColor=ALT)
    data_align = Alignment(vertical='center')

    col_widths = [max(len(c), 4) for c in header_cols]

    for row_idx, obj in enumerate(
        qs.values('pk', *fields).iterator(chunk_size=500), start=2
    ):
        row_vals = [obj.get('pk', '')] + [
            '' if obj.get(f) is None else obj.get(f) for f in fields
        ]
        for col_idx, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font   = data_font
            cell.border = thin
            cell.alignment = data_align
            if row_idx % 2 == 0:
                cell.fill = alt_fill
            col_widths[col_idx - 1] = max(
                col_widths[col_idx - 1], len(str(val)) if val is not None else 0
            )

    # ── Mise en page ────────────────────────────────────────────────────────
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes    = 'A2'

    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = min(w + 3, 40)

    ws.row_dimensions[1].height = 22

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    def _stream():
        while chunk := buf.read(65536):
            yield chunk

    suffix      = '_selection' if pks else ''
    mime        = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    resp        = StreamingHttpResponse(_stream(), content_type=mime)
    resp['Content-Disposition'] = f'attachment; filename="{couche}{suffix}.xlsx"'
    return resp


@api_login_required
@role_required(*_EXP)
@require_POST
def export_geojson(request):
    raise NotImplementedError


@api_login_required
@role_required(*_EXP)
@require_POST
def export_carte(request):
    """
    POST {
      format        : 'A4'|'A3'|'A2'|'A1'|'A0',
      orientation   : 'portrait'|'landscape',
      dpi           : 72|150|300,
      map_image     : 'data:image/png;base64,...',
      bbox          : [xmin, ymin, xmax, ymax]  (EPSG:4326),
      legende_items : [{label, color, geom_type}, ...],
      elements      : {titre, legende, nord, echelle, date}
    }
    → StreamingHttpResponse(application/pdf)
    """
    import base64, io, math
    from datetime import date as _today
    from django.http import StreamingHttpResponse

    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'erreur': 'Corps JSON invalide'}, status=400)

    fmt           = body.get('format', 'A4').upper()
    orientation   = body.get('orientation', 'landscape').lower()
    map_b64       = body.get('map_image', '')
    bbox          = body.get('bbox')
    legende_items = body.get('legende_items', [])
    el            = body.get('elements', {})
    titre         = (el.get('titre') or body.get('titre') or '').strip()
    sous_titre    = (body.get('sous_titre') or '').strip()
    source_txt    = (body.get('source') or '').strip()
    logos_cfg     = body.get('logos', {})
    show_legende  = bool(el.get('legende', True))
    show_nord     = bool(el.get('nord',    True))
    show_echelle  = bool(el.get('echelle', True))
    show_date     = bool(el.get('date',    True))

    try:
        from reportlab.lib.pagesizes import A4, A3, A2, A1, A0
        from reportlab.lib.pagesizes import landscape as rl_land, portrait as rl_port
        from reportlab.lib.units import mm
        from reportlab.pdfgen import canvas as rl_canvas
        from reportlab.lib import colors
        from reportlab.lib.utils import ImageReader
    except ImportError:
        return JsonResponse({'erreur': 'ReportLab non disponible'}, status=500)

    SIZES = {'A4': A4, 'A3': A3, 'A2': A2, 'A1': A1, 'A0': A0}
    psize = SIZES.get(fmt, A4)
    psize = rl_land(psize) if orientation == 'landscape' else rl_port(psize)
    pw, ph = psize

    MARGIN   = 10 * mm
    TITLE_H  = (16 if titre else 0) * mm + (6 if sous_titre else 0) * mm
    FOOTER_H = (16 if (show_legende or show_date or logos_cfg) else 0) * mm + \
               (5 if source_txt else 0) * mm

    mx, my = MARGIN, MARGIN + FOOTER_H
    mw, mh = pw - 2 * MARGIN, ph - 2 * MARGIN - TITLE_H - FOOTER_H

    buf = io.BytesIO()
    c   = rl_canvas.Canvas(buf, pagesize=psize)
    c.setTitle(titre or 'Carte')

    # ── Fond carte ────────────────────────────────────────────────────────────
    if map_b64:
        try:
            raw = map_b64.split(',', 1)[-1] if ',' in map_b64 else map_b64
            img_buf      = io.BytesIO(base64.b64decode(raw))
            img_buf.name = 'map.png'
            c.drawImage(ImageReader(img_buf), mx, my, mw, mh,
                        preserveAspectRatio=False)
        except Exception:
            c.setFillColor(colors.HexColor('#e8e0d0'))
            c.rect(mx, my, mw, mh, fill=1, stroke=0)
    else:
        c.setFillColor(colors.HexColor('#e8e0d0'))
        c.rect(mx, my, mw, mh, fill=1, stroke=0)

    c.setStrokeColor(colors.HexColor('#8b7355'))
    c.setLineWidth(0.7)
    c.rect(mx, my, mw, mh, fill=0, stroke=1)

    # ── Titre + sous-titre ────────────────────────────────────────────────────
    if titre:
        ty = ph - MARGIN - TITLE_H
        c.setFillColor(colors.HexColor('#1A1A2E'))
        c.rect(mx, ty, mw, TITLE_H, fill=1, stroke=0)
        c.setFillColor(colors.white)
        base_h = (16 if titre else 0) * mm
        fsize = 13 if len(titre) < 60 else 10
        c.setFont('Helvetica-Bold', fsize)
        titre_y = ty + base_h / 2 - fsize / 3 + (3 * mm if sous_titre else 0)
        c.drawCentredString(pw / 2, titre_y, titre)
        if sous_titre:
            c.setFont('Helvetica', 8)
            c.drawCentredString(pw / 2, ty + 2 * mm, sous_titre)

    # ── Flèche Nord ───────────────────────────────────────────────────────────
    if show_nord:
        _pdf_north_arrow(c, mx + mw - 14 * mm, my + mh - 22 * mm, 4.5 * mm)

    # ── Barre d'échelle ───────────────────────────────────────────────────────
    if show_echelle and bbox and len(bbox) == 4:
        _pdf_scale_bar(c, mx + 8 * mm, my + 7 * mm, mw * 0.28, bbox)

    # ── Footer ────────────────────────────────────────────────────────────────
    if FOOTER_H:
        c.setStrokeColor(colors.HexColor('#8b7355'))
        c.setLineWidth(0.4)
        c.line(mx, MARGIN + FOOTER_H, mx + mw, MARGIN + FOOTER_H)

        if show_legende and legende_items:
            _pdf_legend(c, mx + 2 * mm, MARGIN + 2 * mm,
                        min(mw * 0.65, 130 * mm), FOOTER_H - 4 * mm,
                        legende_items)

        if show_date:
            c.setFillColor(colors.HexColor('#666666'))
            c.setFont('Helvetica', 7.5)
            c.drawRightString(mx + mw, MARGIN + 2,
                              f"Exporté le {_today.today().strftime('%d/%m/%Y')}")

        if source_txt:
            c.setFillColor(colors.HexColor('#666666'))
            c.setFont('Helvetica-Oblique', 7)
            c.drawString(mx, MARGIN + 2, source_txt)

        # ── Logos ─────────────────────────────────────────────────────────────
        if logos_cfg:
            from django.conf import settings as djsettings
            LOGO_FILES = {
                'hydroplan_icone': 'admin/img/logo1.png',
                'hydroplan_texte': 'admin/img/logo2.png',
                'sgiat':           'admin/img/logo_sgiat.png',
                'iav':             'admin/img/logo_iav.png',
            }
            static_dir = djsettings.STATICFILES_DIRS[0] if djsettings.STATICFILES_DIRS else None
            logo_h_pt  = 20  # hauteur cible en pts
            logo_x     = mx + mw - 2 * mm
            logo_margin = 2 * mm
            for key, enabled in reversed(list(logos_cfg.items())):
                if not enabled or key not in LOGO_FILES:
                    continue
                if not static_dir:
                    continue
                logo_path = str(static_dir / LOGO_FILES[key]) if hasattr(static_dir, '__truediv__') \
                            else str(static_dir) + '/' + LOGO_FILES[key]
                try:
                    img_reader = ImageReader(logo_path)
                    iw, ih     = img_reader.getSize()
                    ratio      = logo_h_pt / ih if ih else 1
                    logo_w     = iw * ratio
                    logo_x    -= logo_w
                    c.drawImage(img_reader, logo_x, MARGIN + 3,
                                width=logo_w, height=logo_h_pt,
                                preserveAspectRatio=True, mask='auto')
                    logo_x -= logo_margin
                except Exception:
                    pass

    # ── Cadre général ─────────────────────────────────────────────────────────
    c.setStrokeColor(colors.HexColor('#8b7355'))
    c.setLineWidth(1.2)
    c.rect(MARGIN / 2, MARGIN / 2, pw - MARGIN, ph - MARGIN, fill=0, stroke=1)

    c.save()
    buf.seek(0)

    fname = f'carte_{fmt.lower()}_{orientation}.pdf'

    def _stream():
        while chunk := buf.read(65536):
            yield chunk

    resp = StreamingHttpResponse(_stream(), content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp


# ── ReportLab helpers (utilisés uniquement par export_carte) ──────────────────

def _pdf_north_arrow(c, cx, cy, r):
    """Flèche bicolore N/S avec lettre N."""
    from reportlab.lib import colors

    tip_y   = cy + r * 2.4
    base_y  = cy - r * 1.0
    half_w  = r * 0.55

    p_dark = c.beginPath()
    p_dark.moveTo(cx, tip_y)
    p_dark.lineTo(cx - half_w, base_y)
    p_dark.lineTo(cx, cy - r * 0.05)
    p_dark.close()
    c.setFillColor(colors.HexColor('#1A1A2E'))
    c.drawPath(p_dark, fill=1, stroke=0)

    p_grey = c.beginPath()
    p_grey.moveTo(cx, tip_y)
    p_grey.lineTo(cx + half_w, base_y)
    p_grey.lineTo(cx, cy - r * 0.05)
    p_grey.close()
    c.setFillColor(colors.HexColor('#aaaaaa'))
    c.drawPath(p_grey, fill=1, stroke=0)

    p_out = c.beginPath()
    p_out.moveTo(cx, tip_y)
    p_out.lineTo(cx - half_w, base_y)
    p_out.lineTo(cx, cy - r * 0.05)
    p_out.lineTo(cx + half_w, base_y)
    p_out.close()
    c.setStrokeColor(colors.HexColor('#1A1A2E'))
    c.setLineWidth(0.4)
    c.drawPath(p_out, fill=0, stroke=1)

    c.setFillColor(colors.HexColor('#1A1A2E'))
    c.setFont('Helvetica-Bold', r * 1.6)
    c.drawCentredString(cx, tip_y + r * 0.4, 'N')


def _pdf_scale_bar(c, x, y, max_w, bbox):
    """Barre d'échelle graphique alternée."""
    import math
    from reportlab.lib import colors
    from reportlab.lib.units import mm

    xmin, ymin, xmax, ymax = bbox
    lat_c  = (ymin + ymax) / 2
    lon_m  = (xmax - xmin) * math.cos(math.radians(lat_c)) * 111_320
    if lon_m <= 0 or max_w <= 0:
        return

    m_per_pt   = lon_m / max_w
    target_m   = m_per_pt * max_w / 3
    if target_m <= 0:
        return
    exp    = math.floor(math.log10(max(target_m, 0.001)))
    step   = 10 ** exp
    bar_m  = min([step, 2*step, 5*step, 10*step], key=lambda s: abs(s - target_m))
    bar_w  = bar_m / m_per_pt
    if bar_w <= 0 or bar_w > max_w:
        return

    bar_h = 2 * mm
    n     = 4
    sw, sm = bar_w / n, bar_m / n

    def _fmt(m):
        if m >= 1000:
            v = m / 1000
            return (f'{v:.0f}' if v == int(v) else f'{v:.1f}') + ' km'
        return f'{int(m)} m'

    for i in range(n):
        c.setFillColor(colors.HexColor('#1A1A2E') if i % 2 == 0 else colors.white)
        c.setStrokeColor(colors.HexColor('#1A1A2E'))
        c.setLineWidth(0.4)
        c.rect(x + i * sw, y, sw, bar_h, fill=1, stroke=1)

    c.setFillColor(colors.HexColor('#1A1A2E'))
    c.setFont('Helvetica', 7)
    c.drawCentredString(x, y - 3.5, '0')
    for i in range(1, n + 1):
        c.drawCentredString(x + i * sw, y - 3.5, _fmt(i * sm))


def _pdf_legend(c, x, y, w, h, items):
    """Légende compacte : symbole + texte par couche."""
    from reportlab.lib import colors
    from reportlab.lib.units import mm

    c.setFont('Helvetica-Bold', 8)
    c.setFillColor(colors.HexColor('#1A1A2E'))
    c.drawString(x, y + h - 2.5 * mm, 'Légende')

    usable = h - 4.5 * mm
    row_h  = usable / max(len(items[:8]), 1)
    sym_s  = min(row_h * 0.55, 2.8 * mm)

    for i, item in enumerate(items[:8]):
        iy  = y + h - 5 * mm - i * row_h
        col = item.get('color', '#888888')
        gt  = (item.get('geom_type') or '').lower()
        lbl = (item.get('label') or '')[:35]

        c.setFillColor(colors.HexColor(col))
        if 'point' in gt or gt in ('', 'geometry'):
            c.circle(x + sym_s / 2, iy + sym_s / 2, sym_s / 2, fill=1, stroke=0)
        elif 'line' in gt:
            c.setStrokeColor(colors.HexColor(col))
            c.setLineWidth(1.5)
            c.line(x, iy + sym_s / 2, x + sym_s, iy + sym_s / 2)
        else:
            c.setStrokeColor(colors.HexColor(col))
            c.setLineWidth(0.3)
            c.rect(x, iy, sym_s, sym_s, fill=1, stroke=1)

        c.setFillColor(colors.HexColor('#333333'))
        c.setFont('Helvetica', 7.5)
        c.drawString(x + sym_s + 3, iy + 1, lbl)


@api_login_required
@role_required(*_EXP)
@require_POST
def export_dashboard(request):
    raise NotImplementedError


# ── §9 FEATURE-C2 — Tableaux enfants ouvrages d'un périmètre ────────────────

# Mapping : type URL → related_name sur Perimetre
_PERIMETRE_OUVRAGE_RELATED = {
    'seuils':           'seuils',
    'murs_protection':  'murs_protection',
    'troncons_seguias': 'seguias',
    'barrages':         'barrages_retenue',
    'khettaras':        'khettaras',
    'forages_puits':    'forages_puits',
    'prises_locales':   'prises_locales',
}
# FK-only fields à ne pas inclure dans les .values() simples
_FK_FIELDS = frozenset({'bassin_versant', 'province', 'perimetre', 'commune'})


@api_login_required
@require_GET
def perimetre_ouvrages(request, pk, type_ouvrage):
    """
    GET /carte/api/perimetre/<pk>/ouvrages/<type>/
    → { type, pk_perimetre, count, fields, ouvrages: [{pk, ...}] }
    """
    from django.shortcuts import get_object_or_404
    from diagnostic.models import Perimetre

    if type_ouvrage not in _PERIMETRE_OUVRAGE_RELATED:
        return JsonResponse({'erreur': f'Type inconnu : {type_ouvrage}'}, status=400)

    p            = get_object_or_404(Perimetre, pk=pk)
    related_name = _PERIMETRE_OUVRAGE_RELATED[type_ouvrage]
    meta         = LAYER_REGISTRY.get(type_ouvrage, {})
    fields       = [f for f in meta.get('fields', []) if f not in _FK_FIELDS]

    qs = getattr(p, related_name).all()

    ouvrages = []
    for obj in qs.values('pk', *fields):
        ouvrages.append(obj)

    return JsonResponse({
        'type':          type_ouvrage,
        'pk_perimetre':  pk,
        'count':         len(ouvrages),
        'fields':        ['pk'] + fields,
        'label':         meta.get('label', type_ouvrage),
        'ouvrages':      ouvrages,
    })


# ── §8 Analyse — Périmètre irrigué ───────────────────────────────────────────

@api_login_required
@require_GET
def perimetre_rendement(request, pk):
    """
    GET /carte/api/perimetre/<pk>/rendement/
    → { rendement_pondere, culture_dominante, total_surface_ha, assolement: [...] }
    """
    from django.shortcuts import get_object_or_404
    from diagnostic.models import Perimetre

    p      = get_object_or_404(Perimetre, pk=pk)
    lignes = list(
        p.assolement.values('culture', 'surface_ha', 'rendement', 'pourcentage', 'unite_rendement')
    )

    total_surface = sum(l['surface_ha'] or 0 for l in lignes)
    total_pond    = sum((l['surface_ha'] or 0) * (l['rendement'] or 0) for l in lignes)

    rendement_pondere = round(total_pond / total_surface, 2) if total_surface else None
    culture_dominante = (
        max(lignes, key=lambda l: l['surface_ha'] or 0)['culture'] if lignes else None
    )

    return JsonResponse({
        'rendement_pondere':  rendement_pondere,
        'culture_dominante':  culture_dominante,
        'total_surface_ha':   round(total_surface, 2),
        'assolement':         lignes,
    })


@api_login_required
@require_GET
def perimetre_tours_eau(request, pk):
    """
    GET /carte/api/perimetre/<pk>/tours-eau/
    → { count, tours: [{ayant_droit, cycle_jours, duree_heures}] }
    """
    from django.shortcuts import get_object_or_404
    from diagnostic.models import Perimetre

    p    = get_object_or_404(Perimetre, pk=pk)
    tours = list(p.tours_eau.values('ayant_droit', 'cycle_jours', 'duree_heures'))
    return JsonResponse({'count': len(tours), 'tours': tours})


@api_login_required
@require_GET
def perimetre_volume_bilan(request, pk):
    """
    GET /carte/api/perimetre/<pk>/volume-bilan/
    → { volume_annee_humide/normale/seche, volume_excedent_deficit_humide/normale/seche }
    """
    from django.shortcuts import get_object_or_404
    from diagnostic.models import Perimetre

    p = get_object_or_404(Perimetre, pk=pk)
    return JsonResponse({
        'volume_annee_humide':             p.volume_annee_humide,
        'volume_annee_normale':            p.volume_annee_normale,
        'volume_annee_seche':              p.volume_annee_seche,
        'volume_excedent_deficit_humide':  p.volume_excedent_deficit_humide,
        'volume_excedent_deficit_normale': p.volume_excedent_deficit_normale,
        'volume_excedent_deficit_seche':   p.volume_excedent_deficit_seche,
    })


# Champ volume par type d'année (outil « Besoin » du panneau droit)
_BESOIN_CHAMP_ANNEE = {
    'humide':  'volume_annee_humide',
    'normale': 'volume_annee_normale',
    'seche':   'volume_annee_seche',
}


@api_login_required
@require_GET
def perimetres_besoin_points(request):
    """
    GET /carte/api/perimetres/besoin/?annee=normale&pks=1,2,3
    → FeatureCollection de Points (point_on_surface, EPSG:4326), un par périmètre
      ayant une valeur de besoin non nulle pour l'année demandée.

    Propriétés : { pk, nom, value, v_humide, v_normale, v_seche }.
    - value      : valeur de l'année demandée (modes point / cercle / choroplèthe)
    - v_*        : les trois valeurs annuelles (mode camembert / « cycle »)
    Le point-sur-surface est garanti à l'intérieur du polygone (≠ centroïde
    qui peut tomber dehors sur une forme concave) — valeur affichée au centre.
    """
    from diagnostic.models import Perimetre

    annee = request.GET.get('annee', 'normale').strip().lower()
    champ = _BESOIN_CHAMP_ANNEE.get(annee)
    if not champ:
        return JsonResponse(
            {'erreur': "annee doit être 'humide', 'normale' ou 'seche'"},
            status=400,
        )

    qs = (
        Perimetre.objects
        .exclude(geometrie__isnull=True)
        .exclude(**{f'{champ}__isnull': True})
    )

    pks_param = request.GET.get('pks')
    if pks_param:
        try:
            pk_list = [int(p.strip()) for p in pks_param.split(',') if p.strip()]
        except ValueError:
            return JsonResponse({'erreur': 'pks invalide — entiers séparés par virgules'}, status=400)
        qs = qs.filter(pk__in=pk_list)

    features = []
    for p in qs.only(
        'pk', 'ksar_village', 'geometrie',
        'volume_annee_humide', 'volume_annee_normale', 'volume_annee_seche',
    ):
        geom = p.geometrie
        try:
            pt = geom.point_on_surface           # garanti à l'intérieur
        except Exception:
            pt = geom.centroid                   # repli
        if pt is None:
            continue
        if pt.srid and pt.srid != 4326:
            pt.transform(4326)
        features.append({
            'type':       'Feature',
            'geometry':   {'type': 'Point', 'coordinates': [pt.x, pt.y]},
            'properties': {
                'pk':        p.pk,
                'nom':       p.ksar_village or f'Périmètre #{p.pk}',
                'value':     getattr(p, champ),
                'v_humide':  p.volume_annee_humide,
                'v_normale': p.volume_annee_normale,
                'v_seche':   p.volume_annee_seche,
            },
        })

    return JsonResponse({
        'type':     'FeatureCollection',
        'annee':    annee,
        'count':    len(features),
        'features': features,
    })


# (champ volume besoin, champ excédent/déficit) par type d'année
_COMPARAISON_CHAMP_ANNEE = {
    'humide':  ('volume_annee_humide',  'volume_excedent_deficit_humide'),
    'normale': ('volume_annee_normale', 'volume_excedent_deficit_normale'),
    'seche':   ('volume_annee_seche',   'volume_excedent_deficit_seche'),
}
_COMPARAISON_MAX = 25   # plafond de l'outil « Comparaison besoin »


@api_login_required
@require_GET
def perimetres_comparaison_besoin(request):
    """
    GET /carte/api/perimetres/comparaison-besoin/?annee=normale&pks=1,2,3
    → { annee, count, total, tronque, perimetres: [{pk, nom, besoin,
         excedent, deficit, solde}, ...] }

    Données de l'outil « Comparaison besoin » (fenêtre flottante du template
    carte) — valeurs lues telles quelles dans diagnostic.Perimetre, aucun calcul.

    Conventions de signe (volume_excedent_deficit_* est signé) :
      - excedent : part positive  (≥ 0)
      - deficit  : part négative  (≤ 0)  — le déficit est TOUJOURS négatif
      - solde    : valeur signée brute (+ excédent / − déficit)
    Plafonné à 25 périmètres (les 25 premiers).
    """
    from diagnostic.models import Perimetre

    annee = request.GET.get('annee', 'normale').strip().lower()
    champs = _COMPARAISON_CHAMP_ANNEE.get(annee)
    if not champs:
        return JsonResponse(
            {'erreur': "annee doit être 'humide', 'normale' ou 'seche'"},
            status=400,
        )
    champ_besoin, champ_ed = champs

    qs = Perimetre.objects.exclude(**{f'{champ_besoin}__isnull': True})

    pks_param = request.GET.get('pks')
    if pks_param:
        try:
            pk_list = [int(p.strip()) for p in pks_param.split(',') if p.strip()]
        except ValueError:
            return JsonResponse({'erreur': 'pks invalide — entiers séparés par virgules'}, status=400)
        qs = qs.filter(pk__in=pk_list)

    qs = qs.order_by('pk')
    total = qs.count()

    perimetres = []
    for p in qs.only('pk', 'ksar_village', champ_besoin, champ_ed)[:_COMPARAISON_MAX]:
        besoin = getattr(p, champ_besoin) or 0.0
        ed = getattr(p, champ_ed) or 0.0
        perimetres.append({
            'pk':       p.pk,
            'nom':      p.ksar_village or f'Périmètre #{p.pk}',
            'besoin':   besoin,
            'excedent': ed if ed > 0 else 0.0,    # ≥ 0
            'deficit':  ed if ed < 0 else 0.0,    # ≤ 0 (toujours négatif)
            'solde':    ed,
        })

    return JsonResponse({
        'annee':      annee,
        'count':      len(perimetres),
        'total':      total,
        'tronque':    total > _COMPARAISON_MAX,
        'perimetres': perimetres,
    })


# ── Box « Hydrologie / Crues » (Lot 1) ────────────────────────────────────────

@api_login_required
@require_GET
def couche_liste(request, nom):
    """
    GET /carte/api/couche/<nom>/liste/ → { options: [{pk, label}] }

    Liste légère (pk + libellé) pour alimenter les menus déroulants des outils,
    y compris les couches `hidden` (ex. bv_ouvrage_tete). Le libellé suit
    `label_field` du registre (ou le 1er champ déclaré).
    """
    if nom not in LAYER_REGISTRY:
        return JsonResponse({'erreur': f'Couche inconnue : {nom}'}, status=404)

    meta = LAYER_REGISTRY[nom]
    app_label, model_name = meta['model'].split('.')
    Model = apps.get_model(app_label, model_name)
    label_field = meta.get('label_field') or (meta.get('fields') or ['pk'])[0]

    qs = Model.objects.all()
    try:
        qs = qs.order_by(label_field)
    except Exception:
        pass

    options = []
    for row in qs.values('pk', label_field):
        options.append({'pk': row['pk'], 'label': str(row.get(label_field) or f"#{row['pk']}")})
    return JsonResponse({'options': options})


@api_login_required
@require_GET
def bv_crue_periodes(request, pk):
    """
    GET /carte/api/bv/<pk>/crue-periodes/
    → { bv, date_analyse, methode, tc_min, periodes: [{T, q}] }

    Lit la DERNIÈRE analyse hydrologique du bassin versant (lecture seule —
    le détail/recalcul reste dans l'app analyse_hydrologique).
    """
    from django.shortcuts import get_object_or_404
    from analyse_hydrologique.models import BassinVersant, ResultatAnalyseHydrologique

    bv = get_object_or_404(BassinVersant, pk=pk)
    analyse = (ResultatAnalyseHydrologique.objects
               .filter(bassin_versant=bv).order_by('-date_analyse').first())
    if not analyse:
        return JsonResponse({'bv': bv.nom, 'periodes': [],
                             'message': "Aucune analyse hydrologique pour ce bassin versant."})

    periodes = [{'T': T, 'q': getattr(analyse, champ)}
                for T, champ in ((10, 'qcrue_t10'), (20, 'qcrue_t20'),
                                 (50, 'qcrue_t50'), (100, 'qcrue_t100'))]
    return JsonResponse({
        'bv':           bv.nom,
        'date_analyse': analyse.date_analyse.strftime('%d/%m/%Y') if analyse.date_analyse else None,
        'methode':      analyse.get_methode_display() if analyse.methode else '—',
        'tc_min':       analyse.temps_concentration,
        'periodes':     periodes,
    })


@api_login_required
@require_GET
def bv_tc(request, pk):
    """
    GET /carte/api/bv/<pk>/tc/
    → { bv, formules: [{nom, tc_min}], moyenne_min, moyenne_h }

    Calcule le temps de concentration par formule via
    analyse_hydrologique.calculs (import des fonctions de l'app).
    """
    from django.shortcuts import get_object_or_404
    from analyse_hydrologique.models import BassinVersant

    bv = get_object_or_404(BassinVersant, pk=pk)
    try:
        from analyse_hydrologique.calculs import bv_to_hydro, FORMULES_TC_DISPONIBLES
        from hydrologie_bv import calculer_tc_bv
        tc = calculer_tc_bv(bv_to_hydro(bv), FORMULES_TC_DISPONIBLES, verbose=False)
    except Exception as exc:
        return JsonResponse({'bv': bv.nom, 'formules': [],
                             'erreur': f"Tc non calculable (données BV incomplètes) : {exc}"})

    moyenne = tc.get('Moyenne')
    formules = [{'nom': k, 'tc_min': round(v, 2)} for k, v in tc.items() if k != 'Moyenne']
    return JsonResponse({
        'bv':          bv.nom,
        'formules':    formules,
        'moyenne_min': round(moyenne, 2) if moyenne is not None else None,
        'moyenne_h':   round(moyenne / 60.0, 3) if moyenne is not None else None,
    })


def _compute_apports_crue(bv, st_param=None, tc_param=None):
    """
    Apports de crue mensuels d'un BV (12 mois Sep→Aoû × 3 années), via
    calculer_apports_crue_sans_prelevement (transposition Francou-Rodier).

      - station auto si `st_param` absent : station hydrométrique dans le BV,
        sinon la plus proche du centroïde (override possible via st_param) ;
      - Tc auto (moyenne des formules) si `tc_param` absent.

    Retourne le dict de calcul enrichi (bv, station, tc_h, auto_station) ou
    {'erreur': …}.
    """
    from analyse_hydrologique.models import StationHydrometrique
    from analyse_hydrologique.calculs import (
        calculer_apports_crue_sans_prelevement, bv_to_hydro, FORMULES_TC_DISPONIBLES,
    )

    if not bv:
        return {'erreur': 'Aucun bassin versant associé à cet ouvrage.'}

    auto = False
    station = StationHydrometrique.objects.filter(pk=st_param).first() if st_param else None
    if station is None:
        auto = True
        if bv.geometrie:
            station = StationHydrometrique.objects.filter(geometrie__within=bv.geometrie).first()
            if station is None:
                from django.contrib.gis.db.models.functions import Distance
                station = (StationHydrometrique.objects
                           .exclude(geometrie__isnull=True)
                           .annotate(_d=Distance('geometrie', bv.geometrie.centroid))
                           .order_by('_d').first())
    if station is None:
        return {'erreur': 'Aucune station hydrométrique de référence disponible.'}

    tc_h = None
    if tc_param:
        try:
            tc_h = float(tc_param)
        except (TypeError, ValueError):
            tc_h = None
    if tc_h is None:
        try:
            from hydrologie_bv import calculer_tc_bv
            m = calculer_tc_bv(bv_to_hydro(bv), FORMULES_TC_DISPONIBLES, verbose=False).get('Moyenne')
            tc_h = (m / 60.0) if m else 1.0
        except Exception:
            tc_h = 1.0

    data = calculer_apports_crue_sans_prelevement(station, tc_h, bv.surface)
    data['bv'] = bv.nom
    data['station'] = station.nom
    data['auto_station'] = auto
    data.setdefault('tc_h', tc_h)
    return data


@api_login_required
@require_GET
def bv_apports_crue(request, pk):
    """
    GET /carte/api/bv/<pk>/apports-crue/?station=<pk>&tc=<heures>
    → apports de crue mensuels du BV cible (12 mois Sep→Aoû) × 3 années.

    station/tc sont désormais OPTIONNELS (station auto, Tc auto) — cf.
    _compute_apports_crue.
    """
    from django.shortcuts import get_object_or_404
    from analyse_hydrologique.models import BassinVersant

    bv = get_object_or_404(BassinVersant, pk=pk)
    data = _compute_apports_crue(bv, request.GET.get('station'), request.GET.get('tc'))
    return JsonResponse(data, status=200 if 'erreur' not in data else 400)


def _apport_volume_capte_m3_an(debit_l_s):
    """Volume capté annuel (m³/an) au droit de l'ouvrage à partir de son débit (l/s)."""
    if debit_l_s in (None, ''):
        return None
    try:
        return round(float(debit_l_s) * 31536.0, 0)   # 1 l/s sur 1 an = 31 536 m³
    except (TypeError, ValueError):
        return None


def _ouvrage_apport_response(request, ouvrage, type_, debit, bv):
    """Réponse apport-crue d'un ouvrage (Q6 = les deux) :
       apport de crue du BV (transposé) + volume capté au droit de l'ouvrage.
       Si l'ouvrage n'a pas de BV lié, renvoie quand même le volume capté."""
    info = {
        'ouvrage':            ouvrage,
        'type':               type_,
        'debit_l_s':          debit,
        'volume_capte_m3_an': _apport_volume_capte_m3_an(debit),
    }
    data = _compute_apports_crue(bv, request.GET.get('station'), request.GET.get('tc'))
    if 'erreur' in data:
        # Pas de BV/apport : on expose tout de même la part captée.
        return JsonResponse({**info, 'message': data['erreur'], 'mois': []})
    data.update(info)
    return JsonResponse(data)


@api_login_required
@require_GET
def seuil_apport_crue(request, pk):
    """GET /carte/api/seuil/<pk>/apport-crue/ — apport BV (transposé) + volume
    capté (debit_mobilise)."""
    from django.shortcuts import get_object_or_404
    from diagnostic.models import Seuil
    s = get_object_or_404(Seuil.objects.select_related('bassin_versant'), pk=pk)
    return _ouvrage_apport_response(request, s.nom_du_seuil or f'Seuil #{s.pk}',
                                    'seuil', s.debit_mobilise, s.bassin_versant)


@api_login_required
@require_GET
def prise_apport_crue(request, pk):
    """GET /carte/api/prise/<pk>/apport-crue/ — apport BV (transposé) + volume
    capté (debit_derive)."""
    from django.shortcuts import get_object_or_404
    from diagnostic.models import PriseLocale
    pl = get_object_or_404(PriseLocale.objects.select_related('bassin_versant'), pk=pk)
    return _ouvrage_apport_response(request, pl.nom or f'Prise #{pl.pk}',
                                    'prise', pl.debit_derive, pl.bassin_versant)


@api_login_required
@require_GET
def barrage_apport_crue(request, pk):
    """GET /carte/api/barrage/<pk>/apport-crue/ — apport BV (transposé) + volume
    dérivé (debit_derive)."""
    from django.shortcuts import get_object_or_404
    from diagnostic.models import BarrageRetenue
    b = get_object_or_404(BarrageRetenue.objects.select_related('bassin_versant'), pk=pk)
    return _ouvrage_apport_response(request, b.nom or f'Barrage #{b.pk}',
                                    'barrage', b.debit_derive, b.bassin_versant)


_CRUE_CHAMP_PERIODE = {10: 'qcrue_t10', 20: 'qcrue_t20', 50: 'qcrue_t50', 100: 'qcrue_t100'}


@api_login_required
@require_GET
def bv_crue_points(request):
    """
    GET /carte/api/bv/crue-points/?t=<10|20|50|100>&pks=<id,…>
    → FeatureCollection de Points (point_on_surface des BV) avec value = Q(T)

    Thématique « Crue de projet » (Module A). Lit la dernière analyse par BV.
    """
    from analyse_hydrologique.models import BassinVersant, ResultatAnalyseHydrologique

    try:
        T = int(request.GET.get('t', 100))
    except ValueError:
        T = 100
    champ = _CRUE_CHAMP_PERIODE.get(T)
    if not champ:
        return JsonResponse({'erreur': 'T doit être 10, 20, 50 ou 100'}, status=400)

    qs = BassinVersant.objects.exclude(geometrie__isnull=True)
    pks_param = request.GET.get('pks')
    if pks_param:
        try:
            pk_list = [int(p) for p in pks_param.split(',') if p.strip()]
            qs = qs.filter(pk__in=pk_list)
        except ValueError:
            return JsonResponse({'erreur': 'pks invalide'}, status=400)

    # Dernière analyse par BV (1 requête, sans N+1)
    latest = {}
    for a in (ResultatAnalyseHydrologique.objects
              .filter(bassin_versant__in=qs)
              .order_by('bassin_versant_id', '-date_analyse')):
        latest.setdefault(a.bassin_versant_id, a)

    features = []
    for bv in qs:
        analyse = latest.get(bv.pk)
        if not analyse:
            continue
        val = getattr(analyse, champ)
        if val is None:
            continue
        geom = bv.geometrie
        try:
            pt = geom.point_on_surface
        except Exception:
            pt = geom.centroid
        if pt is None:
            continue
        if pt.srid and pt.srid != 4326:
            pt.transform(4326)
        features.append({
            'type':       'Feature',
            'geometry':   {'type': 'Point', 'coordinates': [pt.x, pt.y]},
            'properties': {'pk': bv.pk, 'nom': bv.nom, 'value': val},
        })

    return JsonResponse({
        'type':     'FeatureCollection',
        'periode':  T,
        'count':    len(features),
        'features': features,
    })


# ── Box « Bilan eau » (Lot 2) ─────────────────────────────────────────────────

_BILAN_ANNEE_CHAMP = {
    'normale': 'resultats_bilan_normale',
    'humide':  'resultats_bilan_humide',
    'seche':   'resultats_bilan_seche',
}


@api_login_required
@require_GET
def perimetre_bilan_mensuel(request, pk):
    """
    GET /carte/api/perimetre/<pk>/bilan-mensuel/?annee=normale
    → 12 mois (Sep→Aoû) : besoins / ressources / déficit / excédent + totaux

    Lit le DERNIER bilan calculé du périmètre (lecture seule — le calcul reste
    dans l'app Besions_Ressources).
    """
    from django.shortcuts import get_object_or_404
    from diagnostic.models import Perimetre
    from Besions_Ressources.models import BilanBesoinRessources

    annee = request.GET.get('annee', 'normale').strip().lower()
    champ = _BILAN_ANNEE_CHAMP.get(annee)
    if not champ:
        return JsonResponse({'erreur': "annee doit être 'normale', 'humide' ou 'seche'"}, status=400)

    p = get_object_or_404(Perimetre, pk=pk)
    bilan = (BilanBesoinRessources.objects
             .filter(perimetre=p, **{f'{champ}__isnull': False})
             .order_by('-created_at').first())
    if not bilan:
        return JsonResponse({'perimetre': p.ksar_village or f'Périmètre #{p.pk}',
                             'annee': annee, 'mois': [],
                             'message': "Aucun bilan calculé pour ce périmètre (année demandée)."})

    r = getattr(bilan, champ) or {}
    return JsonResponse({
        'perimetre':        p.ksar_village or f'Périmètre #{p.pk}',
        'annee':            annee,
        'mois':             r.get('mois', []),
        'besoins_m3':       r.get('besoins_m3', []),
        'ressources_m3':    r.get('ressources_m3', []),
        'deficit_m3':       r.get('deficit_m3', []),
        'excedent_m3':      r.get('excedent_m3', []),
        'total_besoins':    r.get('total_besoins'),
        'total_ressources': r.get('total_ressources'),
        'total_deficit':    r.get('total_deficit'),
        'total_excedent':   r.get('total_excedent'),
    })


@api_login_required
@require_GET
def perimetres_couverture(request):
    """
    GET /carte/api/perimetres/couverture/?annee=normale&pks=<id,…>
    → FeatureCollection de Points : value = taux de couverture (%) =
      total_ressources / total_besoins × 100, par périmètre (dernier bilan).

    Thématique « Taux de couverture » (Module A).
    """
    from diagnostic.models import Perimetre
    from Besions_Ressources.models import BilanBesoinRessources

    annee = request.GET.get('annee', 'normale').strip().lower()
    champ = _BILAN_ANNEE_CHAMP.get(annee)
    if not champ:
        return JsonResponse({'erreur': "annee invalide"}, status=400)

    qs = Perimetre.objects.exclude(geometrie__isnull=True)
    pks_param = request.GET.get('pks')
    if pks_param:
        try:
            qs = qs.filter(pk__in=[int(p) for p in pks_param.split(',') if p.strip()])
        except ValueError:
            return JsonResponse({'erreur': 'pks invalide'}, status=400)

    # Dernier bilan (avec l'année calculée) par périmètre
    latest = {}
    for b in (BilanBesoinRessources.objects
              .filter(perimetre__in=qs, **{f'{champ}__isnull': False})
              .order_by('perimetre_id', '-created_at')):
        latest.setdefault(b.perimetre_id, b)

    features = []
    for p in qs:
        bilan = latest.get(p.pk)
        if not bilan:
            continue
        r = getattr(bilan, champ) or {}
        bes = r.get('total_besoins') or 0
        ress = r.get('total_ressources') or 0
        if not bes:
            continue
        geom = p.geometrie
        try:
            pt = geom.point_on_surface
        except Exception:
            pt = geom.centroid
        if pt is None:
            continue
        if pt.srid and pt.srid != 4326:
            pt.transform(4326)
        features.append({
            'type':       'Feature',
            'geometry':   {'type': 'Point', 'coordinates': [pt.x, pt.y]},
            'properties': {
                'pk':         p.pk,
                'nom':        p.ksar_village or f'Périmètre #{p.pk}',
                'value':      round(ress / bes * 100, 1),
                'besoins':    bes,
                'ressources': ress,
            },
        })

    return JsonResponse({
        'type':     'FeatureCollection',
        'annee':    annee,
        'count':    len(features),
        'features': features,
    })


@api_login_required
@require_GET
def station_clim_eto(request, pk):
    """
    GET /carte/api/station-clim/<pk>/eto/
    → ET0 mensuel (12 mois Sep→Aoû) via Besions_Ressources.calculs.calculer_eto.
    """
    from django.shortcuts import get_object_or_404
    from Besions_Ressources.models import StationClimatique
    from Besions_Ressources.calculs import calculer_eto

    s = get_object_or_404(StationClimatique, pk=pk)
    temps = s.temperatures_moyennes or []
    insol = s.taux_insolation or []
    if len(temps) != 12 or len(insol) != 12:
        return JsonResponse({'station': s.nom, 'mois': [],
                             'message': "Données mensuelles incomplètes (12 valeurs requises)."})

    data = calculer_eto(temps, insol, s.latitude)
    data['station'] = s.nom
    data['latitude'] = s.latitude
    return JsonResponse(data)


# ── Box « Efficience réseau » (Lot 3) ─────────────────────────────────────────

@api_login_required
@require_GET
def efficiences_liste(request):
    """
    GET /carte/api/efficiences/liste/
    → { options: [{pk, label, globale, principale, secondaire, tertiaire,
         nb_p, nb_s, nb_t, ouvrage, perimetre}] }

    Dernier résultat Efficience par ouvrage de tête (cascade P/S/T → globale).
    """
    from efficiences.models import Efficience

    seen, options = set(), []
    for e in (Efficience.objects.select_related('perimetre')
              .order_by('ouvrage_tete_type', 'ouvrage_tete_id', 'perimetre_id', '-date_calcul')):
        key = (e.ouvrage_tete_type, e.ouvrage_tete_id, e.perimetre_id)
        if key in seen:
            continue
        seen.add(key)
        perim = (e.perimetre.ksar_village if e.perimetre else None) or f'#{e.perimetre_id}'
        g = e.efficience_globale
        options.append({
            'pk':         e.pk,
            'label':      f"{e.get_ouvrage_tete_type_display()} #{e.ouvrage_tete_id} — {perim} ({round(g) if g is not None else '—'}%)",
            'ouvrage':    e.get_ouvrage_tete_type_display(),
            'perimetre':  perim,
            'globale':    round(g, 1) if g is not None else None,
            'principale': round(e.efficience_principale, 1) if e.efficience_principale is not None else None,
            'secondaire': round(e.efficience_secondaire, 1) if e.efficience_secondaire is not None else None,
            'tertiaire':  round(e.efficience_tertiaire, 1) if e.efficience_tertiaire is not None else None,
            'nb_p':       e.nb_troncons_principaux,
            'nb_s':       e.nb_troncons_secondaires,
            'nb_t':       e.nb_troncons_tertiaires,
        })
    return JsonResponse({'options': options})


@api_login_required
@require_GET
def seguias_liste(request):
    """GET /carte/api/seguias/liste/ → { options: [{pk, label}] }."""
    from diagnostic.models import Seguias

    options = []
    for s in Seguias.objects.select_related('perimetre').order_by('nom_de_la_seguia'):
        perim = s.perimetre.ksar_village if s.perimetre else ''
        label = f"{s.nom_de_la_seguia} ({s.get_type_deguia_display()})" + (f" — {perim}" if perim else '')
        options.append({'pk': s.pk, 'label': label})
    return JsonResponse({'options': options})


@api_login_required
@require_GET
def seguia_profil(request, pk):
    """
    GET /carte/api/seguia/<pk>/profil/
    → débit amont→aval + pertes (infiltration / vaporisation) par tronçon,
      avec propagation séquentielle du débit (lecture seule, sans écriture).

    Réutilise efficiences.services.efficience_troncon.calculer_efficience_troncon.
    """
    from django.shortcuts import get_object_or_404
    from diagnostic.models import Seguias, TronconSeguia
    from efficiences.services.efficience_troncon import calculer_efficience_troncon

    seg = get_object_or_404(Seguias, pk=pk)
    troncons = list(TronconSeguia.objects.filter(seguia=seg).select_related('seguia'))

    def _num(t):
        c = (t.troncon or '').upper().replace('TR', '')
        return int(c) if c.isdigit() else 999
    troncons.sort(key=_num)

    if not troncons:
        return JsonResponse({'seguia': seg.nom_de_la_seguia, 'troncons': [],
                             'message': "Aucun tronçon pour cette séguia."})

    q_entree = troncons[0].debit or 0.0
    q = q_entree
    rows = []
    for tr in troncons:
        det = calculer_efficience_troncon(tr, perimetre=seg.perimetre, q_amont=q, persister=False)
        rows.append({
            'troncon':            tr.troncon,
            'debit_amont':        round(det['debit_amont'], 4),
            'debit_aval':         round(det['debit_aval'], 4),
            'perte_infiltration': round(det['perte_infiltration_m3s'], 5),
            'perte_vaporisation': round(det['perte_vaporisation_m3s'], 5),
            'efficience':         round(det['efficience_pourcent'], 1),
        })
        q = det['debit_aval']

    e_seguia = (q / q_entree * 100) if q_entree > 0 else 0.0
    return JsonResponse({
        'seguia':            seg.nom_de_la_seguia,
        'type':              seg.get_type_deguia_display(),
        'efficience_seguia': round(max(0.0, min(100.0, e_seguia)), 1),
        'troncons':          rows,
    })


# ── Box « Diagnostic » (Lot 4) ────────────────────────────────────────────────

@api_login_required
@require_GET
def ouvrages_etat_comparaison(request):
    """
    GET /carte/api/ouvrages/etat-comparaison/?couche=<nom>
    → { couche, label, total, etats: [{valeur, label, count}] }

    Répartition des ouvrages par état général (Etat<X>) — pour barres.
    """
    from django.db.models import Count
    from diagnostic.models import ETAT_CONSTRUCTION_DIAG_CHOICES

    couche = request.GET.get('couche', '').strip()
    if couche not in LAYER_REGISTRY:
        return JsonResponse({'erreur': f'Couche inconnue : {couche}'}, status=404)

    meta = LAYER_REGISTRY[couche]
    etat_lookup = meta.get('etat_lookup')
    if not etat_lookup:
        return JsonResponse({'erreur': f'Couche sans état : {couche}'}, status=400)

    app_label, model_name = meta['model'].split('.')
    Model = apps.get_model(app_label, model_name)

    qs = Model.objects.all()
    pks_param = request.GET.get('pks')
    if pks_param:
        try:
            qs = qs.filter(pk__in=[int(p) for p in pks_param.split(',') if p.strip()])
        except ValueError:
            return JsonResponse({'erreur': 'pks invalide'}, status=400)

    counts = {r[etat_lookup]: r['n']
              for r in qs.values(etat_lookup).annotate(n=Count('pk'))}

    etats = [{'valeur': v, 'label': str(lbl), 'count': counts.get(v, 0)}
             for v, lbl in ETAT_CONSTRUCTION_DIAG_CHOICES]
    if counts.get(None):
        etats.append({'valeur': None, 'label': 'Non renseigné', 'count': counts[None]})

    return JsonResponse({
        'couche': couche, 'label': meta['label'],
        'total':  sum(e['count'] for e in etats), 'etats': etats,
    })


# Champ « débit » + unité par couche d'ouvrage (outil Débit mobilisé)
_OUVRAGE_DEBIT = {
    'seuils':           ('debit_mobilise', 'l/s'),
    'barrages':         ('debit_derive',   'l/s'),
    'khettaras':        ('debit',          'l/s'),
    'forages_puits':    ('debit',          'l/s'),
    'prises_locales':   ('debit_derive',   'l/s'),
    'troncons_seguias': ('debit',          'm³/s'),
}


@api_login_required
@require_GET
def ouvrages_debit_points(request):
    """
    GET /carte/api/ouvrages/debit-points/?couche=<nom>&pks=<id,…>
    → FeatureCollection de Points : value = débit de l'ouvrage.

    Thématique « Débit mobilisé » (Module A, cercles proportionnels).
    """
    couche = request.GET.get('couche', '').strip()
    info = _OUVRAGE_DEBIT.get(couche)
    if couche not in LAYER_REGISTRY or not info:
        return JsonResponse({'erreur': f'Couche sans débit exploitable : {couche}'}, status=400)

    champ, unite = info
    meta = LAYER_REGISTRY[couche]
    app_label, model_name = meta['model'].split('.')
    Model = apps.get_model(app_label, model_name)
    geom_field = meta['geom_field']
    label_field = meta.get('label_field') or (meta.get('fields') or ['pk'])[0]

    qs = (Model.objects
          .exclude(**{f'{geom_field}__isnull': True})
          .exclude(**{f'{champ}__isnull': True}))
    pks_param = request.GET.get('pks')
    if pks_param:
        try:
            qs = qs.filter(pk__in=[int(p) for p in pks_param.split(',') if p.strip()])
        except ValueError:
            return JsonResponse({'erreur': 'pks invalide'}, status=400)

    features = []
    for obj in qs:
        geom = getattr(obj, geom_field)
        try:
            pt = geom if geom.geom_type == 'Point' else geom.point_on_surface
        except Exception:
            pt = getattr(geom, 'centroid', None)
        if pt is None:
            continue
        if pt.srid and pt.srid != 4326:
            pt.transform(4326)
        features.append({
            'type':       'Feature',
            'geometry':   {'type': 'Point', 'coordinates': [pt.x, pt.y]},
            'properties': {
                'pk':    obj.pk,
                'nom':   str(getattr(obj, label_field, None) or f'#{obj.pk}'),
                'value': getattr(obj, champ),
            },
        })

    return JsonResponse({
        'type':  'FeatureCollection',
        'couche': couche, 'champ': champ, 'unite': unite,
        'count': len(features), 'features': features,
    })


# ── §8 Analyse — Bassin versant helpers (Seuil / Prise / Barrage) ─────────────

def _bv_geojson_feature(bv):
    """Serialise un BassinVersant en Feature GeoJSON."""
    geom = json.loads(bv.geometrie.geojson) if bv.geometrie else None
    return {
        'type':       'Feature',
        'geometry':   geom,
        'properties': {
            'pk':       bv.pk,
            'nom':      bv.nom,
            'surface':  bv.surface,
            'z_min':    bv.z_min,
            'z_max':    bv.z_max,
            'thalweg':  bv.thalweg,
        },
    }


def _bv_apport_data(bv):
    """Paramètres hydrologiques d'un BV + résultats du dernier calcul si disponible."""
    from analyse_hydrologique.models import ResultatAnalyseHydrologique

    res = (
        ResultatAnalyseHydrologique.objects
        .filter(bassin_versant=bv, statut='valide')
        .order_by('-date_analyse')
        .first()
    ) or (
        ResultatAnalyseHydrologique.objects
        .filter(bassin_versant=bv)
        .order_by('-date_analyse')
        .first()
    )

    data = {
        'pk':          bv.pk,
        'nom':         bv.nom,
        'surface_km2': bv.surface,
        'thalweg_km':  bv.thalweg,
        'z_min':       bv.z_min,
        'z_max':       bv.z_max,
        'url_calcul':  f'/hydrologie/bv/{bv.pk}/',
    }
    if res:
        data.update({
            'methode':      res.methode,
            'qcrue_t10':    res.qcrue_t10,
            'qcrue_t20':    res.qcrue_t20,
            'qcrue_t50':    res.qcrue_t50,
            'qcrue_t100':   res.qcrue_t100,
            'tc_h':         res.temps_concentration,
            'date_analyse': str(res.date_analyse.date()) if res.date_analyse else None,
        })
    return data


@api_login_required
@require_GET
def seuil_bv_geojson(request, pk):
    """GET /carte/api/seuil/<pk>/bv/ — GeoJSON du BV lié au seuil."""
    from django.shortcuts import get_object_or_404
    from diagnostic.models import Seuil

    s = get_object_or_404(Seuil.objects.select_related('bassin_versant'), pk=pk)
    if not s.bassin_versant:
        return JsonResponse({'erreur': 'Aucun bassin versant lié à ce seuil'}, status=404)
    return JsonResponse(_bv_geojson_feature(s.bassin_versant))


@api_login_required
@require_GET
def seuil_bv_apport(request, pk):
    """GET /carte/api/seuil/<pk>/bv-apport/ — Paramètres hydrologiques du BV du seuil."""
    from django.shortcuts import get_object_or_404
    from diagnostic.models import Seuil

    s = get_object_or_404(Seuil.objects.select_related('bassin_versant'), pk=pk)
    if not s.bassin_versant:
        return JsonResponse({'erreur': 'Aucun bassin versant lié à ce seuil'}, status=404)
    return JsonResponse(_bv_apport_data(s.bassin_versant))


@api_login_required
@require_GET
def prise_bv_geojson(request, pk):
    """GET /carte/api/prise/<pk>/bv/ — GeoJSON du BV lié à la prise locale."""
    from django.shortcuts import get_object_or_404
    from diagnostic.models import PriseLocale

    pl = get_object_or_404(PriseLocale.objects.select_related('bassin_versant'), pk=pk)
    if not pl.bassin_versant:
        return JsonResponse({'erreur': 'Aucun bassin versant lié à cette prise locale'}, status=404)
    return JsonResponse(_bv_geojson_feature(pl.bassin_versant))


@api_login_required
@require_GET
def prise_bv_apport(request, pk):
    """GET /carte/api/prise/<pk>/bv-apport/ — Paramètres hydrologiques du BV de la prise."""
    from django.shortcuts import get_object_or_404
    from diagnostic.models import PriseLocale

    pl = get_object_or_404(PriseLocale.objects.select_related('bassin_versant'), pk=pk)
    if not pl.bassin_versant:
        return JsonResponse({'erreur': 'Aucun bassin versant lié à cette prise locale'}, status=404)
    return JsonResponse(_bv_apport_data(pl.bassin_versant))


@api_login_required
@require_GET
def barrage_bv_geojson(request, pk):
    """GET /carte/api/barrage/<pk>/bv/ — GeoJSON du BV lié au barrage."""
    from django.shortcuts import get_object_or_404
    from diagnostic.models import BarrageRetenue

    b = get_object_or_404(BarrageRetenue.objects.select_related('bassin_versant'), pk=pk)
    if not b.bassin_versant:
        return JsonResponse({'erreur': 'Aucun bassin versant lié à ce barrage'}, status=404)
    return JsonResponse(_bv_geojson_feature(b.bassin_versant))


@api_login_required
@require_GET
def barrage_bv_apport(request, pk):
    """
    GET /carte/api/barrage/<pk>/bv-apport/
    → { capacite_retenue, volume_attribue_irrigation,
        apports_mensuels_humide/normale/seche (si BilanOuvrageAssocie),
        bv: { paramètres hydrologiques } }
    """
    from django.shortcuts import get_object_or_404
    from diagnostic.models import BarrageRetenue

    b = get_object_or_404(BarrageRetenue.objects.select_related('bassin_versant'), pk=pk)

    data = {
        'pk':                         b.pk,
        'nom':                        b.nom,
        'capacite_retenue':           b.capacite_retenue,
        'volume_attribue_irrigation': b.volume_attribue_irrigation,
    }

    # Apports mensuels depuis BilanOuvrageAssocie
    from Besions_Ressources.models import BilanOuvrageAssocie
    bilan_ouvrage = (
        BilanOuvrageAssocie.objects
        .filter(barrage=b, type_ouvrage='barrage')
        .order_by('-id')
        .first()
    )
    if bilan_ouvrage:
        data.update({
            'apports_mensuels_normale': bilan_ouvrage.apports_mensuels_normale,
            'apports_mensuels_humide':  bilan_ouvrage.apports_mensuels_humide,
            'apports_mensuels_seche':   bilan_ouvrage.apports_mensuels_seche,
        })

    if b.bassin_versant:
        data['bv'] = _bv_apport_data(b.bassin_versant)

    return JsonResponse(data)


# ── §R-IP — Indice de priorité d'intervention ─────────────────────────────────

_IP_CLASS_COLORS = {
    1: '#c0392b',   # 80-100% — Intervention urgente
    2: '#e67e22',   # 60-80%  — Priorité haute
    3: '#f1c40f',   # 40-60%  — Priorité modérée
    4: '#2ecc71',   # 20-40%  — À surveiller
    5: '#27ae60',   # 0-20%   — Bon état
}
_IP_CLASS_LABELS = {
    1: 'Intervention urgente',
    2: 'Priorité haute',
    3: 'Priorité modérée',
    4: 'À surveiller',
    5: 'Bon état',
}


@api_login_required
@require_POST
def indice_priorite(request):
    """
    POST { couche, coefficients: {champ: float 0-5, ...} }
    → { scores, classes, max_possible, nb_entites, paint_expression, class_colors, class_labels }

    Calcule pour toutes les entités de la couche le score pondéré normalisé (0-100 %)
    puis classe en 5 niveaux selon les seuils fixes du CDC §R-IP.
    Les notes sont lues dans Etat<X> via la relation `diagnostic_etat`.
    """
    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'erreur': 'Corps JSON invalide'}, status=400)

    couche       = body.get('couche', '').strip()
    coefficients = body.get('coefficients')
    pks          = body.get('pks') or None   # restreint le scoring à une sélection carte

    if couche not in LAYER_REGISTRY:
        return JsonResponse({'erreur': f'Couche inconnue : {couche}'}, status=404)

    meta           = LAYER_REGISTRY[couche]
    scoring_champs = meta.get('scoring_champs')
    join_etat      = meta.get('join_etat')

    if not scoring_champs or not join_etat:
        return JsonResponse({'erreur': f'Couche sans critères de scoring : {couche}'}, status=400)

    if not isinstance(coefficients, dict) or not coefficients:
        return JsonResponse({'erreur': 'coefficients doit être un objet non vide'}, status=400)

    try:
        coefficients = {k: float(v) for k, v in coefficients.items() if k in scoring_champs}
    except (TypeError, ValueError):
        return JsonResponse({'erreur': 'Valeurs de coefficients invalides'}, status=400)

    if not coefficients:
        return JsonResponse({'erreur': 'Aucun champ de scoring valide dans les coefficients'}, status=400)

    score_max = sum(5 * c for c in coefficients.values())
    if score_max == 0:
        return JsonResponse({'erreur': 'Tous les coefficients sont à zéro'}, status=400)

    app_label, model_name = meta['model'].split('.')
    Model = apps.get_model(app_label, model_name)

    scores  = {}
    classes = {}

    qs = Model.objects.select_related(join_etat).all()
    if pks:
        try:
            qs = qs.filter(pk__in=[int(p) for p in pks])
        except (TypeError, ValueError):
            return JsonResponse({'erreur': 'pks invalide'}, status=400)

    for obj in qs:
        etat = getattr(obj, join_etat, None)
        if not etat:
            classes[str(obj.pk)] = None
            continue

        score_brut = 0.0
        n_valid    = 0
        for champ, coeff in coefficients.items():
            note = getattr(etat, champ, None)
            if note is not None:
                score_brut += note * coeff
                n_valid    += 1

        if n_valid == 0:
            classes[str(obj.pk)] = None
            continue

        score_norm = score_brut / score_max * 100
        scores[str(obj.pk)] = round(score_norm, 1)

        if   score_norm >= 80: cls = 1
        elif score_norm >= 60: cls = 2
        elif score_norm >= 40: cls = 3
        elif score_norm >= 20: cls = 4
        else:                  cls = 5
        classes[str(obj.pk)] = cls

    # Expression MapLibre : match sur l'id de l'entité → couleur classe
    paint_args = []
    for pk_str, cls in classes.items():
        if cls is not None:
            paint_args += [int(pk_str), _IP_CLASS_COLORS[cls]]
    paint_expression = ['match', ['get', 'id'], *paint_args, '#95a5a6']

    return JsonResponse({
        'scores':           scores,
        'classes':          classes,
        'max_possible':     score_max,
        'nb_entites':       len(scores) + sum(1 for c in classes.values() if c is None),
        'paint_expression': paint_expression,
        'class_colors':     _IP_CLASS_COLORS,
        'class_labels':     _IP_CLASS_LABELS,
    })


# ── Masque hiérarchique (T4 / T5) ─────────────────────────────────────────────

@api_login_required
@require_GET
def masque_enfants(request, couche_parente, pk, couche_enfant):
    """
    GET /carte/api/masque/<couche_parente>/<pk>/<couche_enfant>/
    Retourne les PKs des entités de couche_enfant liées à l'entité pk de couche_parente.

    Hiérarchie 1 : Province → Commune → Périmètre → Ouvrages  (FK en base)
    Hiérarchie 2 : BassinVersant → Réseau / Ouvrages / Stations (FK + intersection)
    """
    from django.db.models import Q
    from carte.models import Province, Commune
    from analyse_hydrologique.models import (
        BassinVersant, ReseauHydrographique,
        StationPluviometrique, StationHydrometrique,
    )
    from diagnostic.models import (
        Perimetre, Seuil, MurProtection, BarrageRetenue,
        Khettara, ForagePuits, PriseLocale, TronconSeguia,
    )
    from Besions_Ressources.models import StationClimatique

    def _pks(qs):
        return list(qs.values_list('id', flat=True))

    # ── Hiérarchie 1 : Province ──────────────────────────────────────────────
    if couche_parente == 'provinces':
        queries = {
            'communes':         lambda: Commune.objects.filter(province_id=pk),
            'perimetres':       lambda: Perimetre.objects.filter(commune__province_id=pk),
            'seuils':           lambda: Seuil.objects.filter(perimetre__commune__province_id=pk),
            'murs_protection':  lambda: MurProtection.objects.filter(perimetre__commune__province_id=pk),
            'troncons_seguias': lambda: TronconSeguia.objects.filter(seguia__perimetre__commune__province_id=pk),
            'barrages':         lambda: BarrageRetenue.objects.filter(perimetre__commune__province_id=pk),
            'khettaras':        lambda: Khettara.objects.filter(perimetre__commune__province_id=pk),
            'forages_puits':    lambda: ForagePuits.objects.filter(perimetre__commune__province_id=pk),
            'prises_locales':   lambda: PriseLocale.objects.filter(perimetre__commune__province_id=pk),
        }

    # ── Hiérarchie 1 : Commune ───────────────────────────────────────────────
    elif couche_parente == 'communes':
        queries = {
            'perimetres':       lambda: Perimetre.objects.filter(commune_id=pk),
            'seuils':           lambda: Seuil.objects.filter(perimetre__commune_id=pk),
            'murs_protection':  lambda: MurProtection.objects.filter(perimetre__commune_id=pk),
            'troncons_seguias': lambda: TronconSeguia.objects.filter(seguia__perimetre__commune_id=pk),
            'barrages':         lambda: BarrageRetenue.objects.filter(perimetre__commune_id=pk),
            'khettaras':        lambda: Khettara.objects.filter(perimetre__commune_id=pk),
            'forages_puits':    lambda: ForagePuits.objects.filter(perimetre__commune_id=pk),
            'prises_locales':   lambda: PriseLocale.objects.filter(perimetre__commune_id=pk),
        }

    # ── Hiérarchie 1 : Périmètre ─────────────────────────────────────────────
    elif couche_parente == 'perimetres':
        queries = {
            'seuils':           lambda: Seuil.objects.filter(perimetre_id=pk),
            'murs_protection':  lambda: MurProtection.objects.filter(perimetre_id=pk),
            'troncons_seguias': lambda: TronconSeguia.objects.filter(seguia__perimetre_id=pk),
            'barrages':         lambda: BarrageRetenue.objects.filter(perimetre_id=pk),
            'khettaras':        lambda: Khettara.objects.filter(perimetre_id=pk),
            'forages_puits':    lambda: ForagePuits.objects.filter(perimetre_id=pk),
            'prises_locales':   lambda: PriseLocale.objects.filter(perimetre_id=pk),
        }

    # ── Hiérarchie 2 : Bassin Versant ─────────────────────────────────────────
    elif couche_parente == 'bassins_versants':
        try:
            bv = BassinVersant.objects.get(pk=pk)
        except BassinVersant.DoesNotExist:
            return JsonResponse({"erreur": f"BassinVersant {pk} introuvable"}, status=404)

        bv_geom = bv.geometrie
        queries = {
            'reseau_hydrographique': lambda: ReseauHydrographique.objects.filter(bassin_versant_id=pk),
            'seuils':         lambda: Seuil.objects.exclude(geometrie__isnull=True).filter(
                                  Q(bassin_versant_id=pk) | Q(geometrie__intersects=bv_geom)),
            'barrages':       lambda: BarrageRetenue.objects.exclude(geometrie__isnull=True).filter(
                                  Q(bassin_versant_id=pk) | Q(geometrie__intersects=bv_geom)),
            'prises_locales': lambda: PriseLocale.objects.exclude(geometrie__isnull=True).filter(
                                  Q(bassin_versant_id=pk) | Q(geometrie__intersects=bv_geom)),
            'stations_hydro': lambda: StationHydrometrique.objects.exclude(geometrie__isnull=True).filter(
                                  geometrie__intersects=bv_geom),
            'stations_clim':  lambda: StationClimatique.objects.exclude(geometrie__isnull=True).filter(
                                  geometrie__intersects=bv_geom),
            'stations_pluvio': lambda: StationPluviometrique.objects.exclude(geom_point__isnull=True).filter(
                                  geom_point__intersects=bv_geom),
        }

    else:
        return JsonResponse({"erreur": f"Couche parente non supportée : {couche_parente}"}, status=400)

    if couche_enfant not in queries:
        return JsonResponse({"erreur": f"Couche enfant non supportée : {couche_enfant} pour {couche_parente}"}, status=400)

    try:
        pks = _pks(queries[couche_enfant]())
    except Exception as exc:
        return JsonResponse({"erreur": str(exc)}, status=500)

    return JsonResponse({"pks": pks, "count": len(pks)})
