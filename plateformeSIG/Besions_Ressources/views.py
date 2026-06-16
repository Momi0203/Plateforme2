import json
from datetime import datetime

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
from django.contrib.gis.geos import GEOSGeometry
from django.db.models import Prefetch, Q

from diagnostic.models import (
    Perimetre, Seguias, TronconSeguia, Seuil, PriseLocale, BarrageRetenue, Khettara, ForagePuits,
)
from analyse_hydrologique.models import StationHydrometrique
from .models import (
    StationClimatique, Kc_Kr_culture, BilanBesoinRessources, BilanOuvrageAssocie,
    AutreRessource,
)
from .forms import (
    StationClimatique12Form, KcKrCultureForm, BilanBaseForm, BilanOuvrageFormSet,
    AutreRessourceFormSet,
)

# Alias rétrocompat (nom historique du modèle / formulaire)
CulturePerimetre = Kc_Kr_culture
CulturePerimetreForm = KcKrCultureForm
from . import calculs

# Mapping type_ouvrage → modèle Django
OUVRAGE_MODEL_MAP = {
    'seuil':        Seuil,
    'prise_locale': PriseLocale,
    'barrage':      BarrageRetenue,
    'khettara':     Khettara,
    'forage':       ForagePuits,
}


def _find_bv_for_ouvrage(ouvrage):
    """Trouve le bassin versant d'un ouvrage en 2 étapes :

    1. FK directe `ouvrage.bassin_versant` si renseignée.
    2. Sinon, recherche un BV dont `ouvrage_en_tete` correspond au nom de l'ouvrage
       (comparaison insensible à la casse).
    Retourne `None` si rien trouvé.
    """
    from analyse_hydrologique.models import BassinVersant
    bv = getattr(ouvrage, 'bassin_versant', None)
    if bv:
        return bv
    nom = (
        getattr(ouvrage, 'nom_du_seuil', None)
        or getattr(ouvrage, 'nom', None)
    )
    if not nom:
        return None
    return BassinVersant.objects.filter(ouvrage_en_tete__iexact=nom).first()


def _set_station_geom(instance, request):
    wkt = request.POST.get('geometrie', '').strip()
    if not wkt:
        return
    try:
        instance.geometrie = GEOSGeometry(wkt, srid=4326)
    except Exception:
        instance.geometrie = None


MOIS_LABELS = ["Sep", "Oct", "Nov", "Déc", "Jan", "Fév", "Mar", "Avr", "Mai", "Jui", "Jul", "Aoû"]


# ─── Accueil ──────────────────────────────────────────────────────────────────

@login_required
def bilan_home(request):
    bilans = BilanBesoinRessources.objects.select_related('perimetre', 'station_climatique').all()
    perimetres = Perimetre.objects.prefetch_related(
        Prefetch('bilans_ressources', queryset=BilanBesoinRessources.objects.order_by('-created_at').select_related('station_climatique'))
    ).order_by('ksar_village')
    return render(request, 'besions_ressources/home.html', {
        'bilans': bilans,
        'perimetres': perimetres,
        'nb_perimetres': perimetres.count(),
        'nb_stations_clim': StationClimatique.objects.count(),
        'nb_stations_hydro': StationHydrometrique.objects.count(),
        'nb_bilans': bilans.count(),
    })


@login_required
def coeff_culture_home(request):
    """Page d'accueil pour la gestion des coefficients culturaux Kc/Kr"""
    cultures = Kc_Kr_culture.objects.all().order_by('nom')
    return render(request, 'besions_ressources/home_coeff_culture.html', {
        'cultures': cultures,
        'mois_labels': MOIS_LABELS,
    })


# ─── Bilan CRUD ───────────────────────────────────────────────────────────────

def _appliquer_defauts_canal(bilan):
    """Applique les défauts hydrauliques si non renseignés par l'utilisateur.

    Manning : pente=0.0001 et coeff=0.015 si vides. Forme + dimensions canal
    auto-remplies depuis la séguia principale du périmètre si vides :
      - trapezoidale / rectangulaire → b, y, z
      - circulaire                   → diametre, y
    """
    if not bilan.canal_pente:
        bilan.canal_pente = 0.0001
    if not bilan.canal_manning_n:
        bilan.canal_manning_n = 0.015
    # On considère que les dimensions n'ont pas été renseignées si ni b ni
    # diametre ne sont fournis (l'un OU l'autre suffit selon la forme).
    if not bilan.canal_b and not bilan.canal_diametre:
        tr = TronconSeguia.objects.filter(
            seguia__perimetre=bilan.perimetre, seguia__type_deguia='principale'
        ).select_related('seguia').first()
        if tr:
            forme = tr.forme or 'trapezoidale'
            bilan.canal_forme = forme
            bilan.canal_y = tr.hauteur_eau
            if forme == 'circulaire':
                bilan.canal_diametre = tr.diametre
                bilan.canal_b = None
                bilan.canal_z = None
            else:
                bilan.canal_b = tr.largeur_meroire
                bilan.canal_z = tr.fruit_de_berge if forme == 'trapezoidale' else 0
                bilan.canal_diametre = None


def _maj_snapshots_ouvrages(bilan):
    """Recalcule le snapshot (BV, Tc, débit tronçon) pour chaque ouvrage associé,
    puis ré-aligne les paramètres du bilan (BV, Tc, débits, sup_jaugée) sur le
    Tc *le plus récent* (ouvrage le plus récemment créé).

    Le BV de chaque ouvrage est résolu dans l'ordre :
      1. choix explicite de l'utilisateur (`oa.bassin_versant` posté par le form)
      2. FK directe de l'ouvrage source (`seuil.bassin_versant`)
      3. correspondance par nom via `BassinVersant.ouvrage_en_tete`
    """
    ouvrages_seuil_prise = list(
        bilan.ouvrages_associes
            .filter(type_ouvrage__in=['seuil', 'prise_locale'])
            .order_by('-created_at', '-id')
    )
    for oa in bilan.ouvrages_associes.all():
        ouv = oa.ouvrage
        if ouv is None:
            continue
        if oa.type_ouvrage in ('seuil', 'prise_locale'):
            # 1) BV : respecter le choix utilisateur si fourni, sinon auto-détecter
            if oa.bassin_versant is None:
                oa.bassin_versant = _find_bv_for_ouvrage(ouv)
            tc_info = calculs.calculer_tc_pour_bv(oa.bassin_versant)
            if tc_info:
                oa.tc_h = tc_info['tc_h']
                oa.tc_source = tc_info['source']
            else:
                oa.tc_h = None
                oa.tc_source = ''
            oa.debit_troncon_m3s = oa.troncon_amenee.debit if oa.troncon_amenee else None
            # 2e tronçon : seuil uniquement (prise locale → forcé à None)
            if oa.type_ouvrage == 'seuil':
                oa.debit_troncon_2_m3s = (
                    oa.troncon_amenee_2.debit if oa.troncon_amenee_2 else None
                )
            else:
                oa.troncon_amenee_2 = None
                oa.debit_troncon_2_m3s = None
        oa.save()

    # Tc bilan global = celui de l'ouvrage seuil/prise le plus récent
    if ouvrages_seuil_prise:
        recent = ouvrages_seuil_prise[0]
        if recent.bassin_versant:
            bilan.bassin_versant = recent.bassin_versant
            bilan.superficie_bv_jaugee_km2 = recent.bassin_versant.surface
        if recent.tc_h:
            bilan.tc_h = recent.tc_h
        # Débits mensuels depuis la station hydrométrique du bilan (année normale)
        sh = bilan.station_hydrometrique
        if sh and sh.debits_mensuels_annee_normale:
            bilan.debits_mensuels_m3s = list(sh.debits_mensuels_annee_normale)


@login_required
def bilan_creer(request):
    if request.method == 'POST':
        form = BilanBaseForm(request.POST)
        if form.is_valid():
            bilan = form.save(commit=False)
            _appliquer_defauts_canal(bilan)
            bilan.save()

            formset = BilanOuvrageFormSet(request.POST, instance=bilan)
            autres_formset = AutreRessourceFormSet(request.POST, instance=bilan)
            if formset.is_valid() and autres_formset.is_valid():
                formset.save()
                autres_formset.save()
                _maj_snapshots_ouvrages(bilan)
                bilan.save()
                messages.success(request, "Bilan créé. Lancez le calcul depuis le détail.")
                return redirect('besions_ressources:bilan_detail', pk=bilan.pk)
            else:
                # Si l'un des formsets est invalide on annule la création
                bilan.delete()
        else:
            formset = BilanOuvrageFormSet(request.POST)
            autres_formset = AutreRessourceFormSet(request.POST)
    else:
        form = BilanBaseForm()
        formset = BilanOuvrageFormSet()
        autres_formset = AutreRessourceFormSet()
    return render(request, 'besions_ressources/bilan_form.html', {
        'form': form,
        'formset': formset,
        'autres_formset': autres_formset,
        'titre': 'Nouveau bilan',
    })


@login_required
def bilan_modifier(request, pk):
    bilan = get_object_or_404(BilanBesoinRessources, pk=pk)
    if request.method == 'POST':
        form = BilanBaseForm(request.POST, instance=bilan)
        formset = BilanOuvrageFormSet(request.POST, instance=bilan)
        autres_formset = AutreRessourceFormSet(request.POST, instance=bilan)
        if form.is_valid() and formset.is_valid() and autres_formset.is_valid():
            b = form.save(commit=False)
            _appliquer_defauts_canal(b)
            # Réinitialiser les résultats pour forcer un recalcul
            b.resultats_eto = None
            b.resultats_cultures = None
            b.resultats_crue = None
            b.resultats_bilan_normale = None
            b.resultats_bilan_humide = None
            b.date_calcul = None
            b.save()
            formset.save()
            autres_formset.save()
            _maj_snapshots_ouvrages(b)
            b.save()
            messages.success(request, "Bilan mis à jour.")
            return redirect('besions_ressources:bilan_detail', pk=bilan.pk)
    else:
        form = BilanBaseForm(instance=bilan)
        formset = BilanOuvrageFormSet(instance=bilan)
        autres_formset = AutreRessourceFormSet(instance=bilan)
    return render(request, 'besions_ressources/bilan_form.html', {
        'form': form,
        'formset': formset,
        'autres_formset': autres_formset,
        'bilan': bilan,
        'titre': f'Modifier le bilan — {bilan.perimetre}',
    })


@login_required
def bilan_supprimer(request, pk):
    bilan = get_object_or_404(BilanBesoinRessources, pk=pk)
    if request.method == 'POST':
        perimetre_nom = str(bilan.perimetre)
        bilan.delete()
        messages.success(request, f"Bilan de {perimetre_nom} supprimé.")
        return redirect('besions_ressources:home')
    return render(request, 'besions_ressources/bilan_confirm_delete.html', {'bilan': bilan})


@login_required
@require_POST
def valider_bilan(request, pk):
    bilan = get_object_or_404(BilanBesoinRessources, pk=pk)
    bilan.est_valide = True
    bilan.save(update_fields=['est_valide'])

    # Écrire les volumes et excédents/déficits calculés dans le périmètre
    perimetre = bilan.perimetre
    update_fields_perimetre = []

    def _net(res):
        """Excédent − Déficit annuel (positif = excédent, négatif = déficit)."""
        if res is None:
            return None
        return (res.get('total_excedent') or 0) - (res.get('total_deficit') or 0)

    for scenario, vol_field, ed_field in [
        (bilan.resultats_bilan_normale, 'volume_annee_normale',  'volume_excedent_deficit_normale'),
        (bilan.resultats_bilan_humide,  'volume_annee_humide',   'volume_excedent_deficit_humide'),
        (bilan.resultats_bilan_seche,   'volume_annee_seche',    'volume_excedent_deficit_seche'),
    ]:
        if scenario is not None:
            setattr(perimetre, vol_field, scenario.get('total_besoins'))
            setattr(perimetre, ed_field,  _net(scenario))
            update_fields_perimetre += [vol_field, ed_field]

    if update_fields_perimetre:
        perimetre.save(update_fields=update_fields_perimetre)

    return JsonResponse({'ok': True})


# ─── Calcul ───────────────────────────────────────────────────────────────────

@login_required
def bilan_calculer(request, pk):
    """Lance le calcul du bilan pour les 3 types d'année (normale/humide/sèche).

    Le calcul utilise :
      - ETo : station climatique du bilan (Hargreaves simplifié)
      - Pluie efficace : précipitations normales/humides de la station ;
        pour année sèche, fallback sur normale (pas de série dédiée pour l'instant)
      - Besoins cultures : Kc/Kr global × surface (assolement) × ETo × pluie eff
      - Apports : calculer_apports_bilan() agrège seuil/prise/khettara/forage/
        barrage + autres ressources avec efficience par ouvrage
    """
    bilan = get_object_or_404(BilanBesoinRessources, pk=pk)
    station = bilan.station_climatique

    if not station:
        messages.error(request, "Aucune station climatique associée au bilan.")
        return redirect('besions_ressources:bilan_detail', pk=pk)

    # Cultures depuis l'assolement du périmètre
    assols = list(bilan.perimetre.assolement.all())
    if not assols:
        messages.error(request, "Aucun assolement défini pour ce périmètre.")
        return redirect('besions_ressources:bilan_detail', pk=pk)
    noms = [a.culture for a in assols]
    kckr_par_nom = {k.nom: k for k in Kc_Kr_culture.objects.filter(nom__in=noms)}
    cultures = []  # liste de tuples (nom_label, kc, kr, surface_ha)
    for a in assols:
        k = kckr_par_nom.get(a.culture)
        if not k:
            continue
        surface = a.surface_ha or (
            (a.pourcentage or 0) / 100.0 * (bilan.perimetre.superficie_irriguee or 0)
        )
        cultures.append((a.get_culture_display(), k.kc, k.kr, surface))
    if not cultures:
        messages.error(request, "Aucun Kc/Kr défini pour les cultures de ce périmètre.")
        return redirect('besions_ressources:bilan_detail', pk=pk)

    try:
        # ── ETo (commun aux 3 types d'année) ──────────────────────────────────
        eto_data = calculs.calculer_eto(
            station.temperatures_moyennes,
            station.taux_insolation,
            station.latitude,
        )

        # ── Pluie efficace par type d'année ───────────────────────────────────
        precip_norm = station.precipitations_normales or [0.0] * 12
        precip_hum = station.precipitations_humides or precip_norm
        precip_sec = precip_norm  # fallback : pas de série "sèche" sur la station
        pluies_eff = {
            'normale': calculs.pluie_efficace(precip_norm),
            'humide':  calculs.pluie_efficace(precip_hum),
            'seche':   calculs.pluie_efficace(precip_sec),
        }

        # ── Besoins par culture (efficience = 1.0 ; elle s'applique côté apport) ─
        besoins_par_type = {'normale': [], 'humide': [], 'seche': []}
        for annee_type, pluie_eff in pluies_eff.items():
            for nom_label, kc, kr, surface in cultures:
                res = calculs.calculer_besoins_culture(
                    nom_label, kc, kr, surface, 1.0,
                    eto_data['eto_mm_j'], pluie_eff,
                )
                besoins_par_type[annee_type].append(res)

        # ── Apports par ouvrage (et autres ressources) par type d'année ──────
        apports_par_type = {
            annee: calculs.calculer_apports_bilan(bilan, annee)
            for annee in calculs.ANNEES_TYPES
        }

        # ── Bilan global pour chaque type d'année ──────────────────────────────
        def _build_bilan(annee_type):
            cult = besoins_par_type[annee_type]
            apports = apports_par_type[annee_type]
            ressources = list(apports['total_m3'])
            besoins = calculs.besoins_globaux_m3(cult)
            bilan_m = [r - b for r, b in zip(ressources, besoins)]
            deficit = [max(0.0, -x) for x in bilan_m]
            excedent = [max(0.0, x) for x in bilan_m]
            return {
                'annee_type': annee_type,
                'mois': calculs.MOIS_SEP_AOU,
                'pluie_efficace': pluies_eff[annee_type],
                'cultures': cult,
                'apports': apports,
                'besoins_m3':   [round(v, 0) for v in besoins],
                'ressources_m3':[round(v, 0) for v in ressources],
                'bilan_m3':     [round(v, 0) for v in bilan_m],
                'deficit_m3':   [round(v, 0) for v in deficit],
                'excedent_m3':  [round(v, 0) for v in excedent],
                'total_besoins':    round(sum(besoins), 0),
                'total_ressources': round(sum(ressources), 0),
                'total_deficit':    round(sum(deficit), 0),
                'total_excedent':   round(sum(excedent), 0),
            }

        bilan_norm  = _build_bilan('normale')
        bilan_hum   = _build_bilan('humide')
        bilan_sec   = _build_bilan('seche')

        # Sauvegarde
        bilan.resultats_eto = eto_data
        bilan.resultats_cultures = {
            'normale': besoins_par_type['normale'],
            'humide':  besoins_par_type['humide'],
            'seche':   besoins_par_type['seche'],
        }
        bilan.resultats_crue = None  # remplacé par le détail des apports par ouvrage
        bilan.resultats_bilan_normale = bilan_norm
        bilan.resultats_bilan_humide  = bilan_hum
        bilan.resultats_bilan_seche   = bilan_sec
        bilan.date_calcul = datetime.now()
        bilan.save()
        messages.success(request, "Calcul effectué avec succès.")

    except Exception as exc:
        messages.error(request, f"Erreur de calcul : {exc}")

    return redirect('besions_ressources:bilan_detail', pk=pk)


# ─── Détail bilan ─────────────────────────────────────────────────────────────

@login_required
def bilan_detail(request, pk):
    bilan = get_object_or_404(
        BilanBesoinRessources.objects.select_related('perimetre', 'station_climatique', 'bassin_versant'),
        pk=pk,
    )
    noms = list(bilan.perimetre.assolement.values_list('culture', flat=True))
    cultures = Kc_Kr_culture.objects.filter(nom__in=noms)

    # IMPORTANT : `.geojson` retourne une chaîne JSON. On la parse en dict
    # pour qu'elle traverse `json_script` correctement (sinon double-encodage
    # et Leaflet reçoit une chaîne au lieu d'un objet GeoJSON).
    geojson_perimetre = None
    _geom = getattr(bilan.perimetre, 'geometrie', None)
    if _geom:
        try:
            geojson_perimetre = json.loads(_geom.geojson)
        except (TypeError, ValueError):
            geojson_perimetre = None

    # Données JSON pour Chart.js — un payload par type d'année
    payload = {
        'eto':       bilan.resultats_eto,
        'normale':   bilan.resultats_bilan_normale,
        'humide':    bilan.resultats_bilan_humide,
        'seche':     bilan.resultats_bilan_seche,
        'mois':      calculs.MOIS_SEP_AOU,
    }

    ctx = {
        'bilan': bilan,
        'cultures': cultures,
        'geojson_perimetre': geojson_perimetre,
        'mois': calculs.MOIS_SEP_AOU,
        'resultats_eto': bilan.resultats_eto,
        'resultats_bilan_normale': bilan.resultats_bilan_normale,
        'resultats_bilan_humide':  bilan.resultats_bilan_humide,
        'resultats_bilan_seche':   bilan.resultats_bilan_seche,
        # `json_script` dans le template sérialise déjà en JSON :
        # passer un dict ici évite le double encodage (string JSON dans JSON).
        'json_payload': payload,
    }
    return render(request, 'besions_ressources/bilan_detail.html', ctx)


# ─── API JSON pour hydrogramme détaillé ───────────────────────────────────────

@login_required
def api_hydrogramme(request, pk):
    bilan = get_object_or_404(BilanBesoinRessources, pk=pk)
    mois_idx = int(request.GET.get('mois', 0))
    if not bilan.resultats_crue or not bilan.debits_mensuels_m3s:
        return JsonResponse({'error': 'Données de crue non disponibles'}, status=400)

    debits_transp = bilan.resultats_crue.get('debits_transposes_m3s', [0.0] * 12)
    Qp = debits_transp[mois_idx] if mois_idx < len(debits_transp) else 0.0
    Qdmax = bilan.resultats_crue.get('qdmax_m3s', 0.0)
    tc = bilan.tc_h or 1.0
    data = calculs.hydrogramme_detail(Qp, tc, Qdmax)
    data['mois'] = calculs.MOIS_SEP_AOU[mois_idx]
    data['Qp'] = Qp
    data['Qdmax'] = Qdmax
    return JsonResponse(data)


# ─── Stations climatiques CRUD ────────────────────────────────────────────────

@login_required
def station_list(request):
    stations = StationClimatique.objects.all()
    return render(request, 'besions_ressources/station_list.html', {'stations': stations})


@login_required
def station_creer(request):
    if request.method == 'POST':
        form = StationClimatique12Form(request.POST)
        if form.is_valid():
            station = form.save(commit=False)
            _set_station_geom(station, request)
            station.save()
            messages.success(request, "Station climatique créée.")
            return redirect('besions_ressources:station_list')
    else:
        form = StationClimatique12Form()
    return render(request, 'besions_ressources/station_form.html', {
        'form': form,
        'titre': 'Nouvelle station climatique',
        'insolation_auto': form.insolation_auto,
        'geojson_station': None,
    })


@login_required
def station_modifier(request, pk):
    station = get_object_or_404(StationClimatique, pk=pk)
    if request.method == 'POST':
        form = StationClimatique12Form(request.POST, instance=station)
        if form.is_valid():
            station_obj = form.save(commit=False)
            _set_station_geom(station_obj, request)
            station_obj.save()
            messages.success(request, "Station mise à jour.")
            return redirect('besions_ressources:station_list')
    else:
        form = StationClimatique12Form(instance=station)
    return render(request, 'besions_ressources/station_form.html', {
        'form': form,
        'station': station,
        'titre': f'Modifier — {station.nom}',
        'insolation_auto': form.insolation_auto,
        'geojson_station': station.geometrie.geojson if station.geometrie else None,
    })


@login_required
def station_insolation_auto(request):
    raw_lat = request.GET.get('latitude', '').strip()
    if not raw_lat:
        return JsonResponse({'error': 'Latitude manquante.'}, status=400)
    try:
        latitude = float(raw_lat.replace(',', '.'))
    except ValueError:
        return JsonResponse({'error': 'Latitude invalide.'}, status=400)
    if latitude < -90 or latitude > 90:
        return JsonResponse({'error': 'Latitude hors intervalle [-90, 90].'}, status=400)
    return JsonResponse({'latitude': latitude, 'insolation': calculs.taux_insolation_par_latitude(latitude)})


@login_required
def station_supprimer(request, pk):
    station = get_object_or_404(StationClimatique, pk=pk)
    if request.method == 'POST':
        station.delete()
        messages.success(request, "Station supprimée.")
        return redirect('besions_ressources:station_list')
    return render(request, 'besions_ressources/station_confirm_delete.html', {'station': station})


# ─── Cultures CRUD ────────────────────────────────────────────────────────────

@login_required
def culture_creer(request):
    """Ajoute Kc/Kr pour une culture (référentiel global, sans périmètre)."""
    if request.method == 'POST':
        form = KcKrCultureForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Culture (Kc/Kr) ajoutée.")
            return redirect(request.GET.get('next', 'besions_ressources:coeff_culture_home'))
    else:
        form = KcKrCultureForm()
    return render(request, 'besions_ressources/culture_form.html', {
        'form': form, 'titre': 'Ajouter une culture (Kc/Kr)',
    })


@login_required
def culture_modifier(request, pk):
    culture = get_object_or_404(Kc_Kr_culture, pk=pk)
    if request.method == 'POST':
        form = KcKrCultureForm(request.POST, instance=culture)
        if form.is_valid():
            form.save()
            messages.success(request, "Culture mise à jour.")
            return redirect(request.GET.get('next', 'besions_ressources:coeff_culture_home'))
    else:
        form = KcKrCultureForm(instance=culture)
    return render(request, 'besions_ressources/culture_form.html', {
        'form': form, 'culture': culture,
        'titre': f'Modifier — {culture.get_nom_display()}',
    })


@login_required
def culture_supprimer(request, pk):
    culture = get_object_or_404(Kc_Kr_culture, pk=pk)
    if request.method == 'POST':
        next_url = request.GET.get('next', 'besions_ressources:coeff_culture_home')
        culture.delete()
        messages.success(request, "Culture supprimée.")
        return redirect(next_url)
    return render(request, 'besions_ressources/culture_confirm_delete.html', {'culture': culture})


# ─── Export Excel ─────────────────────────────────────────────────────────────

@login_required
def bilan_exporter_excel(request, pk):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    bilan = get_object_or_404(BilanBesoinRessources, pk=pk)

    wb = Workbook()
    C_DARK, C_GOLD, C_GREY = "1A1A2E", "F0A500", "F8F4EE"
    thin = Side(style='thin', color="D0C8BC")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
    LEFT   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    F_HDR  = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    F_LBL  = Font(name="Calibri", bold=True, size=10, color=C_DARK)
    F_VAL  = Font(name="Calibri",             size=10, color=C_DARK)
    F_GOLD = Font(name="Calibri", bold=True, size=11, color=C_GOLD)
    F_TTL  = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    FILL_D = PatternFill("solid", fgColor=C_DARK)
    FILL_G = PatternFill("solid", fgColor=C_GREY)
    FILL_N = PatternFill("solid", fgColor="FFF8EE")
    FILL_H = PatternFill("solid", fgColor="EBF5FB")
    FILL_S = PatternFill("solid", fgColor="FDEDEC")
    FILL_R = PatternFill("solid", fgColor="D4EDDA")

    def _ttl(ws, text, ncols, row=1):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        c = ws.cell(row=row, column=1, value=text)
        c.font = F_TTL; c.fill = FILL_D; c.alignment = CENTER
        ws.row_dimensions[row].height = 24

    def _sec(ws, text, ncols, row):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        c = ws.cell(row=row, column=1, value=text)
        c.font = F_HDR; c.fill = FILL_D; c.alignment = CENTER
        ws.row_dimensions[row].height = 18

    def _hdr(ws, row, col, text, fill=None):
        c = ws.cell(row=row, column=col, value=text)
        c.font = F_HDR; c.fill = fill or FILL_D; c.alignment = CENTER; c.border = BORDER

    def _lbl(ws, row, col, text, fill=None):
        c = ws.cell(row=row, column=col, value=text)
        c.font = F_LBL; c.fill = fill or FILL_G; c.alignment = LEFT; c.border = BORDER

    def _val(ws, row, col, v, fill=None, gold=False):
        c = ws.cell(row=row, column=col, value=v)
        c.font = F_GOLD if gold else F_VAL
        c.fill = fill or PatternFill(); c.alignment = CENTER; c.border = BORDER

    bn = bilan.resultats_bilan_normale or {}
    bh = bilan.resultats_bilan_humide  or {}
    bs = bilan.resultats_bilan_seche   or {}
    eto = bilan.resultats_eto          or {}

    # ══════════════════════════════════════════════════════════════════════════
    # Feuille 1 — Synthèse
    # ══════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Synthèse"
    ws1.column_dimensions["A"].width = 38
    ws1.column_dimensions["B"].width = 28
    _ttl(ws1, "RAPPORT BILAN BESOINS-RESSOURCES — HydroPlan SIG", 2)
    _sec(ws1, "Informations générales", 2, 3)
    meta = [
        ("Périmètre irrigué",       str(bilan.perimetre)),
        ("Superficie irriguée",     f"{bilan.perimetre.superficie_irriguee or '—'} ha"),
        ("Station climatique",      str(bilan.station_climatique) if bilan.station_climatique else "—"),
        ("Station hydrométrique",   str(bilan.station_hydrometrique) if bilan.station_hydrometrique else "—"),
        ("Date de calcul",          bilan.date_calcul.strftime("%d/%m/%Y  %H:%M") if bilan.date_calcul else "—"),
        ("Statut",                  "Validé" if bilan.est_valide else ("Calculé" if bilan.est_calcule else "Non calculé")),
    ]
    for ri, (k, v) in enumerate(meta, start=4):
        _lbl(ws1, ri, 1, k); _val(ws1, ri, 2, v)
    if bilan.est_calcule:
        _sec(ws1, "KPI — Année normale", 2, 11)
        kpi = [
            ("Besoin total annuel (m³)",    bn.get('total_besoins')),
            ("Ressources totales (m³)",     bn.get('total_ressources')),
            ("Déficit cumulé (m³)",         bn.get('total_deficit')),
            ("Excédent cumulé (m³)",        bn.get('total_excedent')),
        ]
        for ri, (k, v) in enumerate(kpi, start=12):
            _lbl(ws1, ri, 1, k)
            _val(ws1, ri, 2, round(v, 0) if v is not None else "—", gold=True)

    # ══════════════════════════════════════════════════════════════════════════
    # Feuille 2 — ETo & Données climatiques
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("ETo & Climat")
    ws2.column_dimensions["A"].width = 26
    for _c in ["B","C","D","E","F","G","H","I","J","K","L","M"]:
        ws2.column_dimensions[_c].width = 9
    _ttl(ws2, f"ETo et données climatiques — Station : {bilan.station_climatique or '—'}", 13)
    _hdr(ws2, 3, 1, "Paramètre")
    for ci, m in enumerate(MOIS_LABELS, 2):
        _hdr(ws2, 3, ci, m)
    clim_rows = [
        ("Température moy. (°C)",   eto.get('temperatures', [])),
        ("n/N — Insolation",        eto.get('taux_insolation', [])),
        ("ETo journalière (mm/j)",  eto.get('eto_mm_j', [])),
        ("ETo mensuelle (mm/mois)", eto.get('eto_mm_mois', [])),
    ]
    for ri, (lbl_txt, vals) in enumerate(clim_rows, start=4):
        _lbl(ws2, ri, 1, lbl_txt)
        for ci, v in enumerate(vals[:12], start=2):
            _val(ws2, ri, ci, round(v, 2) if v else 0)

    # ══════════════════════════════════════════════════════════════════════════
    # Feuille 3 — Besoins culturaux (année normale)
    # ══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Besoins culturaux")
    ws3.column_dimensions["A"].width = 24
    ws3.column_dimensions["B"].width = 12
    for _c in ["C","D","E","F","G","H","I","J","K","L","M","N"]:
        ws3.column_dimensions[_c].width = 9
    _ttl(ws3, "Besoins en eau par culture — Année normale (m³/mois)", 14)
    _hdr(ws3, 3, 1, "Culture"); _hdr(ws3, 3, 2, "Surface (ha)")
    for ci, m in enumerate(MOIS_LABELS, 3):
        _hdr(ws3, 3, ci, m)
    cultures = bn.get('cultures', [])
    for ri, cult in enumerate(cultures, start=4):
        _lbl(ws3, ri, 1, cult.get('nom', '—'))
        _val(ws3, ri, 2, round(cult.get('surface_ha', 0) or 0, 2))
        for ci, v in enumerate(cult.get('besoins_reel_m3_mois', [])[:12], start=3):
            _val(ws3, ri, ci, round(v, 0) if v else 0)
    if not cultures:
        ws3.cell(row=4, column=1, value="Aucune culture enregistrée.").font = Font(italic=True, size=9, color="888888")

    # ══════════════════════════════════════════════════════════════════════════
    # Feuille 4 — Bilan complet (3 scénarios)
    # ══════════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Bilan 3 scénarios")
    ws4.column_dimensions["A"].width = 26
    ws4.column_dimensions["B"].width = 16
    for _c in ["C","D","E","F","G","H","I","J","K","L","M","N"]:
        ws4.column_dimensions[_c].width = 9
    ws4.column_dimensions["O"].width = 13
    _ttl(ws4, "Bilan Besoins vs Ressources — 3 scénarios (m³/mois)", 15)
    _hdr(ws4, 3, 1, "Scénario"); _hdr(ws4, 3, 2, "Paramètre")
    for ci, m in enumerate(MOIS_LABELS, 3):
        _hdr(ws4, 3, ci, m)
    _hdr(ws4, 3, 15, "Total (m³)")

    SCEN_BILAN = [
        ("Année normale", bn, FILL_N),
        ("Année humide",  bh, FILL_H),
        ("Année sèche",   bs, FILL_S),
    ]
    BILAN_ROWS = [
        ("Besoins (m³)",    "besoins_m3",    "total_besoins"),
        ("Ressources (m³)", "ressources_m3", "total_ressources"),
        ("Bilan (m³)",      "bilan_m3",      None),
        ("Déficit (m³)",    "deficit_m3",    "total_deficit"),
        ("Excédent (m³)",   "excedent_m3",   "total_excedent"),
    ]
    cur_row = 4
    for scen_lbl, scen_data, scen_fill in SCEN_BILAN:
        if not scen_data:
            continue
        ws4.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row + len(BILAN_ROWS) - 1, end_column=1)
        c = ws4.cell(row=cur_row, column=1, value=scen_lbl)
        c.font = F_HDR; c.fill = FILL_D; c.alignment = CENTER; c.border = BORDER
        for i, (row_lbl, data_key, total_key) in enumerate(BILAN_ROWS):
            ri = cur_row + i
            _lbl(ws4, ri, 2, row_lbl, fill=scen_fill)
            vals_list = scen_data.get(data_key, [])
            for ci, v in enumerate(vals_list[:12], start=3):
                _val(ws4, ri, ci, round(v, 0) if v is not None else 0, fill=scen_fill)
            if total_key and scen_data.get(total_key) is not None:
                _val(ws4, ri, 15, round(scen_data[total_key], 0), fill=scen_fill, gold=True)
        cur_row += len(BILAN_ROWS) + 1

    # ══════════════════════════════════════════════════════════════════════════
    # Feuille 5 — Ressources par ouvrage (année normale)
    # ══════════════════════════════════════════════════════════════════════════
    ws5 = wb.create_sheet("Ressources")
    ws5.column_dimensions["A"].width = 28
    ws5.column_dimensions["B"].width = 16
    for _c in ["C","D","E","F","G","H","I","J","K","L","M","N"]:
        ws5.column_dimensions[_c].width = 9
    ws5.column_dimensions["O"].width = 13
    _ttl(ws5, "Apports par ouvrage / ressource — Année normale (m³/mois)", 15)
    _hdr(ws5, 3, 1, "Source"); _hdr(ws5, 3, 2, "Type")
    for ci, m in enumerate(MOIS_LABELS, 3):
        _hdr(ws5, 3, ci, m)
    _hdr(ws5, 3, 15, "Total (m³)")
    apports_n = (bn.get('apports') or {})
    ri = 4
    for src in apports_n.get('par_ouvrage', []):
        _lbl(ws5, ri, 1, src.get('nom', '—'))
        _val(ws5, ri, 2, src.get('type', '—'))
        tot = 0
        for ci, v in enumerate((src.get('apports_m3') or [])[:12], start=3):
            _val(ws5, ri, ci, round(v, 0) if v else 0)
            tot += v or 0
        _val(ws5, ri, 15, round(tot, 0), gold=True)
        ri += 1
    for src in apports_n.get('par_autre_ressource', []):
        _lbl(ws5, ri, 1, src.get('nom', '—'))
        _val(ws5, ri, 2, "Autre")
        tot = 0
        for ci, v in enumerate((src.get('apports_m3') or [])[:12], start=3):
            _val(ws5, ri, ci, round(v, 0) if v else 0)
            tot += v or 0
        _val(ws5, ri, 15, round(tot, 0), gold=True)
        ri += 1
    if ri == 4:
        ws5.cell(row=4, column=1, value="Aucune ressource associée.").font = Font(italic=True, size=9, color="888888")

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    nom_fichier = f"bilan_{str(bilan.perimetre).replace(' ', '_')}_{bilan.created_at.strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{nom_fichier}"'
    wb.save(response)
    return response


# ─── Export PDF ───────────────────────────────────────────────────────────────

@login_required
def bilan_exporter_pdf(request, pk):
    import io as _io
    from datetime import date as _date
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable

    bilan = get_object_or_404(BilanBesoinRessources, pk=pk)

    buf = _io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=1.8*cm, rightMargin=1.8*cm,
        topMargin=1.8*cm,  bottomMargin=1.8*cm,
        title=f"Bilan B/R — {bilan.perimetre}",
    )
    PW = 16.9 * cm

    DARK  = colors.HexColor("#1A1A2E")
    GOLD  = colors.HexColor("#F0A500")
    GREY  = colors.HexColor("#F8F4EE")
    NORM  = colors.HexColor("#FFF8EE")
    HUM   = colors.HexColor("#EBF5FB")
    SECHE = colors.HexColor("#FDEDEC")
    GRNN  = colors.HexColor("#D4EDDA")
    RED   = colors.HexColor("#C0392B")
    TEAL  = colors.HexColor("#0C7A8A")

    S_WHITE = ParagraphStyle("w", fontSize=14, fontName="Helvetica-Bold", textColor=colors.white, alignment=1)
    S_HEAD  = ParagraphStyle("h", fontSize=11, fontName="Helvetica-Bold", textColor=colors.white, alignment=1)
    S_LABEL = ParagraphStyle("lb",fontSize=8.5, fontName="Helvetica-Bold", textColor=DARK)
    S_VALUE = ParagraphStyle("v", fontSize=8.5, fontName="Helvetica",      textColor=colors.HexColor("#333333"))
    S_GOLD  = ParagraphStyle("g", fontSize=10,  fontName="Helvetica-Bold", textColor=GOLD, alignment=1)
    S_SMALL = ParagraphStyle("sm",fontSize=7.5, fontName="Helvetica",      textColor=colors.grey)
    S_CENT  = ParagraphStyle("c", fontSize=8.5, fontName="Helvetica",      textColor=DARK, alignment=1)
    PAD = [("LEFTPADDING",(0,0),(-1,-1),5), ("RIGHTPADDING",(0,0),(-1,-1),5),
           ("TOPPADDING",(0,0),(-1,-1),4),  ("BOTTOMPADDING",(0,0),(-1,-1),4),
           ("VALIGN",(0,0),(-1,-1),"MIDDLE")]

    def sec(text):
        t = Table([[Paragraph(text, S_HEAD)]], colWidths=[PW])
        t.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),DARK),
                                *PAD,("TOPPADDING",(0,0),(-1,-1),6),("BOTTOMPADDING",(0,0),(-1,-1),6)]))
        return t

    def kv(rows, w1=6*cm):
        data = [[Paragraph(k, S_LABEL), Paragraph(str(v), S_VALUE)] for k, v in rows]
        t = Table(data, colWidths=[w1, PW - w1])
        t.setStyle(TableStyle([*PAD,("BACKGROUND",(0,0),(0,-1),GREY),
                                ("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#E0D0C0"))]))
        return t

    def gh(txt):
        return Paragraph(txt, ParagraphStyle("gh",fontSize=8,fontName="Helvetica-Bold",
               textColor=colors.white,alignment=1))

    fmt = lambda v: f"{int(round(v)):,}".replace(",", " ") if v is not None else "—"
    story = []

    # ── Bandeau titre
    t = Table([[Paragraph("RAPPORT BILAN BESOINS-RESSOURCES", S_WHITE),
                Paragraph(f"Bilan #{pk}", S_SMALL)]],
              colWidths=[PW * 0.78, PW * 0.22])
    t.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),DARK), *PAD,
                            ("TOPPADDING",(0,0),(-1,-1),10), ("BOTTOMPADDING",(0,0),(-1,-1),10),
                            ("ALIGN",(-1,0),(-1,0),"RIGHT"), ("VALIGN",(0,0),(-1,-1),"MIDDLE")]))
    story += [t, Spacer(1, .35*cm)]

    # ── 1. Informations générales
    story += [sec("1. Informations générales"), Spacer(1, .1*cm)]
    story.append(kv([
        ("Périmètre irrigué",      str(bilan.perimetre)),
        ("Superficie irriguée",    f"{bilan.perimetre.superficie_irriguee or '—'} ha"),
        ("Station climatique",     str(bilan.station_climatique) if bilan.station_climatique else "—"),
        ("Station hydrométrique",  str(bilan.station_hydrometrique) if bilan.station_hydrometrique else "—"),
        ("Date de calcul",         bilan.date_calcul.strftime("%d/%m/%Y  %H:%M") if bilan.date_calcul else "Non calculé"),
        ("Statut",                 "Validé" if bilan.est_valide else ("Calculé" if bilan.est_calcule else "Non calculé")),
    ]))

    bn = bilan.resultats_bilan_normale or {}
    bh = bilan.resultats_bilan_humide  or {}
    bs = bilan.resultats_bilan_seche   or {}

    if not bilan.est_calcule:
        story += [Spacer(1, .5*cm),
                  Paragraph("Le bilan n'a pas encore été calculé. Lancez le calcul pour obtenir les résultats.", S_SMALL)]
        doc.build(story)
        buf.seek(0)
        slug = f"bilan_{pk}_{_date.today().strftime('%Y%m%d')}.pdf"
        resp = HttpResponse(buf.read(), content_type="application/pdf")
        resp["Content-Disposition"] = f'attachment; filename="{slug}"'
        return resp

    # ── 2. KPI synthèse (tableau 2×4)
    story += [Spacer(1, .3*cm), sec("2. KPI — Synthèse par scénario"), Spacer(1, .1*cm)]
    kpi_hdr = [gh(h) for h in ["Indicateur", "Année normale", "Année humide", "Année sèche"]]
    kpi_data = [kpi_hdr]
    kpi_rows = [
        ("Besoins totaux (m³)",    "total_besoins",    False),
        ("Ressources totales (m³)","total_ressources", False),
        ("Déficit cumulé (m³)",    "total_deficit",    True),
        ("Excédent cumulé (m³)",   "total_excedent",   False),
    ]
    for lbl_txt, key, is_deficit in kpi_rows:
        row = [Paragraph(lbl_txt, S_LABEL)]
        for scen_data in [bn, bh, bs]:
            v = scen_data.get(key)
            style = ParagraphStyle("kv", fontSize=9, fontName="Helvetica-Bold",
                                   textColor=RED if (is_deficit and v and v > 0) else GOLD, alignment=1)
            row.append(Paragraph(fmt(v), style))
        kpi_data.append(row)
    t_kpi = Table(kpi_data, colWidths=[5.5*cm, 3.7*cm, 3.7*cm, 4*cm])
    t_kpi.setStyle(TableStyle([*PAD,
        ("BACKGROUND",(0,0),(-1,0),DARK),
        ("BACKGROUND",(0,1),(0,-1),GREY),
        ("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#E0D0C0")),
        ("ALIGN",(1,0),(-1,-1),"CENTER")]))
    story.append(t_kpi)

    # ── 3. Tableau bilan mensuel — chaque scénario
    SCEN_DEF = [
        ("3. Bilan mensuel — Année normale", bn, NORM),
        ("4. Bilan mensuel — Année humide",  bh, HUM),
        ("5. Bilan mensuel — Année sèche",   bs, SECHE),
    ]
    for sec_title, scen_data, fill_color in SCEN_DEF:
        if not scen_data:
            continue
        story += [Spacer(1, .3*cm), sec(sec_title), Spacer(1, .1*cm)]
        mois = MOIS_LABELS
        cw_mois = [3.5*cm] + [1.1*cm] * 12 + [1.5*cm]
        hdr_row = [gh("Indicateur")] + [gh(m) for m in mois] + [gh("Total")]
        bilan_rows_def = [
            ("Besoins (m³)",    "besoins_m3",    "total_besoins"),
            ("Ressources (m³)", "ressources_m3", "total_ressources"),
            ("Bilan (m³)",      "bilan_m3",      None),
            ("Déficit (m³)",    "deficit_m3",    "total_deficit"),
            ("Excédent (m³)",   "excedent_m3",   "total_excedent"),
        ]
        tbl_data = [hdr_row]
        for row_lbl, data_key, total_key in bilan_rows_def:
            vals = scen_data.get(data_key) or []
            total = scen_data.get(total_key) if total_key else None
            row = [Paragraph(row_lbl, S_LABEL)]
            for v in (vals[:12] + [0] * max(0, 12 - len(vals))):
                row.append(Paragraph(fmt(v), S_CENT))
            row.append(Paragraph(fmt(total), ParagraphStyle("tot",fontSize=8.5,fontName="Helvetica-Bold",
                                                             textColor=GOLD,alignment=1)))
            tbl_data.append(row)
        t_bilan = Table(tbl_data, colWidths=cw_mois)
        t_bilan.setStyle(TableStyle([*PAD,
            ("BACKGROUND",(0,0),(-1,0),DARK),
            ("BACKGROUND",(0,1),(0,-1),fill_color),
            ("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#E0D0C0")),
            ("ALIGN",(1,0),(-1,-1),"CENTER"),
            ("FONTSIZE",(0,0),(-1,-1),7.5)]))
        story.append(t_bilan)

    # ── Pied de page
    story += [Spacer(1, .4*cm),
              HRFlowable(width="100%", thickness=0.5, color=GOLD),
              Spacer(1, .08*cm),
              Paragraph(f"HydroPlan SIG  ·  Généré le {_date.today().strftime('%d/%m/%Y')}  ·  Bilan #{pk}", S_SMALL)]

    doc.build(story)
    buf.seek(0)
    slug = f"bilan_{pk}_{_date.today().strftime('%Y%m%d')}.pdf"
    resp = HttpResponse(buf.read(), content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="{slug}"'
    return resp


# ─── API JSON — Auto-fill du formulaire de bilan ──────────────────────────────

@login_required
def api_perimetre_info(request, perimetre_id):
    """Renvoie les infos auto-remplies pour Div 2 : surface, assolements
    avec leur Kc/Kr (défini ou non), et tours d'eau."""
    perimetre = get_object_or_404(Perimetre, pk=perimetre_id)
    assolements = list(perimetre.assolement.all())
    kckr_nom = set(Kc_Kr_culture.objects.filter(
        nom__in=[a.culture for a in assolements]
    ).values_list('nom', flat=True))

    superficie = perimetre.superficie_irriguee or 0.0
    cultures = []
    for a in assolements:
        surface = a.surface_ha if a.surface_ha is not None else (
            ((a.pourcentage or 0) / 100.0) * superficie
        )
        cultures.append({
            'culture':       a.culture,
            'culture_label': a.get_culture_display(),
            'pourcentage':   a.pourcentage,
            'surface_ha':    round(surface, 3) if surface else 0.0,
            'kc_kr_defini':  a.culture in kckr_nom,
        })

    tours_eau = [
        {
            'ayant_droit': t.ayant_droit,
            'cycle_jours': t.cycle_jours,
            'duree_heures': t.duree_heures,
        }
        for t in perimetre.tours_eau.all()
    ]

    return JsonResponse({
        'perimetre_id': perimetre.id,
        'nom': str(perimetre),
        'superficie_irriguee': superficie,
        'efficiance_reseau':   perimetre.efficiance_reseau,
        'cultures': cultures,
        'tours_eau': tours_eau,
    })


@login_required
def api_ouvrage_details(request, type_ouvrage, ouvrage_id):
    """Renvoie les paramètres par défaut d'un ouvrage diagnostic
    (efficience réseau notamment) pour pré-remplir le sous-formulaire bilan.

    type_ouvrage ∈ {seuil, prise_locale, barrage, khettara, forage}
    """
    model = OUVRAGE_MODEL_MAP.get(type_ouvrage)
    if model is None:
        return JsonResponse({'error': f"Type d'ouvrage inconnu : {type_ouvrage}"}, status=400)
    ouvrage = get_object_or_404(model, pk=ouvrage_id)
    # Le champ efficience_reseaux a un default 0.75 — on l'expose tel quel
    efficience = getattr(ouvrage, 'efficience_reseaux', None)
    if efficience is None:
        efficience = 0.75
    return JsonResponse({
        'type_ouvrage':       type_ouvrage,
        'id':                 ouvrage.id,
        'efficience_reseau':  efficience,
    })


@login_required
def api_ouvrages_perimetre(request, perimetre_id):
    """Liste tous les ouvrages déjà définis pour ce périmètre, groupés par type."""
    perimetre = get_object_or_404(Perimetre, pk=perimetre_id)
    return JsonResponse({
        'seuil':        [{'id': o.id, 'nom': o.nom_du_seuil} for o in perimetre.seuils.all()],
        'prise_locale': [{'id': o.id, 'nom': o.nom} for o in perimetre.prises_locales.all()],
        'barrage':      [{'id': o.id, 'nom': o.nom} for o in perimetre.barrages_retenue.all()],
        'khettara':     [{'id': o.id, 'nom': o.nom} for o in perimetre.khettaras.all()],
        'forage':       [{'id': o.id, 'nom': o.nom} for o in perimetre.forages_puits.all()],
    })


def _serialize_troncon(troncon):
    return {
        'id': troncon.id,
        'nom': f"{troncon.seguia.nom_de_la_seguia} — {troncon.troncon}",
        'type_deguia': troncon.seguia.type_deguia,
        'debit_m3s': troncon.debit,
    }


def _serialize_bv(bv):
    return {'id': bv.id, 'nom': bv.nom, 'surface_km2': bv.surface}


def _build_ouvrage_response(ouvrage, nom, troncons_associes_qs):
    """Construit la réponse JSON commune à `api_seuil_info` et `api_prise_info`.

    Renvoie :
      - `bv` : BV auto-détecté (FK directe ou match par nom), peut être null
      - `bv_source` : 'fk' / 'nom' / 'none'
      - `bvs_disponibles` : liste de tous les BV pour permettre un choix manuel
      - `tc` : info Tc calculée pour le BV détecté
      - `troncons` : tronçons spécifiquement associés à l'ouvrage
      - `troncons_perimetre` : toutes les séguias du périmètre de l'ouvrage (fallback)
    """
    from analyse_hydrologique.models import BassinVersant

    bv = _find_bv_for_ouvrage(ouvrage)
    if getattr(ouvrage, 'bassin_versant', None):
        bv_source = 'fk'
    elif bv is not None:
        bv_source = 'nom'
    else:
        bv_source = 'none'
    tc_info = calculs.calculer_tc_pour_bv(bv) if bv else None

    troncons_perimetre_qs = TronconSeguia.objects.filter(
        seguia__perimetre=ouvrage.perimetre
    ).select_related('seguia').order_by('seguia__nom_de_la_seguia', 'troncon')

    return {
        'nom': nom,
        'bv': _serialize_bv(bv) if bv else None,
        'bv_source': bv_source,
        'bvs_disponibles': [_serialize_bv(b) for b in BassinVersant.objects.all().order_by('nom')],
        'tc': tc_info,
        'troncons': [_serialize_troncon(s) for s in troncons_associes_qs],
        'troncons_perimetre': [_serialize_troncon(s) for s in troncons_perimetre_qs],
    }


@login_required
def api_seuil_info(request, seuil_id):
    """Auto-fill : BV, Tc, tronçons d'amenée disponibles pour un seuil."""
    seuil = get_object_or_404(Seuil, pk=seuil_id)
    troncons = TronconSeguia.objects.filter(
        seguia__ouvrages_tete_associes__FK_seuil=seuil
    ).select_related('seguia').distinct()
    payload = _build_ouvrage_response(seuil, seuil.nom_du_seuil, troncons)
    payload['seuil_id'] = seuil.id
    return JsonResponse(payload)


@login_required
def api_prise_info(request, prise_id):
    """Auto-fill : BV, Tc, tronçons d'amenée disponibles pour une prise locale."""
    prise = get_object_or_404(PriseLocale, pk=prise_id)
    troncons = TronconSeguia.objects.filter(
        seguia__ouvrages_tete_associes__FK_prise_locale=prise
    ).select_related('seguia').distinct()
    payload = _build_ouvrage_response(prise, prise.nom, troncons)
    payload['prise_id'] = prise.id
    return JsonResponse(payload)


@login_required
def api_bv_tc(request, bv_id):
    """Renvoie le Tc calculé pour un BV donné (sans contexte d'ouvrage).

    Utilisé côté front quand l'utilisateur change manuellement le BV
    d'un ouvrage : on recalcule le Tc en conséquence.
    """
    from analyse_hydrologique.models import BassinVersant
    bv = get_object_or_404(BassinVersant, pk=bv_id)
    tc_info = calculs.calculer_tc_pour_bv(bv)
    return JsonResponse({'bv': _serialize_bv(bv), 'tc': tc_info})


@login_required
def api_troncon_info(request, troncon_id):
    """Renvoie le débit d'un tronçon de séguia."""
    troncon = get_object_or_404(TronconSeguia.objects.select_related('seguia'), pk=troncon_id)
    return JsonResponse(_serialize_troncon(troncon))
