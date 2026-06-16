from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q, Count, Sum
from django.http import HttpResponse, JsonResponse

import io
import os
from datetime import date, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import json

from .models import (
    PlanAmenagement, ActionPlan, TYPE_ACTION_CHOICES,
    CalendrierIntervention, TacheIntervention,
    SuiviAvancement, PieceJustificative,
)
from .forms import (
    PlanAmenagementForm, ActionPlanForm,
    CalendrierInterventionForm, TacheFormSet,
    SuiviAvancementForm,
)
from .utils import has_cycle, compute_cpm
from .decorators import require_role, ROLES_PLAN


# ─── Axe 1 — Plans d'aménagement ─────────────────────────────────────────────

@login_required
@require_role(*ROLES_PLAN)
def plan_list(request):
    qs = PlanAmenagement.objects.annotate(
        nb_actions=Count('actions'),
        budget_engage=Sum('actions__budget_prevu'),
    ).order_by('-annee')

    # Filtres GET
    annee = request.GET.get('annee', '')
    statut = request.GET.get('statut', '')
    q = request.GET.get('q', '').strip()

    if annee:
        qs = qs.filter(annee=annee)
    if statut:
        qs = qs.filter(statut=statut)
    if q:
        qs = qs.filter(Q(titre__icontains=q) | Q(description__icontains=q))

    paginator = Paginator(qs, 12)
    page = paginator.get_page(request.GET.get('page'))

    annees_dispo = PlanAmenagement.objects.values_list('annee', flat=True).distinct().order_by('-annee')

    return render(request, 'plan_action/plan_list.html', {
        'page_obj': page,
        'annees_dispo': annees_dispo,
        'filtre_annee': annee,
        'filtre_statut': statut,
        'q': q,
    })


@login_required
@require_role('operateur', 'administrateur')
def plan_create(request):
    if request.method == 'POST':
        form = PlanAmenagementForm(request.POST)
        if form.is_valid():
            plan = form.save(commit=False)
            plan.cree_par = request.user
            plan.save()
            messages.success(request, f"Plan {plan.annee} — « {plan.titre} » créé.")
            return redirect('plan_action:plan_detail', pk=plan.pk)
    else:
        form = PlanAmenagementForm()
    return render(request, 'plan_action/plan_form.html', {'form': form, 'titre': "Créer un plan d'aménagement"})


@login_required
@require_role(*ROLES_PLAN)
def plan_detail(request, pk):
    plan = get_object_or_404(PlanAmenagement, pk=pk)
    actions = plan.actions.select_related('commune', 'perimetre').order_by('priorite', 'type_action')

    # Filtres actions
    type_action = request.GET.get('type_action', '')
    statut_action = request.GET.get('statut_action', '')
    commune_id = request.GET.get('commune', '')
    priorite = request.GET.get('priorite', '')

    if type_action:
        actions = actions.filter(type_action=type_action)
    if statut_action:
        actions = actions.filter(statut=statut_action)
    if commune_id:
        actions = actions.filter(commune_id=commune_id)
    if priorite:
        actions = actions.filter(priorite=priorite)

    # Synthèse budgétaire
    total_actions = plan.actions.count()
    nb_realises = plan.actions.filter(statut='realise').count()
    taux = round(nb_realises * 100 / total_actions) if total_actions else 0
    budget_engage = plan.actions.aggregate(s=Sum('budget_prevu'))['s'] or 0

    # Communes présentes dans le plan (pour filtre)
    communes_plan = plan.actions.select_related('commune').values_list(
        'commune__id', 'commune__nom_fr'
    ).distinct().order_by('commune__nom_fr')

    return render(request, 'plan_action/plan_detail.html', {
        'plan': plan,
        'actions': actions,
        'total_actions': total_actions,
        'nb_realises': nb_realises,
        'taux': taux,
        'budget_engage': budget_engage,
        'type_action_choices': TYPE_ACTION_CHOICES,
        'communes_plan': communes_plan,
        'filtre_type_action': type_action,
        'filtre_statut_action': statut_action,
        'filtre_commune': commune_id,
        'filtre_priorite': priorite,
    })


@login_required
@require_role('operateur', 'administrateur')
def plan_update(request, pk):
    plan = get_object_or_404(PlanAmenagement, pk=pk)
    if request.method == 'POST':
        form = PlanAmenagementForm(request.POST, instance=plan)
        if form.is_valid():
            form.save()
            messages.success(request, "Plan mis à jour.")
            return redirect('plan_action:plan_detail', pk=plan.pk)
    else:
        form = PlanAmenagementForm(instance=plan)
    return render(request, 'plan_action/plan_form.html', {
        'form': form,
        'titre': f"Modifier — Plan {plan.annee}",
        'plan': plan,
    })


@login_required
@require_role('administrateur')
def plan_delete(request, pk):
    plan = get_object_or_404(PlanAmenagement, pk=pk)
    if request.method == 'POST':
        annee = plan.annee
        plan.delete()
        messages.success(request, f"Plan {annee} supprimé.")
        return redirect('plan_action:plan_list')
    return render(request, 'plan_action/plan_confirm_delete.html', {'plan': plan})


# ─── Actions du plan ─────────────────────────────────────────────────────────

@login_required
@require_role('operateur', 'administrateur')
def action_create(request, plan_pk):
    plan = get_object_or_404(PlanAmenagement, pk=plan_pk)
    if request.method == 'POST':
        form = ActionPlanForm(request.POST)
        if form.is_valid():
            action = form.save(commit=False)
            action.plan = plan
            action.save()
            messages.success(request, "Action ajoutée au plan.")
            return redirect('plan_action:plan_detail', pk=plan.pk)
    else:
        form = ActionPlanForm()
    return render(request, 'plan_action/action_form.html', {
        'form': form,
        'plan': plan,
        'titre': "Ajouter une action",
    })


@login_required
@require_role('operateur', 'administrateur')
def action_update(request, pk):
    action = get_object_or_404(ActionPlan, pk=pk)
    if request.method == 'POST':
        form = ActionPlanForm(request.POST, instance=action)
        if form.is_valid():
            form.save()
            messages.success(request, "Action mise à jour.")
            return redirect('plan_action:plan_detail', pk=action.plan.pk)
    else:
        form = ActionPlanForm(instance=action)
    return render(request, 'plan_action/action_form.html', {
        'form': form,
        'plan': action.plan,
        'titre': f"Modifier — {action.get_type_action_display()}",
        'action': action,
    })


@login_required
@require_role('administrateur')
def action_delete(request, pk):
    action = get_object_or_404(ActionPlan, pk=pk)
    plan_pk = action.plan.pk
    if request.method == 'POST':
        action.delete()
        messages.success(request, "Action supprimée.")
        return redirect('plan_action:plan_detail', pk=plan_pk)
    return render(request, 'plan_action/action_confirm_delete.html', {'action': action})


# ─── Axe 1 — Synthèse budgétaire ─────────────────────────────────────────────

@login_required
@require_role(*ROLES_PLAN)
def plan_synthese(request, pk):
    plan = get_object_or_404(PlanAmenagement, pk=pk)
    actions = plan.actions.select_related('commune__province').all()

    # Tableau croisé communes × types d'action (budget prévu)
    communes = sorted(
        {a.commune for a in actions},
        key=lambda c: c.nom_fr,
    )
    types_presents = sorted(
        {a.type_action for a in actions},
        key=lambda t: t,
    )
    type_labels = {code: label for code, label in TYPE_ACTION_CHOICES}

    # Matrice {commune_id: {type_action: budget}}
    matrice = {}
    for commune in communes:
        matrice[commune.pk] = {}
        for ta in types_presents:
            total = sum(
                a.budget_prevu
                for a in actions
                if a.commune_id == commune.pk and a.type_action == ta
            )
            matrice[commune.pk][ta] = total

    # Totaux par ligne (commune) et par colonne (type)
    totaux_commune = {
        c.pk: sum(matrice[c.pk].values()) for c in communes
    }
    totaux_type = {
        ta: sum(matrice[c.pk][ta] for c in communes)
        for ta in types_presents
    }
    total_general = sum(totaux_type.values())

    # Données pour le camembert (Chart.js) — par type d'action
    chart_labels = [type_labels.get(ta, ta) for ta in types_presents]
    chart_data = [float(totaux_type[ta]) for ta in types_presents]
    PALETTE = [
        '#f0a500', '#3498db', '#27ae60', '#e74c3c', '#9b59b6',
        '#1abc9c', '#e67e22', '#2c3e50', '#16a085', '#8e44ad',
        '#d35400', '#2980b9', '#c0392b', '#27ae60', '#f39c12', '#e91e63',
    ]
    chart_colors = PALETTE[:len(types_presents)]

    # Répartition par statut (pour second graphique)
    statuts_data = {
        'programme': actions.filter(statut='programme').count(),
        'en_cours':  actions.filter(statut='en_cours').count(),
        'realise':   actions.filter(statut='realise').count(),
        'annule':    actions.filter(statut='annule').count(),
    }

    # G1 — Répartition budget par commune (bar horizontale, triée budget desc)
    communes_sorted_budget = sorted(communes, key=lambda c: totaux_commune[c.pk], reverse=True)
    g1_labels = [c.nom_fr for c in communes_sorted_budget]
    g1_data   = [float(totaux_commune[c.pk]) for c in communes_sorted_budget]

    # G2 — Répartition budget par province (donut)
    prov_map = {}
    for a in actions:
        pnom = a.commune.province.nom_fr if a.commune.province_id else 'Non définie'
        prov_map[pnom] = float(prov_map.get(pnom, 0)) + float(a.budget_prevu)
    prov_sorted = sorted(prov_map.items(), key=lambda x: x[1], reverse=True)
    g2_labels = [p[0] for p in prov_sorted]
    g2_data   = [p[1] for p in prov_sorted]
    PROV_COLORS = ['#1a1a2e', '#3498db', '#27ae60', '#9b59b6', '#e67e22', '#16a085']
    g2_colors = PROV_COLORS[:len(g2_labels)]

    return render(request, 'plan_action/plan_synthese.html', {
        'plan': plan,
        'communes': communes,
        'types_presents': types_presents,
        'type_labels': type_labels,
        'matrice': matrice,
        'totaux_commune': totaux_commune,
        'totaux_type': totaux_type,
        'total_general': total_general,
        'chart_labels': chart_labels,
        'chart_data': chart_data,
        'chart_colors': chart_colors,
        'statuts_data': statuts_data,
        'g1_labels': g1_labels,
        'g1_data': g1_data,
        'g2_labels': g2_labels,
        'g2_data': g2_data,
        'g2_colors': g2_colors,
    })


@login_required
@require_role(*ROLES_PLAN)
def synthese_data(request, pk):
    """Endpoint JSON — données G1 (commune) et G2 (province) pour un plan."""
    plan = get_object_or_404(PlanAmenagement, pk=pk)
    actions = plan.actions.select_related('commune__province').all()

    commune_map = {}
    prov_map = {}
    for a in actions:
        cnom = a.commune.nom_fr
        pnom = a.commune.province.nom_fr if a.commune.province_id else 'Non définie'
        commune_map[cnom] = float(commune_map.get(cnom, 0)) + float(a.budget_prevu)
        prov_map[pnom]    = float(prov_map.get(pnom, 0))    + float(a.budget_prevu)

    comm_sorted = sorted(commune_map.items(), key=lambda x: x[1], reverse=True)
    prov_sorted = sorted(prov_map.items(),    key=lambda x: x[1], reverse=True)

    return JsonResponse({
        'plan': plan.annee,
        'commune': {
            'labels': [x[0] for x in comm_sorted],
            'data':   [x[1] for x in comm_sorted],
        },
        'province': {
            'labels': [x[0] for x in prov_sorted],
            'data':   [x[1] for x in prov_sorted],
        },
    })


@login_required
@require_role(*ROLES_PLAN)
def synthese_comparaison(request):
    """G3 — Comparaison inter-plans : budget prévu vs réalisé."""
    all_plans = PlanAmenagement.objects.order_by('annee')
    selected_ids = [int(i) for i in request.GET.getlist('plans') if i.isdigit()]

    comp_data = []
    if selected_ids:
        plans_sel = all_plans.filter(pk__in=selected_ids)
        for p in plans_sel:
            acts = p.actions.all()
            prevu   = float(acts.aggregate(s=Sum('budget_prevu'))['s'] or 0)
            realise = float(
                acts.filter(statut='realise').aggregate(s=Sum('budget_prevu'))['s'] or 0
            )
            en_cours = float(
                acts.filter(statut='en_cours').aggregate(s=Sum('budget_prevu'))['s'] or 0
            )
            comp_data.append({
                'annee': str(p.annee),
                'titre': p.titre,
                'prevu': prevu,
                'realise': realise,
                'en_cours': en_cours,
                'taux': round(realise * 100 / prevu, 1) if prevu else 0,
            })

    return render(request, 'plan_action/synthese_comparaison.html', {
        'all_plans': all_plans,
        'selected_ids': selected_ids,
        'comp_data': comp_data,
    })


# ─── Axe 1 — Export Excel ────────────────────────────────────────────────────

# Styles openpyxl réutilisables
_HDR_FILL   = PatternFill('solid', fgColor='1A1A2E')
_HDR_FONT   = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
_ACCENT_FILL = PatternFill('solid', fgColor='F0A500')
_ACCENT_FONT = Font(name='Calibri', bold=True, color='1A1A2E', size=11)
_TOTAL_FILL  = PatternFill('solid', fgColor='FFF3CD')
_TOTAL_FONT  = Font(name='Calibri', bold=True, color='856404', size=11)
_THIN_BORDER = Border(
    left=Side(style='thin', color='E0D0C0'),
    right=Side(style='thin', color='E0D0C0'),
    top=Side(style='thin', color='E0D0C0'),
    bottom=Side(style='thin', color='E0D0C0'),
)
_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
_LEFT   = Alignment(horizontal='left', vertical='center', wrap_text=True)


def _style_header(cell, accent=False):
    cell.fill = _ACCENT_FILL if accent else _HDR_FILL
    cell.font = _ACCENT_FONT if accent else _HDR_FONT
    cell.alignment = _CENTER
    cell.border = _THIN_BORDER


def _style_total(cell):
    cell.fill = _TOTAL_FILL
    cell.font = _TOTAL_FONT
    cell.border = _THIN_BORDER
    cell.alignment = _CENTER


def _border(cell):
    cell.border = _THIN_BORDER
    cell.alignment = _LEFT


@login_required
@require_role(*ROLES_PLAN)
def export_plan_excel(request, pk):
    plan = get_object_or_404(PlanAmenagement, pk=pk)
    actions = plan.actions.select_related('commune', 'perimetre').order_by('commune__nom_fr', 'type_action')
    type_labels = {code: label for code, label in TYPE_ACTION_CHOICES}

    wb = openpyxl.Workbook()

    # ── Feuille 1 : Liste complète des actions ────────────────────────────────
    ws1 = wb.active
    ws1.title = f"Plan {plan.annee}"
    ws1.sheet_view.showGridLines = False

    # Titre fusionné
    ws1.merge_cells('A1:J1')
    title_cell = ws1['A1']
    title_cell.value = f"Plan d'aménagement PMH {plan.annee} — {plan.titre}"
    title_cell.font = Font(name='Calibri', bold=True, size=14, color='1A1A2E')
    title_cell.fill = PatternFill('solid', fgColor='FDF8F0')
    title_cell.alignment = _CENTER
    ws1.row_dimensions[1].height = 28

    # Sous-titre
    ws1.merge_cells('A2:J2')
    sub = ws1['A2']
    sub.value = (
        f"Source : {plan.get_source_financement_display()}   |   "
        f"Budget total : {plan.budget_total:,.0f} MAD   |   "
        f"Statut : {plan.get_statut_display()}"
    )
    sub.font = Font(name='Calibri', italic=True, size=10, color='888888')
    sub.alignment = _CENTER
    ws1.row_dimensions[2].height = 18

    ws1.row_dimensions[3].height = 6  # espaceur

    # En-têtes colonnes
    headers = [
        'N°', 'Commune', 'Périmètre', "Type d'action",
        'Description', 'Budget prévu (MAD)',
        'Superficie (ha)', 'Longueur (ml)',
        'Priorité', 'Statut',
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws1.cell(row=4, column=col, value=h)
        _style_header(cell)
    ws1.row_dimensions[4].height = 22

    # Données
    PRIORITE_LABELS = {1: 'Haute', 2: 'Moyenne', 3: 'Basse'}
    STATUT_LABELS = {
        'programme': 'Programmé', 'en_cours': 'En cours',
        'realise': 'Réalisé', 'annule': 'Annulé',
    }
    for row_idx, action in enumerate(actions, start=5):
        row = [
            row_idx - 4,
            action.commune.nom_fr,
            str(action.perimetre) if action.perimetre else '',
            type_labels.get(action.type_action, action.type_action),
            action.description,
            float(action.budget_prevu),
            float(action.superficie_concernee) if action.superficie_concernee else '',
            float(action.longueur_prevue) if action.longueur_prevue else '',
            PRIORITE_LABELS.get(action.priorite, ''),
            STATUT_LABELS.get(action.statut, action.statut),
        ]
        for col, val in enumerate(row, start=1):
            cell = ws1.cell(row=row_idx, column=col, value=val)
            _border(cell)
            if col == 6 and isinstance(val, float):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
            if row_idx % 2 == 0:
                cell.fill = PatternFill('solid', fgColor='FDF8F0')
        ws1.row_dimensions[row_idx].height = 16

    # Ligne total
    total_row = 5 + actions.count()
    ws1.cell(row=total_row, column=1, value='TOTAL')
    ws1.merge_cells(f'A{total_row}:E{total_row}')
    total_budget = float(sum(a.budget_prevu for a in actions))
    budget_cell = ws1.cell(row=total_row, column=6, value=total_budget)
    budget_cell.number_format = '#,##0.00'
    for col in range(1, 11):
        _style_total(ws1.cell(row=total_row, column=col))
    ws1.row_dimensions[total_row].height = 20

    # Largeurs colonnes feuille 1
    col_widths = [6, 20, 18, 36, 42, 20, 14, 14, 12, 14]
    for col, w in enumerate(col_widths, start=1):
        ws1.column_dimensions[get_column_letter(col)].width = w

    # ── Feuille 2 : Synthèse commune × type d'action ─────────────────────────
    ws2 = wb.create_sheet("Synthèse budgétaire")
    ws2.sheet_view.showGridLines = False

    # Listes
    communes = sorted(
        {a.commune for a in actions},
        key=lambda c: c.nom_fr,
    )
    types_presents = sorted({a.type_action for a in actions})

    # Titre
    ncols = len(types_presents) + 2
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    t = ws2.cell(row=1, column=1, value=f"Synthèse budgétaire par commune × type d'action — Plan PMH {plan.annee}")
    t.font = Font(name='Calibri', bold=True, size=13, color='1A1A2E')
    t.fill = PatternFill('solid', fgColor='FDF8F0')
    t.alignment = _CENTER
    ws2.row_dimensions[1].height = 26

    ws2.row_dimensions[2].height = 6

    # En-têtes : commune + types + total
    ws2.cell(row=3, column=1, value='Commune')
    _style_header(ws2.cell(row=3, column=1))
    for col, ta in enumerate(types_presents, start=2):
        cell = ws2.cell(row=3, column=col, value=type_labels.get(ta, ta))
        _style_header(cell)
    ws2.cell(row=3, column=len(types_presents) + 2, value='TOTAL')
    _style_header(ws2.cell(row=3, column=len(types_presents) + 2), accent=True)
    ws2.row_dimensions[3].height = 40

    # Matrice
    matrice = {}
    for commune in communes:
        matrice[commune.pk] = {}
        for ta in types_presents:
            matrice[commune.pk][ta] = float(sum(
                a.budget_prevu
                for a in actions
                if a.commune_id == commune.pk and a.type_action == ta
            ))

    for row_idx, commune in enumerate(communes, start=4):
        ws2.cell(row=row_idx, column=1, value=commune.nom_fr)
        _border(ws2.cell(row=row_idx, column=1))
        row_total = 0.0
        for col, ta in enumerate(types_presents, start=2):
            val = matrice[commune.pk][ta]
            row_total += val
            cell = ws2.cell(row=row_idx, column=col, value=val if val else '')
            _border(cell)
            if val:
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')
            if row_idx % 2 == 0:
                cell.fill = PatternFill('solid', fgColor='FDF8F0')
        tot_cell = ws2.cell(row=row_idx, column=len(types_presents) + 2, value=row_total)
        tot_cell.number_format = '#,##0'
        _style_total(tot_cell)
        ws2.row_dimensions[row_idx].height = 16

    # Ligne total colonnes
    total_row2 = 4 + len(communes)
    ws2.cell(row=total_row2, column=1, value='TOTAL')
    grand_total = 0.0
    for col, ta in enumerate(types_presents, start=2):
        col_total = sum(matrice[c.pk][ta] for c in communes)
        grand_total += col_total
        cell = ws2.cell(row=total_row2, column=col, value=col_total)
        cell.number_format = '#,##0'
        _style_total(cell)
    gt_cell = ws2.cell(row=total_row2, column=len(types_presents) + 2, value=grand_total)
    gt_cell.number_format = '#,##0'
    _style_total(gt_cell)
    ws2.row_dimensions[total_row2].height = 20

    # Largeurs feuille 2
    ws2.column_dimensions['A'].width = 22
    for col in range(2, len(types_presents) + 3):
        ws2.column_dimensions[get_column_letter(col)].width = 18

    # ── Réponse HTTP ─────────────────────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"Plan_PMH_{plan.annee}_{plan.titre[:30].replace(' ', '_')}.xlsx"
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ─── Axe 2 — Calendrier d'intervention ───────────────────────────────────────

@login_required
@require_role(*ROLES_PLAN)
def calendrier_list(request):
    calendriers = CalendrierIntervention.objects.select_related(
        'action__plan', 'action__commune', 'chef_projet'
    ).order_by('-action__plan__annee', 'action__commune__nom_fr')
    return render(request, 'plan_action/calendrier_list.html', {'calendriers': calendriers})


@login_required
@require_role('operateur', 'administrateur')
def calendrier_form(request, action_pk):
    action = get_object_or_404(ActionPlan, pk=action_pk)
    plan = action.plan

    try:
        calendrier = action.calendrier
        is_new = False
    except CalendrierIntervention.DoesNotExist:
        calendrier = None
        is_new = True

    # Bloc édition si calendrier validé (non-admin)
    if calendrier and calendrier.statut_calendrier == 'valide' and not request.user.is_superuser:
        messages.warning(request, "Ce calendrier est validé — lecture seule.")
        return redirect('plan_action:plan_detail', pk=plan.pk)

    # Queryset de tâches pour le formset (ordre stable pour le mapping indice→tâche)
    taches_qs = calendrier.taches.order_by('pk') if calendrier else TacheIntervention.objects.none()

    # Données d'initialisation des antérieures pour JavaScript (mode édition)
    tache_pk_to_idx = {t.pk: i for i, t in enumerate(taches_qs)}
    anterieures_init = {}
    if calendrier:
        for i, tache in enumerate(taches_qs):
            anterieures_init[i] = [
                tache_pk_to_idx[a.pk]
                for a in tache.taches_anterieures.all()
                if a.pk in tache_pk_to_idx
            ]

    if request.method == 'POST':
        cal_form = CalendrierInterventionForm(request.POST, instance=calendrier)
        formset = TacheFormSet(request.POST, queryset=taches_qs)

        cal_valid = cal_form.is_valid()
        fs_valid = formset.is_valid()

        if cal_valid and fs_valid:
            total_forms = int(request.POST.get('form-TOTAL_FORMS', 0))

            # Extraire les antérieures depuis POST (indices de formulaire)
            anterieures_map = {}  # {idx_successeur: [idx_prédécesseur, ...]}
            active_indices = set()
            for i in range(total_forms):
                if request.POST.get(f'form-{i}-DELETE') == 'on':
                    continue
                active_indices.add(i)
                anterieures_map[i] = [
                    int(v) for v in request.POST.getlist(f'anterieures-{i}')
                    if v.isdigit()
                ]

            # Détection de cycle (avant toute sauvegarde)
            edges = [
                (pred, succ)
                for succ, preds in anterieures_map.items()
                for pred in preds
                if pred in active_indices
            ]
            if has_cycle(edges, active_indices):
                cal_form.add_error(None, "Dépendance cyclique détectée dans les antériorités.")
            else:
                # 1. Sauvegarder le calendrier
                cal = cal_form.save(commit=False)
                cal.action = action
                cal.save()

                # 2. Sauvegarder les tâches (sans M2M)
                taches_objs = formset.save(commit=False)
                for t in taches_objs:
                    t.calendrier = cal
                    t.save()
                for deleted_obj in formset.deleted_objects:
                    deleted_obj.delete()

                # 3. Mapper indice formulaire → instance TacheIntervention sauvegardée
                index_to_tache = {}
                for i in range(total_forms):
                    if i not in active_indices:
                        continue
                    pk_str = request.POST.get(f'form-{i}-id', '').strip()
                    code   = request.POST.get(f'form-{i}-code_tache', '').strip()
                    if pk_str:
                        try:
                            index_to_tache[i] = TacheIntervention.objects.get(pk=int(pk_str), calendrier=cal)
                        except TacheIntervention.DoesNotExist:
                            pass
                    elif code:
                        try:
                            index_to_tache[i] = cal.taches.get(code_tache=code)
                        except TacheIntervention.DoesNotExist:
                            pass

                # 4. Définir les antérieures (M2M)
                for i, tache in index_to_tache.items():
                    ant_taches = [
                        index_to_tache[j]
                        for j in anterieures_map.get(i, [])
                        if j in index_to_tache
                    ]
                    tache.taches_anterieures.set(ant_taches)

                messages.success(request, "Calendrier d'intervention enregistré.")
                return redirect('plan_action:plan_detail', pk=plan.pk)

    else:
        cal_form = CalendrierInterventionForm(instance=calendrier)
        formset = TacheFormSet(queryset=taches_qs)

    return render(request, 'plan_action/calendrier_form.html', {
        'cal_form': cal_form,
        'formset': formset,
        'action': action,
        'plan': plan,
        'calendrier': calendrier,
        'is_new': is_new,
        'anterieures_init_json': json.dumps(anterieures_init),
        'nb_existing': taches_qs.count(),
    })


@login_required
@require_role('administrateur')
def valider_calendrier(request, action_pk):
    action = get_object_or_404(ActionPlan, pk=action_pk)
    calendrier = get_object_or_404(CalendrierIntervention, action=action)

    if request.method == 'POST':
        from django.utils import timezone
        calendrier.statut_calendrier = 'valide'
        calendrier.valide_par = request.user
        calendrier.date_validation = timezone.now()
        calendrier.save()
        messages.success(request, f"Calendrier de « {action} » validé.")

    return redirect('plan_action:plan_detail', pk=action.plan.pk)


# ─── Axe 2 — Gantt (Frappe Gantt) ────────────────────────────────────────────

@login_required
@require_role(*ROLES_PLAN)
def calendrier_gantt(request, action_pk):
    action = get_object_or_404(ActionPlan, pk=action_pk)
    calendrier = get_object_or_404(CalendrierIntervention, action=action)
    plan = action.plan
    nb_taches = calendrier.taches.count()
    return render(request, 'plan_action/calendrier_gantt.html', {
        'action': action,
        'plan': plan,
        'calendrier': calendrier,
        'nb_taches': nb_taches,
    })


@login_required
@require_role(*ROLES_PLAN)
def gantt_data(request, action_pk):
    action = get_object_or_404(ActionPlan, pk=action_pk)
    calendrier = get_object_or_404(CalendrierIntervention, action=action)

    taches = list(
        calendrier.taches
        .prefetch_related('taches_anterieures', 'suivis')
        .order_by('code_tache')
    )

    # Chemin critique (CPM)
    try:
        cpm = compute_cpm(calendrier.taches.all())
    except ValueError:
        cpm = {}

    tasks_data = []
    for tache in taches:
        # Avancement : dernier rapport de suivi ou 100 si terminée
        latest = tache.suivis.order_by('-date_rapport').first()
        if tache.statut_tache == 'terminee':
            progress = 100
        elif latest:
            progress = latest.avancement_pct
        else:
            progress = 0

        # Dépendances FS (code_tache des antérieures)
        deps = ','.join(a.code_tache for a in tache.taches_anterieures.all())

        # Infos CPM
        cpm_info = cpm.get(tache.pk, {})
        is_critical = cpm_info.get('is_critical', False)

        # Détection retard (SC-08) — tâche non terminée en avance réelle < théorique-10%
        today_gantt = date.today()
        av_theo_gantt = _avancement_theorique(tache, today_gantt)
        en_retard = (
            tache.statut_tache != 'terminee'
            and (tache.date_fin_prevue < today_gantt or progress < max(0, av_theo_gantt - 10))
        )

        # Classe CSS pour Frappe Gantt
        custom_class = f'gantt-statut-{tache.statut_tache}'
        if is_critical:
            custom_class += ' gantt-critical'
        if en_retard:
            custom_class += ' gantt-retard'

        tasks_data.append({
            'id': tache.code_tache,
            'name': f'{tache.code_tache} — {tache.nom_tache}',
            'start': tache.date_debut_prevue.strftime('%Y-%m-%d'),
            'end': tache.date_fin_prevue.strftime('%Y-%m-%d'),
            'progress': progress,
            'dependencies': deps,
            'custom_class': custom_class,
            # Données extra pour la popup
            '_statut': tache.statut_tache,
            '_statut_label': tache.get_statut_tache_display(),
            '_responsable': tache.responsable.get_full_name() or tache.responsable.username,
            '_type_suivi': tache.get_type_suivi_display(),
            '_duree': tache.duree_prevue,
            '_is_critical': is_critical,
            '_marge': cpm_info.get('marge', ''),
            '_en_retard': en_retard,
        })

    return JsonResponse({
        'tasks': tasks_data,
        'calendrier': {
            'id': calendrier.pk,
            'date_debut': calendrier.date_debut_prevue.strftime('%Y-%m-%d'),
            'date_fin': calendrier.date_fin_prevue.strftime('%Y-%m-%d'),
            'statut': calendrier.statut_calendrier,
            'nb_taches': len(taches),
        },
    })


# ─── Axe 2 — PERT (vis.js Network + CPM) ─────────────────────────────────────

@login_required
@require_role(*ROLES_PLAN)
def calendrier_pert(request, action_pk):
    action = get_object_or_404(ActionPlan, pk=action_pk)
    calendrier = get_object_or_404(CalendrierIntervention, action=action)
    plan = action.plan
    nb_taches = calendrier.taches.count()
    return render(request, 'plan_action/calendrier_pert.html', {
        'action': action,
        'plan': plan,
        'calendrier': calendrier,
        'nb_taches': nb_taches,
    })


@login_required
@require_role(*ROLES_PLAN)
def pert_data(request, action_pk):
    action = get_object_or_404(ActionPlan, pk=action_pk)
    calendrier = get_object_or_404(CalendrierIntervention, action=action)

    taches = list(
        calendrier.taches
        .prefetch_related('taches_anterieures')
        .order_by('code_tache')
    )

    if not taches:
        return JsonResponse({'nodes': [], 'edges': [], 'project_duration': 0, 'nb_taches': 0, 'nb_critiques': 0})

    try:
        cpm = compute_cpm(calendrier.taches.all())
    except ValueError as exc:
        return JsonResponse({'error': str(exc)}, status=400)

    project_duration = max((v['EF'] for v in cpm.values()), default=0)
    nb_critiques = sum(1 for v in cpm.values() if v['is_critical'])

    # Nœuds
    nodes = []
    for tache in taches:
        info = cpm.get(tache.pk, {})
        is_crit = info.get('is_critical', False)
        nodes.append({
            'id': tache.code_tache,
            'code': tache.code_tache,
            'nom': tache.nom_tache,
            'duree': tache.duree_prevue,
            'statut': tache.statut_tache,
            'statut_label': tache.get_statut_tache_display(),
            'responsable': tache.responsable.get_full_name() or tache.responsable.username,
            'is_critical': is_crit,
            'cpm': {
                'ES': info.get('ES', 0),
                'EF': info.get('EF', 0),
                'LS': info.get('LS', 0),
                'LF': info.get('LF', 0),
                'marge': info.get('marge', 0),
            },
        })

    # Arcs — critique si les deux extrémités sont critiques
    edges = []
    for tache in taches:
        is_crit_succ = cpm.get(tache.pk, {}).get('is_critical', False)
        for ant in tache.taches_anterieures.all():
            is_crit_pred = cpm.get(ant.pk, {}).get('is_critical', False)
            edges.append({
                'from': ant.code_tache,
                'to': tache.code_tache,
                'is_critical': is_crit_pred and is_crit_succ,
            })

    return JsonResponse({
        'nodes': nodes,
        'edges': edges,
        'project_duration': project_duration,
        'nb_taches': len(taches),
        'nb_critiques': nb_critiques,
    })


# ─── Axe 3 — Suivi d'avancement ──────────────────────────────────────────────

_ALLOWED_MIME = frozenset({
    'application/pdf',
    'image/jpeg',
    'image/png',
    'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
})
_MAX_UPLOAD_SIZE = 10 * 1024 * 1024  # 10 Mo


def _avancement_theorique(tache, today):
    """Retourne l'avancement théorique en % basé sur le temps écoulé."""
    total = (tache.date_fin_prevue - tache.date_debut_prevue).days or 1
    elapsed = max(0, min(total, (today - tache.date_debut_prevue).days))
    return round(elapsed * 100 / total)


@login_required
@require_role(*ROLES_PLAN)
def suivi_dashboard(request, action_pk):
    action = get_object_or_404(ActionPlan, pk=action_pk)
    calendrier = get_object_or_404(CalendrierIntervention, action=action)
    plan = action.plan
    today = date.today()

    taches = calendrier.taches.prefetch_related('suivis').order_by('code_tache')

    taches_data = []
    for tache in taches:
        latest = tache.suivis.order_by('-date_rapport').first()
        av_reel = latest.avancement_pct if latest else 0
        av_theo = _avancement_theorique(tache, today)

        en_retard = (
            tache.statut_tache != 'terminee'
            and (
                tache.date_fin_prevue < today
                or av_reel < max(0, av_theo - 10)
            )
        )

        taches_data.append({
            'tache': tache,
            'latest': latest,
            'av_reel': av_reel,
            'av_theo': av_theo,
            'en_retard': en_retard,
            'nb_suivis': tache.suivis.count(),
            'is_responsable': request.user == tache.responsable or request.user.is_superuser,
        })

    # KPI globaux
    nb_total = len(taches_data)
    nb_retard = sum(1 for t in taches_data if t['en_retard'])
    nb_terminees = sum(1 for t in taches_data if t['tache'].statut_tache == 'terminee')
    av_moyen = round(sum(t['av_reel'] for t in taches_data) / nb_total) if nb_total else 0

    return render(request, 'plan_action/suivi_dashboard.html', {
        'action': action,
        'plan': plan,
        'calendrier': calendrier,
        'taches_data': taches_data,
        'today': today,
        'nb_total': nb_total,
        'nb_retard': nb_retard,
        'nb_terminees': nb_terminees,
        'av_moyen': av_moyen,
    })


@login_required
@require_role('operateur', 'editeur', 'administrateur')
def suivi_form(request, tache_pk):
    tache = get_object_or_404(TacheIntervention, pk=tache_pk)
    action = tache.calendrier.action
    plan = action.plan

    # F-A3-02 : contrôle responsable côté serveur (SEC-02)
    if request.user != tache.responsable and not request.user.is_superuser:
        return render(request, '403.html', {'message': "Seul le responsable de la tâche peut saisir un rapport d'avancement."}, status=403)

    if request.method == 'POST':
        form = SuiviAvancementForm(request.POST)
        if form.is_valid():
            suivi = form.save(commit=False)
            suivi.tache = tache
            suivi.auteur = request.user
            suivi.save()

            # F-A3-03 : upload multi-fichiers avec contrôle MIME + taille
            type_piece = request.POST.get('type_piece_global', 'autre')
            upload_errors = []
            for f in request.FILES.getlist('fichiers'):
                if f.content_type not in _ALLOWED_MIME:
                    upload_errors.append(f"Type non autorisé : {f.name}")
                    continue
                if f.size > _MAX_UPLOAD_SIZE:
                    upload_errors.append(f"Trop volumineux (>10 Mo) : {f.name}")
                    continue
                PieceJustificative.objects.create(
                    suivi=suivi,
                    type_piece=type_piece,
                    fichier=f,
                    libelle=f.name,
                    uploade_par=request.user,
                )
            for err in upload_errors:
                messages.warning(request, err)

            messages.success(request, "Rapport d'avancement enregistré.")
            return redirect('plan_action:suivi_historique', tache_pk=tache.pk)
    else:
        form = SuiviAvancementForm(initial={'date_rapport': date.today()})

    return render(request, 'plan_action/suivi_form.html', {
        'form': form,
        'tache': tache,
        'action': action,
        'plan': plan,
    })


@login_required
@require_role(*ROLES_PLAN)
def suivi_historique(request, tache_pk):
    tache = get_object_or_404(TacheIntervention, pk=tache_pk)
    action = tache.calendrier.action
    plan = action.plan

    suivis = tache.suivis.prefetch_related('pieces').order_by('-date_rapport')

    # Photos filtrées pour Lightbox2 (F-A3-04)
    photos = PieceJustificative.objects.filter(
        suivi__tache=tache,
        type_piece='photo_chantier',
    ).select_related('suivi').order_by('-suivi__date_rapport')

    is_responsable = request.user == tache.responsable or request.user.is_superuser

    return render(request, 'plan_action/suivi_historique.html', {
        'tache': tache,
        'action': action,
        'plan': plan,
        'suivis': suivis,
        'photos': photos,
        'is_responsable': is_responsable,
    })


@login_required
@require_role('operateur', 'editeur', 'administrateur')
def piece_delete(request, piece_pk):
    piece = get_object_or_404(PieceJustificative, pk=piece_pk)
    tache = piece.suivi.tache

    if request.user != piece.suivi.auteur and not request.user.is_superuser:
        messages.error(request, "Action non autorisée.")
        return redirect('plan_action:suivi_historique', tache_pk=tache.pk)

    if request.method == 'POST':
        piece.delete()  # signal post_delete supprime le fichier physique
        messages.success(request, "Pièce justificative supprimée.")

    return redirect('plan_action:suivi_historique', tache_pk=tache.pk)


# ─── Axe 3 — Courbe S (JSON) ─────────────────────────────────────────────────

@login_required
@require_role(*ROLES_PLAN)
def courbe_s_data(request, action_pk):
    """Endpoint JSON pour la courbe S Chart.js : avancement prévu vs réel."""
    action = get_object_or_404(ActionPlan, pk=action_pk)
    calendrier = get_object_or_404(CalendrierIntervention, action=action)

    taches = list(calendrier.taches.prefetch_related('suivis').order_by('code_tache'))
    if not taches:
        return JsonResponse({'labels': [], 'planned': [], 'actual': [], 'today': ''})

    today = date.today()
    start = calendrier.date_debut_prevue
    end_cal = calendrier.date_fin_prevue
    end = max(end_cal, today)

    total_days = (end - start).days or 1
    step = max(1, total_days // 28)  # ~28 points

    # Pré-chargement des suivis par tâche
    suivis_by_tache = {}
    for t in taches:
        suivis_by_tache[t.pk] = sorted(t.suivis.all(), key=lambda s: s.date_rapport)

    labels, planned_series, actual_series = [], [], []

    d = start
    while d <= end:
        labels.append(d.strftime('%Y-%m-%d'))

        # Avancement prévu : progression linéaire par tâche
        p_total = 0.0
        for t in taches:
            if d <= t.date_debut_prevue:
                p_total += 0.0
            elif d >= t.date_fin_prevue:
                p_total += 100.0
            else:
                dur = (t.date_fin_prevue - t.date_debut_prevue).days or 1
                p_total += (d - t.date_debut_prevue).days * 100.0 / dur
        planned_series.append(round(p_total / len(taches), 1))

        # Avancement réel : dernier rapport connu à la date d (≤ today)
        if d <= today:
            a_total = 0.0
            for t in taches:
                latest_av = 0
                for s in reversed(suivis_by_tache.get(t.pk, [])):
                    if s.date_rapport <= d:
                        latest_av = s.avancement_pct
                        break
                a_total += latest_av
            actual_series.append(round(a_total / len(taches), 1))
        else:
            actual_series.append(None)

        d += timedelta(days=step)

    # Garantir l'inclusion de la date de fin du calendrier
    if labels[-1] != end_cal.strftime('%Y-%m-%d'):
        labels.append(end_cal.strftime('%Y-%m-%d'))
        planned_series.append(100.0)
        actual_series.append(None)

    return JsonResponse({
        'labels': labels,
        'planned': planned_series,
        'actual': actual_series,
        'today': today.strftime('%Y-%m-%d'),
    })


# ─── Axe 3 — Suivi global (toutes actions avec calendrier) ───────────────────

@login_required
@require_role(*ROLES_PLAN)
def suivi_global(request):
    """Vue synthèse globale : avancement moyen pondéré par tâches, toutes actions."""
    today = date.today()

    plans = PlanAmenagement.objects.prefetch_related(
        'actions__calendrier__taches__suivis'
    ).order_by('-annee')

    plans_data = []
    total_actions = 0
    total_taches = 0
    total_av_weighted = 0.0
    total_weight = 0
    total_retards = 0

    for plan in plans:
        actions_data = []
        for action in plan.actions.order_by('priorite', 'type_action'):
            try:
                cal = action.calendrier
            except CalendrierIntervention.DoesNotExist:
                continue

            taches = list(cal.taches.prefetch_related('suivis').all())
            nb = len(taches)
            if nb == 0:
                continue

            av_sum = 0
            nb_retard = 0
            for t in taches:
                latest = t.suivis.order_by('-date_rapport').first()
                av = latest.avancement_pct if latest else 0
                av_sum += av
                av_theo = _avancement_theorique(t, today)
                if t.statut_tache != 'terminee' and (
                    t.date_fin_prevue < today or av < max(0, av_theo - 10)
                ):
                    nb_retard += 1

            av_moyen = round(av_sum / nb)
            actions_data.append({
                'action': action,
                'cal': cal,
                'av_moyen': av_moyen,
                'nb_taches': nb,
                'nb_retard': nb_retard,
            })
            total_actions += 1
            total_taches += nb
            total_av_weighted += av_moyen * nb
            total_weight += nb
            total_retards += nb_retard

        if actions_data:
            plan_weight = sum(a['nb_taches'] for a in actions_data)
            av_plan = round(
                sum(a['av_moyen'] * a['nb_taches'] for a in actions_data) / plan_weight
            ) if plan_weight else 0
            plans_data.append({
                'plan': plan,
                'actions_data': actions_data,
                'av_plan': av_plan,
                'nb_retard_plan': sum(a['nb_retard'] for a in actions_data),
            })

    av_global = round(total_av_weighted / total_weight) if total_weight else 0

    return render(request, 'plan_action/suivi_global.html', {
        'plans_data': plans_data,
        'total_actions': total_actions,
        'total_taches': total_taches,
        'av_global': av_global,
        'total_retards': total_retards,
        'today': today,
    })
